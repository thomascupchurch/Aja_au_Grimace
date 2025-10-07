import os, tempfile, pytest
from PyQt6.QtWidgets import QApplication
from main import ProjectDataModel, CostEstimatesView

try:
    import openpyxl  # noqa: F401
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

@pytest.fixture(scope="module")
def qapp():
    return QApplication.instance() or QApplication([])

@pytest.mark.skipif(not HAVE_OPENPYXL, reason="openpyxl not installed")
def test_selected_subset_xlsx_export(qapp):
    model = ProjectDataModel()
    model.rows = [
        {"Project Part":"Alpha","Parent":"","Production Cost":100,"Installation Cost":50,"Production Price":200,"Installation Price":80},
        {"Project Part":"Beta","Parent":"","Production Cost":150,"Installation Cost":70,"Production Price":260,"Installation Price":100},
        {"Project Part":"Gamma","Parent":"","Production Cost":120,"Installation Cost":60,"Production Price":210,"Installation Price":90},
    ]
    view = CostEstimatesView(model)
    view.refresh()
    view.table.selectRow(0)
    view.table.selectRow(2)
    view.chk_selected_only.setChecked(True)
    from PyQt6.QtWidgets import QFileDialog
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx'); tmp.close()
    orig = QFileDialog.getSaveFileName
    try:
        QFileDialog.getSaveFileName = staticmethod(lambda *a, **k: (tmp.name, 'Excel Workbook (*.xlsx)'))
        view._export_xlsx()
    finally:
        QFileDialog.getSaveFileName = orig
    import openpyxl
    wb = openpyxl.load_workbook(tmp.name, data_only=True)
    costs = wb['Costs']; meta = wb['_Meta']
    rows_written = costs.max_row - 1
    subset_val = None
    for row in meta.iter_rows(values_only=True):
        if row[0] == 'Subset':
            subset_val = row[1]; break
    wb.close(); os.unlink(tmp.name)
    assert rows_written == 2
    assert subset_val == 'Selected'
