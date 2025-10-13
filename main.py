"""Main application module (PyQt6 native).

All legacy PyQt5 compatibility shim code has been removed as the project now
depends directly on PyQt6. Any remaining lazy imports referencing PyQt5 will
be migrated; if found they should be updated to their PyQt6 counterparts.
"""
from PyQt6.QtWidgets import (
    QDialog, QFormLayout, QLineEdit, QTextEdit, QComboBox, QDateEdit, QPushButton,
    QFileDialog, QLabel, QHBoxLayout, QMessageBox, QWidget, QVBoxLayout, QMainWindow,
    QApplication, QListWidget, QTreeWidget, QGraphicsScene, QStackedWidget, QTreeWidgetItem,
    QAbstractItemView
)
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtGui import QPixmap
import os

# (Removed transitional PyQt5 compatibility shim – codebase now fully PyQt6 native)

# --- Dependency cycle detection helper ---
def _would_create_cycle(model, part_name: str, new_deps: set[str]) -> bool:
    try:
        if not part_name or not new_deps:
            return False
        graph = {}
        for r in getattr(model, 'rows', []):
            n = r.get('Project Part','')
            deps = [d.strip() for d in (r.get('Dependencies') or '').split(',') if d.strip()]
            graph[n] = set(deps)
        graph.setdefault(part_name, set()).update(new_deps)
        visiting=set(); visited=set()
        def dfs(n):
            if n in visiting: return True
            if n in visited: return False
            visiting.add(n)
            for d in graph.get(n, ()):  # n depends on d
                if dfs(d): return True
            visiting.remove(n); visited.add(n)
            return False
        return dfs(part_name)
    except Exception:
        return False

# Unified helper for smooth pixmap scaling mode (centralizes fallback logic)
def _smooth_mode():
    try:
        return Qt.TransformationMode.SmoothTransformation
    except Exception:
        return getattr(Qt, 'SmoothTransformation', 1)

# Unified helper for aspect ratio mode (PyQt6-safe)
def _keep_ar():
    try:
        return Qt.AspectRatioMode.KeepAspectRatio
    except Exception:
        return getattr(Qt, 'KeepAspectRatio', 1)

# --- Central JSON lines logger -------------------------------------------------
# Lightweight, dependency-free structured logging. Writes JSON objects one per
# line to app.log (sibling to the active DB file). Rotates when file exceeds
# ~1MB (simple single-level rotation). Safe for concurrent appenders on local/
# network FS (best-effort; no hard locking). Use log_event(category, event, **kv)
def log_event(category: str, event: str, **fields):
    try:
        import os, json, datetime, socket, getpass, threading
        ts = datetime.datetime.utcnow().isoformat(timespec='seconds') + 'Z'
        try:
            user = getpass.getuser()
        except Exception:
            user = 'unknown'
        try:
            host = socket.gethostname()
        except Exception:
            host = 'host'
        base_dir = None
        # Attempt to colocate with DB if model global exists later; fallback to script dir
        try:
            # Lazy import of __main__ to introspect model if already created
            import __main__ as _m
            if hasattr(_m, 'model') and getattr(_m, 'model') and hasattr(getattr(_m, 'model'), 'DB_FILE'):
                base_dir = os.path.dirname(os.path.abspath(getattr(_m.model, 'DB_FILE')))
        except Exception:
            base_dir = None
        if not base_dir:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        log_path = os.path.join(base_dir, 'app.log')
        # Simple rotation
        try:
            if os.path.exists(log_path) and os.path.getsize(log_path) > 1_000_000:
                bak = log_path + '.1'
                try:
                    if os.path.exists(bak):
                        os.remove(bak)
                except Exception:
                    pass
                try:
                    os.replace(log_path, bak)
                except Exception:
                    pass
        except Exception:
            pass
        rec = {
            'ts': ts,
            'pid': os.getpid(),
            'thread': threading.current_thread().name,
            'user': user,
            'host': host,
            'category': category,
            'event': event,
        }
        # Merge user fields (stringify anything not JSON-serializable)
        for k, v in fields.items():
            try:
                json.dumps(v)
                rec[k] = v
            except Exception:
                rec[k] = repr(v)
        line = json.dumps(rec, ensure_ascii=False)
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception:
        # Never raise from logger
        pass
import shutil

# --- Resource path resolution helper ---
def resolve_resource_path(path: str) -> str:
    """Return an absolute path to a resource that may live next to the script,
    next to the frozen executable, or inside PyInstaller's _MEIPASS (one-file temp dir).
    Checks in this order: absolute -> _MEIPASS -> exe dir -> source dir.
    Returns the first existing candidate, otherwise the first candidate for best effort.
    """
    import os, sys
    if not path:
        return path
    if os.path.isabs(path):
        return path
    candidates = []
    meipass = getattr(sys, '_MEIPASS', None)
    if meipass:
        candidates.append(os.path.join(meipass, path))
    if getattr(sys, 'frozen', False):
        candidates.append(os.path.join(os.path.dirname(sys.executable), path))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), path))
    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[0] if candidates else path

# --- Holidays helper (business calendar) ---
HOLIDAYS_FILE = "holidays.json"
def _holidays_path():
    import os
    try:
        from PyQt6.QtCore import QSettings
        db_path = QSettings('LSI','ProjectApp').value('DB/path', '')
        if db_path:
            base_dir = os.path.dirname(os.path.abspath(db_path))
            return os.path.join(base_dir, HOLIDAYS_FILE)
    except Exception:
        pass
    # Fallback to app folder
    base_dir = os.path.dirname(resolve_resource_path("."))
    return os.path.join(base_dir, HOLIDAYS_FILE)

def load_holiday_dates():
    """Return a set of datetime.date objects for holidays stored in holidays.json (MM-dd-YYYY)."""
    import json, os, datetime
    p = _holidays_path()
    out = set()
    try:
        if os.path.exists(p):
            data = json.load(open(p, "r", encoding="utf-8"))
            for s in data if isinstance(data, list) else []:
                try:
                    dt = datetime.datetime.strptime(s, "%m-%d-%Y").date()
                    out.add(dt)
                except Exception:
                    pass
    except Exception:
        return set()
    return out

def save_holiday_dates(dates):
    """Persist a set/iterable of date strings formatted MM-dd-YYYY to holidays.json."""
    import json
    p = _holidays_path()
    try:
        json.dump(sorted(list(dates)), open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    except Exception:
        pass

# --- Export Settings Dialog (format/page size/orientation/margins) ---
class ExportSettingsDialog(QDialog):
    """Persistent export settings used by PNG/PDF exports.
    Stored under QSettings("LSI", "ProjectPlanner") with keys:
      Export/format -> 'PNG' | 'PDF'
      Export/page_size -> 'A4' | 'Letter' | 'Legal' | 'Tabloid'
      Export/orientation -> 'Portrait' | 'Landscape'
      Export/margin_left_mm, Export/margin_top_mm, Export/margin_right_mm, Export/margin_bottom_mm (float mm)
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        from PyQt6.QtWidgets import QFormLayout, QComboBox, QDialogButtonBox
        from PyQt6.QtWidgets import QDoubleSpinBox
        from PyQt6.QtCore import QSettings
        self.setWindowTitle("Export Settings")
        self.resize(380, 220)
        self.form = QFormLayout(self)
        self.format_combo = QComboBox(); self.format_combo.addItems(["PNG", "PDF"])
        self.size_combo = QComboBox(); self.size_combo.addItems(["A4", "Letter", "Legal", "Tabloid"])
        self.orientation_combo = QComboBox(); self.orientation_combo.addItems(["Portrait", "Landscape"])
        def mkspin():
            sb = QDoubleSpinBox(); sb.setRange(0.0, 50.0); sb.setDecimals(1); sb.setSingleStep(0.5); sb.setSuffix(" mm"); return sb
        self.margin_left = mkspin(); self.margin_top = mkspin(); self.margin_right = mkspin(); self.margin_bottom = mkspin()
        self.form.addRow("Format", self.format_combo)
        self.form.addRow("Page Size (PDF)", self.size_combo)
        self.form.addRow("Orientation (PDF)", self.orientation_combo)
        self.form.addRow("Left Margin", self.margin_left)
        self.form.addRow("Top Margin", self.margin_top)
        self.form.addRow("Right Margin", self.margin_right)
        self.form.addRow("Bottom Margin", self.margin_bottom)
        from PyQt6.QtWidgets import QCheckBox
        self.include_header_cb = QCheckBox("Include Header Graphic")
        self.include_header_cb.setToolTip("If unchecked, exports omit the header.svg/header.png banner.")
        self.form.addRow("Header", self.include_header_cb)
        # Buttons
        try:
            std = QDialogButtonBox.StandardButton
            self.buttons = QDialogButtonBox(std.Ok | std.Cancel)
        except Exception:
            self.buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)
        self.form.addRow(self.buttons)
        # Load settings
        s = QSettings("LSI", "ProjectPlanner")
        fmt = s.value("Export/format", "PNG")
        ps = s.value("Export/page_size", "A4")
        orient = s.value("Export/orientation", "Portrait")
        ml = float(s.value("Export/margin_left_mm", 8.0))
        mt = float(s.value("Export/margin_top_mm", 8.0))
        mr = float(s.value("Export/margin_right_mm", 8.0))
        mb = float(s.value("Export/margin_bottom_mm", 8.0))
        self.format_combo.setCurrentText(fmt if fmt in ("PNG","PDF") else "PNG")
        if ps not in ("A4","Letter","Legal","Tabloid"): ps = "A4"
        self.size_combo.setCurrentText(ps)
        if orient not in ("Portrait","Landscape"): orient = "Portrait"
        self.orientation_combo.setCurrentText(orient)
        self.margin_left.setValue(ml); self.margin_top.setValue(mt); self.margin_right.setValue(mr); self.margin_bottom.setValue(mb)
        inc_header = s.value("Export/include_header", True)
        if isinstance(inc_header, str):
            inc_header = (inc_header.lower() in ("1","true","yes"))
        self.include_header_cb.setChecked(bool(inc_header))
        # Disable PDF-only fields when PNG selected
        def update_pdf_only():
            is_pdf = (self.format_combo.currentText() == "PDF")
            self.size_combo.setEnabled(is_pdf)
            self.orientation_combo.setEnabled(is_pdf)
        self.format_combo.currentTextChanged.connect(lambda _: update_pdf_only())
        update_pdf_only()
    def accept(self):
        try:
            from PyQt6.QtCore import QSettings
            s = QSettings("LSI", "ProjectPlanner")
            s.setValue("Export/format", self.format_combo.currentText())
            s.setValue("Export/page_size", self.size_combo.currentText())
            s.setValue("Export/orientation", self.orientation_combo.currentText())
            s.setValue("Export/margin_left_mm", float(self.margin_left.value()))
            s.setValue("Export/margin_top_mm", float(self.margin_top.value()))
            s.setValue("Export/margin_right_mm", float(self.margin_right.value()))
            s.setValue("Export/margin_bottom_mm", float(self.margin_bottom.value()))
            s.setValue("Export/include_header", bool(self.include_header_cb.isChecked()))
        except Exception:
            pass
        return super().accept()

class PricingSettingsDialog(QDialog):
    """Dialog to configure pricing guidance: target margin and default labor rates."""
    def __init__(self, parent=None):
        super().__init__(parent)
        from PyQt6.QtWidgets import QFormLayout, QDialogButtonBox, QDoubleSpinBox
        from PyQt6.QtCore import QSettings
        self.setWindowTitle("Pricing Settings")
        self.resize(360, 180)
        form = QFormLayout(self)
        def mkspin(minv, maxv, step, dec=1, suffix=""):
            sb = QDoubleSpinBox(); sb.setRange(minv,maxv); sb.setDecimals(dec); sb.setSingleStep(step)
            if suffix: sb.setSuffix(" "+suffix)
            return sb
        self.target_margin = mkspin(0, 95, 1, 1, "%")
        self.labor_rate = mkspin(0, 1000, 5, 2, "$ /h")
        self.install_labor_rate = mkspin(0, 1000, 5, 2, "$ /h")
        form.addRow("Target Margin %", self.target_margin)
        form.addRow("Fabrication Labor Rate", self.labor_rate)
        form.addRow("Install Labor Rate", self.install_labor_rate)
        s = QSettings("LSI","ProjectPlanner")
        try:
            self.target_margin.setValue(float(s.value("Pricing/target_margin", 35)))
            self.labor_rate.setValue(float(s.value("Pricing/labor_rate", 55)))
            self.install_labor_rate.setValue(float(s.value("Pricing/install_labor_rate", 65)))
        except Exception:
            pass
        try:
            std = QDialogButtonBox.StandardButton
            buttons = QDialogButtonBox(std.Ok | std.Cancel)
        except Exception:
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        form.addWidget(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
    def accept(self):
        from PyQt6.QtCore import QSettings
        try:
            s = QSettings("LSI","ProjectPlanner")
            s.setValue("Pricing/target_margin", float(self.target_margin.value()))
            s.setValue("Pricing/labor_rate", float(self.labor_rate.value()))
            s.setValue("Pricing/install_labor_rate", float(self.install_labor_rate.value()))
        except Exception:
            pass
        return super().accept()

# --- First Run / Empty DB Onboarding Dialog ---
class FirstRunDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        from PyQt6.QtWidgets import QVBoxLayout, QLabel, QPushButton, QHBoxLayout, QCheckBox
        self.setWindowTitle("Welcome – Getting Started")
        self.setModal(True)
        self.resize(520, 320)
        lay = QVBoxLayout(self)
        msg = QLabel("It looks like this database is empty. Choose how you'd like to proceed.")
        msg.setWordWrap(True)
        lay.addWidget(msg)
        # Action buttons row
        def mkbtn(text, tooltip):
            b = QPushButton(text); b.setToolTip(tooltip); b.setMinimumHeight(46); return b
        self.btn_switch = mkbtn("Switch to Existing DB…", "Select another SQLite file (shared or local)")
        self.btn_open   = mkbtn("Open Data Folder", "Open the current working folder in your file browser")
        for b in (self.btn_switch, self.btn_open):
            lay.addWidget(b)
        # Read-only tip
        tip = QLabel("Tip: Use Tools → Read-Only Mode to safely view a shared DB. Disable it to edit (acquiring the edit lock). You can generate sample data later via Tools → Generate Sample Data if needed.")
        tip.setWordWrap(True)
        tip.setStyleSheet("color:#bbb; font-size:11px; margin-top:8px;")
        lay.addWidget(tip)
        # Don't show again
        self.chk_hide = QCheckBox("Don't show this again on empty databases")
        lay.addWidget(self.chk_hide)
        # Close row
        row = QHBoxLayout(); row.addStretch(1)
        self.btn_close = QPushButton("Close")
        row.addWidget(self.btn_close)
        lay.addLayout(row)
        self.btn_close.clicked.connect(self.accept)
        self.selected_action = None  # 'switch' | 'open'
        self.btn_switch.clicked.connect(lambda: self._choose('switch'))
        self.btn_open.clicked.connect(lambda: self._choose('open'))
    def _choose(self, which):
        self.selected_action = which
        self.accept()
    def hide_future(self):
        return self.chk_hide.isChecked()

class ImageCellWidget(QWidget):
    def __init__(self, parent, row, col, model, on_data_changed=None):
        super().__init__(parent)
        self.row = row
        self.col = col
        self.model = model
        self.on_data_changed = on_data_changed
        layout = QHBoxLayout()
        self.img_label = QLabel()
        self.img_label.setFixedSize(48, 48)
        layout.addWidget(self.img_label)
        self.btn = QPushButton("Upload")
        self.btn.clicked.connect(self.upload_image)
        layout.addWidget(self.btn)
        self.setLayout(layout)
        self.refresh()

    def upload_image(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Select Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp *.gif)")
        if fname:
            # Store images next to the executable/script so packaged app can load them
            base_dir = os.path.dirname(resolve_resource_path("."))
            images_dir = os.path.join(base_dir, "images")
            if not os.path.exists(images_dir):
                os.makedirs(images_dir)
            base = os.path.basename(fname)
            dest = os.path.join(images_dir, base)
            count = 1
            orig_base, ext = os.path.splitext(base)
            while os.path.exists(dest):
                dest = os.path.join(images_dir, f"{orig_base}_{count}{ext}")
                count += 1
            shutil.copy2(fname, dest)
            rel_path = os.path.relpath(dest, base_dir)
            self.model.rows[self.row][self.model.COLUMNS[self.col]] = rel_path
            self.model.save_to_db()
            self.refresh()

    def refresh(self):
        img_path = self.model.rows[self.row].get(self.model.COLUMNS[self.col], "")
        if img_path:
            # Resolve relative path via resource resolver
            base_dir = os.path.dirname(resolve_resource_path("."))
            img_path_full = os.path.join(base_dir, img_path)
            pixmap = QPixmap(img_path_full)
            if not pixmap.isNull():
                self.img_label.setPixmap(pixmap.scaled(48, 48, _keep_ar(), _smooth_mode()))
                try:
                    self.img_label.setCursor(getattr(Qt,'CursorShape', Qt).PointingHandCursor)
                except Exception:
                    pass
                self.img_label.mousePressEvent = lambda event, p=img_path_full: self.show_full_image(p)
            else:
                self.img_label.setText("[Image not found]")
                self.img_label.setCursor(getattr(Qt,'CursorShape', Qt).ArrowCursor)
                self.img_label.mousePressEvent = None
        else:
            self.img_label.setText("")
            self.img_label.setCursor(getattr(Qt,'CursorShape', Qt).ArrowCursor)
            self.img_label.mousePressEvent = None

    def show_full_image(self, img_path_full):
        dlg = QDialog(self)
        dlg.setWindowTitle("Image Preview")
        vbox = QVBoxLayout(dlg)
        lbl = QLabel()
        pixmap = QPixmap(img_path_full)
        if not pixmap.isNull():
            try:
                _smooth = Qt.TransformationMode.SmoothTransformation
            except Exception:
                _smooth = getattr(Qt, 'SmoothTransformation', 1)
            lbl.setPixmap(pixmap.scaledToWidth(600, _smooth))
        else:
            lbl.setText("[Image not found]")
        vbox.addWidget(lbl)
        dlg.setLayout(vbox)
        dlg.exec()

class ProjectDataModel:
    # NOTE: Append-only pattern; new progress-related columns added at end to avoid breaking older rows
    COLUMNS = [
        "Project Part", "Parent", "Children", "Start Date", "Duration (days)", "Internal/External", "Dependencies", "Type", "Calculated End Date", "Resources", "Notes", "Responsible", "Images", "Pace Link", "Attachments",
        # Progress tracking fields
        "% Complete",            # Integer 0-100 (leaf editable, parents rolled up)
        "Status",                 # Planned | In Progress | Blocked | Done | Deferred
        "Actual Start Date",      # Set when Status transitions to In Progress
        "Actual Finish Date",     # Set when Status transitions to Done
        "Baseline Start Date",    # Captured first time valid start/duration appear
        "Baseline End Date",      # Derived from baseline start + duration (working days not yet applied)
        # Cost tracking fields
        "Production Cost",        # Internal estimated production cost (materials + fabrication labor)
        "Installation Cost",      # Internal estimated install cost (crew labor + equipment)
        "Production Price",       # Decimal number (price to be charged for production)
        "Installation Price",     # Decimal number (price to be charged for installation)
        # Extended cost breakdown (append-only)
        "Material Cost",          # Materials-only direct cost
        "Fabrication Labor Hours",# Hours for shop fabrication
        "Installation Labor Hours",# Hours for field install
        "Labor Rate",             # Default blended labor rate (can be overridden per row)
        "Install Labor Rate"      # Field install labor rate
        ,"Equipment Cost"         # Lift / crane / equipment rental cost
        ,"Permit/Eng Cost"        # Permitting or engineering fees
        ,"Contingency %"          # Applied percentage buffer (on cost basis)
        ,"Warranty Reserve %"     # Percentage of price allocated to warranty reserve
        ,"Risk Level"             # Low | Medium | High affects margin target (future)
        ,"Quote Version"          # Current working quote version label
        ,"Frozen Production Cost" # Snapshot baseline cost
        ,"Frozen Installation Cost"
        ,"Frozen Production Price"
        ,"Frozen Installation Price"
    ]
    DB_FILE = "project_data.db"

    def __init__(self):
        self.rows = []  # Each row is a dict with keys as COLUMNS
        # collaborative mode: prevent writes on viewer machines (persisted via QSettings)
        try:
            from PyQt6.QtCore import QSettings
            _qs = QSettings('LSI','ProjectApp')
            ro = _qs.value('DB/read_only', False)
            if isinstance(ro, str):
                ro = ro.lower() in ('1','true','yes','on')
            self.read_only = bool(ro)
        except Exception:
            self.read_only = False
        # Allow environment to force read-only explicitly (e.g., for viewers)
        try:
            import os
            env_ro = os.environ.get('PROJECTAPP_READ_ONLY','')
            if env_ro and str(env_ro).lower() in ('1','true','yes','on'):
                self.read_only = True
        except Exception:
            pass
        # Resolve DB path with simple override mechanisms suitable for a shared network DB scenario.
        # Precedence:
        #  1) Environment variable PROJECT_DB_PATH (UNC or local; supports %VAR% and ~)
        #  2) db_path.txt file in the working directory containing a path
        #  3) db_path.txt adjacent to this script
        #  4) Default: project_data.db in current working directory
        try:
            import os
            def _norm(p):
                if not p:
                    return p
                p = os.path.expanduser(os.path.expandvars(p.strip()))
                return os.path.normpath(p)
            env_path = os.environ.get("PROJECT_DB_PATH")
            if env_path and env_path.strip():
                # Strip UTF-8 BOM if present
                if env_path and env_path[:1] == '\ufeff':
                    env_path = env_path.lstrip('\ufeff')
                self.DB_FILE = _norm(env_path)
            else:
                # db_path.txt in CWD
                cwd_cfg = os.path.join(os.getcwd(), "db_path.txt")
                script_cfg = os.path.join(os.path.dirname(os.path.abspath(__file__)), "db_path.txt")
                use_cfg = cwd_cfg if os.path.exists(cwd_cfg) else (script_cfg if os.path.exists(script_cfg) else None)
                if use_cfg:
                    try:
                        with open(use_cfg, "r", encoding="utf-8") as f:
                            cfg_path = f.read()
                            # Remove BOM and whitespace/newlines
                            if cfg_path and cfg_path[:1] == '\ufeff':
                                cfg_path = cfg_path.lstrip('\ufeff')
                            cfg_path = cfg_path.strip()
                            if cfg_path:
                                self.DB_FILE = _norm(cfg_path)
                    except Exception:
                        pass
        except Exception:
            pass

        # Persist resolved DB path for helpers (e.g., holidays path)
        try:
            from PyQt6.QtCore import QSettings
            QSettings('LSI','ProjectApp').setValue('DB/path', self.DB_FILE)
        except Exception:
            pass

        # If running in a PyInstaller one-file bundle, the bundled DB is read-only inside the temp extraction dir.
        # We want user edits to persist next to the executable (current working directory) if no DB exists yet.
        try:
            import sys, os, shutil
            if not os.path.exists(self.DB_FILE):
                # Detect PyInstaller one-file _MEIPASS path (only present in one-file mode)
                bundle_dir = getattr(sys, '_MEIPASS', None)
                if bundle_dir:
                    candidate = os.path.join(bundle_dir, 'project_data.db')
                    if os.path.exists(candidate):
                        shutil.copy2(candidate, self.DB_FILE)
        except Exception:
            pass
        # Ensure parent directory exists for DB file when using a custom path
        try:
            import os
            db_dir = os.path.dirname(os.path.abspath(self.DB_FILE))
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir, exist_ok=True)
        except Exception:
            pass
        try:
            self.ensure_schema()
            self.load_from_db()
        except Exception as e:
            # Provide better diagnostics when DB cannot be opened
            import traceback
            print(f"ERROR: Failed to open database '{self.DB_FILE}': {e}")
            traceback.print_exc()
            # Fallback: if path looks like a network/UNC path, retry with a local DB to allow app to start
            try:
                import os
                path_lower = str(self.DB_FILE).lower()
                if path_lower.startswith('\\\\') or path_lower.startswith('\\') or ':' in path_lower and not os.path.exists(os.path.dirname(os.path.abspath(self.DB_FILE))):
                    local_db = 'project_data.db'
                    if local_db != self.DB_FILE:
                        print(f"WARN: Falling back to local '{local_db}' due to inaccessible primary DB path.")
                        self.DB_FILE = local_db
                        try:
                            from PyQt6.QtCore import QSettings as _QS
                            _QS('LSI','ProjectApp').setValue('DB/path', self.DB_FILE)
                        except Exception:
                            pass
                        try:
                            self.ensure_schema()
                            self.load_from_db()
                            print("INFO: Local fallback DB initialized.")
                            return
                        except Exception as _e2:
                            print(f"ERROR: Fallback local DB also failed: {_e2}")
            except Exception:
                pass
            raise

    def _connect(self):
        """Return an sqlite3 connection with WAL, busy timeout and slightly safer cache settings.
        For network drives like OneDrive/SharePoint, WAL reduces lock contention but conflicts can still occur.
        """
        import sqlite3
        # Ensure parent exists (defensive guard for scenarios where DB path dir was missing)
        try:
            db_dir = os.path.dirname(os.path.abspath(self.DB_FILE))
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir, exist_ok=True)
        except Exception:
            pass
        # Use check_same_thread=False to allow background UI operations if needed (Qt timers)
        if getattr(self, 'read_only', False):
            try:
                # Attempt read-only connection via URI
                uri = f"file:{self.DB_FILE}?mode=ro"
                conn = sqlite3.connect(uri, uri=True, timeout=5.0, check_same_thread=False)
                try:
                    cur = conn.cursor()
                    cur.execute("PRAGMA foreign_keys=ON")
                    cur.execute("PRAGMA busy_timeout=5000")
                    conn.commit()
                except Exception:
                    pass
                return conn
            except Exception:
                # Fallback to normal connect (may still be readable)
                return sqlite3.connect(self.DB_FILE, timeout=5.0, check_same_thread=False)
        else:
            conn = sqlite3.connect(self.DB_FILE, timeout=5.0, check_same_thread=False)
            try:
                cur = conn.cursor()
                cur.execute("PRAGMA journal_mode=WAL")
                cur.execute("PRAGMA synchronous=NORMAL")
                cur.execute("PRAGMA busy_timeout=5000")
                cur.execute("PRAGMA foreign_keys=ON")
                conn.commit()
            except Exception:
                pass
            return conn

    # --- Schema migration to add progress columns if missing ---
    def ensure_schema(self):
        import sqlite3, os
        if not os.path.exists(self.DB_FILE):
            # Table will be created later in create_table()
            try:
                log_event('schema','db_missing', path=self.DB_FILE)
            except Exception:
                pass
            return
        # In read-only mode, skip schema migrations to avoid errors on shared DBs
        if getattr(self, 'read_only', False):
            try:
                log_event('schema','skip_read_only')
            except Exception:
                pass
            return
        with self._connect() as conn:
            c = conn.cursor()
            # Inspect existing columns
            try:
                c.execute("PRAGMA table_info(project_parts)")
                existing = [row[1] for row in c.fetchall()]  # name in 2nd column
                try:
                    log_event('schema','existing_columns', count=len(existing))
                except Exception:
                    pass
            except Exception:
                existing = []
            to_add = [col for col in self.COLUMNS if col not in existing]
            for col in to_add:
                # Decide type based on semantic
                if col == "% Complete":
                    col_type = "INTEGER"
                elif col in ("Production Price", "Installation Price", "Production Cost", "Installation Cost", "Material Cost", "Labor Rate", "Install Labor Rate", "Equipment Cost", "Permit/Eng Cost", "Frozen Production Cost", "Frozen Installation Cost", "Frozen Production Price", "Frozen Installation Price"):
                    col_type = "REAL"
                elif col in ("Fabrication Labor Hours", "Installation Labor Hours"):
                    col_type = "REAL"
                elif col in ("Contingency %", "Warranty Reserve %"):
                    col_type = "REAL"
                elif col == "Attachments":
                    col_type = "TEXT"  # JSON list of relative paths
                else:
                    col_type = "TEXT"
                try:
                    c.execute(f'ALTER TABLE project_parts ADD COLUMN "{col}" {col_type}')
                except Exception:
                    pass
            # Add optimistic concurrency columns if missing
            if 'row_version' not in existing:
                try:
                    c.execute('ALTER TABLE project_parts ADD COLUMN row_version INTEGER DEFAULT 0')
                except Exception:
                    pass
            if 'last_modified_utc' not in existing:
                try:
                    c.execute('ALTER TABLE project_parts ADD COLUMN last_modified_utc TEXT DEFAULT ""')
                except Exception:
                    pass
            # Baselines table for named snapshots
            try:
                c.execute(
                    """
                    CREATE TABLE IF NOT EXISTS baselines (
                        baseline_name TEXT NOT NULL,
                        part_name TEXT NOT NULL,
                        start_date TEXT,
                        end_date TEXT,
                        PRIMARY KEY (baseline_name, part_name)
                    )
                    """
                )
            except Exception:
                pass
            # Quote versions table for frozen pricing snapshots
            try:
                c.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quote_versions (
                        version_name TEXT NOT NULL,
                        part_name TEXT NOT NULL,
                        production_cost REAL,
                        installation_cost REAL,
                        production_price REAL,
                        installation_price REAL,
                        PRIMARY KEY (version_name, part_name)
                    )
                    """
                )
            except Exception:
                pass
            # Changes audit log table
            try:
                c.execute(
                    """
                    CREATE TABLE IF NOT EXISTS changes (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        when_utc TEXT NOT NULL,
                        user TEXT,
                        part_name TEXT NOT NULL,
                        field TEXT NOT NULL,
                        old_value TEXT,
                        new_value TEXT
                    )
                    """
                )
            except Exception:
                pass
            conn.commit()
        try:
            log_event('schema','ensure_schema_complete', added=to_add, has_row_version=('row_version' in existing), has_last_modified=('last_modified_utc' in existing))
        except Exception:
            pass

    def add_row(self, data, parent=None):
        row = {col: val for col, val in zip(self.COLUMNS, data)}
        row['Parent'] = parent
        self.rows.append(row)
        return len(self.rows) - 1

    def update_row(self, idx, data):
        # Legacy in-memory update (used before persistence commit); does not apply concurrency logic.
        for i, col in enumerate(self.COLUMNS):
            self.rows[idx][col] = data[i]

    def update_part_values(self, part_name: str, new_values: dict, expected_version: int):
        """Optimistic concurrency update by Project Part name.
        new_values: dict of column->value (must be subset of self.COLUMNS)
        expected_version: caller's last known row_version
        Returns (True, new_version) on success; (False, reason) on conflict/error."""
        import datetime, os
        if getattr(self, 'read_only', False):
            try: log_event('concurrency','update_read_only', part=part_name)
            except Exception: pass
            return False, "Read-only mode (shared lock active)"
        if not part_name or not new_values:
            try: log_event('concurrency','update_invalid_params', part=part_name)
            except Exception: pass
            return False, "Invalid parameters"
        if not os.path.exists(self.DB_FILE):
            try: log_event('concurrency','update_db_missing', part=part_name)
            except Exception: pass
            return False, "DB missing"
        # Sanitize keys
        valid = {k: v for k, v in new_values.items() if k in self.COLUMNS}
        if not valid:
            try: log_event('concurrency','update_no_valid_fields', part=part_name)
            except Exception: pass
            return False, "No valid fields"
        # Build SQL
        set_fragments = []
        params = []
        for k, v in valid.items():
            set_fragments.append(f'"{k}"=?')
            params.append(v)
        set_fragments.append('last_modified_utc=?')
        now_iso = datetime.datetime.utcnow().isoformat(timespec='seconds')
        params.append(now_iso)
        set_fragments.append('row_version = row_version + 1')
        sql = f'UPDATE project_parts SET {", ".join(set_fragments)} WHERE "Project Part"=? AND row_version=?'
        params.extend([part_name, expected_version])
        try:
            with self._connect() as conn:
                cur = conn.cursor()
                cur.execute('BEGIN IMMEDIATE')
                cur.execute(sql, params)
                if cur.rowcount == 0:
                    try: log_event('concurrency','conflict', part=part_name, expected=expected_version)
                    except Exception: pass
                    return False, "Conflict"
                # Fetch new version
                cur.execute('SELECT row_version FROM project_parts WHERE "Project Part"=?', (part_name,))
                new_ver = cur.fetchone()[0]
                conn.commit()
            # Update in-memory copy
            for r in self.rows:
                if r.get('Project Part') == part_name:
                    r.update(valid)
                    r['row_version'] = new_ver
                    r['last_modified_utc'] = now_iso
                    break
            try: log_event('concurrency','update_success', part=part_name, new_version=new_ver, fields=list(valid.keys()))
            except Exception: pass
            return True, new_ver
        except Exception as e:
            try: log_event('concurrency','update_exception', part=part_name, error=str(e))
            except Exception: pass
            return False, str(e)

    def delete_row(self, idx):
        # Remove children recursively
        children = [i for i, r in enumerate(self.rows) if r.get('Parent') == idx]
        for c in sorted(children, reverse=True):
            self.delete_row(c)
        del self.rows[idx]
        # Update parent indices
        for r in self.rows:
            if r.get('Parent') is not None and isinstance(r.get('Parent'), int) and r.get('Parent') > idx:
                r['Parent'] -= 1

    def get_tree(self):
        # Returns a list of (row, children) tuples for tree rendering
        def collect(parent_idx):
            nodes = []
            for i, r in enumerate(self.rows):
                if r.get('parent') == parent_idx:
                    nodes.append((i, collect(i)))
            return nodes
        return collect(None)

    def get_flat(self):
        return self.rows

    def load_from_db(self):
        import os
        import sqlite3
        self.rows.clear()
        if not os.path.exists(self.DB_FILE):
            # In read-only mode, don't attempt to create a new DB/schema
            if getattr(self, 'read_only', False):
                try:
                    log_event('db','skip_create_missing_read_only', path=self.DB_FILE)
                except Exception:
                    pass
                return
            self.create_table()
            return
        with self._connect() as conn:
            c = conn.cursor()
            # Build quoted column list without nested f-strings/backslashes (macOS Python parser-safe)
            cols_quoted = ", ".join(['"{}"'.format(col) for col in self.COLUMNS])
            # Also pull concurrency columns if they exist
            try:
                c.execute('PRAGMA table_info(project_parts)')
                existing_cols = [r[1] for r in c.fetchall()]
            except Exception:
                existing_cols = []
            concurrency_select = []
            if 'row_version' in existing_cols:
                concurrency_select.append('row_version')
            if 'last_modified_utc' in existing_cols:
                concurrency_select.append('last_modified_utc')
            extra_sql = (', ' + ', '.join(concurrency_select)) if concurrency_select else ''
            c.execute(f"SELECT {cols_quoted}{extra_sql} FROM project_parts")
            for row in c.fetchall():
                base_part = row[:len(self.COLUMNS)]
                row_dict = {col: val for col, val in zip(self.COLUMNS, base_part)}
                if concurrency_select:
                    # Append concurrency fields by order appended
                    tail = row[len(self.COLUMNS):]
                    for name, val in zip(concurrency_select, tail):
                        row_dict[name] = val
                # Default missing progress fields (older rows) if any are absent or None
                if row_dict.get("% Complete") in (None, ""):
                    row_dict["% Complete"] = 0
                if not row_dict.get("Status"):
                    row_dict["Status"] = "Planned"
                # Normalize attachments field to JSON list string
                import json as _json_att
                att_val = row_dict.get("Attachments")
                if att_val in (None, ""):
                    row_dict["Attachments"] = "[]"
                else:
                    try:
                        parsed = _json_att.loads(att_val)
                        if not isinstance(parsed, list):
                            row_dict["Attachments"] = _json_att.dumps([att_val])
                    except Exception:
                        row_dict["Attachments"] = _json_att.dumps([att_val])
                print(f"Loaded from DB: {row_dict}")
                self.rows.append(row_dict)
        self.update_calculated_end_dates()
        # After loading & computing end dates, establish baseline if missing
        self.capture_missing_baselines()

    def get_row_snapshot(self, part_name: str):
        """Return a fresh DB snapshot dict for the given part including row_version/last_modified if present, or None."""
        if not part_name:
            return None
        import os, sqlite3
        if not os.path.exists(self.DB_FILE):
            return None
        with self._connect() as conn:
            cur = conn.cursor()
            try:
                cur.execute('PRAGMA table_info(project_parts)')
                cols = [r[1] for r in cur.fetchall()]
            except Exception:
                cols = []
            select_cols = list(self.COLUMNS)
            if 'row_version' in cols:
                select_cols.append('row_version')
            if 'last_modified_utc' in cols:
                select_cols.append('last_modified_utc')
            quoted = ', '.join(['"{}"'.format(c) for c in select_cols])
            try:
                cur.execute(f'SELECT {quoted} FROM project_parts WHERE "Project Part"=? LIMIT 1', (part_name,))
                res = cur.fetchone()
                if not res:
                    return None
                snap = {c: v for c, v in zip(select_cols, res)}
                return snap
            except Exception:
                return None

    def capture_missing_baselines(self):
        import datetime
        for r in self.rows:
            start = r.get("Start Date")
            dur = r.get("Duration (days)")
            if start and dur and (not r.get("Baseline Start Date") or not r.get("Baseline End Date")):
                try:
                    sd = datetime.datetime.strptime(start, "%m-%d-%Y")
                    d = int(dur)
                    end = sd + datetime.timedelta(days=d)
                    if not r.get("Baseline Start Date"):
                        r["Baseline Start Date"] = sd.strftime("%m-%d-%Y")
                    if not r.get("Baseline End Date"):
                        r["Baseline End Date"] = end.strftime("%m-%d-%Y")
                except Exception:
                    pass

    def save_to_db(self):
        import sqlite3, os, socket, getpass, json
        if getattr(self, 'read_only', False):
            # Skip save in read-only mode (collaborative viewer)
            try: log_event('db','save_skipped_read_only')
            except Exception: pass
            return
            # Enforce single-editor lock: if a lock file exists and we are not the owner, prevent writes
        try:
            dbp = getattr(self, 'DB_FILE', None) or ''
            if dbp:
                lock_path = os.path.abspath(dbp) + '.lock.json'
                if os.path.exists(lock_path):
                    try:
                        info = json.load(open(lock_path, 'r', encoding='utf-8'))
                    except Exception:
                        info = None
                    me = None
                    try:
                        me = f"{getpass.getuser()}@{socket.gethostname()}"
                    except Exception:
                        me = 'unknown@host'
                    owner = (info or {}).get('owner') if isinstance(info, dict) else None
                    if owner and owner != me:
                        # Hard block to guarantee only the lock owner can write
                        print(f"Save blocked: edit lock held by {owner}")
                        try: log_event('db','save_blocked_lock', owner=owner)
                        except Exception: pass
                        return
        except Exception:
            pass
        self.update_calculated_end_dates()
        # Roll-ups before save to persist auto-calculated parent progress
        self.rollup_progress()
        self.create_table()
        # Full rewrite (legacy). Preserve row_version/last_modified_utc if present in table.
        with self._connect() as conn:
            c = conn.cursor()
            try:
                c.execute("BEGIN IMMEDIATE")
            except Exception:
                pass
            # Determine if concurrency columns exist
            try:
                c.execute('PRAGMA table_info(project_parts)')
                cols_exist = [r[1] for r in c.fetchall()]
            except Exception:
                cols_exist = []
            has_rv = 'row_version' in cols_exist
            has_lm = 'last_modified_utc' in cols_exist
            c.execute("DELETE FROM project_parts")
            base_cols = [col for col in self.COLUMNS]
            extra_cols = []
            if has_rv:
                extra_cols.append('row_version')
            if has_lm:
                extra_cols.append('last_modified_utc')
            all_cols = base_cols + extra_cols
            columns_sql = ", ".join(['"{}"'.format(col) for col in all_cols])
            placeholders = ", ".join(["?" for _ in all_cols])
            for row in self.rows:
                vals = [row.get(c, "") for c in base_cols]
                if has_rv:
                    vals.append(row.get('row_version', 0))
                if has_lm:
                    vals.append(row.get('last_modified_utc', ''))
                c.execute(f"INSERT INTO project_parts ({columns_sql}) VALUES ({placeholders})", vals)
            conn.commit()
        try: log_event('db','save_complete', rows=len(self.rows))
        except Exception: pass

    def create_table(self):
        import sqlite3
        with self._connect() as conn:
            c = conn.cursor()
            fields = ", ".join([f'"{col}" TEXT' for col in self.COLUMNS])
            c.execute(f"""
                CREATE TABLE IF NOT EXISTS project_parts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    {fields}
                )
            """)
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS baselines (
                    baseline_name TEXT NOT NULL,
                    part_name TEXT NOT NULL,
                    start_date TEXT,
                    end_date TEXT,
                    PRIMARY KEY (baseline_name, part_name)
                )
                """
            )
            conn.commit()

    # --- Baseline snapshots API ---
    def list_baselines(self):
        import sqlite3, os
        if not os.path.exists(self.DB_FILE):
            return []
        with self._connect() as conn:
            c = conn.cursor()
            try:
                c.execute("SELECT DISTINCT baseline_name FROM baselines ORDER BY baseline_name")
                return [r[0] for r in c.fetchall()]
            except Exception:
                return []

    def save_baseline(self, name: str):
        import sqlite3, datetime
        if not name:
            return
        with self._connect() as conn:
            c = conn.cursor()
            for r in self.rows:
                part = r.get("Project Part", "")
                s = r.get("Start Date", "")
                e = r.get("Calculated End Date", "")
                if not e and s and r.get("Duration (days)"):
                    try:
                        sd = datetime.datetime.strptime(s, "%m-%d-%Y")
                        d = int(r.get("Duration (days)") or 0)
                        e = (sd + datetime.timedelta(days=d)).strftime("%m-%d-%Y")
                    except Exception:
                        e = ""
                try:
                    c.execute(
                        """
                        INSERT INTO baselines (baseline_name, part_name, start_date, end_date)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(baseline_name, part_name)
                        DO UPDATE SET start_date=excluded.start_date, end_date=excluded.end_date
                        """,
                        (name, part, s, e)
                    )
                except Exception:
                    pass
            conn.commit()

    # --- Quote Versioning API ---
    def save_quote_version(self, version_name: str):
        if not version_name:
            return
        import os
        if not os.path.exists(self.DB_FILE):
            return
        with self._connect() as conn:
            cur = conn.cursor()
            for r in self.rows:
                try:
                    cur.execute(
                        """
                        INSERT INTO quote_versions (version_name, part_name, production_cost, installation_cost, production_price, installation_price)
                        VALUES (?,?,?,?,?,?)
                        ON CONFLICT(version_name, part_name) DO UPDATE SET
                          production_cost=excluded.production_cost,
                          installation_cost=excluded.installation_cost,
                          production_price=excluded.production_price,
                          installation_price=excluded.installation_price
                        """,
                        (
                            version_name,
                            r.get('Project Part',''),
                            float(r.get('Production Cost') or 0),
                            float(r.get('Installation Cost') or 0),
                            float(r.get('Production Price') or 0),
                            float(r.get('Installation Price') or 0)
                        )
                    )
                except Exception:
                    pass
            conn.commit()

    def list_quote_versions(self):
        import os
        if not os.path.exists(self.DB_FILE):
            return []
        with self._connect() as conn:
            cur = conn.cursor()
            try:
                cur.execute("SELECT DISTINCT version_name FROM quote_versions ORDER BY version_name")
                return [r[0] for r in cur.fetchall()]
            except Exception:
                return []

    def delete_quote_version(self, version_name: str):
        import os
        if not version_name or not os.path.exists(self.DB_FILE):
            return False
        with self._connect() as conn:
            cur = conn.cursor()
            try:
                cur.execute("DELETE FROM quote_versions WHERE version_name=?", (version_name,))
                conn.commit()
                return True
            except Exception:
                return False

    def rename_quote_version(self, old_name: str, new_name: str):
        """Rename a quote version label. Returns True on success.
        Fails (returns False) if destination already exists or source missing."""
        import os
        if (not old_name or not new_name or old_name == new_name or
                not os.path.exists(self.DB_FILE)):
            return False
        with self._connect() as conn:
            cur = conn.cursor()
            try:
                # Abort if destination exists
                cur.execute("SELECT 1 FROM quote_versions WHERE version_name=? LIMIT 1", (new_name,))
                if cur.fetchone():
                    return False
                cur.execute("UPDATE quote_versions SET version_name=? WHERE version_name=?", (new_name, old_name))
                conn.commit()
                return cur.rowcount > 0
            except Exception:
                return False

    def load_quote_version_map(self, version_name: str):
        import os
        if not version_name or not os.path.exists(self.DB_FILE):
            return {}
        with self._connect() as conn:
            cur = conn.cursor()
            try:
                cur.execute("SELECT part_name, production_cost, installation_cost, production_price, installation_price FROM quote_versions WHERE version_name=?", (version_name,))
                out = {}
                for part, pc, ic, pp, ip in cur.fetchall():
                    out[part] = (pc or 0.0, ic or 0.0, pp or 0.0, ip or 0.0)
                return out
            except Exception:
                return {}

    def load_baseline_map(self, name: str):
        import sqlite3, os
        if not name:
            return {}
        if not os.path.exists(self.DB_FILE):
            return {}
        with self._connect() as conn:
            c = conn.cursor()
            try:
                c.execute("SELECT part_name, start_date, end_date FROM baselines WHERE baseline_name=?", (name,))
                out = {}
                for part, s, e in c.fetchall():
                    out[part] = (s or "", e or "")
                return out
            except Exception:
                return {}

    def update_calculated_end_dates(self):
        import datetime
        for row in self.rows:
            start = row.get("Start Date", "")
            duration = row.get("Duration (days)", "")
            try:
                if start and duration:
                    start_date = datetime.datetime.strptime(start, "%m-%d-%Y")
                    days = int(duration)
                    current_date = start_date
                    added_days = 0
                    while added_days < days:
                        current_date += datetime.timedelta(days=1)
                        # Skip Saturday (5) and Sunday (6)
                        if current_date.weekday() < 5:
                            added_days += 1
                    row["Calculated End Date"] = current_date.strftime("%m-%d-%Y")
                else:
                    row["Calculated End Date"] = ""
            except Exception:
                row["Calculated End Date"] = ""

    # --- Progress Roll-up Logic ---
    def rollup_progress(self):
        # Build children mapping by parent part name (string)
        name_to_row = {r.get("Project Part", ""): r for r in self.rows}
        children = {}
        for r in self.rows:
            p = r.get("Parent") or ""
            if p:
                children.setdefault(p, []).append(r)

        # Depth-first post-order to compute parent % Complete
        visited = set()
        def dfs(name):
            if name in visited:
                return
            visited.add(name)
            row = name_to_row.get(name)
            if not row:
                return
            # Leaf: ensure numeric % Complete & Status defaults
            if name not in children:
                try:
                    pc = int(row.get("% Complete") or 0)
                except Exception:
                    pc = 0
                row["% Complete"] = max(0, min(100, pc))
                if row.get("Status") == "Done" and row["% Complete"] < 100:
                    row["% Complete"] = 100
                return
            # Recurse children first
            total_weight = 0
            weighted = 0
            all_done = True
            any_in_progress = False
            any_blocked = False
            for child in children[name]:
                cname = child.get("Project Part", "")
                dfs(cname)
                try:
                    dur = int(child.get("Duration (days)") or 0)
                except Exception:
                    dur = 0
                try:
                    cpc = int(child.get("% Complete") or 0)
                except Exception:
                    cpc = 0
                weighted += cpc * dur
                total_weight += dur
                st = child.get("Status") or "Planned"
                if st != "Done":
                    all_done = False
                if st == "In Progress":
                    any_in_progress = True
                if st == "Blocked":
                    any_blocked = True
            if total_weight > 0:
                row["% Complete"] = int(round(weighted / total_weight))
            else:
                # No duration children: average raw
                vals = []
                for child in children[name]:
                    try:
                        vals.append(int(child.get("% Complete") or 0))
                    except Exception:
                        pass
                row["% Complete"] = int(round(sum(vals)/len(vals))) if vals else 0
            # Derive parent status
            if all_done and children[name]:
                row["Status"] = "Done"
                row["% Complete"] = 100
            else:
                # Preserve explicit Blocked if all children blocked
                if any_blocked and not any_in_progress:
                    row["Status"] = "Blocked"
                elif any_in_progress:
                    row["Status"] = "In Progress"
                else:
                    # Keep existing or default
                    row["Status"] = row.get("Status") or "Planned"

        # Start DFS from top-level rows (no Parent or blank)
        for r in self.rows:
            if not (r.get("Parent") or ""):
                dfs(r.get("Project Part", ""))

    # --- Aggregate metrics helper for dashboard ---
    def load_sample_data(self):
        """Populate model & DB with a small sample hierarchy.
        Appends if existing rows present. Intended for first-run onboarding.
        """
        try:
            import datetime as _dt, random as _rnd
            base = _dt.date.today()
            # (name, start_offset_days, duration_days, parent_name, status, percent)
            spec = [
                ("Sample Project", 0, 30, None, "In Progress", 35),
                ("Planning", 0, 5, "Sample Project", "Done", 100),
                ("Design", 5, 7, "Sample Project", "In Progress", 50),
                ("Implementation", 12, 18, "Sample Project", "In Progress", 20),
                ("Module A", 12, 8, "Implementation", "In Progress", 40),
                ("Module B", 14, 10, "Implementation", "Planned", 0),
                ("Testing", 30, 6, "Sample Project", "Planned", 0),
                ("Rollout", 36, 4, None, "Planned", 0)
            ]
            # Build name->id map for parent linking (store part names as 'Parent')
            name_to_row = {r.get("Project Part",""): r for r in self.rows}
            for name, off, dur, parent, status, pc in spec:
                start = base + _dt.timedelta(days=off)
                # Convert to existing date string format MM-dd-YYYY
                start_str = start.strftime("%m-%d-%Y")
                # Duration stored as text
                end_calc = (start + _dt.timedelta(days=dur))
                end_str = end_calc.strftime("%m-%d-%Y")
                row = {col: "" for col in self.COLUMNS}
                row["Project Part"] = name
                row["Parent"] = parent
                row["Start Date"] = start_str
                row["Duration (days)"] = str(dur)
                row["Calculated End Date"] = end_str
                row["Status"] = status
                row["% Complete"] = str(pc)
                row["Notes"] = f"Autogenerated sample task '{name}'."
                # Minimal baseline capture
                row["Baseline Start Date"] = row["Start Date"]
                row["Baseline End Date"] = row["Calculated End Date"]
                self.rows.append(row)
            # Recalculate roll-ups & derived fields
            self.rollup_progress()
            self.update_calculated_end_dates()
            try:
                self.save_to_db()
            except Exception:
                pass
        except Exception as e:
            print(f"load_sample_data failed: {e}")

    def progress_metrics(self):
        import datetime
        total_tasks = 0
        sum_weighted = 0
        total_weight = 0
        critical_tasks = 0
        critical_weighted = 0
        critical_weight = 0
        done = 0
        today = datetime.datetime.today().date()
        overdue = 0
        at_risk = 0
        # Identify critical path quickly (reuse minimal logic)
        try:
            name_to_row = {r.get("Project Part", ""): r for r in self.rows}
            graph = {}
            duration_map = {}
            min_date = None
            for r in self.rows:
                n = r.get("Project Part", "")
                deps = [d.strip() for d in (r.get("Dependencies", "") or '').split(',') if d.strip()]
                graph[n] = deps
                try:
                    duration_map[n] = int(r.get("Duration (days)") or 0)
                except Exception:
                    duration_map[n] = 0
                try:
                    sd = r.get("Start Date", "")
                    if sd:
                        dt = datetime.datetime.strptime(sd, "%m-%d-%Y")
                        if min_date is None or dt < min_date:
                            min_date = dt
                except Exception:
                    pass
            visited = set(); order = []
            def dfs(n):
                if n in visited: return
                for d in graph.get(n, []): dfs(d)
                visited.add(n); order.append(n)
            for n in graph: dfs(n)
            earliest_finish = {}; earliest_start = {}
            base_min = min_date or datetime.datetime.today()
            for n in order:
                deps = graph.get(n, [])
                if not deps:
                    row = name_to_row.get(n, {})
                    try:
                        earliest_start[n] = datetime.datetime.strptime(row.get("Start Date", ""), "%m-%d-%Y")
                    except Exception:
                        earliest_start[n] = base_min
                else:
                    earliest_start[n] = max([earliest_finish.get(d, base_min) for d in deps])
                earliest_finish[n] = earliest_start[n] + datetime.timedelta(days=duration_map.get(n,0))
            project_finish = max(earliest_finish.values()) if earliest_finish else base_min
            latest_start = {}; latest_finish = {}
            for n in reversed(order):
                succs = [k for k,v in graph.items() if n in v]
                if not succs:
                    latest_finish[n] = project_finish
                else:
                    latest_finish[n] = min([latest_start[s] for s in succs]) if succs else project_finish
                latest_start[n] = latest_finish[n] - datetime.timedelta(days=duration_map.get(n,0))
            critical_set = {n for n in order if abs((earliest_start[n]-latest_start[n]).days) <= 0}
        except Exception:
            critical_set = set()
        for r in self.rows:
            # Skip parent aggregator tasks for EV style metrics: treat non-leaf if it has children with durations
            name = r.get("Project Part", "")
            has_child = any(ch.get("Parent", "") == name for ch in self.rows if ch is not r)
            try:
                dur = int(r.get("Duration (days)") or 0)
            except Exception:
                dur = 0
            try:
                pc = int(r.get("% Complete") or 0)
            except Exception:
                pc = 0
            status_val = (r.get("Status") or "").strip()
            if dur and not has_child:
                total_tasks += 1
                total_weight += dur
                sum_weighted += pc * dur
                if status_val == "Done":
                    done += 1
                # Overdue / at-risk logic (mirrors Gantt drawing)
                try:
                    end_calc = r.get("Calculated End Date", "")
                    if end_calc:
                        end_dt = datetime.datetime.strptime(end_calc, "%m-%d-%Y").date()
                    else:
                        start_dt = datetime.datetime.strptime(r.get("Start Date", ""), "%m-%d-%Y").date()
                        end_dt = start_dt + datetime.timedelta(days=dur)
                    if pc < 100 and today > end_dt:
                        overdue += 1
                    elif pc == 0 and today > start_dt and status_val in ("Planned", "Blocked"):
                        at_risk += 1
                except Exception:
                    pass
            if name in critical_set and dur and not has_child:
                critical_tasks += 1
                critical_weight += dur
                critical_weighted += pc * dur
        overall_pc = (sum_weighted/total_weight) if total_weight else 0
        critical_pc = (critical_weighted/critical_weight) if critical_weight else 0
        return {
            "overall_percent": round(overall_pc, 1),
            "critical_percent": round(critical_pc, 1),
            "leaf_count": total_tasks,
            "done_count": done,
            "overdue": overdue,
            "at_risk": at_risk,
            "critical_leaf_count": critical_tasks
        }

class ProgressDashboard(QWidget):
    def __init__(self, model):
        super().__init__()
        self.model = model
        from PyQt6.QtWidgets import QHBoxLayout, QCheckBox, QTableWidget, QTableWidgetItem, QHeaderView
        from PyQt6.QtWidgets import QMenu, QProgressBar
        self.vbox = QVBoxLayout(self)
        header_row = QHBoxLayout()
        title = QLabel("Progress Dashboard")
        title.setStyleSheet("font-weight:600; font-size:15px")
        header_row.addWidget(title)
        header_row.addStretch(1)
        # Export (CSV/PDF) menu button
        self.export_btn = QPushButton("Export")
        self.export_btn.setToolTip("Export a snapshot of current dashboard metrics (CSV / PDF / PNG)")
        menu = QMenu(self.export_btn)
        act_csv = menu.addAction("CSV Snapshot…")
        act_pdf = menu.addAction("PDF Snapshot…")
        act_png = menu.addAction("PNG Snapshot…")
        act_csv.triggered.connect(self._export_csv_snapshot)
        act_pdf.triggered.connect(self._export_pdf_snapshot)
        act_png.triggered.connect(self._export_png_snapshot)
        self.export_btn.setMenu(menu)
        header_row.addWidget(self.export_btn)
        self.auto_chk = QCheckBox("Auto-Refresh")
        self.auto_chk.setChecked(True)
        header_row.addWidget(self.auto_chk)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh)
        header_row.addWidget(refresh_btn)
        self.vbox.addLayout(header_row)
        # Summary metrics
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("QLabel { font-family: Consolas, monospace; }")
        self.vbox.addWidget(self.summary_label)
        # Progress gauges row
        gauges = QHBoxLayout()
        self.overall_bar = QProgressBar(); self.overall_bar.setRange(0,100); self.overall_bar.setFormat("Overall %p%")
        self.critical_bar = QProgressBar(); self.critical_bar.setRange(0,100); self.critical_bar.setFormat("Critical %p%")
        for bar in (self.overall_bar, self.critical_bar):
            bar.setMinimumWidth(180)
            bar.setToolTip("Visual gauge for progress percentage")
            gauges.addWidget(bar)
        gauges.addStretch(1)
        self.vbox.addLayout(gauges)
        # Status distribution table
        self.status_table = QTableWidget(0, 3)
        self.status_table.setHorizontalHeaderLabels(["Status","Count","% of Leaf Tasks"])
        try:
            self.status_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        except Exception:
            # Fallback for any environments still exposing legacy attribute
            self.status_table.horizontalHeader().setSectionResizeMode(0, getattr(QHeaderView, 'Stretch', 1))
        # PyQt6: enums moved under QHeaderView.ResizeMode; provide robust fallback
        try:
            resize_mode = QHeaderView.ResizeMode
            self.status_table.horizontalHeader().setSectionResizeMode(1, resize_mode.ResizeToContents)
            self.status_table.horizontalHeader().setSectionResizeMode(2, resize_mode.ResizeToContents)
        except Exception:
            # Fallback: attempt legacy attribute name or integer constant 2 (ResizeToContents)
            legacy_rtc = getattr(QHeaderView, 'ResizeToContents', 2)
            self.status_table.horizontalHeader().setSectionResizeMode(1, legacy_rtc)
            self.status_table.horizontalHeader().setSectionResizeMode(2, legacy_rtc)
        # PyQt6: Edit trigger enums live under QAbstractItemView.EditTrigger
        try:
            no_edit = QAbstractItemView.EditTrigger.NoEditTriggers
        except Exception:
            # Fallback to legacy attribute if still present
            no_edit = getattr(QAbstractItemView, 'NoEditTriggers', 0)
        self.status_table.setEditTriggers(no_edit)
        self.status_table.setAlternatingRowColors(True)
        self.vbox.addWidget(self.status_table)
        # Overdue / At-Risk lists
        lists_row = QHBoxLayout()
        self.overdue_label = QLabel()
        self.overdue_label.setText("Overdue: (none)")
        self.overdue_label.setStyleSheet("font-family: Consolas; font-size:12px")
        self.at_risk_label = QLabel("At Risk: (none)")
        self.at_risk_label.setStyleSheet("font-family: Consolas; font-size:12px")
        lists_row.addWidget(self.overdue_label,1)
        lists_row.addWidget(self.at_risk_label,1)
        self.vbox.addLayout(lists_row)
        # Critical path tasks list (simple comma string)
        self.critical_label = QLabel("Critical Path: (computing)")
        self.critical_label.setWordWrap(True)
        self.critical_label.setStyleSheet("font-family: Consolas; font-size:11px")
        self.vbox.addWidget(self.critical_label)
        # Simple history sparkline (store last N overall % values)
        self._history = []
        self.history_label = QLabel("History: -")
        self.history_label.setStyleSheet("font-family: Consolas; font-size:11px")
        self.vbox.addWidget(self.history_label)
        # Internal caches for export
        self._status_counts = {}
        self._last_overdue_list = []
        self._last_at_risk_list = []
        self._last_critical_path = []
        # Timer for auto refresh
        from PyQt6.QtCore import QTimer
        self._timer = QTimer(self)
        self._timer.setInterval(5000)
        self._timer.timeout.connect(self._auto_tick)
        self._timer.start()
        self.refresh()

    def _auto_tick(self):
        if self.auto_chk.isChecked():
            self.refresh()

    def refresh(self):
        m = self.model.progress_metrics()
        # Update history (keep last 25 points)
        try:
            self._history.append(m['overall_percent'])
            if len(self._history) > 25:
                self._history = self._history[-25:]
        except Exception:
            pass
        hist_str = ''.join(self._spark_chars(self._history)) if self._history else '-'
        self.history_label.setText(f"History: {hist_str}")
        self.summary_label.setText(
            f"Overall %: {m['overall_percent']}%  |  Critical %: {m['critical_percent']}%  |  Leaf: {m['leaf_count']}  Done: {m['done_count']}  Overdue: {m['overdue']}  At Risk: {m['at_risk']}  Critical Leafs: {m['critical_leaf_count']}"
        )
        # Gauges
        try:
            self.overall_bar.setValue(int(round(m['overall_percent'])))
            self.critical_bar.setValue(int(round(m['critical_percent'])))
        except Exception:
            pass
        self._populate_status_distribution()
        self._populate_overdue_lists()
        self._populate_critical_path()

    def _spark_chars(self, vals):
        # Unicode sparkline blocks (▁▂▃▄▅▆▇█)
        if not vals:
            return []
        mn = min(vals); mx = max(vals); span = (mx - mn) or 1.0
        blocks = ['▁','▂','▃','▄','▅','▆','▇','█']
        out=[]
        for v in vals:
            idx = int(round((v - mn)/span * (len(blocks)-1)))
            out.append(blocks[idx])
        return out

    def _populate_status_distribution(self):
        try:
            status_counts = {}
            leaf_set = set()
            for r in self.model.rows:
                name = r.get('Project Part','')
                has_child = any(ch.get('Parent','') == name for ch in self.model.rows if ch is not r)
                if has_child:
                    continue
                leaf_set.add(name)
                st = (r.get('Status') or 'Planned').strip() or 'Planned'
                status_counts[st] = status_counts.get(st,0)+1
            total_leaf = len(leaf_set) or 1
            rows = sorted(status_counts.items())
            self.status_table.setRowCount(len(rows))
            for i,(st,c) in enumerate(rows):
                self.status_table.setItem(i,0,QTableWidgetItem(st))
                self.status_table.setItem(i,1,QTableWidgetItem(str(c)))
                pct = f"{(c/total_leaf)*100:.1f}%"
                self.status_table.setItem(i,2,QTableWidgetItem(pct))
            self._status_counts = dict(status_counts)
        except Exception:
            pass

    def _populate_overdue_lists(self):
        import datetime
        overdue=[]; at_risk=[]
        today = datetime.date.today()
        for r in self.model.rows:
            name = r.get('Project Part','')
            if not name: continue
            has_child = any(ch.get('Parent','') == name for ch in self.model.rows if ch is not r)
            if has_child: continue
            try:
                dur = int(r.get('Duration (days)') or 0)
            except Exception:
                dur = 0
            try:
                pc = int(r.get('% Complete') or 0)
            except Exception:
                pc = 0
            status_val = (r.get('Status') or '').strip()
            try:
                if r.get('Calculated End Date'):
                    end_dt = datetime.datetime.strptime(r.get('Calculated End Date'),'%m-%d-%Y').date()
                else:
                    start_dt = datetime.datetime.strptime(r.get('Start Date') or '', '%m-%d-%Y').date()
                    end_dt = start_dt + datetime.timedelta(days=dur)
            except Exception:
                end_dt = None
            try:
                start_dt = datetime.datetime.strptime(r.get('Start Date') or '', '%m-%d-%Y').date()
            except Exception:
                start_dt = None
            if pc < 100 and end_dt and today > end_dt:
                overdue.append(name)
            elif pc == 0 and start_dt and today > start_dt and status_val in ('Planned','Blocked'):
                at_risk.append(name)
        self.overdue_label.setText("Overdue: " + (", ".join(overdue[:12]) + ("…" if len(overdue)>12 else "") if overdue else "(none)"))
        self.at_risk_label.setText("At Risk: " + (", ".join(at_risk[:12]) + ("…" if len(at_risk)>12 else "") if at_risk else "(none)"))
        self._last_overdue_list = list(overdue)
        self._last_at_risk_list = list(at_risk)

    def _populate_critical_path(self):
        # Reuse quick critical set derivation from progress_metrics (duplicate minimal logic to avoid hidden coupling)
        import datetime
        try:
            name_to_row = {r.get('Project Part',''): r for r in self.model.rows}
            graph={}; duration={}
            for r in self.model.rows:
                n=r.get('Project Part','');
                deps=[d.strip() for d in (r.get('Dependencies') or '').split(',') if d.strip()]
                graph[n]=deps
                try: duration[n]=int(r.get('Duration (days)') or 0)
                except Exception: duration[n]=0
            visited=set(); order=[]
            def dfs(n):
                if n in visited: return
                for d in graph.get(n,[]): dfs(d)
                visited.add(n); order.append(n)
            for n in graph: dfs(n)
            earliest_finish={}; earliest_start={}
            base=datetime.datetime.today()
            for n in order:
                deps=graph.get(n,[])
                if not deps:
                    try:
                        earliest_start[n]=datetime.datetime.strptime(name_to_row.get(n,{}).get('Start Date',''),'%m-%d-%Y')
                    except Exception: earliest_start[n]=base
                else:
                    earliest_start[n]=max([earliest_finish.get(d,base) for d in deps])
                earliest_finish[n]=earliest_start[n]+datetime.timedelta(days=duration.get(n,0))
            project_finish = max(earliest_finish.values()) if earliest_finish else base
            latest_start={}; latest_finish={}
            for n in reversed(order):
                succ=[k for k,v in graph.items() if n in v]
                if not succ:
                    latest_finish[n]=project_finish
                else:
                    latest_finish[n]=min([latest_start[s] for s in succ]) if succ else project_finish
                latest_start[n]=latest_finish[n]-datetime.timedelta(days=duration.get(n,0))
            critical = [n for n in order if abs((earliest_start[n]-latest_start[n]).days) <= 0]
            self.critical_label.setText("Critical Path: " + (" → ".join(critical) if critical else "(none)"))
            self._last_critical_path = list(critical)
        except Exception:
            self.critical_label.setText("Critical Path: (error)")
            self._last_critical_path = []

    # --- Export helpers -----------------------------------------------------
    def _export_csv_snapshot(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import csv, datetime, json
        path, _ = QFileDialog.getSaveFileName(self, "Export Dashboard Snapshot (CSV)", "dashboard_snapshot.csv", "CSV Files (*.csv)")
        if not path:
            return
        try:
            now = datetime.datetime.now().isoformat(timespec='seconds')
            # Refresh once more to ensure latest values
            self.refresh()
            metrics = self.model.progress_metrics()
            with open(path, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(["Metric","Value"])
                for k,v in metrics.items():
                    w.writerow([k, v])
                # Status distribution
                w.writerow([]); w.writerow(["Status Distribution",""])            
                for st,count in sorted(self._status_counts.items()):
                    w.writerow([f"status:{st}", count])
                # Lists
                w.writerow([]); w.writerow(["Overdue Tasks", "; ".join(self._last_overdue_list) or "<none>"])
                w.writerow(["At Risk Tasks", "; ".join(self._last_at_risk_list) or "<none>"])
                w.writerow(["Critical Path", " -> ".join(self._last_critical_path) or "<none>"])
                # Raw JSON for automation
                snapshot = {
                    'timestamp': now,
                    'metrics': metrics,
                    'status_counts': self._status_counts,
                    'overdue': self._last_overdue_list,
                    'at_risk': self._last_at_risk_list,
                    'critical_path': self._last_critical_path
                }
                w.writerow([]); w.writerow(["snapshot_json", json.dumps(snapshot)])
            if self.parent() and self.parent().window().statusBar():
                self.parent().window().statusBar().showMessage(f"Dashboard CSV exported: {path}", 4000)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    def _export_pdf_snapshot(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from PyQt6.QtGui import QPainter
        try:
            from PyQt6.QtPrintSupport import QPrinter
        except Exception:
            QPrinter = None
        import datetime, os
        path, _ = QFileDialog.getSaveFileName(self, "Export Dashboard Snapshot (PDF)", "dashboard_snapshot.pdf", "PDF Files (*.pdf)")
        if not path:
            return
        if not path.lower().endswith('.pdf'):
            path += '.pdf'
        try:
            # Make sure data is fresh
            self.refresh()
            if QPrinter is None:
                raise RuntimeError("QPrinter not available in this build")
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setFullPage(True)
            printer.setOutputFileName(path)
            # Margins (mm) from export settings if present
            try:
                from PyQt6.QtCore import QSettings
                s = QSettings('LSI','ProjectPlanner')
                ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
            except Exception:
                ml=mr=mt=mb=8.0
            painter = QPainter(printer)
            try:
                # Convert mm margins to device units
                dpi = printer.resolution()
                def mm(v): return v/25.4 * dpi
                page_rect = printer.pageRect()
                avail_w = page_rect.width() - (mm(ml)+mm(mr))
                avail_h = page_rect.height() - (mm(mt)+mm(mb))
                # Grab current widget pixmap
                pix = self.grab()
                if pix.isNull():
                    raise RuntimeError('Failed to grab dashboard contents')
                scale = min(avail_w / pix.width(), avail_h / pix.height())
                target_w = pix.width()*scale
                target_h = pix.height()*scale
                tx = page_rect.left() + mm(ml) + (avail_w - target_w)/2
                ty = page_rect.top() + mm(mt)
                painter.drawPixmap(int(tx), int(ty), int(target_w), int(target_h), pix)
                # Footer timestamp
                ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                painter.setPen(Qt.black)
                painter.drawText(int(page_rect.left()+mm(ml)), int(page_rect.bottom()-mm(mb/2)), f"Dashboard Snapshot – {ts}")
            finally:
                painter.end()
            if self.parent() and self.parent().window().statusBar():
                self.parent().window().statusBar().showMessage(f"Dashboard PDF exported: {path}", 4000)
        except Exception as e:
            try:
                QMessageBox.critical(self, "Export Failed", str(e))
            except Exception:
                pass

    def _export_png_snapshot(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import datetime
        path, _ = QFileDialog.getSaveFileName(self, "Export Dashboard Snapshot (PNG)", "dashboard_snapshot.png", "PNG Files (*.png)")
        if not path:
            return
        if not path.lower().endswith('.png'):
            path += '.png'
        try:
            self.refresh()
            pix = self.grab()
            if pix.isNull():
                raise RuntimeError('Failed to grab dashboard contents')
            # Write directly; optionally annotate timestamp in future
            if not pix.save(path, 'PNG'):
                raise RuntimeError('Save failed')
            if self.parent() and self.parent().window().statusBar():
                self.parent().window().statusBar().showMessage(f"Dashboard PNG exported: {path}", 4000)
        except Exception as e:
            try:
                QMessageBox.critical(self, "Export Failed", str(e))
            except Exception:
                pass

# --- Conflict Resolution Dialog -------------------------------------------------
class ConflictResolutionDialog(QDialog):
    """Dialog shown when an optimistic concurrency update detects a conflict.

    Presents side-by-side comparison of each changed field:
      - Local Pending (user edits)
      - Remote Current (fresh DB snapshot)
      - Original Local (value before edit, if supplied)

    Actions:
      - Keep Remote (discard local pending)
      - Overwrite Remote (force save local pending using new expected version)
      - Merge Field-by-Field (user selects which side per field) then save
    """
    def __init__(self, part_name: str, original: dict, pending: dict, remote: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Conflict – {part_name}")
        self.part_name = part_name
        self.original = original or {}
        self.pending = pending or {}
        self.remote = remote or {}
        self.merged = dict(self.remote)  # start from remote baseline
        from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QScrollArea, QWidget, QGridLayout, QRadioButton, QButtonGroup
        layout = QVBoxLayout(self)
        info = QLabel("Another user modified this row before your save completed. Resolve each differing field.")
        info.setWordWrap(True)
        layout.addWidget(info)
        # Scrollable area for fields
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        container = QWidget(); grid = QGridLayout(container)
        headers = ["Field","Original","Remote","Pending","Use"]
        for c,h in enumerate(headers):
            lbl = QLabel(f"<b>{h}</b>")
            grid.addWidget(lbl,0,c)
        self._choice_groups = {}
        diff_fields = sorted(set(list(self.pending.keys()) + list(self.remote.keys())))
        row_idx = 1
        for field in diff_fields:
            if field == 'Project Part':
                continue
            orig_v = self.original.get(field)
            remote_v = self.remote.get(field)
            pend_v = self.pending.get(field)
            # Only show if there is a meaningful difference
            if str(remote_v) == str(pend_v):
                continue
            grid.addWidget(QLabel(field), row_idx, 0)
            grid.addWidget(QLabel(str(orig_v) if orig_v not in (None,'') else '—'), row_idx, 1)
            grid.addWidget(QLabel(str(remote_v) if remote_v not in (None,'') else '—'), row_idx, 2)
            grid.addWidget(QLabel(str(pend_v) if pend_v not in (None,'') else '—'), row_idx, 3)
            grp = QButtonGroup(self)
            r_remote = QRadioButton("Remote")
            r_local = QRadioButton("Local")
            r_remote.setChecked(True)
            grp.addButton(r_remote, 0); grp.addButton(r_local, 1)
            hl = QHBoxLayout(); hl.addWidget(r_remote); hl.addWidget(r_local); hl.addStretch(1)
            cell = QWidget(); cell.setLayout(hl)
            grid.addWidget(cell, row_idx, 4)
            self._choice_groups[field] = grp
            row_idx += 1
        if row_idx == 1:
            # No differing fields – fallback simple notice
            grid.addWidget(QLabel("No differing fields – you can keep remote copy."), row_idx, 0, 1, 5)
        scroll.setWidget(container)
        layout.addWidget(scroll,1)
        # Buttons
        btn_row = QHBoxLayout()
        self.btn_keep = QPushButton("Keep Remote")
        self.btn_overwrite = QPushButton("Overwrite Remote")
        self.btn_merge = QPushButton("Merge & Save")
        self.btn_cancel = QPushButton("Cancel")
        btn_row.addWidget(self.btn_keep)
        btn_row.addWidget(self.btn_overwrite)
        btn_row.addWidget(self.btn_merge)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_cancel)
        layout.addLayout(btn_row)
        self.choice = None  # 'keep' | 'overwrite' | 'merge'
        self.btn_keep.clicked.connect(self._keep)
        self.btn_overwrite.clicked.connect(self._overwrite)
        self.btn_merge.clicked.connect(self._merge)
        self.btn_cancel.clicked.connect(self.reject)
        self.setMinimumSize(760, 420)
    def _keep(self):
        self.choice = 'keep'
        self.accept()
    def _overwrite(self):
        self.choice = 'overwrite'
        self.accept()
    def _merge(self):
        # Build merged dict starting from remote + field-level picks
        merged = dict(self.remote)
        for field, grp in self._choice_groups.items():
            sel = grp.checkedId()  # 0 remote, 1 local
            if sel == 1:  # local/pending
                merged[field] = self.pending.get(field)
        self.merged = merged
        self.choice = 'merge'
        self.accept()

class CostEstimatesView(QWidget):
    """Enhanced cost & margin estimation view.
    Columns:
      Project Part | Parent | Prod Cost | Inst Cost | Total Cost | Prod Price | Inst Price | Total Price | Profit $ | Margin % | % of Total Price
    Features:
      - Aggregated totals (leaf-only or all rows)
      - Filtering (name substring, internal/external, min total price)
      - Highlight top-N expensive parts & low/negative margins
      - CSV export
    """
    def __init__(self, model):
        super().__init__()
        self.model = model
        from PyQt6.QtWidgets import QVBoxLayout, QTableWidget, QTableWidgetItem, QLabel, QPushButton, QHBoxLayout, QLineEdit, QComboBox, QSpinBox, QFileDialog, QCheckBox
        self.vbox = QVBoxLayout(self)
        # --- Header / controls ---
        header = QHBoxLayout()
        title = QLabel("Cost & Margin Estimates")
        title.setStyleSheet("font-weight:600; font-size:15px")
        header.addWidget(title)
        header.addStretch(1)
        self.le_filter = QLineEdit(); self.le_filter.setPlaceholderText("Filter name…")
        self.le_filter.textChanged.connect(self.refresh)
        header.addWidget(self.le_filter)
        self.combo_int_ext = QComboBox(); self.combo_int_ext.addItems(["All","Internal","External"]); self.combo_int_ext.currentIndexChanged.connect(self.refresh)
        header.addWidget(self.combo_int_ext)
        self.min_price_spin = QSpinBox(); self.min_price_spin.setRange(0, 10_000_000); self.min_price_spin.setPrefix(">$ "); self.min_price_spin.setSingleStep(500)
        self.min_price_spin.setToolTip("Minimum total price filter")
        self.min_price_spin.valueChanged.connect(self.refresh)
        header.addWidget(self.min_price_spin)
        self.chk_leaf_only = QCheckBox("Leaf Only")
        self.chk_leaf_only.setToolTip("Show leaf rows only (exclude parent aggregators)")
        self.chk_leaf_only.stateChanged.connect(self.refresh)
        header.addWidget(self.chk_leaf_only)
        self.chk_rollup = QCheckBox("Roll-up Parents")
        self.chk_rollup.setToolTip("Aggregate descendant costs & prices into parent rows (shown even if leaf-only mode).")
        self.chk_rollup.stateChanged.connect(self.refresh)
        header.addWidget(self.chk_rollup)
        self.chk_compact = QCheckBox("Compact")
        self.chk_compact.stateChanged.connect(self._apply_compact_mode)
        header.addWidget(self.chk_compact)
        export_btn = QPushButton("Export…")
        export_btn.setToolTip("Open unified export dialog (CSV / XLSX / PDF / PNG)")
        export_btn.clicked.connect(self._open_export_dialog)
        header.addWidget(export_btn)
        self.chk_selected_only = QCheckBox("Selected Only")
        self.chk_selected_only.setToolTip("When checked, only selected table rows are exported (CSV/XLSX/PDF/PNG). If none selected, falls back to all.")
        header.addWidget(self.chk_selected_only)
    # (Legacy individual export buttons removed in favor of unified dialog)
        # Column visibility / layouts button
        self.columns_btn = QPushButton("Columns…")
        self.columns_btn.setToolTip("Show/hide columns and manage saved visibility layouts")
        self.columns_btn.clicked.connect(self._open_columns_dialog)
        header.addWidget(self.columns_btn)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh)
        header.addWidget(refresh_btn)
        self.vbox.addLayout(header)
        # Table
        # Version comparison controls
        self.version_combo = QComboBox(); self.version_combo.addItem("<None>")
        self.version_combo.setToolTip("Select a saved quote version to compare deltas")
        self.version_combo.currentIndexChanged.connect(self.refresh)
        self.freeze_btn = QPushButton("Freeze Version…")
        self.delete_version_btn = QPushButton("Delete Version")
        self.rename_version_btn = QPushButton("Rename Version")
        def do_freeze():
            from PyQt6.QtWidgets import QInputDialog, QMessageBox
            name, ok = QInputDialog.getText(self, "Freeze Quote Version", "Version name:")
            if ok and name.strip():
                try:
                    ver_name = name.strip()
                    existing = set(self.model.list_quote_versions())
                    if ver_name in existing:
                        resp = QMessageBox.question(self, "Overwrite Version?", f"A version named '{ver_name}' already exists. Overwrite it?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if resp != QMessageBox.Yes:
                            return
                        # Delete existing first so save acts like replace
                        try:
                            self.model.delete_quote_version(ver_name)
                        except Exception:
                            pass
                    self.model.save_quote_version(ver_name)
                    self._reload_versions()
                    self.version_combo.setCurrentText(ver_name)
                    if self.parent() and self.parent().window().statusBar():
                        action = "overwritten" if ver_name in existing else "saved"
                        self.parent().window().statusBar().showMessage(f"Quote version '{ver_name}' {action}", 3000)
                except Exception as e:
                    QMessageBox.critical(self, "Freeze Failed", str(e))
        self.freeze_btn.clicked.connect(do_freeze)
        def do_delete_version():
            from PyQt6.QtWidgets import QMessageBox
            ver = self.version_combo.currentText()
            if not ver or ver in ("<None>",""):
                return
            resp = QMessageBox.question(self, "Delete Quote Version", f"Delete version '{ver}'? This cannot be undone.", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if resp == QMessageBox.Yes:
                if self.model.delete_quote_version(ver):
                    if self.parent() and self.parent().window().statusBar():
                        self.parent().window().statusBar().showMessage(f"Deleted quote version '{ver}'",3000)
                    self._reload_versions()
                else:
                    QMessageBox.critical(self, "Delete Failed", f"Could not delete version '{ver}'.")
        self.delete_version_btn.clicked.connect(do_delete_version)
        def do_rename_version():
            from PyQt6.QtWidgets import QInputDialog, QMessageBox
            cur = self.version_combo.currentText()
            if not cur or cur in ("<None>", ""):
                return
            new_name, ok = QInputDialog.getText(self, "Rename Quote Version", f"Rename '{cur}' to:")
            if not ok or not new_name.strip() or new_name.strip() == cur:
                return
            new_name = new_name.strip()
            # If destination exists ask to overwrite via delete+rename path
            existing = set(self.model.list_quote_versions())
            if new_name in existing:
                resp = QMessageBox.question(self, "Overwrite Existing?", f"A version named '{new_name}' exists. Overwrite it with '{cur}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if resp != QMessageBox.Yes:
                    return
                # Delete new_name then rename
                try:
                    self.model.delete_quote_version(new_name)
                except Exception:
                    pass
            if self.model.rename_quote_version(cur, new_name):
                self._reload_versions()
                self.version_combo.setCurrentText(new_name)
                if self.parent() and self.parent().window().statusBar():
                    self.parent().window().statusBar().showMessage(f"Renamed version '{cur}' → '{new_name}'", 4000)
            else:
                QMessageBox.critical(self, "Rename Failed", f"Could not rename '{cur}' to '{new_name}'.")
        self.rename_version_btn.clicked.connect(do_rename_version)
        header.addWidget(self.version_combo)
        header.addWidget(self.freeze_btn)
        header.addWidget(self.delete_version_btn)
        header.addWidget(self.rename_version_btn)
        self.table = QTableWidget(0, 14)
        self.table.setHorizontalHeaderLabels([
            "Project Part","Parent","Prod Cost","Inst Cost","Total Cost","Prod Price","Inst Price","Total Price","Profit $","Margin %","Δ Price %","Δ Margin pts","% of Total Price","Internal/External"
        ])
        try:
            no_edit2 = QAbstractItemView.EditTrigger.NoEditTriggers
        except Exception:
            no_edit2 = getattr(QAbstractItemView, 'NoEditTriggers', 0)
        self.table.setEditTriggers(no_edit2)
        try:
            sel_rows = QAbstractItemView.SelectionBehavior.SelectRows
        except Exception:
            sel_rows = getattr(QAbstractItemView, 'SelectRows', 0)
        self.table.setSelectionBehavior(sel_rows)
        self.table.setAlternatingRowColors(True)
        try:
            self.table.horizontalHeader().setSectionsMovable(True)
            self.table.horizontalHeader().setSectionsClickable(True)
        except Exception:
            pass
        self.vbox.addWidget(self.table, 1)
        # Totals footer
        self.totals_label = QLabel()
        self.totals_label.setStyleSheet("font-family:Consolas,monospace; font-size:12px; margin-top:4px")
        self.vbox.addWidget(self.totals_label)
        # Restore previously used column layout if any
        self._apply_last_layout()
        self.refresh()

    def _num(self, v):
        try:
            if v is None: return 0.0
            if isinstance(v,(int,float)): return float(v)
            s=str(v).strip().replace('$','').replace(',','')
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    def _is_leaf(self, row):
        name = row.get("Project Part","")
        return not any(r.get("Parent") == name for r in self.model.rows if r is not row)

    def _apply_compact_mode(self):
        try:
            if self.chk_compact.isChecked():
                self.table.verticalHeader().setDefaultSectionSize(18)
                self.table.setStyleSheet("QTableWidget { font-size:11px; }")
            else:
                self.table.verticalHeader().setDefaultSectionSize(24)
                self.table.setStyleSheet("")
        except Exception:
            pass

    def _export_csv(self):
        from PyQt6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(self, "Export Cost Table", "cost_estimates.csv", "CSV Files (*.csv)")
        if not path:
            return
        import csv
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
                writer.writerow(headers)
                selected_rows = set()
                if hasattr(self, 'chk_selected_only') and self.chk_selected_only.isChecked():
                    selected_rows = {idx.row() for idx in self.table.selectionModel().selectedRows()}
                # Fallback to all rows if none selected
                if selected_rows:
                    row_iter = sorted(selected_rows)
                else:
                    row_iter = range(self.table.rowCount())
                for r in row_iter:
                    writer.writerow([self.table.item(r,c).text() if self.table.item(r,c) else '' for c in range(self.table.columnCount())])
            print(f"Exported CSV -> {path}")
        except Exception as e:
            print(f"CSV export failed: {e}")

    def _open_export_dialog(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QCheckBox, QComboBox, QFileDialog, QMessageBox
        dlg = QDialog(self); dlg.setWindowTitle("Export Cost Data")
        v = QVBoxLayout(dlg)
        v.addWidget(QLabel("Choose an export format. Options honor current filters, visible columns, and (optionally) selected rows."))
        # Format choices
        fmt_row = QHBoxLayout(); v.addLayout(fmt_row)
        fmt_label = QLabel("Format:"); fmt_row.addWidget(fmt_label)
        fmt_combo = QComboBox(); fmt_combo.addItems(["CSV","XLSX","PDF","PNG"]); fmt_row.addWidget(fmt_combo,1)
        # Selected only toggle mirrors main checkbox
        sel_chk = QCheckBox("Selected Only")
        sel_chk.setChecked(hasattr(self,'chk_selected_only') and self.chk_selected_only.isChecked())
        v.addWidget(sel_chk)
        # Include header (only relevant for PDF/PNG)
        hdr_chk = QCheckBox("Include Header Banner (PDF/PNG)")
        hdr_chk.setChecked(True)
        v.addWidget(hdr_chk)
        # Buttons
        btn_row = QHBoxLayout(); v.addLayout(btn_row)
        export_btn = QPushButton("Export")
        cancel_btn = QPushButton("Cancel")
        btn_row.addStretch(1); btn_row.addWidget(export_btn); btn_row.addWidget(cancel_btn)
        cancel_btn.clicked.connect(dlg.reject)
        def do_export():
            fmt = fmt_combo.currentText()
            # propagate selection choice back to main toggle for reuse in lower-level exporters
            if hasattr(self,'chk_selected_only'):
                self.chk_selected_only.setChecked(sel_chk.isChecked())
            if fmt == 'CSV':
                dlg.accept(); self._export_csv(); return
            if fmt == 'XLSX':
                dlg.accept(); self._export_xlsx(); return
            if fmt in ('PDF','PNG'):
                # Temporarily set export format preferences via QSettings so _export_render uses them
                from PyQt6.QtCore import QSettings
                s = QSettings('LSI','ProjectPlanner')
                s.setValue('Export/format', fmt)
                s.setValue('Export/include_header', hdr_chk.isChecked())
                dlg.accept(); self._export_render(); return
            QMessageBox.warning(self, 'Unsupported', f'Format {fmt} not implemented.')
        export_btn.clicked.connect(do_export)
        # Finalize and run dialog
        dlg.setLayout(v)
        dlg.exec()

    def _export_render(self):
        """Render the QTableWidget to PDF or PNG using export settings & header/footer branding."""
        from PyQt6.QtCore import QSettings, QRectF
        from PyQt6.QtWidgets import QFileDialog, QApplication
        from PyQt6.QtGui import QPixmap, QPainter
        import os
        s = QSettings('LSI','ProjectPlanner')
        fmt = s.value('Export/format','PDF')
        include_header = s.value('Export/include_header', True)
        if isinstance(include_header, str): include_header = include_header.lower() in ('1','true','yes','on')
        ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
        init_name = f"cost_estimates.{ 'pdf' if fmt=='PDF' else 'png'}"
        filters = 'PDF Files (*.pdf);;PNG Files (*.png)' if fmt=='PDF' else 'PNG Files (*.png);;PDF Files (*.pdf)'
        path, chosen = QFileDialog.getSaveFileName(self, 'Export Cost Estimates', init_name, filters)
        if not path: return
        is_pdf = path.lower().endswith('.pdf') or (chosen and 'PDF' in chosen and not path.lower().endswith('.png'))
        if is_pdf and not path.lower().endswith('.pdf'): path += '.pdf'
        if (not is_pdf) and not path.lower().endswith('.png'): path += '.png'
        # Ensure columns sized to contents for export
        try: self.table.resizeColumnsToContents()
        except Exception: pass
        # Compute total table size (sum of column widths + header height + row heights)
        h_header = self.table.horizontalHeader(); v_header = self.table.verticalHeader()
        width = sum(self.table.columnWidth(c) for c in range(self.table.columnCount())) + v_header.width()
        # Determine row set (Selected Only logic)
        selected_rows = set()
        if hasattr(self, 'chk_selected_only') and self.chk_selected_only.isChecked():
            try:
                selected_rows = {idx.row() for idx in self.table.selectionModel().selectedRows()}
            except Exception:
                selected_rows = set()
        if selected_rows:
            row_iter = sorted(selected_rows)
        else:
            row_iter = range(self.table.rowCount())
        height = h_header.height() + sum(self.table.rowHeight(r) for r in row_iter)
        if width <=0 or height <=0:
            print('Empty table; abort export.')
            return
        header_path = resolve_resource_path('header.png')
        header_pix = QPixmap(header_path) if os.path.exists(header_path) else None
        svg_path = resolve_resource_path('header.svg')
        header_is_svg=False; header_svg_renderer=None
        try:
            if os.path.exists(svg_path):
                from PyQt6.QtSvg import QSvgRenderer
                r = QSvgRenderer(svg_path)
                if r.isValid(): header_is_svg=True; header_svg_renderer=r
        except Exception:
            pass
        # Footer metadata: timestamp + preset summary (page size, orientation, margins)
        try:
            import datetime as _dt_footer
            page_size = s.value('Export/page_size','A4')
            orientation = s.value('Export/orientation','Portrait')
            ml_s, mt_s, mr_s, mb_s = s.value('Export/margin_left_mm',8.0), s.value('Export/margin_top_mm',8.0), s.value('Export/margin_right_mm',8.0), s.value('Export/margin_bottom_mm',8.0)
            try:
                ml_s = float(ml_s); mt_s = float(mt_s); mr_s = float(mr_s); mb_s = float(mb_s)
            except Exception:
                ml_s, mt_s, mr_s, mb_s = 8.0, 8.0, 8.0, 8.0
            ts = _dt_footer.datetime.now().strftime('%Y-%m-%d %H:%M')
            preset = f"{page_size} {orientation}, margins {ml_s:.1f}/{mt_s:.1f}/{mr_s:.1f}/{mb_s:.1f} mm"
            footer_text = f"Cost Estimates • {ts} • {preset}"
        except Exception:
            footer_text = "© 2025 LSI – For Internal Use Only"
        if is_pdf:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtCore import QMarginsF
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.PdfFormat)
            page_size = s.value('Export/page_size','A4'); orientation = s.value('Export/orientation','Portrait')
            size_map={'A4':QPrinter.A4,'Letter':QPrinter.Letter,'Legal':QPrinter.Legal,'Tabloid':QPrinter.Tabloid}
            printer.setPaperSize(size_map.get(page_size, QPrinter.A4))
            printer.setOrientation(QPrinter.Portrait if orientation=='Portrait' else QPrinter.Landscape)
            try: printer.setPageMargins(QMarginsF(ml,mt,mr,mb))
            except Exception: pass
            painter = QPainter(printer)
            page_rect = printer.pageRect()
            # Header height determination
            def _svg_h(renderer, tw, default_ratio=0.12, min_h=40):
                try:
                    ds=renderer.defaultSize(); w,h=ds.width(), ds.height()
                    if w<=0 or h<=0:
                        vb=renderer.viewBoxF(); w,h=vb.width(), vb.height()
                    if w>0 and h>0 and tw>0:
                        return int(round(h*(tw/float(w))))
                except Exception:
                    pass
                return max(min_h, int(round(tw*default_ratio))) if tw>0 else None
            y_offset=0
            if include_header:
                if header_is_svg and header_svg_renderer:
                    hh=_svg_h(header_svg_renderer, page_rect.width())
                    if hh: header_svg_renderer.render(painter, QRectF(0,0,page_rect.width(),hh)); y_offset=hh+8
                elif header_pix and not header_pix.isNull():
                    try:
                        _smooth = Qt.TransformationMode.SmoothTransformation
                    except Exception:
                        _smooth = getattr(Qt, 'SmoothTransformation', 1)
                    sh=header_pix.scaledToWidth(page_rect.width(), _smooth); painter.drawPixmap((page_rect.width()-sh.width())//2,0,sh); y_offset=sh.height()+8
            # Pagination: draw rows until page full, then newPage
            cur_y = y_offset
            from PyQt6.QtGui import QFont
            font = painter.font(); fm = painter.fontMetrics()
            row_h = fm.height()+4
            # Draw header row
            def draw_header(y):
                painter.save(); painter.setFont(font)
                x=0
                for c in range(self.table.columnCount()):
                    w = self.table.columnWidth(c)
                    txt = self.table.horizontalHeaderItem(c).text()
                    painter.drawText(x+2, y+fm.ascent()+2, txt)
                    x += w
                painter.restore()
            def draw_row(r, y):
                painter.save(); x=0
                for c in range(self.table.columnCount()):
                    w = self.table.columnWidth(c)
                    it = self.table.item(r,c)
                    txt = it.text() if it else ''
                    align_right = (c>=2 and c not in (9,11))
                    if align_right:
                        tw = fm.width(txt); painter.drawText(x+w-tw-2, y+fm.ascent()+2, txt)
                    else:
                        painter.drawText(x+2, y+fm.ascent()+2, txt)
                    x += w
                painter.restore()
            draw_header(cur_y); cur_y += row_h
            for r in row_iter:
                if cur_y + row_h + 24 > page_rect.height():  # leave room for footer
                    # footer before new page
                    try:
                        painter.save(); f=QFont(font); f.setPointSizeF(f.pointSizeF()*0.85); painter.setFont(f)
                        painter.drawText(QRectF(0,page_rect.height()-18,page_rect.width(),16), Qt.AlignmentFlag.AlignCenter, footer_text)
                        painter.restore()
                    except Exception: pass
                    printer.newPage(); cur_y=0
                    if include_header:
                        if header_is_svg and header_svg_renderer:
                            hh=_svg_h(header_svg_renderer, page_rect.width())
                            if hh: header_svg_renderer.render(painter, QRectF(0,0,page_rect.width(),hh)); cur_y=hh+8
                        elif header_pix and not header_pix.isNull():
                            try:
                                _smooth = Qt.TransformationMode.SmoothTransformation
                            except Exception:
                                _smooth = getattr(Qt, 'SmoothTransformation', 1)
                            sh=header_pix.scaledToWidth(page_rect.width(), _smooth); painter.drawPixmap((page_rect.width()-sh.width())//2,0,sh); cur_y=sh.height()+8
                    draw_header(cur_y); cur_y += row_h
                draw_row(r, cur_y); cur_y += row_h
            # Final footer
            try:
                painter.save(); f=QFont(font); f.setPointSizeF(f.pointSizeF()*0.85); painter.setFont(f)
                painter.drawText(QRectF(0,page_rect.height()-18,page_rect.width(),16), Qt.AlignmentFlag.AlignCenter, footer_text)
                painter.restore()
            except Exception: pass
            painter.end(); print(f'Exported PDF -> {path}'); return
        # PNG branch: render full table to pixmap (single tall image)
        screen = QApplication.primaryScreen(); dpi = screen.logicalDotsPerInch() if screen else 96.0
        def mm_to_px(mm): return int(round((mm/25.4)*dpi))
        pad_l,pad_t,pad_r,pad_b=[mm_to_px(v) for v in (ml,mt,mr,mb)]
        table_pix = QPixmap(width+pad_l+pad_r, height+pad_t+pad_b); table_pix.fill()
        painter = QPainter(table_pix)
        painter.translate(pad_l, pad_t)
        # Manual paint (similar to PDF)
        from PyQt6.QtGui import QFont
        font = painter.font(); fm = painter.fontMetrics(); row_h = fm.height()+4
        x=0
        for c in range(self.table.columnCount()):
            w=self.table.columnWidth(c)
            txt=self.table.horizontalHeaderItem(c).text()
            painter.drawText(x+2,fm.ascent()+2,txt)
            x+=w
        y=row_h
        for r in row_iter:
            x=0
            for c in range(self.table.columnCount()):
                w=self.table.columnWidth(c); it=self.table.item(r,c); txt=it.text() if it else ''
                align_right = (c>=2 and c not in (9,11))
                if align_right:
                    tw=fm.width(txt); painter.drawText(x+w-tw-2, y+fm.ascent()+2, txt)
                else:
                    painter.drawText(x+2, y+fm.ascent()+2, txt)
                x+=w
            y+=row_h
        painter.end()
        if not include_header:
            table_pix.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        # With header/footer
        if header_is_svg and header_svg_renderer:
            def _svg_h(renderer, tw, default_ratio=0.12, min_h=40):
                try:
                    ds=renderer.defaultSize(); w,h=ds.width(), ds.height()
                    if w<=0 or h<=0:
                        vb=renderer.viewBoxF(); w,h=vb.width(), vb.height()
                    if w>0 and h>0 and tw>0:
                        return int(round(h*(tw/float(w))))
                except Exception:
                    pass
                return max(min_h, int(round(tw*default_ratio))) if tw>0 else None
            tw=table_pix.width(); hh=_svg_h(header_svg_renderer, tw)
            combo = QPixmap(tw, hh+table_pix.height()); combo.fill(); painter=QPainter(combo)
            header_svg_renderer.render(painter, QRectF(0,0,tw,hh)); painter.drawPixmap(0,hh,table_pix)
            try:
                from PyQt6.QtGui import QFont
                f=QFont(); f.setPointSizeF(f.pointSizeF()*0.85); painter.setFont(f)
                painter.drawText(0, hh+table_pix.height()-18, tw, 16, Qt.AlignmentFlag.AlignCenter, footer_text)
            except Exception: pass
            painter.end(); combo.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        if header_pix and not header_pix.isNull():
            cw=max(header_pix.width(), table_pix.width()); combo=QPixmap(cw, header_pix.height()+table_pix.height()); combo.fill(); painter=QPainter(combo)
            painter.drawPixmap((cw-header_pix.width())//2,0, header_pix); painter.drawPixmap(0, header_pix.height(), table_pix)
            try:
                from PyQt6.QtGui import QFont
                f=QFont(); f.setPointSizeF(f.pointSizeF()*0.85); painter.setFont(f)
                painter.drawText(0, header_pix.height()+table_pix.height()-18, cw, 16, Qt.AlignmentFlag.AlignCenter, footer_text)
            except Exception: pass
            painter.end(); combo.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        # Footer only
        painter = QPainter(table_pix)
        try:
            from PyQt6.QtGui import QFont
            f=QFont(); f.setPointSizeF(f.pointSizeF()*0.85); painter.setFont(f)
            painter.drawText(0, table_pix.height()-18, table_pix.width(), 16, Qt.AlignmentFlag.AlignCenter, footer_text)
        except Exception: pass
        painter.end(); table_pix.save(path,'PNG'); print(f'Exported PNG -> {path}')

    def _export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Missing Dependency", "openpyxl not installed. Please install requirements.")
            return
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        path, _ = QFileDialog.getSaveFileName(self, "Export Cost Table (XLSX)", "cost_estimates.xlsx", "Excel Workbook (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
        wb = Workbook(); ws = wb.active; ws.title = 'Costs'
        # Gather headers & data (only visible columns)
        visible_cols = [c for c in range(self.table.columnCount()) if not self.table.isColumnHidden(c)]
        headers = [self.table.horizontalHeaderItem(c).text() for c in visible_cols]
        ws.append(headers)
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
        # Rows
        currency_cols = set()
        percent_cols = set()
        selected_rows = set()
        if hasattr(self,'chk_selected_only') and self.chk_selected_only.isChecked():
            selected_rows = {idx.row() for idx in self.table.selectionModel().selectedRows()}
        if selected_rows:
            row_range = sorted(selected_rows)
        else:
            row_range = range(self.table.rowCount())
        for r in row_range:
            row_vals = []
            for c in visible_cols:
                it = self.table.item(r,c)
                txt = it.text() if it else ''
                # Numeric detection (strip %,$, commas)
                raw = txt.replace(',','').replace('$','').strip()
                if raw.endswith('%'):
                    try:
                        val = float(raw[:-1]) / 100.0
                        row_vals.append(val)
                        percent_cols.add(len(row_vals))  # 1-based within written row
                        continue
                    except Exception:
                        pass
                try:
                    if raw:
                        val = float(raw)
                        row_vals.append(val)
                        # Heuristic: treat price/cost/profit columns as currency if header matches
                        hdr = headers[len(row_vals)-1].lower()
                        if any(k in hdr for k in ["cost","price","profit"]):
                            currency_cols.add(len(row_vals))
                        continue
                except Exception:
                    pass
                row_vals.append(txt)
            ws.append(row_vals)
        # Apply number formats (skip header row which is row 1)
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row, start=1):
                if idx in currency_cols and isinstance(cell.value,(int,float)):
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                elif idx in percent_cols and isinstance(cell.value,(int,float)):
                    cell.number_format = '0.0%'
        # Autosize columns (simple heuristic)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[letter].width = min(60, max_len + 2)
        # Add metadata sheet
        meta = wb.create_sheet('_Meta')
        from datetime import datetime
        meta.append(["Generated", datetime.now().isoformat(timespec='seconds')])
        meta.append(["Version Selected", self.version_combo.currentText() if hasattr(self,'version_combo') else ''])
        subset = "Selected" if selected_rows else "All"
        meta.append(["Filters", f"Name='{self.le_filter.text()}', Int/Ext='{self.combo_int_ext.currentText()}', MinPrice>{self.min_price_spin.value()}"])
        meta.append(["Subset", subset])
        try:
            wb.save(path)
            print(f"Exported XLSX -> {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))

    # --------------- Column Visibility & Layout Management ---------------
    def _open_columns_dialog(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QListWidget, QListWidgetItem, QPushButton, QCheckBox, QLineEdit, QLabel, QMessageBox
        from PyQt6.QtCore import QSettings
        dlg = QDialog(self)
        dlg.setWindowTitle("Columns & Layouts")
        dlg.resize(540, 420)
        vbox = QVBoxLayout(dlg)
        # Instruction
        info = QLabel("Toggle visibility of columns below. Save named layouts for quick recall.")
        info.setWordWrap(True)
        vbox.addWidget(info)
        # Column list with checkboxes
        from PyQt6.QtWidgets import QWidget, QScrollArea, QGridLayout
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        container = QWidget(); grid = QGridLayout(container)
        self._col_checkboxes = []
        for c in range(self.table.columnCount()):
            name = self.table.horizontalHeaderItem(c).text()
            cb = QCheckBox(name)
            cb.setChecked(not self.table.isColumnHidden(c))
            grid.addWidget(cb, c // 2, c % 2)  # two columns layout
            self._col_checkboxes.append(cb)
        scroll.setWidget(container)
        vbox.addWidget(scroll, 1)
        # Layout management
        layout_row = QHBoxLayout()
        self.layout_name_edit = QLineEdit(); self.layout_name_edit.setPlaceholderText("Layout name…")
        layout_row.addWidget(self.layout_name_edit, 1)
        save_btn = QPushButton("Save Layout")
        apply_btn = QPushButton("Apply Layout")
        delete_btn = QPushButton("Delete Layout")
        layout_row.addWidget(save_btn); layout_row.addWidget(apply_btn); layout_row.addWidget(delete_btn)
        vbox.addLayout(layout_row)
        # Existing layouts list
        self.layouts_list = QListWidget()
        vbox.addWidget(QLabel("Saved Layouts:"))
        vbox.addWidget(self.layouts_list, 1)
        # Close row
        btn_row = QHBoxLayout(); btn_row.addStretch(1)
        close_btn = QPushButton("Close")
        btn_row.addWidget(close_btn)
        vbox.addLayout(btn_row)
        # Load existing layouts from QSettings
        s = QSettings("LSI", "ProjectPlanner")
        raw = s.value("Columns/layouts", "{}")
        import json
        try:
            layouts = json.loads(raw) if raw else {}
            if not isinstance(layouts, dict):
                layouts = {}
        except Exception:
            layouts = {}
        def refresh_layout_list():
            self.layouts_list.clear()
            for name in sorted(layouts.keys()):
                item = QListWidgetItem(name)
                self.layouts_list.addItem(item)
        refresh_layout_list()
        # Handlers
        def _current_order():
            try:
                header = self.table.horizontalHeader()
                order = [header.logicalIndex(i) for i in range(header.count())]
                return order
            except Exception:
                return list(range(self.table.columnCount()))
        def _apply_order(order_list):
            try:
                header = self.table.horizontalHeader()
                for visual_pos, logical in enumerate(order_list):
                    cur_logical = header.logicalIndex(visual_pos)
                    if cur_logical != logical:
                        header.moveSection(header.visualIndex(logical), visual_pos)
            except Exception:
                pass
        def save_layout():
            name = self.layout_name_edit.text().strip()
            if not name:
                QMessageBox.warning(dlg, "Name Required", "Enter a layout name.")
                return
            # Capture visibility: 1 = visible, 0 = hidden
            vis = [1 if cb.isChecked() else 0 for cb in self._col_checkboxes]
            order = _current_order()
            # Store current column header names to allow graceful re-mapping later
            col_names = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            layouts[name] = {"visibility": vis, "order": order, "columns": col_names, "v": 2}
            s.setValue("Columns/layouts", json.dumps(layouts))
            refresh_layout_list()
            if self.parent() and self.parent().window().statusBar():
                self.parent().window().statusBar().showMessage(f"Saved layout '{name}'", 3000)
        def apply_layout():
            items = self.layouts_list.selectedItems()
            if not items:
                QMessageBox.information(dlg, "Select Layout", "Select a saved layout to apply.")
                return
            name = items[0].text()
            cfg = layouts.get(name)
            if not cfg:
                QMessageBox.warning(dlg, "Layout Missing", "Layout data not found.")
                return
            vis = cfg.get("visibility") if isinstance(cfg, dict) else None
            order = cfg.get("order") if isinstance(cfg, dict) else None
            saved_cols = cfg.get("columns") if isinstance(cfg, dict) else None
            current_cols = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            reindex_map = None
            if saved_cols and len(saved_cols) != len(current_cols):
                # Attempt name-based mapping
                name_to_cur = {n: i for i, n in enumerate(current_cols)}
                reindex_map = []
                for saved_idx, saved_name in enumerate(saved_cols):
                    cur_idx = name_to_cur.get(saved_name)
                    if cur_idx is not None:
                        reindex_map.append(cur_idx)
                # If mapping incomplete, warn but continue with best effort
                if len(reindex_map) != len(saved_cols):
                    QMessageBox.information(dlg, "Partial Layout Mapping", "Some columns from this layout are missing or renamed; applying what matches.")
            # Apply visibility
            if vis:
                if reindex_map:  # remap vis sequence to current order
                    for saved_pos, cur_idx in enumerate(reindex_map):
                        flag = vis[saved_pos] if saved_pos < len(vis) else 1
                        self.table.setColumnHidden(cur_idx, flag == 0)
                        self._col_checkboxes[cur_idx].setChecked(flag == 1)
                elif len(vis) == len(current_cols):
                    for idx, flag in enumerate(vis):
                        self.table.setColumnHidden(idx, flag == 0)
                        self._col_checkboxes[idx].setChecked(flag == 1)
                else:
                    QMessageBox.warning(dlg, "Layout Incompatible", "Cannot apply visibility (size mismatch).")
            # Apply order only if sizes match exactly and no missing columns
            if order and len(order) == len(current_cols) and not reindex_map:
                _apply_order(order)
            # Persist last used
            s.setValue("Columns/last_layout", name)
            self.refresh()  # ensure totals etc recalc if hidden cols influence width
            if self.parent() and self.parent().window().statusBar():
                self.parent().window().statusBar().showMessage(f"Applied layout '{name}'", 3000)
        def delete_layout():
            items = self.layouts_list.selectedItems()
            if not items:
                return
            name = items[0].text()
            if name in layouts:
                del layouts[name]
                s.setValue("Columns/layouts", json.dumps(layouts))
                refresh_layout_list()
                if self.parent() and self.parent().window().statusBar():
                    self.parent().window().statusBar().showMessage(f"Deleted layout '{name}'", 3000)
        def checkbox_changed():
            # Apply immediate visibility changes
            for idx, cb in enumerate(self._col_checkboxes):
                self.table.setColumnHidden(idx, not cb.isChecked())
            self.refresh()
        for cb in self._col_checkboxes:
            cb.stateChanged.connect(checkbox_changed)
        save_btn.clicked.connect(save_layout)
        apply_btn.clicked.connect(apply_layout)
        delete_btn.clicked.connect(delete_layout)
        close_btn.clicked.connect(dlg.accept)
        dlg.exec()

    def _apply_last_layout(self):
        # Called after construction to restore last used layout
        from PyQt6.QtCore import QSettings
        import json
        try:
            s = QSettings("LSI", "ProjectPlanner")
            last = s.value("Columns/last_layout", "")
            if not last:
                return
            raw = s.value("Columns/layouts", "{}")
            layouts = json.loads(raw) if raw else {}
            cfg = layouts.get(last)
            if isinstance(cfg, dict):
                vis = cfg.get("visibility")
                order = cfg.get("order")
                # Attempt graceful application on startup (silent if mismatch)
                if vis and len(vis) == self.table.columnCount():
                    for idx, flag in enumerate(vis):
                        self.table.setColumnHidden(idx, flag == 0)
                if order and len(order) == self.table.columnCount():
                    # Apply after a short delay if needed to ensure header exists
                    try:
                        from PyQt6.QtCore import QTimer
                        QTimer.singleShot(0, lambda o=order: [self.table.horizontalHeader().moveSection(self.table.horizontalHeader().visualIndex(log), pos) for pos, log in enumerate(o)])
                    except Exception:
                        pass
        except Exception:
            pass

    # --- Programmatic layout API (for tests / automation) ---
    def save_layout_programmatic(self, name: str):
        """Save current layout (visibility + order + column names) under name. Returns True if saved."""
        from PyQt6.QtCore import QSettings
        import json
        if not name:
            return False
        try:
            s = QSettings("LSI", "ProjectPlanner")
            raw = s.value("Columns/layouts", "{}")
            layouts = json.loads(raw) if raw else {}
            vis = [0 if self.table.isColumnHidden(i) else 1 for i in range(self.table.columnCount())]
            header = self.table.horizontalHeader()
            order = [header.logicalIndex(i) for i in range(header.count())]
            col_names = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            layouts[name] = {"visibility": vis, "order": order, "columns": col_names, "v": 2}
            s.setValue("Columns/layouts", json.dumps(layouts))
            s.setValue("Columns/last_layout", name)
            return True
        except Exception:
            return False

    def apply_layout_programmatic(self, name: str):
        """Apply a saved layout by name. Returns tuple(success, message)."""
        from PyQt6.QtCore import QSettings
        import json
        try:
            s = QSettings("LSI", "ProjectPlanner")
            raw = s.value("Columns/layouts", "{}")
            layouts = json.loads(raw) if raw else {}
            cfg = layouts.get(name)
            if not cfg:
                return False, "Layout not found"
            vis = cfg.get("visibility")
            order = cfg.get("order")
            saved_cols = cfg.get("columns")
            current_cols = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            if saved_cols and len(saved_cols) != len(current_cols):
                # Attempt name-based remap
                name_to_cur = {n: i for i, n in enumerate(current_cols)}
                for saved_idx, saved_name in enumerate(saved_cols):
                    cur_idx = name_to_cur.get(saved_name)
                    if cur_idx is not None and vis and saved_idx < len(vis):
                        flag = vis[saved_idx]
                        self.table.setColumnHidden(cur_idx, flag == 0)
                # Skip order if mismatch
                return True, "Applied partial (column count changed)"
            # Full match
            if vis and len(vis) == len(current_cols):
                for idx, flag in enumerate(vis):
                    self.table.setColumnHidden(idx, flag == 0)
            if order and len(order) == len(current_cols):
                header = self.table.horizontalHeader()
                for visual_pos, logical in enumerate(order):
                    if header.logicalIndex(visual_pos) != logical:
                        header.moveSection(header.visualIndex(logical), visual_pos)
            s.setValue("Columns/last_layout", name)
            return True, "Applied"
        except Exception as e:
            return False, str(e)

    def refresh(self):
        rows = getattr(self.model,'rows', [])
        # Version map
        selected_version = self.version_combo.currentText() if hasattr(self, 'version_combo') else '<None>'
        version_map = {}
        if selected_version and selected_version not in ("<None>", ""):
            try:
                version_map = self.model.load_quote_version_map(selected_version)
            except Exception:
                version_map = {}
        name_filter = (self.le_filter.text() or '').strip().lower()
        int_ext_filter = self.combo_int_ext.currentText()
        min_price = float(self.min_price_spin.value())
        leaf_only = self.chk_leaf_only.isChecked()
        rollup = self.chk_rollup.isChecked()
        data = []
        total_cost = total_price = total_profit = 0.0
        # Pre-build child mapping for rollups
        children_map = {}
        if rollup:
            for r in rows:
                p = (r.get('Parent') or '').strip()
                if p:
                    children_map.setdefault(p, []).append(r)
        # Helper to accumulate descendants
        def accumulate(row):
            stack = [row]
            cost_prod = cost_inst = price_prod = price_inst = 0.0
            while stack:
                cur = stack.pop()
                pcost = self._num(cur.get('Production Cost', 0))
                icost = self._num(cur.get('Installation Cost', 0))
                pprice = self._num(cur.get('Production Price', 0))
                iprice = self._num(cur.get('Installation Price', 0))
                cost_prod += pcost; cost_inst += icost; price_prod += pprice; price_inst += iprice
                for ch in children_map.get(cur.get('Project Part', ''), []):
                    stack.append(ch)
            return cost_prod, cost_inst, price_prod, price_inst
        # Build data
        for r in rows:
            if leaf_only and not self._is_leaf(r):
                # If rollup enabled and parent row: still include using aggregated numbers
                if not rollup:
                    continue
            name = r.get('Project Part','')
            if name_filter and name_filter not in name.lower():
                continue
            ie = r.get('Internal/External','') or ''
            if int_ext_filter != 'All' and ie != int_ext_filter:
                continue
            if rollup:
                pcost, icost, pprice, iprice = accumulate(r)
            else:
                pcost = self._num(r.get('Production Cost',0))
                icost = self._num(r.get('Installation Cost',0))
                pprice = self._num(r.get('Production Price',0))
                iprice = self._num(r.get('Installation Price',0))
            tcost = pcost + icost
            tprice = pprice + iprice
            if tprice < min_price:
                continue
            profit = tprice - tcost
            margin_pct = (profit / tprice * 100.0) if tprice > 0 else 0.0
            data.append((name, r.get('Parent','') or '', pcost, icost, tcost, pprice, iprice, tprice, profit, margin_pct, ie))
            total_cost += tcost
            total_price += tprice
            total_profit += profit
        pct_base = total_price if total_price>0 else 1.0
        self.table.setRowCount(len(data))
        # Determine thresholds for highlighting top-N (top 10% by total price)
        import math
        sorted_prices = sorted([d[7] for d in data], reverse=True)
        top_n = max(1, math.ceil(len(sorted_prices)*0.10)) if sorted_prices else 0
        top_cut = sorted_prices[top_n-1] if sorted_prices and top_n<=len(sorted_prices) else None
        for row_idx, (name,parent,pcost,icost,tcost,pprice,iprice,tprice,profit,margin_pct,ie) in enumerate(data):
            base = version_map.get(name)
            if base:
                b_pc, b_ic, b_pp, b_ip = base
                base_total_price = (b_pp or 0)+(b_ip or 0)
                cur_total_price = tprice
                price_delta_pct = ((cur_total_price - base_total_price)/base_total_price*100.0) if base_total_price>0 else 0.0
                # margin delta
                base_profit = base_total_price - ((b_pc or 0)+(b_ic or 0))
                base_margin_pct = (base_profit/base_total_price*100.0) if base_total_price>0 else 0.0
                margin_delta_pts = margin_pct - base_margin_pct
            else:
                price_delta_pct = 0.0; margin_delta_pts = 0.0
            values = [
                name, parent,
                f"{pcost:,.2f}", f"{icost:,.2f}", f"{tcost:,.2f}",
                f"{pprice:,.2f}", f"{iprice:,.2f}", f"{tprice:,.2f}",
                f"{profit:,.2f}", f"{margin_pct:,.1f}%", f"{price_delta_pct:,.1f}%", f"{margin_delta_pts:,.1f}", f"{(tprice/pct_base)*100:,.1f}%", ie
            ]
            for col_idx, val in enumerate(values):
                item = QTableWidgetItem(val)
                if col_idx >=2 and col_idx not in (9,11):  # margin & delta margin are textual with % / pts
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                # Conditional formatting
                from PyQt6.QtGui import QColor
                if col_idx == 8:  # profit
                    if profit < 0:
                        item.setBackground(QColor(120,0,0))
                        item.setForeground(QColor('white'))
                    elif profit == 0:
                        item.setBackground(QColor(80,80,80))
                        item.setForeground(QColor('white'))
                if col_idx == 9:  # margin %
                    if margin_pct < 5:
                        item.setBackground(QColor(130,0,0))
                        item.setForeground(QColor('white'))
                    elif margin_pct < 15:
                        item.setBackground(QColor(110,70,0))
                        item.setForeground(QColor('white'))
                if col_idx == 10:  # Δ Price % highlight
                    if price_delta_pct > 0:
                        item.setForeground(QColor('#3CB371'))
                    elif price_delta_pct < 0:
                        item.setForeground(QColor('#FF6347'))
                if col_idx == 11:  # Δ Margin pts
                    if margin_delta_pts > 0:
                        item.setForeground(QColor('#3CB371'))
                    elif margin_delta_pts < 0:
                        item.setForeground(QColor('#FF6347'))
                if top_cut is not None and tprice >= top_cut and col_idx == 7:
                    item.setBackground(QColor(60,60,0))
                    item.setForeground(QColor('#FFE066'))
                self.table.setItem(row_idx, col_idx, item)
        try:
            self.table.resizeColumnsToContents()
        except Exception:
            pass
        blended_margin = (total_profit/total_price*100.0) if total_price>0 else 0.0
        avg_margin = (sum(d[9] for d in data)/len(data)) if data else 0.0
        self.totals_label.setText(
            f"Cost: ${total_cost:,.2f}  Price: ${total_price:,.2f}  Profit: ${total_profit:,.2f}  Blended Margin: {blended_margin:,.1f}%  Avg Margin: {avg_margin:,.1f}%  Rows: {len(data)}"
        )
        self._apply_compact_mode()
        self._reload_versions()

    def _reload_versions(self):
        try:
            existing = set(self.list_versions_cache) if hasattr(self,'list_versions_cache') else set()
            versions = self.model.list_quote_versions()
            if set(versions) != existing:
                cur = self.version_combo.currentText() if self.version_combo.count() else "<None>"
                self.version_combo.blockSignals(True)
                self.version_combo.clear(); self.version_combo.addItem("<None>")
                for v in versions:
                    self.version_combo.addItem(v)
                # restore if possible
                idx = self.version_combo.findText(cur)
                if idx >= 0:
                    self.version_combo.setCurrentIndex(idx)
                self.version_combo.blockSignals(False)
                self.list_versions_cache = set(versions)
        except Exception:
            pass

class ProjectTreeView(QWidget):
    """Horizontal left-to-right branching tree visualization (graphics based).
    Replaces prior multi-column QTreeWidget with a schematic layout closer to the web app's
    horizontally indented appearance, but rendered as a node graph for clarity.
    on_jump_to_gantt: optional callback invoked when user selects Jump To In Gantt.
    """
    def __init__(self, model, on_part_selected=None, on_jump_to_gantt=None):
        super().__init__()
        self.model = model
        self.on_part_selected = on_part_selected
        self.on_jump_to_gantt = on_jump_to_gantt
        self._name_to_row = {}
        self._name_to_item = {}
        self._hover_preview_enabled = True
        # tiny LRU cache for preview pixmaps
        self._preview_cache = {}
        self._preview_cache_order = []  # MRU at end
        self._preview_cache_cap = 64
        layout = QVBoxLayout()
        header = QHBoxLayout()
        title = QLabel("Project Tree (Horizontal)")
        title.setStyleSheet("font-weight:600; padding:2px 4px;")
        header.addWidget(title)
        from PyQt6.QtWidgets import QPushButton, QCheckBox
        fit_btn = QPushButton("Fit")
        # Add explicit zoom controls for Project Tree
        zoom_in_btn = QPushButton("Zoom In")
        zoom_in_btn.setToolTip("Zoom in (also Ctrl + '+')")
        zoom_out_btn = QPushButton("Zoom Out")
        zoom_out_btn.setToolTip("Zoom out (also Ctrl + '-')")
        from PyQt6.QtWidgets import QLabel as _QLabel
        zoom_label = _QLabel("100%")
        zoom_label.setToolTip("Current zoom")
        refresh_btn = QPushButton("Refresh")
        toggle_img_btn = QPushButton("Previews: On")
        export_btn = QPushButton("Export")
        reset_btn = QPushButton("Reset View")
        clear_cache_btn = QPushButton("Clear Cache")
        settings_btn = QPushButton("Settings…")
        fit_btn.setToolTip("Fit entire tree in view")
        refresh_btn.setToolTip("Rebuild layout from model")
        toggle_img_btn.setToolTip("Toggle image hover previews")
        export_btn.setToolTip("Export Project Tree")
        settings_btn.setToolTip("Open export settings dialog")
        reset_btn.setToolTip("Reset zoom and fit entire tree")
        clear_cache_btn.setToolTip("Clear preview image cache")
        # Links toggle (persisted, shared with other views)
        from PyQt6.QtWidgets import QPushButton as _QPB_links
        try:
            from PyQt6.QtCore import QSettings as _QS_links
            _ps_links = _QS_links('LSI','ProjectPlanner')
            val_links = _ps_links.value('UI/ShowLinks', 'true')
            def _b_l(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return True
            self._show_links = _b_l(val_links)
        except Exception:
            self._show_links = True
        links_btn = _QPB_links('Links: On' if self._show_links else 'Links: Off')
        links_btn.setCheckable(True)
        links_btn.setChecked(self._show_links)
        links_btn.setToolTip('Show link indicators; click nodes to open Pace Link')
        def _toggle_links_tree():
            self._show_links = links_btn.isChecked()
            links_btn.setText('Links: On' if self._show_links else 'Links: Off')
            try:
                from PyQt6.QtCore import QSettings as _QS_links2
                _QS_links2('LSI','ProjectPlanner').setValue('UI/ShowLinks', self._show_links)
            except Exception:
                pass
            # Propagate to other views
            try:
                mw = self.window()
                if hasattr(mw, 'gantt_chart_view') and mw.gantt_chart_view:
                    mw.gantt_chart_view._show_links = self._show_links
                    mw.gantt_chart_view.links_checkbox.setChecked(self._show_links)
                    mw.gantt_chart_view.refresh_gantt()
                if hasattr(mw, 'timeline_view') and mw.timeline_view:
                    mw.timeline_view._show_links = self._show_links
                    mw.timeline_view._sync_links_checkbox()
                    mw.timeline_view.render_timeline()
            except Exception:
                pass
            self.refresh()
        links_btn.clicked.connect(_toggle_links_tree)
        def do_fit():
            if hasattr(self, 'view'):
                r = self.scene.itemsBoundingRect()
                if not r.isNull():
                    self.view.fitInView(r, _keep_ar())
                    try:
                        sf = float(self.view.transform().m11())
                        from math import isfinite
                        if isfinite(sf):
                            zoom_label.setText(f"{int(round(sf*100))}%")
                    except Exception:
                        pass
        def do_refresh():
            self.refresh()
        def do_toggle_preview():
            self._hover_preview_enabled = not self._hover_preview_enabled
            is_on = self._hover_preview_enabled
            toggle_img_btn.setText(f"Previews: {'On' if is_on else 'Off'}")
            if not is_on:
                try:
                    self.preview_label.clear()
                except Exception:
                    pass
            else:
                # If turning on, try to show preview for the item under cursor
                try:
                    from PyQt6.QtGui import QCursor
                    vp = self.view.mapFromGlobal(QCursor.pos())
                    sp = self.view.mapToScene(vp)
                    item = self.scene.itemAt(sp, self.view.transform())
                    if item:
                        name = item.data(0)
                        if isinstance(name, str):
                            row = self._name_to_row.get(name)
                            if row:
                                self._show_image_for_row(row)
                except Exception:
                    pass
        fit_btn.clicked.connect(do_fit)
        refresh_btn.clicked.connect(do_refresh)
        toggle_img_btn.clicked.connect(do_toggle_preview)
        export_btn.clicked.connect(lambda: self._export_tree())
        def do_reset():
            try:
                if hasattr(self, 'view') and hasattr(self.view, 'resetZoom'):
                    self.view.resetZoom()
                # Fit to deterministic default rectangle (captured on first refresh)
                target_rect = getattr(self, '_initial_view_rect', None)
                if target_rect is None or target_rect.isNull():
                    # fallback to current scene rect with padding
                    r = self.scene.sceneRect()
                    if not r.isNull():
                        target_rect = r
                if target_rect is not None and not target_rect.isNull():
                    self.view.fitInView(target_rect, _keep_ar())
                    # Persist the resulting zoom factor after fit
                    if hasattr(self.view, '_persist_zoom'):
                        self.view._persist_zoom()
                    # Update zoom label to reflect fit transform
                    try:
                        sf = float(self.view.transform().m11())
                        from math import isfinite
                        if isfinite(sf):
                            zoom_label.setText(f"{int(round(sf*100))}%")
                    except Exception:
                        pass
            except Exception:
                pass
        reset_btn.clicked.connect(do_reset)
        def do_clear_cache():
            try:
                if hasattr(self, '_preview_cache'):
                    self._preview_cache.clear()
                if hasattr(self, '_preview_cache_order'):
                    self._preview_cache_order.clear()
                self.preview_label.clear()
            except Exception:
                pass
        clear_cache_btn.clicked.connect(do_clear_cache)
        def open_settings():
            try:
                dlg = ExportSettingsDialog(self)
                dlg.exec()
            except Exception as e:
                print(f"Tree export settings failed: {e}")
        settings_btn.clicked.connect(open_settings)
        header.addStretch(1)
        header.addWidget(fit_btn)
        header.addWidget(zoom_in_btn)
        header.addWidget(zoom_out_btn)
        header.addWidget(refresh_btn)
        header.addWidget(toggle_img_btn)
        header.addWidget(export_btn)
        header.addWidget(reset_btn)
        header.addWidget(clear_cache_btn)
        header.addWidget(settings_btn)
        header.addWidget(links_btn)
        header.addWidget(zoom_label)
        # Panel visibility toggles
        self.preview_panel_cb = QCheckBox("Preview")
        self.minimap_panel_cb = QCheckBox("Minimap")
        from PyQt6.QtCore import QSettings
        try:
            _ps = QSettings('LSI','ProjectApp')
            pv = _ps.value('TreeShowPreviewPanel', 'true')
            mm = _ps.value('TreeShowMinimap', 'true')
            def _b(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return True
            self.preview_panel_cb.setChecked(_b(pv))
            self.minimap_panel_cb.setChecked(_b(mm))
        except Exception:
            self.preview_panel_cb.setChecked(True)
            self.minimap_panel_cb.setChecked(True)
        header.addWidget(self.preview_panel_cb)
        header.addWidget(self.minimap_panel_cb)
        layout.addLayout(header)
        # Graphics view for nodes
        self.scene = QGraphicsScene()
        self.view = ZoomableGraphicsView()
        self.view.setScene(self.scene)
        self.view.setRenderHints(self.view.renderHints())
        # Enable hover tracking so we receive GraphicsSceneHoverMove without pressing mouse buttons
        try:
            self.view.setMouseTracking(True)
            self.view.viewport().setMouseTracking(True)
        except Exception:
            pass
        # Assign settings key to persist zoom
        try:
            self.view.setSettingsKey('TreeZoom')
        except Exception:
            pass
        # Wire up zoom buttons now that view exists
        try:
            zoom_in_btn.clicked.connect(self.view.zoomIn)
            zoom_out_btn.clicked.connect(self.view.zoomOut)
        except Exception:
            pass
        # Update label on zoom changes
        try:
            def _set_zoom_label(sf: float|int):
                try:
                    pct = int(round(float(sf)*100))
                    zoom_label.setText(f"{pct}%")
                except Exception:
                    pass
            self.view.zoomChanged.connect(_set_zoom_label)
            # Seed with restored value
            _set_zoom_label(self.view.transform().m11())
        except Exception:
            pass
        layout.addWidget(self.view, 1)
        # Keyboard shortcuts for zoom
        try:
            from PyQt6.QtWidgets import QShortcut
            from PyQt6.QtGui import QKeySequence
            QShortcut(QKeySequence.ZoomIn, self.view, activated=self.view.zoomIn)
            QShortcut(QKeySequence.ZoomOut, self.view, activated=self.view.zoomOut)
            QShortcut(QKeySequence("Ctrl+0"), self.view, activated=self.view.resetZoom)
        except Exception:
            pass
        # Preview label (shares style with others)
        self.preview_label = QLabel()
        self.preview_label.setFixedHeight(140)
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_label.setStyleSheet("border:1px solid #666; background:#222;")
        layout.addWidget(self.preview_label)
        self.setLayout(layout)
        self._collapsed = set()  # names of collapsed nodes
        self._minimap_view = None
        # Load persisted collapsed state
        try:
            from PyQt6.QtCore import QSettings
            _ts = QSettings('LSI','ProjectApp')
            saved = _ts.value('TreeCollapsed', [])
            if isinstance(saved, list):
                self._collapsed = set(saved)
        except Exception:
            pass
        # Minimap container (lightweight)
        from PyQt6.QtWidgets import QFrame, QGraphicsView
        self._mini_frame = QFrame()
        self._mini_frame.setFixedHeight(120)
        self._mini_frame.setStyleSheet("QFrame { border:1px solid #555; background:#111; }")
        mini_layout = QVBoxLayout(self._mini_frame)
        mini_layout.setContentsMargins(2,2,2,2)
        # Minimap zoom controls (+ / -) affecting main tree view
        try:
            mini_controls = QHBoxLayout()
            btn_zoom_in_mini = QPushButton("+")
            btn_zoom_in_mini.setFixedWidth(24)
            btn_zoom_in_mini.setToolTip("Zoom in")
            btn_zoom_out_mini = QPushButton("-")
            btn_zoom_out_mini.setFixedWidth(24)
            btn_zoom_out_mini.setToolTip("Zoom out")
            btn_zoom_in_mini.clicked.connect(lambda: getattr(self.view, 'zoomIn', lambda: None)())
            btn_zoom_out_mini.clicked.connect(lambda: getattr(self.view, 'zoomOut', lambda: None)())
            mini_controls.addWidget(btn_zoom_in_mini)
            mini_controls.addWidget(btn_zoom_out_mini)
            mini_controls.addStretch(1)
            mini_layout.addLayout(mini_controls)
        except Exception:
            pass
        self._mini_scene = QGraphicsScene()
        self._mini_view = QGraphicsView(self._mini_scene)
        try:
            self._mini_view.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            self._mini_view.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        except Exception:
            # Fallback if enum namespace differs
            self._mini_view.setHorizontalScrollBarPolicy(getattr(Qt, 'ScrollBarAlwaysOff', 1))
            self._mini_view.setVerticalScrollBarPolicy(getattr(Qt, 'ScrollBarAlwaysOff', 1))
        self._mini_view.setRenderHints(self._mini_view.renderHints())
        mini_layout.addWidget(self._mini_view)
        layout.addWidget(self._mini_frame)
        # Apply initial visibility
        try:
            self.preview_label.setVisible(self.preview_panel_cb.isChecked())
            self._mini_frame.setVisible(self.minimap_panel_cb.isChecked())
        except Exception:
            pass
        def _persist_panels():
            try:
                from PyQt6.QtCore import QSettings
                _ps = QSettings('LSI','ProjectApp')
                _ps.setValue('TreeShowPreviewPanel', self.preview_panel_cb.isChecked())
                _ps.setValue('TreeShowMinimap', self.minimap_panel_cb.isChecked())
            except Exception:
                pass
        def _apply_panel_visibility():
            show_prev = self.preview_panel_cb.isChecked()
            show_map = self.minimap_panel_cb.isChecked()
            # Simple visibility
            self.preview_label.setVisible(show_prev)
            self._mini_frame.setVisible(show_map)
            # Collapse space when hidden
            if show_prev:
                self.preview_label.setMaximumHeight(16777215)
            else:
                self.preview_label.setMaximumHeight(0)
            if show_map:
                self._mini_frame.setMaximumHeight(16777215)
            else:
                self._mini_frame.setMaximumHeight(0)
        _apply_panel_visibility()
        self.preview_panel_cb.stateChanged.connect(lambda _s: (_apply_panel_visibility(), _persist_panels()))
        self.minimap_panel_cb.stateChanged.connect(lambda _s: (_apply_panel_visibility(), _persist_panels()))
        self.refresh()
        # Connect viewport + interactions for minimap live updates
        try:
            self.view.viewport().installEventFilter(self)
            # Debounced minimap updates
            from PyQt6.QtCore import QTimer
            self._minimap_timer = QTimer(self)
            self._minimap_timer.setSingleShot(True)
            self._minimap_timer.timeout.connect(self._update_minimap)
            def schedule_minimap():
                try:
                    self._minimap_timer.start(120)
                except Exception:
                    pass
            self.view.horizontalScrollBar().valueChanged.connect(schedule_minimap)
            self.view.verticalScrollBar().valueChanged.connect(schedule_minimap)
        except Exception:
            pass
        # Minimap click -> center main view
        try:
            orig_press = self._mini_view.mousePressEvent
            def mini_click(ev):
                try:
                    _left_btn = Qt.MouseButton.LeftButton
                except Exception:
                    _left_btn = getattr(Qt, 'LeftButton', 1)
                if ev.button() == _left_btn:
                    scene_pt = self._mini_view.mapToScene(ev.pos())
                    # scale factor used in _update_minimap
                    scale_factor = 0.12
                    from PyQt6.QtCore import QRectF
                    # Center main view around corresponding point
                    center_target = QPointF(scene_pt.x()/scale_factor, scene_pt.y()/scale_factor)
                    vr = self.view.mapToScene(self.view.viewport().rect()).boundingRect()
                    new_rect = QRectF(center_target.x() - vr.width()/2, center_target.y() - vr.height()/2, vr.width(), vr.height())
                    try:
                        if hasattr(self.view, 'smoothFocusRect'):
                            self.view.smoothFocusRect(new_rect)
                        else:
                            self.view.fitInView(new_rect, _keep_ar())
                    except Exception:
                        self.view.fitInView(new_rect, _keep_ar())
                    # Update zoom label to reflect any transform changes
                    try:
                        sf = float(self.view.transform().m11())
                        from math import isfinite
                        if isfinite(sf):
                            zoom_label.setText(f"{int(round(sf*100))}%")
                    except Exception:
                        pass
                    self._update_minimap()
                orig_press(ev)
            from PyQt6.QtCore import QPointF
            self._mini_view.mousePressEvent = mini_click
        except Exception:
            pass

    # -------- Data & Layout --------
    def _build_hierarchy(self):
        rows = getattr(self.model, 'rows', []) or []
        self._name_to_row = {r.get('Project Part',''): r for r in rows}
        children = {}
        roots = []
        for r in rows:
            name = r.get('Project Part','')
            parent = (r.get('Parent') or '').strip()
            if parent and parent in self._name_to_row and parent != name:
                children.setdefault(parent, []).append(name)
            else:
                roots.append(name)
        return roots, children

    def _compute_layout(self, roots, children):
        # Tidy-ish layout: allocate vertical space proportional to leaf count
        node_w, node_h = 180, 46
        h_gap, v_gap = 80, 24
        leaf_spacing = node_h + v_gap
        positions = {}
        def count_leaves(name, visited=None):
            if visited is None: visited = set()
            if name in visited: return 1
            visited.add(name)
            ch = children.get(name, [])
            if not ch: return 1
            return sum(count_leaves(c, visited.copy()) for c in ch)
        y_cursor = 0
        def place(name, depth, y_start):
            ch = children.get(name, [])
            if name in self._collapsed:
                ch = []  # treat as leaf visually
            if not ch:
                y_center = y_start + node_h/2
                positions[name] = (depth*(node_w+h_gap), y_center - node_h/2)
                return y_center, y_start + leaf_spacing
            # place children first
            child_centers = []
            cur_y = y_start
            for c in sorted(ch, key=lambda s: s.lower()):
                cc, cur_y = place(c, depth+1, cur_y)
                child_centers.append(cc)
            y_center = sum(child_centers)/len(child_centers)
            positions[name] = (depth*(node_w+h_gap), y_center - node_h/2)
            return y_center, cur_y
        for r in sorted(roots, key=lambda s: s.lower()):
            _c, y_cursor = place(r, 0, y_cursor)
        return positions, (node_w, node_h)

    # -------- Rendering --------
    def refresh(self):
        self.scene.clear()
        self._name_to_item.clear()
        roots, children = self._build_hierarchy()
        if not roots:
            self.scene.addText("(No data)")
            return
        positions, (node_w, node_h) = self._compute_layout(roots, children)
        from PyQt6.QtGui import QPen, QColor, QBrush, QFont
        # Draw connectors first
        pen_conn = QPen(QColor(150,150,150))
        pen_conn.setWidth(2)
        for parent, chs in children.items():
            if parent not in positions: continue
            px, py = positions[parent]
            for c in chs:
                if c not in positions: continue
                if parent in self._collapsed:
                    continue
                cx, cy = positions[c]
                # Parent right middle to child left middle via elbow
                p_mid = (px+node_w, py + node_h/2)
                c_mid = (cx, cy + node_h/2)
                mid_x = (p_mid[0] + c_mid[0]) / 2
                self.scene.addLine(p_mid[0], p_mid[1], mid_x, p_mid[1], pen_conn)
                self.scene.addLine(mid_x, p_mid[1], mid_x, c_mid[1], pen_conn)
                self.scene.addLine(mid_x, c_mid[1], c_mid[0], c_mid[1], pen_conn)
        # Draw nodes
        font = self.font()
        for name, (x, y) in positions.items():
            row = self._name_to_row.get(name, {})
            pc = 0
            try:
                pc = int(row.get('% Complete') or 0)
            except Exception:
                pc = 0
            status = (row.get('Status') or '').strip()
            base_color = QColor('#2f2f2f')
            if status.lower() == 'done':
                base_color = QColor('#235f23')
            elif status.lower() in ('blocked','deferred'):
                base_color = QColor('#5f2323')
            elif status.lower() in ('in progress', 'at risk'):
                base_color = QColor('#5f4a23')
            rect_item = self.scene.addRect(x, y, node_w, node_h, QPen(QColor('#888')), QBrush(base_color))
            rect_item.setData(0, name)
            # Pace Link indicator on node if present
            try:
                pace_link = (row.get('Pace Link') or '').strip()
                has_link = pace_link.lower().startswith('http://') or pace_link.lower().startswith('https://')
            except Exception:
                pace_link = ''; has_link = False
            if has_link and getattr(self, '_show_links', True):
                try:
                    from PyQt6.QtWidgets import QGraphicsSimpleTextItem
                    icon = QGraphicsSimpleTextItem('🔗', rect_item)
                    icon.setBrush(QColor('white'))
                    icon.setPos(node_w - 14, 2)
                    icon.setZValue(rect_item.zValue() + 2)
                    rect_item.setToolTip(pace_link)
                    rect_item.setCursor(getattr(Qt,'CursorShape', Qt).PointingHandCursor)
                except Exception:
                    pass
            try:
                from PyQt6.QtWidgets import QGraphicsItem
                rect_item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, True)
            except Exception:
                # Fallback attempt using legacy attribute
                if hasattr(rect_item, 'ItemIsSelectable'):
                    rect_item.setFlag(rect_item.ItemIsSelectable, True)
            try:
                # Ensure hover events fire at scene level
                rect_item.setAcceptHoverEvents(True)
            except Exception:
                pass
            # Collapse/expand indicator if has children
            if name in children and children[name]:
                tri_w = 12; tri_h = 12
                from PyQt6.QtGui import QPolygonF
                from PyQt6.QtCore import QPointF
                if name in self._collapsed:
                    pts = [QPointF(x+6, y+node_h/2 - tri_h/2), QPointF(x+6, y+node_h/2 + tri_h/2), QPointF(x+6+tri_w, y+node_h/2)]
                else:
                    pts = [QPointF(x+6, y+node_h/2 - tri_h/2), QPointF(x+6+tri_w, y+node_h/2 - tri_h/2), QPointF(x+6+tri_w/2, y+node_h/2 + tri_h/2)]
                poly = self.scene.addPolygon(QPolygonF(pts), QPen(QColor('#ddd')), QBrush(QColor('#ddd')))
                poly.setData(0, f"__toggle__::{name}")
                poly.setZValue(rect_item.zValue()+3)
                try:
                    poly.setAcceptHoverEvents(True)
                except Exception:
                    pass
            # Progress overlay
            if pc > 0:
                prog_w = int((pc/100)*node_w)
                try:
                    no_pen = Qt.PenStyle.NoPen
                except Exception:
                    no_pen = getattr(Qt, 'NoPen', 0)
                prog = self.scene.addRect(x, y+node_h-8, prog_w, 8, QPen(no_pen), QBrush(QColor('#FF8200')))
                prog.setZValue(rect_item.zValue()+1)
                try:
                    prog.setData(0, name)
                    prog.setAcceptHoverEvents(True)
                except Exception:
                    pass
            # Text (wrap / elide simple)
            txt = name
            if len(txt) > 40:
                txt = txt[:37] + '…'
            text_item = self.scene.addText(txt)
            f = QFont(font)
            f.setPointSize(f.pointSize()-1)
            f.setBold(True)
            text_item.setFont(f)
            text_item.setDefaultTextColor(QColor('white'))
            br = text_item.boundingRect()
            text_item.setPos(x + (node_w - br.width())/2, y + 6)
            text_item.setZValue(rect_item.zValue()+2)
            try:
                # Carry the name so hover/click on text resolves the part
                text_item.setData(0, name)
                text_item.setAcceptHoverEvents(True)
            except Exception:
                pass
            # Status / percent line
            meta_line = f"{status or ''}  {pc}%".strip()
            if meta_line:
                meta_item = self.scene.addText(meta_line)
                f2 = QFont(font); f2.setPointSize(f2.pointSize()-2)
                meta_item.setFont(f2)
                meta_item.setDefaultTextColor(QColor('#dddddd'))
                mbr = meta_item.boundingRect()
                meta_item.setPos(x + (node_w - mbr.width())/2, y + node_h - mbr.height() - 12)
                meta_item.setZValue(rect_item.zValue()+2)
                try:
                    meta_item.setData(0, name)
                    meta_item.setAcceptHoverEvents(True)
                except Exception:
                    pass
            self._name_to_item[name] = rect_item
        # Interaction
        self.scene.installEventFilter(self)
        # Autosize scene rect
        r = self.scene.itemsBoundingRect()
        padded = r.adjusted(-40, -40, 40, 40)
        self.scene.setSceneRect(padded)
        # Capture deterministic initial view rectangle once
        try:
            if getattr(self, '_initial_view_rect', None) is None or getattr(self._initial_view_rect, 'isNull', lambda: False)():
                self._initial_view_rect = padded
        except Exception:
            pass
        # Initial fit
        try:
            try:
                self.view.fitInView(self.scene.sceneRect(), _keep_ar())
            except Exception:
                self.view.fitInView(self.scene.sceneRect(), _keep_ar())
        except Exception:
            pass
        # After initial fit, apply persisted zoom scale (if user had manual zoom). Re-run restore.
        try:
            self.view._restore_zoom()
        except Exception:
            pass
        self._update_minimap()

    # -------- Event Handling --------
    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        qevent_type = getattr(QEvent, 'Type', QEvent)
        if event.type() in (qevent_type.Wheel, qevent_type.Resize):
            try:
                if hasattr(self, '_minimap_timer'):
                    self._minimap_timer.start(120)
                else:
                    self._update_minimap()
            except Exception:
                pass
        if event.type() == qevent_type.GraphicsSceneMousePress:
            item = self.scene.itemAt(event.scenePos(), self.view.transform())
            if item:
                name = item.data(0)
                if isinstance(name, str) and name:
                    # If node has a Pace Link, open on left-click
                    try:
                        _left = getattr(Qt, 'MouseButton', Qt).LeftButton
                    except Exception:
                        _left = 1
                    if hasattr(event, 'button') and event.button() == _left:
                        row = self._name_to_row.get(name, {})
                        url = (row.get('Pace Link') or '').strip()
                        if getattr(self, '_show_links', True) and url and (url.lower().startswith('http://') or url.lower().startswith('https://')):
                            try:
                                from PyQt6.QtGui import QDesktopServices
                                from PyQt6.QtCore import QUrl
                                QDesktopServices.openUrl(QUrl(url))
                            except Exception:
                                import webbrowser; webbrowser.open(url)
                            return True
                    if name.startswith("__toggle__::"):
                        target = name.split("::",1)[1]
                        if target in self._collapsed:
                            self._collapsed.remove(target)
                        else:
                            self._collapsed.add(target)
                        # Persist collapsed
                        try:
                            from PyQt6.QtCore import QSettings
                            _ts = QSettings('LSI','ProjectApp')
                            _ts.setValue('TreeCollapsed', list(self._collapsed))
                        except Exception:
                            pass
                        self.refresh()
                        return True
                    # selection + callback
                    if self.on_part_selected:
                        self.on_part_selected(name)
                    # Also sync highlight in gantt if available (without switching view)
                    try:
                        mw = self.window()
                        if hasattr(mw, 'gantt_chart_view') and hasattr(mw.gantt_chart_view, 'highlight_bar'):
                            # Ensure gantt model is current
                            mw.gantt_chart_view.render_gantt(mw.model)
                            mw.gantt_chart_view.highlight_bar(name)
                    except Exception:
                        pass
                    # zoom to node
                    try:
                        r = item.sceneBoundingRect()
                        if not r.isNull():
                            rect_focus = r.adjusted(-80,-60,80,60)
                            try:
                                if hasattr(self.view, 'smoothFocusRect'):
                                    self.view.smoothFocusRect(rect_focus)
                                else:
                                    self.view.fitInView(rect_focus, _keep_ar())
                            except Exception:
                                self.view.fitInView(rect_focus, _keep_ar())
                    except Exception:
                        pass
                    # preview image
                    if self._hover_preview_enabled:
                        row = self._name_to_row.get(name)
                        if row:
                            self._show_image_for_row(row)
            if event.button() == 2:  # Right click fallback if item not matched
                self._show_context_menu(event.screenPos(), None)
        elif event.type() == qevent_type.GraphicsSceneContextMenu:
            item = self.scene.itemAt(event.scenePos(), self.view.transform())
            target_name = None
            if item:
                d = item.data(0)
                if isinstance(d, str) and d and not d.startswith("__toggle__::"):
                    target_name = d
            self._show_context_menu(event.screenPos(), target_name)
        elif event.type() == qevent_type.GraphicsSceneHoverMove:
            if not self._hover_preview_enabled:
                self.preview_label.clear()
            else:
                item = self.scene.itemAt(event.scenePos(), self.view.transform())
                if item:
                    name = item.data(0)
                    if isinstance(name, str):
                        row = self._name_to_row.get(name)
                        if row:
                            self._show_image_for_row(row)
        elif event.type() == qevent_type.GraphicsSceneMouseMove:
            # Fallback: some items might not emit hover events; use mouse move to drive previews
            if not self._hover_preview_enabled:
                self.preview_label.clear()
            else:
                try:
                    item = self.scene.itemAt(event.scenePos(), self.view.transform())
                    if item:
                        name = item.data(0)
                        if isinstance(name, str):
                            row = self._name_to_row.get(name)
                            if row:
                                self._show_image_for_row(row)
                except Exception:
                    pass
        elif event.type() == qevent_type.GraphicsSceneHoverLeave:
            self.preview_label.clear()
        return super().eventFilter(obj, event)

    # -------- Minimap --------
    def _update_minimap(self):
        if not hasattr(self, '_mini_scene'):
            return
        try:
            if self.scene is None:
                return
        except Exception:
            return
        try:
            self._mini_scene.clear()
        except Exception:
            return
        # Render simplified rectangles from main scene
        scale_factor = 0.12
        for it in self.scene.items():
            try:
                if it.data(0):
                    from PyQt6.QtGui import QBrush, QColor
                    r = it.sceneBoundingRect()
                    rect = self._mini_scene.addRect(r.x()*scale_factor, r.y()*scale_factor, r.width()*scale_factor, r.height()*scale_factor,
                                                    pen=(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), brush=QBrush(QColor(255,130,0,90)))
            except Exception:
                pass
        # Viewport box
        try:
            from PyQt6.QtGui import QPen, QColor, QBrush
            vr = self.view.mapToScene(self.view.viewport().rect()).boundingRect()
            no_brush = getattr(getattr(Qt,'BrushStyle', Qt), 'NoBrush', getattr(Qt,'NoBrush', 0))
            vp = self._mini_scene.addRect(vr.x()*scale_factor, vr.y()*scale_factor, vr.width()*scale_factor, vr.height()*scale_factor,
                                          pen=QPen(QColor('#ffffff')), brush=QBrush(no_brush))
            vp.setZValue(999)
        except Exception:
            pass
        # Fit minimap view
        try:
            self._mini_view.fitInView(self._mini_scene.itemsBoundingRect().adjusted(-4,-4,4,4), _keep_ar())
        except Exception:
            pass

    # -------- Context Menu --------
    def _show_context_menu(self, screen_pos, name):
        from PyQt6.QtWidgets import QMenu, QAction, QApplication
        menu = QMenu()
        act_open = QAction("Open Details", menu)
        act_jump = QAction("Jump To In Gantt", menu)
        act_copy = QAction("Copy Name", menu)
        act_set_parent = QAction("Set Parent…", menu)
        act_expand = QAction("Expand Subtree", menu)
        act_collapse = QAction("Collapse Subtree", menu)
        for a in (act_open, act_jump, act_copy, act_set_parent, act_expand, act_collapse):
            menu.addAction(a)
        if not name:
            act_open.setEnabled(False); act_jump.setEnabled(False); act_copy.setEnabled(False); act_set_parent.setEnabled(False); act_expand.setEnabled(False); act_collapse.setEnabled(False)
        chosen = menu.exec(screen_pos)
        if not chosen or not name:
            return
        if chosen == act_open:
            if self.on_part_selected:
                self.on_part_selected(name)
        elif chosen == act_jump:
            self._emit_jump_to_gantt(name)
        elif chosen == act_copy:
            QApplication.clipboard().setText(name)
        elif chosen == act_set_parent:
            self._set_parent_dialog(name)
        elif chosen == act_expand:
            self._expand_subtree(name)
        elif chosen == act_collapse:
            self._collapse_subtree(name)

    def _expand_subtree(self, name):
        queue = [name]
        while queue:
            n = queue.pop()
            if n in self._collapsed:
                self._collapsed.remove(n)
            # add children
            for r in self.model.rows:
                if r.get('Parent') == n:
                    queue.append(r.get('Project Part',''))
        # Persist
        try:
            from PyQt6.QtCore import QSettings
            _ts = QSettings('LSI','ProjectApp')
            _ts.setValue('TreeCollapsed', list(self._collapsed))
        except Exception:
            pass
        self.refresh()

    def _collapse_subtree(self, name):
        queue = [name]
        while queue:
            n = queue.pop()
            self._collapsed.add(n)
            for r in self.model.rows:
                if r.get('Parent') == n:
                    queue.append(r.get('Project Part',''))
        try:
            from PyQt6.QtCore import QSettings
            _ts = QSettings('LSI','ProjectApp')
            _ts.setValue('TreeCollapsed', list(self._collapsed))
        except Exception:
            pass
        self.refresh()

    def _emit_jump_to_gantt(self, name):
        if hasattr(self, 'on_jump_to_gantt') and self.on_jump_to_gantt:
            try:
                self.on_jump_to_gantt(name)
                return
            except Exception:
                pass
        try:
            print(f"[JumpToGantt] {name}")
        except Exception:
            pass

    # -------- Reparent Dialog --------
    def _set_parent_dialog(self, target_name):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QComboBox, QDialogButtonBox
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Set Parent - {target_name}")
        v = QVBoxLayout(dlg)
        v.addWidget(QLabel("Choose new parent (or blank for top level):"))
        combo = QComboBox(); combo.addItem("<None>")
        # Candidate parents exclude target and its descendants
        descendants = self._collect_descendants(target_name)
        for r in self.model.rows:
            n = r.get('Project Part','')
            if n and n != target_name and n not in descendants:
                combo.addItem(n)
        v.addWidget(combo)
        try:
            std = QDialogButtonBox.StandardButton
            buttons = QDialogButtonBox(std.Ok | std.Cancel)
        except Exception:
            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        v.addWidget(buttons)
        def apply():
            new_parent = combo.currentText()
            if new_parent == "<None>":
                new_parent = None
            for r in self.model.rows:
                if r.get('Project Part') == target_name:
                    r['Parent'] = new_parent
                    break
            try:
                if hasattr(self.model, 'save_to_db'):
                    self.model.save_to_db()
            except Exception as e:
                print(f"Reparent save failed: {e}")
            self.refresh()
            dlg.accept()
        buttons.accepted.connect(apply)
        buttons.rejected.connect(dlg.reject)
        dlg.exec()

    def _collect_descendants(self, name):
        out = set(); queue = [name]
        while queue:
            n = queue.pop()
            for r in self.model.rows:
                if r.get('Parent') == n:
                    child = r.get('Project Part','')
                    if child and child not in out:
                        out.add(child); queue.append(child)
        return out

    def _show_image_for_row(self, row):
        img_path = row.get('Images','')
        if img_path and str(img_path).strip():
            from PyQt6.QtGui import QPixmap
            full = resolve_resource_path(img_path)
            # LRU cache lookup
            pm = None
            try:
                if full in self._preview_cache:
                    pm = self._preview_cache[full]
                    # move to MRU
                    try:
                        self._preview_cache_order.remove(full)
                    except ValueError:
                        pass
                    self._preview_cache_order.append(full)
                else:
                    pm = QPixmap(full)
                    if not pm.isNull():
                        self._preview_cache[full] = pm
                        self._preview_cache_order.append(full)
                        if len(self._preview_cache_order) > self._preview_cache_cap:
                            # evict LRU
                            lru = self._preview_cache_order.pop(0)
                            self._preview_cache.pop(lru, None)
            except Exception:
                pm = QPixmap(full)
            if not pm.isNull():
                self.preview_label.setPixmap(pm.scaledToHeight(120, _smooth_mode()))
                self.preview_label.setText("")
                return
        self.preview_label.setText("")
        self.preview_label.clear()
    # Export tree scene wrapper
    def _export_tree(self):
        try:
            self._export_scene_with_header(self.scene, title='Project Tree')
        except Exception as e:
            print(f'Tree export failed: {e}')
    # Simple reuse: delegate to Gantt export implementation style if available else fallback
    def _export_scene_with_header(self, scene, title='Export'):
        # Minimal reuse by instantiating a temporary GanttChartView exporter if complexity grows; for now simple call to existing logic
        from PyQt6.QtCore import QSettings
        from PyQt6.QtGui import QPainter, QPixmap
        from PyQt6.QtWidgets import QFileDialog, QApplication
        import os
        s = QSettings('LSI','ProjectPlanner')
        pref_format = s.value('Export/format','PNG')
        ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
        include_header = s.value('Export/include_header', True)
        if isinstance(include_header, str): include_header = include_header.lower() in ('1','true','yes','on')
        init_name = f"{title.lower().replace(' ','_')}.{ 'pdf' if pref_format=='PDF' else 'png'}"
        filters = 'PDF Files (*.pdf);;PNG Files (*.png)' if pref_format=='PDF' else 'PNG Files (*.png);;PDF Files (*.pdf)'
        path, chosen = QFileDialog.getSaveFileName(self, f'Export {title}', init_name, filters)
        if not path: return
        rect = scene.sceneRect().toRect()
        if rect.isEmpty():
            print('Scene empty; abort export.')
            return
        is_pdf = path.lower().endswith('.pdf') or (chosen and 'PDF' in chosen and not path.lower().endswith('.png'))
        if is_pdf and not path.lower().endswith('.pdf'): path += '.pdf'
        if (not is_pdf) and not path.lower().endswith('.png'): path += '.png'
        header_path = resolve_resource_path('header.png')
        header_pixmap = QPixmap(header_path) if os.path.exists(header_path) else None
        svg_path = resolve_resource_path('header.svg')
        header_is_svg=False; header_svg_renderer=None
        try:
            if os.path.exists(svg_path):
                from PyQt6.QtSvg import QSvgRenderer
                r = QSvgRenderer(svg_path)
                if r.isValid():
                    header_is_svg=True; header_svg_renderer=r
        except Exception:
            pass
        footer_text = "© 2025 LSI – For Internal Use Only"
        if is_pdf:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtCore import QMarginsF, QRectF
            from math import ceil
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.PdfFormat)
            # Use page size/orientation from settings
            page_size = s.value('Export/page_size','A4'); orientation = s.value('Export/orientation','Portrait')
            size_map={'A4':QPrinter.A4,'Letter':QPrinter.Letter,'Legal':QPrinter.Legal,'Tabloid':QPrinter.Tabloid}
            printer.setPaperSize(size_map.get(page_size, QPrinter.A4))
            printer.setOrientation(QPrinter.Portrait if orientation=='Portrait' else QPrinter.Landscape)
            try: printer.setPageMargins(QMarginsF(ml,mt,mr,mb))
            except Exception: pass
            painter = QPainter(printer)
            page_rect = printer.pageRect(); y_offset=0
            def _svg_h(renderer, tw, default_ratio=0.12, min_h=40):
                try:
                    ds=renderer.defaultSize(); w,h=ds.width(), ds.height()
                    if w<=0 or h<=0:
                        vb=renderer.viewBoxF(); w,h=vb.width(), vb.height()
                    if w>0 and h>0 and tw>0:
                        return int(round(h*(tw/float(w))))
                except Exception:
                    pass
                # Fallback: reasonable banner height as a fraction of page width
                return max(min_h, int(round(tw*default_ratio))) if tw>0 else None
            if include_header:
                if header_is_svg and header_svg_renderer:
                    tw=page_rect.width(); hh=_svg_h(header_svg_renderer, tw)
                    if hh:
                        header_svg_renderer.render(painter, QRectF(0,0,tw,hh)); y_offset=hh+10
                elif header_pixmap and not header_pixmap.isNull():
                    try:
                        _smooth = Qt.TransformationMode.SmoothTransformation
                    except Exception:
                        _smooth = getattr(Qt, 'SmoothTransformation', 1)
                    sh=header_pixmap.scaledToWidth(page_rect.width(), _smooth); painter.drawPixmap((page_rect.width()-sh.width())//2,0,sh); y_offset=sh.height()+10
            avail_h=max(1,page_rect.height()-y_offset); scale=avail_h/rect.height(); from math import ceil
            scaled_total_w=max(1.0, rect.width()*scale); cols=max(1,int(ceil(scaled_total_w/page_rect.width())))
            for col in range(cols):
                if col>0:
                    printer.newPage()
                    if include_header:
                        if header_is_svg and header_svg_renderer:
                            tw=page_rect.width(); hh=_svg_h(header_svg_renderer, tw)
                            if hh: header_svg_renderer.render(painter, QRectF(0,0,tw,hh))
                        elif header_pixmap and not header_pixmap.isNull():
                            try:
                                _smooth = Qt.TransformationMode.SmoothTransformation
                            except Exception:
                                _smooth = getattr(Qt, 'SmoothTransformation', 1)
                            sh=header_pixmap.scaledToWidth(page_rect.width(), _smooth); painter.drawPixmap((page_rect.width()-sh.width())//2,0,sh)
                painter.save(); painter.translate(0,y_offset); painter.scale(scale,scale)
                source_x=(col*page_rect.width())/scale
                scene.render(painter, target=QRectF(0,0,page_rect.width(), rect.height()*scale), source=QRectF(source_x,0,page_rect.width()/scale, rect.height()))
                painter.restore()
                # Footer
                try:
                    from PyQt6.QtGui import QFont
                    painter.save()
                    f = QFont(); f.setPointSizeF(f.pointSizeF()*0.85)
                    painter.setFont(f)
                    # Compose footer with page numbers: Page X of Y
                    try:
                        page_num = col + 1
                        page_total = cols
                        right_text = f"Page {page_num} of {page_total}"
                    except Exception:
                        right_text = ""
                    footer_y = page_rect.height() - 14
                    # Left: small copyright, Center: metadata, Right: page X of Y
                    left_txt = "© 2025 LSI"
                    center_txt = footer_text
                    painter.drawText(QRectF(8, footer_y, page_rect.width()/3 - 12, 14), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, left_txt)
                    painter.drawText(QRectF(page_rect.width()/3, footer_y, page_rect.width()/3, 14), Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter, center_txt)
                    painter.drawText(QRectF(page_rect.width()*2/3, footer_y, page_rect.width()/3 - 8, 14), Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, right_text)
                    painter.restore()
                except Exception:
                    pass
            painter.end(); print(f'Exported PDF -> {path}'); return
        # PNG branch
        screen = QApplication.primaryScreen(); dpi = screen.logicalDotsPerInch() if screen else 96.0
        def mm_to_px(mm): return int(round((mm/25.4)*dpi))
        pad_l,pad_t,pad_r,pad_b=[mm_to_px(v) for v in (ml,mt,mr,mb)]
        content_pix = QPixmap(rect.width()+pad_l+pad_r, rect.height()+pad_t+pad_b); content_pix.fill()
        painter = QPainter(content_pix); painter.translate(pad_l,pad_t); scene.render(painter); painter.end()
        if not include_header:
            content_pix.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        # Combine with header (pref SVG)
        if header_is_svg and header_svg_renderer:
            def _svg_h(renderer, tw, default_ratio=0.12, min_h=40):
                try:
                    ds=renderer.defaultSize(); w,h=ds.width(), ds.height()
                    if w<=0 or h<=0:
                        vb=renderer.viewBoxF(); w,h=vb.width(), vb.height()
                    if w>0 and h>0 and tw>0:
                        return int(round(h*(tw/float(w))))
                except Exception:
                    pass
                return max(min_h, int(round(tw*default_ratio))) if tw>0 else None
            tw = content_pix.width(); hh=_svg_h(header_svg_renderer, tw)
            combo = QPixmap(tw, hh+content_pix.height()); combo.fill(); painter=QPainter(combo)
            from PyQt6.QtCore import QRectF
            header_svg_renderer.render(painter, QRectF(0,0,tw,hh)); painter.drawPixmap(0,hh,content_pix);
            # Footer (PNG)
            try:
                from PyQt6.QtGui import QFont
                f = QFont(); f.setPointSizeF(f.pointSizeF()*0.85)
                painter.setFont(f)
                left_txt = "© 2025 LSI"
                center_txt = footer_text
                painter.drawText(8, hh+content_pix.height()-18, tw//3 - 12, 16, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, left_txt)
                painter.drawText(tw//3, hh+content_pix.height()-18, tw//3, 16, Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter, center_txt)
            except Exception:
                pass
            painter.end(); combo.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        if header_pixmap and not header_pixmap.isNull():
            cw = max(header_pixmap.width(), content_pix.width()); combo = QPixmap(cw, header_pixmap.height()+content_pix.height()); combo.fill()
            painter = QPainter(combo); hx=(cw-header_pixmap.width())//2; painter.drawPixmap(hx,0, header_pixmap); painter.drawPixmap(0, header_pixmap.height(), content_pix)
            try:
                from PyQt6.QtGui import QFont
                f = QFont(); f.setPointSizeF(f.pointSizeF()*0.85)
                painter.setFont(f)
                left_txt = "© 2025 LSI"
                center_txt = footer_text
                painter.drawText(8, header_pixmap.height()+content_pix.height()-18, cw//3 - 12, 16, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, left_txt)
                painter.drawText(cw//3, header_pixmap.height()+content_pix.height()-18, cw//3, 16, Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter, center_txt)
            except Exception:
                pass
            painter.end(); combo.save(path,'PNG'); print(f'Exported PNG -> {path}'); return
        # Footer only (no header version)
        painter = QPainter(content_pix)
        try:
            from PyQt6.QtGui import QFont
            f = QFont(); f.setPointSizeF(f.pointSizeF()*0.85)
            painter.setFont(f)
            left_txt = "© 2025 LSI"
            center_txt = footer_text
            painter.drawText(8, content_pix.height()-18, content_pix.width()//3 - 12, 16, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, left_txt)
            painter.drawText(content_pix.width()//3, content_pix.height()-18, content_pix.width()//3, 16, Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter, center_txt)
        except Exception:
            pass
        painter.end(); content_pix.save(path,'PNG'); print(f'Exported PNG -> {path}')


# Add a custom QGraphicsView subclass for zooming
from PyQt6.QtWidgets import QGraphicsView
from PyQt6.QtCore import Qt, pyqtSignal
# Backward compat: provide PyQt5-style attribute name if missing
try:
    if not hasattr(QGraphicsView, 'ScrollHandDrag') and hasattr(QGraphicsView, 'DragMode') and hasattr(QGraphicsView.DragMode, 'ScrollHandDrag'):
        QGraphicsView.ScrollHandDrag = QGraphicsView.DragMode.ScrollHandDrag  # type: ignore[attr-defined]
except Exception:
    pass

class ZoomableGraphicsView(QGraphicsView):
    zoomChanged = pyqtSignal(float)
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._zoom = 0
        # PyQt6 enum namespace change
        try:
            self.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        except Exception:
            # Fallback: if DragMode enum missing (very old PyQt), ignore
            try:
                m = getattr(QGraphicsView, 'DragMode', None)
                if m and hasattr(m, 'ScrollHandDrag'):
                    self.setDragMode(m.ScrollHandDrag)
            except Exception:
                pass
        self._settings_key = None  # e.g., 'GanttZoom' or 'TimelineZoom'
        # Enable custom context menu
        try:
            self.setContextMenuPolicy(Qt.ContextMenuPolicy.DefaultContextMenu)
        except Exception:
            pass

    def wheelEvent(self, event):
        try:
            _ctrl_mod = Qt.KeyboardModifier.ControlModifier
        except Exception:
            _ctrl_mod = getattr(Qt, 'ControlModifier', 0)
        if event.modifiers() & _ctrl_mod:
            if event.angleDelta().y() > 0:
                self.zoomIn()
            else:
                self.zoomOut()
        else:
            super().wheelEvent(event)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Plus or event.key() == Qt.Key_Equal:
            self.zoomIn()
        elif event.key() == Qt.Key_Minus:
            self.zoomOut()
        else:
            super().keyPressEvent(event)

    def zoomIn(self):
        self._zoom += 1
        self.scale(1.2, 1.2)
        self._persist_zoom()
        try:
            self.zoomChanged.emit(float(self.transform().m11()))
        except Exception:
            pass

    def zoomOut(self):
        self._zoom -= 1
        self.scale(1/1.2, 1/1.2)
        self._persist_zoom()
        try:
            self.zoomChanged.emit(float(self.transform().m11()))
        except Exception:
            pass

    def resetZoom(self):
        self.resetTransform()
        self._zoom = 0
        self._persist_zoom()
        try:
            self.zoomChanged.emit(float(self.transform().m11()))
        except Exception:
            pass

    def setSettingsKey(self, key: str):
        self._settings_key = key
        self._restore_zoom()

    def _persist_zoom(self):
        if not self._settings_key:
            return
        try:
            from PyQt6.QtCore import QSettings
            s = QSettings("LSI", "ProjectPlanner")
            # store current scale factor from transform
            s.setValue(self._settings_key, float(self.transform().m11()))
        except Exception:
            pass

    def _restore_zoom(self):
        if not self._settings_key:
            return
        try:
            from PyQt6.QtCore import QSettings
            s = QSettings("LSI", "ProjectPlanner")
            val = s.value(self._settings_key, None)
            if val is not None:
                try:
                    scale_factor = float(val)
                    self.resetTransform()
                    # apply uniform scale on both axes
                    self.scale(scale_factor, scale_factor)
                    try:
                        self.zoomChanged.emit(float(self.transform().m11()))
                    except Exception:
                        pass
                except Exception:
                    pass
        except Exception:
            pass

    # Context menu with quick zoom/fit actions
    def contextMenuEvent(self, event):
        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        act_fit_all = menu.addAction("Fit All")
        act_fit_sel = menu.addAction("Fit to Selection")
        act_reset = menu.addAction("Reset Zoom")
        act_100 = menu.addAction("Zoom 100%")
        chosen = menu.exec(event.globalPos())
        if not chosen:
            return
        try:
            if chosen == act_fit_all:
                scn = self.scene()
                if scn is not None:
                    r = scn.itemsBoundingRect()
                    if not r.isNull():
                        self.fitInView(r, _keep_ar())
            elif chosen == act_fit_sel:
                scn = self.scene()
                if scn is not None:
                    items = list(scn.selectedItems())
                    if items:
                        from PyQt6.QtCore import QRectF
                        rect = QRectF()
                        for it in items:
                            rect = rect.united(it.sceneBoundingRect())
                        if not rect.isNull():
                            self.fitInView(rect, _keep_ar())
                    else:
                        r = scn.itemsBoundingRect()
                        if not r.isNull():
                            self.fitInView(r, _keep_ar())
            elif chosen == act_reset:
                self.resetZoom()
                return
            elif chosen == act_100:
                self.resetTransform(); self._zoom = 0
            # Persist and emit
            try:
                if hasattr(self, '_persist_zoom'):
                    self._persist_zoom()
                self.zoomChanged.emit(float(self.transform().m11()))
            except Exception:
                pass
        except Exception:
            pass


class GanttChartView(QWidget):
    # --- Filtering Support (extensible for future filter panel) ---
    def _init_filters(self):
        # Stored criteria; None / empty means no filtering
        self._filter_statuses = None          # set of status strings
        self._filter_internal_external = None # set like {"Internal","External"}
        self._filter_responsible_substr = None # lowercase substring
        self._filter_critical_only = False     # boolean
        self._filter_risk_only = False         # boolean (overdue OR at-risk)
        self._current_critical_set = set()     # populated during render
        # Feature toggles (default; may be overridden by settings/checkbox in __init__)
        if not hasattr(self, '_show_unscheduled'):
            self._show_unscheduled = True

    def set_filters(self, statuses=None, internal_external=None, responsible_substr=None,
                    critical_only=None, risk_only=None, show_unscheduled=None):
        """Update filter criteria and refresh the Gantt chart.
        Parameters are optional; pass None to leave unchanged, pass empty iterable/string to clear.
        """
        if statuses is not None:
            self._filter_statuses = set(s.strip() for s in statuses if s.strip()) if statuses else None
        if internal_external is not None:
            self._filter_internal_external = set(s.strip() for s in internal_external if s.strip()) if internal_external else None
        if responsible_substr is not None:
            rs = responsible_substr.strip()
            self._filter_responsible_substr = rs.lower() if rs else None
        if critical_only is not None:
            self._filter_critical_only = bool(critical_only)
        if risk_only is not None:
            self._filter_risk_only = bool(risk_only)
        if show_unscheduled is not None:
            self._show_unscheduled = bool(show_unscheduled)
        if hasattr(self, 'model') and self.model:
            self.render_gantt(self.model)

    def _passes_filters(self, row):
        try:
            if self._filter_statuses and (row.get("Status") or "").strip() not in self._filter_statuses:
                return False
            if self._filter_internal_external and (row.get("Internal/External") or "").strip() not in self._filter_internal_external:
                return False
            if self._filter_responsible_substr:
                resp = (row.get("Responsible") or "").lower()
                if self._filter_responsible_substr not in resp:
                    return False
            # Critical path filter
            if self._filter_critical_only:
                name = row.get("Project Part", "")
                if name not in self._current_critical_set:
                    return False
            # Risk filter (overdue OR at-risk)
            if self._filter_risk_only:
                import datetime as _dt_rf
                overdue = False; at_risk = False
                try:
                    # Derive start & end
                    start_str = row.get("Start Date", "")
                    dur = int(row.get("Duration (days)") or 0)
                    if start_str:
                        start_dt = _dt_rf.datetime.strptime(start_str, "%m-%d-%Y")
                    else:
                        start_dt = None
                    if row.get("Calculated End Date"):
                        scheduled_end = _dt_rf.datetime.strptime(row.get("Calculated End Date"), "%m-%d-%Y")
                    elif start_dt and dur:
                        scheduled_end = start_dt + _dt_rf.timedelta(days=dur)
                    else:
                        scheduled_end = None
                    today = _dt_rf.datetime.today()
                    pc_val = int(row.get("% Complete") or 0)
                    status_val = (row.get("Status") or "").strip()
                    if scheduled_end and pc_val < 100 and today.date() > scheduled_end.date():
                        overdue = True
                    elif start_dt and pc_val == 0 and status_val in ("Planned", "Blocked") and today.date() > start_dt.date():
                        at_risk = True
                except Exception:
                    pass
                if not (overdue or at_risk):
                    return False
            return True
        except Exception:
            return True

    # --- Public helper to highlight & scroll to a bar by name (used by search/jump) ---
    def highlight_bar(self, part_name):
        if not part_name:
            return
        # Clear previous bar highlight pen styling
        try:
            for item in getattr(self, '_name_to_rect', {}).values():
                if item and hasattr(item, 'setPen'):
                    from PyQt6.QtGui import QPen
                    item.setPen(item.data(99) or QPen(item.pen()))
        except Exception:
            pass
        rect_item = getattr(self, '_name_to_rect', {}).get(part_name)
        if rect_item:
            from PyQt6.QtGui import QPen, QColor
            # Store original pen once
            if rect_item.data(99) is None:
                rect_item.setData(99, rect_item.pen())
            pen = QPen(QColor('#00BFFF'))
            pen.setWidth(3)
            rect_item.setPen(pen)
            # Center the view on the rectangle
            if hasattr(self, 'view') and self.view:
                center_pt = rect_item.sceneBoundingRect().center()
                self.view.centerOn(center_pt)
            # Also use existing connector/label highlight logic
            self._highlight_connectors(part_name, True)
        else:
            print(f"highlight_bar: No bar found for '{part_name}' (may be filtered out)")
   
    def show_edit_dialog(self, row):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit Project Part: {row.get('Project Part', '')}")
        layout = QFormLayout(dialog)
        edits = {}
        # Early QLabel alias to simplify guarded usages later
        try:
            from PyQt6.QtWidgets import QLabel as _QLocalLabel
        except Exception:
            _QLocalLabel = None
        # Load pricing settings for suggestions
        try:
            from PyQt6.QtCore import QSettings
            _ps = QSettings('LSI','ProjectPlanner')
            target_margin = float(_ps.value('Pricing/target_margin', 35.0)) / 100.0
        except Exception:
            target_margin = 0.35
        suggested_prod = None; suggested_inst = None
        for col in self.model.COLUMNS:
            val = row.get(col, "")
            if col == "Dependencies":
                # Enhanced Dependencies picker with filter persistence, IDs, helper buttons, cycle detection
                from PyQt6.QtWidgets import (
                    QLineEdit as _QLineEdit, QPushButton as _QPushButton, QDialog as _DepDlg,
                    QVBoxLayout as _VB, QListWidget, QListWidgetItem, QHBoxLayout as _HB,
                    QLabel as _Lbl, QMessageBox, QHBoxLayout as _QHBoxLayout
                )
                from PyQt6.QtCore import QSettings as _QS
                dep_edit = _QLineEdit(str(val) if val else "")
                dep_edit.setPlaceholderText("Comma-separated part names or numeric IDs")
                edits[col] = dep_edit
                def open_dep_picker():
                    d = _DepDlg(dialog); d.setWindowTitle("Select Dependencies"); vb = _VB(d)
                    filter_edit = _QLineEdit(); filter_edit.setPlaceholderText("Filter…")
                    # restore last filter
                    try:
                        prev = _QS('LSI','ProjectPlanner').value('DepsPicker/filter','')
                        if prev: filter_edit.setText(prev)
                    except Exception: pass
                    vb.addWidget(filter_edit)
                    lst = QListWidget(); lst.setSelectionMode(QListWidget.MultiSelection); vb.addWidget(lst,1)
                    cur_tokens = {t.strip() for t in (dep_edit.text() or '').split(',') if t.strip()}
                    # id mapping
                    name_to_id = {}
                    try:
                        import os, sqlite3
                        if os.path.exists(self.model.DB_FILE):
                            with self.model._connect() as _c:
                                cur = _c.cursor(); cur.execute('SELECT id, "Project Part" FROM project_parts')
                                for rid, pname in cur.fetchall(): name_to_id[pname]=rid
                    except Exception: pass
                    current_name = row.get('Project Part','')
                    for r2 in self.model.rows:
                        pname = r2.get('Project Part','')
                        if not pname or pname == current_name: continue
                        disp = f"{pname} ({name_to_id[pname]})" if pname in name_to_id else pname
                        it = QListWidgetItem(disp, lst)
                        if pname in cur_tokens: it.setSelected(True)
                        it.setData(Qt.UserRole, pname)
                    def apply_filter():
                        q = filter_edit.text().strip().lower()
                        for i in range(lst.count()):
                            it = lst.item(i); it.setHidden(q not in it.text().lower())
                    filter_edit.textChanged.connect(apply_filter)
                    helpers = _HB(); btn_par = _QPushButton('Select Parents'); btn_cp = _QPushButton('Select Critical Path'); btn_clear = _QPushButton('Clear')
                    for b in (btn_par, btn_cp, btn_clear): helpers.addWidget(b)
                    helpers.addStretch(1); vb.addLayout(helpers)
                    def do_par():
                        target = row.get('Parent') or ''; name_map = {r.get('Project Part',''): r for r in self.model.rows}; names=set()
                        while target:
                            names.add(target); target = name_map.get(target,{}).get('Parent') or ''
                        for i in range(lst.count()):
                            it = lst.item(i)
                            if it.data(Qt.UserRole) in names: it.setSelected(True)
                    btn_par.clicked.connect(do_par)
                    def do_cp():
                        crit = getattr(self, '_current_critical_set', set())
                        for i in range(lst.count()):
                            it = lst.item(i)
                            if it.data(Qt.UserRole) in crit: it.setSelected(True)
                    btn_cp.clicked.connect(do_cp)
                    def do_cl():
                        for i in range(lst.count()): lst.item(i).setSelected(False)
                    btn_clear.clicked.connect(do_cl)
                    btn_row = _HB(); ok=_QPushButton('OK'); canc=_QPushButton('Cancel'); btn_row.addStretch(1); btn_row.addWidget(ok); btn_row.addWidget(canc); vb.addLayout(btn_row)
                    def accept():
                        sels=[]
                        for i in range(lst.count()):
                            it = lst.item(i)
                            if it.isSelected(): sels.append(it.data(Qt.UserRole))
                        if sels and _would_create_cycle(self.model, row.get('Project Part',''), set(sels)):
                            if QMessageBox.warning(d,'Cycle Detected','Cycle introduced. Proceed?', QMessageBox.Yes|QMessageBox.No, QMessageBox.No) != QMessageBox.Yes:
                                return
                        dep_edit.setText(', '.join(sorted(sels)))
                        try: _QS('LSI','ProjectPlanner').setValue('DepsPicker/filter', filter_edit.text())
                        except Exception: pass
                        d.accept()
                    ok.clicked.connect(accept); canc.clicked.connect(d.reject)
                    d.setLayout(vb); d.resize(420,500); d.exec()
                pick_btn = _QPushButton('Select…'); pick_btn.clicked.connect(open_dep_picker)
                # Use direct class if globally imported, else fallback to alias to avoid UnboundLocalError
                try:
                    hb = QHBoxLayout()
                except Exception:
                    try:
                        hb = _QHBoxLayout()
                    except Exception:
                        from PyQt6.QtWidgets import QHBoxLayout as __QL
                        hb = __QL()
                hb.addWidget(dep_edit,1); hb.addWidget(pick_btn); layout.addRow(col, hb)
                continue
            if col in ("Start Date", "Calculated End Date"):
                date_edit = QDateEdit()
                date_edit.setCalendarPopup(True)
                min_blank = QDate(1753, 1, 1)
                date_edit.setMinimumDate(min_blank)
                date_edit.setSpecialValueText("")
                if val:
                    date = QDate.fromString(val, "MM-dd-yyyy")
                    if date.isValid() and date != QDate(1752, 9, 14) and date != min_blank:
                        date_edit.setDate(date)
                    else:
                        date_edit.setDate(min_blank)
                else:
                    date_edit.setDate(min_blank)
                edits[col] = date_edit
                layout.addRow(col, date_edit)
            elif col == "Type":
                combo = QComboBox()
                combo.addItems(["Milestone", "Phase", "Feature", "Item"])
                if val:
                    combo.setCurrentText(val)
                edits[col] = combo
                layout.addRow(col, combo)
            elif col == "Internal/External":
                combo = QComboBox()
                combo.addItems(["Internal", "External"])
                if val:
                    combo.setCurrentText(val)
                edits[col] = combo
                layout.addRow(col, combo)
            elif col == "% Complete":
                from PyQt6.QtWidgets import QSpinBox
                spin = QSpinBox()
                spin.setRange(0, 100)
                try:
                    spin.setValue(int(val or 0))
                except Exception:
                    spin.setValue(0)
                # Disable if this row is a parent (rolled up)
                name = row.get("Project Part", "")
                has_children = any(r.get("Parent", "") == name for r in self.model.rows if r is not row)
                if has_children:
                    spin.setEnabled(False)
                    spin.setToolTip("Parent progress rolls up from children.")
                edits[col] = spin
                layout.addRow(col, spin)
            elif col == "Status":
                combo = QComboBox()
                combo.addItems(["Planned", "In Progress", "Blocked", "Done", "Deferred"])
                if val:
                    combo.setCurrentText(str(val))
                name = row.get("Project Part", "")
                has_children = any(r.get("Parent", "") == name for r in self.model.rows if r is not row)
                if has_children:
                    combo.setEnabled(False)
                    combo.setToolTip("Parent status is derived from children.")
                edits[col] = combo
                layout.addRow(col, combo)
            elif col in ("Actual Start Date", "Actual Finish Date", "Baseline Start Date", "Baseline End Date"):
                # Show read-only line edits for audit trail
                from PyQt6.QtWidgets import QLineEdit as _QLineEdit
                le = _QLineEdit(str(val) if val else "")
                le.setReadOnly(True)
                le.setStyleSheet("QLineEdit { background-color: #222; color: #bbb; }")
                edits[col] = le
                layout.addRow(col, le)
            elif col == "Notes":
                text = QTextEdit()
                text.setPlainText(val)
                edits[col] = text
                layout.addRow(col, text)
            elif col in ("Fabrication Labor Hours", "Installation Labor Hours"):
                from PyQt6.QtWidgets import QDoubleSpinBox
                hrs = QDoubleSpinBox()
                hrs.setRange(0.0, 10000.0)
                hrs.setDecimals(1)
                hrs.setSingleStep(0.5)
                try:
                    hrs.setValue(float(val) if val not in (None, "") else 0.0)
                except Exception:
                    hrs.setValue(0.0)
                hrs.setSuffix(" h")
                hrs.setToolTip("Estimated {} labor hours".format("fabrication" if col.startswith("Fabrication") else "installation"))
                edits[col] = hrs
                layout.addRow(col, hrs)
            elif col == "Images":
                # Ensure QHBoxLayout available in this scope (PyQt6 namespaced import patterns)
                try:
                    from PyQt6.QtWidgets import QHBoxLayout as _LocalHBox
                except Exception:
                    _LocalHBox = None
                # Ensure QLabel available (avoid UnboundLocalError if overshadowed)
                try:
                    from PyQt6.QtWidgets import QLabel as _LocalQLabel
                except Exception:
                    _LocalQLabel = None
                try:
                    hbox = QHBoxLayout()
                except Exception:
                    if _LocalHBox:
                        hbox = _LocalHBox()
                    else:
                        from PyQt6.QtWidgets import QHBoxLayout as _HB2
                        hbox = _HB2()
                try:
                    img_label = QLabel()
                except Exception:
                    if _LocalQLabel:
                        img_label = _LocalQLabel()
                    else:
                        from PyQt6.QtWidgets import QLabel as _QL2
                        img_label = _QL2()
                if val:
                    import os
                    from PyQt6.QtGui import QPixmap
                    if not os.path.isabs(val):
                        base_dir = os.path.dirname(os.path.abspath(__file__))
                        img_path_full = os.path.join(base_dir, val)
                    else:
                        img_path_full = val
                    pixmap = QPixmap(img_path_full)
                    if not pixmap.isNull():
                        try:
                            _smooth = Qt.TransformationMode.SmoothTransformation
                        except Exception:
                            _smooth = getattr(Qt, 'SmoothTransformation', 1)
                        img_label.setPixmap(pixmap.scaledToHeight(48, _smooth))
                btn = QPushButton("Change Image")
                def pick_image():
                    fname, _ = QFileDialog.getOpenFileName(dialog, "Select Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp *.gif)")
                    if fname:
                        try:
                            _smooth = Qt.TransformationMode.SmoothTransformation
                        except Exception:
                            _smooth = getattr(Qt, 'SmoothTransformation', 1)
                        img_label.setPixmap(QPixmap(fname).scaledToHeight(48, _smooth))
                        edits[col].setText(fname)
                btn.clicked.connect(pick_image)
                img_path_edit = QLineEdit(val)
                edits[col] = img_path_edit
                hbox.addWidget(img_label)
                hbox.addWidget(img_path_edit)
                hbox.addWidget(btn)
                layout.addRow(col, hbox)
            elif col == "Pace Link":
                link_edit = QLineEdit(val)
                edits[col] = link_edit
                # Guarded QLabel acquisition to avoid UnboundLocalError in mixed import contexts
                try:
                    from PyQt6.QtWidgets import QLabel as _LinkQLabel
                except Exception:
                    _LinkQLabel = None
                try:
                    link_label = QLabel()
                except Exception:
                    if _LinkQLabel:
                        try:
                            link_label = _LinkQLabel()
                        except Exception:
                            from PyQt6.QtWidgets import QLabel as _QLFallback
                            link_label = _QLFallback()
                    else:
                        from PyQt6.QtWidgets import QLabel as _QLFallback
                        link_label = _QLFallback()
                if val and (val.startswith("http://") or val.startswith("https://")):
                    link_label.setText(f'<a href="{val}">{val}</a>')
                    link_label.setOpenExternalLinks(True)
                else:
                    link_label.setText("")
                layout.addRow(col, link_edit)
                layout.addRow("Link Preview", link_label)
                def update_link_label():
                    v = link_edit.text()
                    if v and (v.startswith("http://") or v.startswith("https://")):
                        link_label.setText(f'<a href="{v}">{v}</a>')
                        link_label.setOpenExternalLinks(True)
                    else:
                        link_label.setText("")
                link_edit.textChanged.connect(update_link_label)
            elif col in ("Production Cost", "Installation Cost", "Production Price", "Installation Price", "Material Cost", "Labor Rate", "Install Labor Rate", "Equipment Cost", "Permit/Eng Cost"):
                from PyQt6.QtWidgets import QDoubleSpinBox
                sb = QDoubleSpinBox()
                sb.setRange(0.0, 10_000_000.0)
                sb.setDecimals(2)
                sb.setSingleStep(50.0)
                try:
                    sb.setValue(float(val) if val not in (None, "") else 0.0)
                except Exception:
                    sb.setValue(0.0)
                sb.setPrefix("$")
                if col in ("Production Cost","Installation Cost","Material Cost","Equipment Cost","Permit/Eng Cost"):
                    if col == "Production Cost":
                        sb.setToolTip("Internal production cost (auto-derived if Material + Labor provided)")
                    elif col == "Installation Cost":
                        sb.setToolTip("Internal install cost (labor + equipment + permit/eng if provided)")
                    elif col == "Material Cost":
                        sb.setToolTip("Direct materials cost")
                    elif col == "Equipment Cost":
                        sb.setToolTip("Equipment rental or usage cost")
                    elif col == "Permit/Eng Cost":
                        sb.setToolTip("Permitting or engineering fees")
                else:
                    sb.setToolTip("Charge amount for {}".format("production" if "Production" in col else "installation"))
                edits[col] = sb
                layout.addRow(col, sb)
            elif col in ("Contingency %","Warranty Reserve %"):
                from PyQt6.QtWidgets import QDoubleSpinBox
                sb = QDoubleSpinBox(); sb.setRange(0.0,100.0); sb.setDecimals(1); sb.setSingleStep(1.0); sb.setSuffix(" %")
                try:
                    sb.setValue(float(val) if val not in (None,"") else 0.0)
                except Exception:
                    sb.setValue(0.0)
                if col == "Contingency %":
                    sb.setToolTip("Percentage buffer applied to internal cost before margin calc for suggestions")
                else:
                    sb.setToolTip("Percentage of price reserved (reduces effective profit)")
                edits[col]=sb; layout.addRow(col,sb)
            elif col == "Risk Level":
                combo = QComboBox(); combo.addItems(["Low","Medium","High"])
                if val: combo.setCurrentText(str(val))
                combo.setToolTip("Qualitative risk indicator")
                edits[col]=combo; layout.addRow(col, combo)
            elif col == "Quote Version":
                from PyQt6.QtWidgets import QLineEdit as _QLE
                le=_QLE(str(val) if val else ""); le.setReadOnly(True); le.setStyleSheet("QLineEdit { background:#222; color:#bbb; }")
                edits[col]=le; layout.addRow(col, le)
            elif col.startswith("Frozen "):
                from PyQt6.QtWidgets import QLineEdit as _QLE
                le=_QLE(str(val) if val else ""); le.setReadOnly(True); le.setStyleSheet("QLineEdit { background:#222; color:#777; }")
                edits[col]=le; layout.addRow(col, le)
            else:
                # Fallback generic text field; ensure string conversion
                line = QLineEdit(str(val) if val is not None else "")
                edits[col] = line
                layout.addRow(col, line)
        save_btn = QPushButton("Save")
        cancel_btn = QPushButton("Cancel")
        # Suggestion row (added just before buttons once widgets exist)
        def _compute_suggestions():
            nonlocal suggested_prod, suggested_inst
            try:
                prod_cost_w = edits.get('Production Cost'); inst_cost_w = edits.get('Installation Cost')
                pc = float(prod_cost_w.value()) if prod_cost_w else float(row.get('Production Cost') or 0)
                ic = float(inst_cost_w.value()) if inst_cost_w else float(row.get('Installation Cost') or 0)
                # Derive production cost if zero using material + fab labor
                try:
                    if pc == 0:
                        mat = float(edits.get('Material Cost').value()) if 'Material Cost' in edits else float(row.get('Material Cost') or 0)
                        fab_hrs = float(edits.get('Fabrication Labor Hours').value()) if 'Fabrication Labor Hours' in edits else float(row.get('Fabrication Labor Hours') or 0)
                        labor_rate = float(edits.get('Labor Rate').value()) if 'Labor Rate' in edits else float(row.get('Labor Rate') or 0)
                        pc = mat + fab_hrs * labor_rate
                except Exception:
                    pass
                # Derive installation cost if zero using install hours + rate + equipment + permit/eng
                try:
                    if ic == 0:
                        inst_hrs = float(edits.get('Installation Labor Hours').value()) if 'Installation Labor Hours' in edits else float(row.get('Installation Labor Hours') or 0)
                        inst_rate = float(edits.get('Install Labor Rate').value()) if 'Install Labor Rate' in edits else float(row.get('Install Labor Rate') or 0)
                        equip = float(edits.get('Equipment Cost').value()) if 'Equipment Cost' in edits else float(row.get('Equipment Cost') or 0)
                        permit = float(edits.get('Permit/Eng Cost').value()) if 'Permit/Eng Cost' in edits else float(row.get('Permit/Eng Cost') or 0)
                        ic = inst_hrs * inst_rate + equip + permit
                except Exception:
                    pass
                contingency_pct = 0.0
                try:
                    contingency_pct = float(edits.get('Contingency %').value())/100.0 if 'Contingency %' in edits else float(row.get('Contingency %') or 0)/100.0
                except Exception:
                    contingency_pct = 0.0
                # Risk-based margin adjustment (Low: -2 pts, Medium: baseline, High: +3 pts)
                try:
                    risk = edits.get('Risk Level').currentText() if 'Risk Level' in edits else (row.get('Risk Level') or '')
                except Exception:
                    risk = ''
                adj_margin = target_margin
                try:
                    if isinstance(risk, str):
                        if risk.lower().startswith('low'):
                            adj_margin = max(0.01, target_margin - 0.02)
                        elif risk.lower().startswith('high'):
                            adj_margin = min(0.95, target_margin + 0.03)
                except Exception:
                    pass
                if pc > 0:
                    suggested_prod = (pc * (1.0 + contingency_pct)) / (1.0 - adj_margin)
                else:
                    suggested_prod = None
                if ic > 0:
                    suggested_inst = (ic * (1.0 + contingency_pct)) / (1.0 - adj_margin)
                else:
                    suggested_inst = None
                if suggest_label:
                    txt=[]
                    if suggested_prod is not None: txt.append(f"Prod Suggest: ${suggested_prod:,.2f}")
                    if suggested_inst is not None: txt.append(f"Inst Suggest: ${suggested_inst:,.2f}")
                    try:
                        if adj_margin != target_margin:
                            delta_pts = (adj_margin - target_margin)*100.0
                            txt.append(f"Risk Adj: {'+' if delta_pts>=0 else ''}{delta_pts:.1f} pts")
                    except Exception:
                        pass
                    # Effective margin after warranty reserve estimate
                    try:
                        warr_pct = float(edits.get('Warranty Reserve %').value())/100.0 if 'Warranty Reserve %' in edits else float(row.get('Warranty Reserve %') or 0)/100.0
                        if warr_pct>0 and (suggested_prod or 0)+(suggested_inst or 0) > 0 and (pc+ic)>0:
                            total_price = (suggested_prod or 0)+(suggested_inst or 0)
                            gross_profit = total_price - ((pc+ic)*(1.0+contingency_pct))
                            eff_profit = gross_profit - total_price*warr_pct
                            eff_margin = eff_profit/total_price*100.0 if total_price>0 else 0.0
                            txt.append(f"Eff Margin(after reserve): {eff_margin:,.1f}%")
                    except Exception:
                        pass
                    suggest_label.setText(" | ".join(txt) if txt else "")
            except Exception:
                pass
        from PyQt6.QtWidgets import QLabel, QHBoxLayout
        suggest_label = QLabel("")
        suggest_label.setStyleSheet("color:#bbb; font-size:11px")
        apply_box = QHBoxLayout()
        apply_prod_btn = QPushButton("Apply Prod Suggest")
        apply_inst_btn = QPushButton("Apply Inst Suggest")
        apply_prod_btn.setEnabled(False); apply_inst_btn.setEnabled(False)
        def _enable_apply():
            apply_prod_btn.setEnabled(suggested_prod is not None and 'Production Price' in edits)
            apply_inst_btn.setEnabled(suggested_inst is not None and 'Installation Price' in edits)
        def _recalc_and_enable():
            _compute_suggestions(); _enable_apply()
        apply_prod_btn.clicked.connect(lambda: (edits['Production Price'].setValue(suggested_prod) if suggested_prod and 'Production Price' in edits else None))
        apply_inst_btn.clicked.connect(lambda: (edits['Installation Price'].setValue(suggested_inst) if suggested_inst and 'Installation Price' in edits else None))
        apply_box.addWidget(suggest_label)
        apply_box.addStretch(1)
        apply_box.addWidget(apply_prod_btn)
        apply_box.addWidget(apply_inst_btn)
        def save():
            try:
                for col in self.model.COLUMNS:
                    widget = edits[col]
                    if isinstance(widget, QLineEdit):
                        row[col] = widget.text()
                    elif isinstance(widget, QComboBox):
                        row[col] = widget.currentText()
                        if col == "Status" and row[col] == "Done" and str(row.get("% Complete")) != "100":
                            row["% Complete"] = 100
                            import datetime as _dt
                            if not row.get("Actual Start Date"):
                                row["Actual Start Date"] = _dt.datetime.today().strftime("%m-%d-%Y")
                            if not row.get("Actual Finish Date"):
                                row["Actual Finish Date"] = _dt.datetime.today().strftime("%m-%d-%Y")
                        if col == "Status" and row[col] == "In Progress" and not row.get("Actual Start Date"):
                            import datetime as _dt
                            row["Actual Start Date"] = _dt.datetime.today().strftime("%m-%d-%Y")
                    elif isinstance(widget, QDateEdit):
                        d = widget.date()
                        min_blank = QDate(1753, 1, 1)
                        if not d.isValid() or d == min_blank:
                            row[col] = ""
                        else:
                            row[col] = d.toString("MM-dd-yyyy")
                    elif hasattr(widget, 'value') and col == "% Complete":
                        try:
                            row[col] = int(widget.value())
                            if int(row[col]) >= 100:
                                row[col] = 100
                                if row.get("Status") != "Done":
                                    row["Status"] = "Done"
                                    import datetime as _dt
                                    if not row.get("Actual Start Date"):
                                        row["Actual Start Date"] = _dt.datetime.today().strftime("%m-%d-%Y")
                                    if not row.get("Actual Finish Date"):
                                        row["Actual Finish Date"] = _dt.datetime.today().strftime("%m-%d-%Y")
                        except Exception:
                            row[col] = 0
                    elif isinstance(widget, QTextEdit):
                        row[col] = widget.toPlainText()
                    elif hasattr(widget, 'value') and col in ("Fabrication Labor Hours", "Installation Labor Hours"):
                        try:
                            row[col] = f"{float(widget.value()):.1f}"
                        except Exception:
                            row[col] = "0.0"
                    elif hasattr(widget, 'value') and col in ("Production Price", "Installation Price", "Production Cost", "Installation Cost", "Material Cost", "Labor Rate", "Install Labor Rate", "Equipment Cost", "Permit/Eng Cost"):
                        try:
                            row[col] = f"{float(widget.value()):.2f}"
                        except Exception:
                            row[col] = "0.00"
                    elif hasattr(widget,'value') and col in ("Contingency %","Warranty Reserve %"):
                        try:
                            row[col] = f"{float(widget.value()):.1f}"
                        except Exception:
                            row[col] = "0.0"
                # Derive production / installation cost if material & labor hour info present
                try:
                    from math import isnan
                    mat = float(row.get('Material Cost') or 0)
                    fab_h = float(row.get('Fabrication Labor Hours') or 0)
                    inst_h = float(row.get('Installation Labor Hours') or 0)
                    rate = float(row.get('Labor Rate') or 0)
                    inst_rate = float(row.get('Install Labor Rate') or rate)
                    equip = float(row.get('Equipment Cost') or 0)
                    permit = float(row.get('Permit/Eng Cost') or 0)
                    if (mat or fab_h or inst_h) and (not row.get('Production Cost') or not row.get('Installation Cost')):
                        prod_cost_calc = mat + fab_h * rate
                        inst_cost_calc = inst_h * inst_rate + equip + permit
                        if not row.get('Production Cost'):
                            row['Production Cost'] = f"{prod_cost_calc:.2f}"
                        if not row.get('Installation Cost'):
                            row['Installation Cost'] = f"{inst_cost_calc:.2f}"
                except Exception:
                    pass
                # Validation: price below cost warning
                try:
                    from PyQt6.QtWidgets import QMessageBox as _QB
                    pcost = float(row.get('Production Cost') or 0)
                    icost = float(row.get('Installation Cost') or 0)
                    pprice = float(row.get('Production Price') or 0)
                    iprice = float(row.get('Installation Price') or 0)
                    warn_msgs = []
                    if pprice and pcost and pprice < pcost:
                        warn_msgs.append(f"Production price ${pprice:,.2f} < cost ${pcost:,.2f}")
                    if iprice and icost and iprice < icost:
                        warn_msgs.append(f"Installation price ${iprice:,.2f} < cost ${icost:,.2f}")
                    if warn_msgs:
                        resp = _QB.question(dialog, "Below-Cost Pricing", "\n".join(warn_msgs) + "\nContinue anyway?", _QB.Yes | _QB.No, _QB.No)
                        if resp != _QB.Yes:
                            return
                except Exception:
                    pass
                self.model.save_to_db()
                self.render_gantt(self.model)
                dialog.accept()
            except Exception as e:
                import traceback
                print(f"ERROR in GanttChartView.save(): {e}")
                traceback.print_exc()
                QMessageBox.critical(dialog, "Save Error", f"An error occurred while saving: {e}")
        save_btn.clicked.connect(save)
        cancel_btn.clicked.connect(dialog.reject)
        btn_hbox = QHBoxLayout()
        btn_hbox.addWidget(save_btn)
        btn_hbox.addWidget(cancel_btn)
        # Insert suggestion row before buttons
        layout.addRow(apply_box)
        layout.addRow(btn_hbox)
        dialog.setLayout(layout)
        # Initial suggestion compute after dialog constructed
        _recalc_and_enable()
        # Recompute on cost field edits
        for key in ('Production Cost','Installation Cost'):
            w = edits.get(key)
            if w and hasattr(w,'valueChanged'):
                w.valueChanged.connect(_recalc_and_enable)
        dialog.exec()
    def highlight_group(self, part_name):
        # Find parent and children for the given part_name
        parent = None
        children = set()
        for row in self.model.rows:
            if row["Project Part"] == part_name:
                parent = row.get("Parent")
            if row.get("Parent") == part_name:
                children.add(row["Project Part"])
        group = {part_name}
        if parent:
            group.add(parent)
        group.update(children)
        # Highlight bars in group
        from PyQt6.QtGui import QPen, QColor
        highlight_color = QColor("#00BFFF")
        for item in self.scene.items():
            if hasattr(item, 'data') and callable(item.data):
                if item.data(0) in group:
                    item.setPen(QPen(highlight_color, 3))
                else:
                    item.setPen(QPen())
            elif hasattr(item, 'toGraphicsObject') and hasattr(item, 'setDefaultTextColor'):
                if hasattr(item, 'toPlainText') and item.toPlainText() in group:
                    item.setDefaultTextColor(highlight_color)
                else:
                    item.setDefaultTextColor(QColor("black"))

    def export_gantt_chart(self):
        self._export_scene_with_header(self.scene, title="Gantt Chart")

    def _export_scene_with_header(self, scene, title="Export"):
        from PyQt6.QtGui import QPainter
        import os
        from PyQt6.QtCore import QSettings
        # Prefer SVG header from repo root; fallback to PNG
        svg_path = resolve_resource_path("header.svg")
        header_is_svg = False
        header_svg_renderer = None
        try:
            if os.path.exists(svg_path):
                from PyQt6.QtSvg import QSvgRenderer  # type: ignore
                r = QSvgRenderer(svg_path)
                if r.isValid():
                    header_is_svg = True
                    header_svg_renderer = r
        except Exception:
            header_is_svg = False
            header_svg_renderer = None
        # Read persisted export settings
        s = QSettings("LSI", "ProjectPlanner")
        pref_format = s.value("Export/format", "PNG")
        page_size = s.value("Export/page_size", "A4")
        orientation = s.value("Export/orientation", "Portrait")
        ml = float(s.value("Export/margin_left_mm", 8.0))
        mt = float(s.value("Export/margin_top_mm", 8.0))
        mr = float(s.value("Export/margin_right_mm", 8.0))
        mb = float(s.value("Export/margin_bottom_mm", 8.0))
        include_header = s.value("Export/include_header", True)
        if isinstance(include_header, str):
            include_header = include_header.lower() in ("1","true","yes","on")
        # Ask for target path with default filter based on preferred format
        init_name = f"{title.lower().replace(' ', '_')}.{ 'pdf' if pref_format=='PDF' else 'png' }"
        filters = "PDF Files (*.pdf);;PNG Files (*.png)" if pref_format=="PDF" else "PNG Files (*.png);;PDF Files (*.pdf)"
        path, chosen = QFileDialog.getSaveFileName(self, f"Export {title}", init_name, filters)
        if not path:
            return
        rect = scene.sceneRect().toRect()
        if rect.isEmpty():
            print("Scene is empty; nothing to export.")
            return
        # Respect chosen filter if extension missing/mismatch
        is_pdf = path.lower().endswith('.pdf') or (chosen and 'PDF' in chosen and not path.lower().endswith('.png'))
        if is_pdf and not path.lower().endswith('.pdf'):
            path += '.pdf'
        if (not is_pdf) and not path.lower().endswith('.png'):
            path += '.png'
        header_path = resolve_resource_path("header.png")
        header_pixmap = QPixmap(header_path) if os.path.exists(header_path) else None
        if is_pdf:
            # Use QPrinter for PDF
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtCore import QRectF
            from math import ceil
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.PdfFormat)
            # Page size mapping
            size_map = {
                'A4': QPrinter.A4,
                'Letter': QPrinter.Letter,
                'Legal': QPrinter.Legal,
                'Tabloid': QPrinter.Tabloid,
            }
            printer.setPaperSize(size_map.get(page_size, QPrinter.A4))
            printer.setOrientation(QPrinter.Portrait if orientation == 'Portrait' else QPrinter.Landscape)
            painter = QPainter(printer)
            # Apply margins (mm)
            from PyQt6.QtCore import QMarginsF
            try:
                m = QMarginsF(ml, mt, mr, mb)
                printer.setPageMargins(m)
            except Exception:
                pass
            page_rect = printer.pageRect()
            y_offset = 0
            # Draw header (SVG preferred)
            def _svg_header_h(renderer, target_w):
                try:
                    ds = renderer.defaultSize(); w, h = ds.width(), ds.height()
                    if w <= 0 or h <= 0:
                        vb = renderer.viewBoxF(); w, h = vb.width(), vb.height()
                    if w > 0 and h > 0 and target_w > 0:
                        return int(round(h * (target_w / w)))
                except Exception:
                    return None
                return None
            if include_header and header_is_svg and header_svg_renderer:
                target_w = page_rect.width()
                header_h = _svg_header_h(header_svg_renderer, target_w)
                if header_h:
                    target_rect = QRectF(0, 0, target_w, header_h)
                    header_svg_renderer.render(painter, target_rect)
                    y_offset = header_h + 10
                elif header_pixmap and not header_pixmap.isNull():
                    header_w = page_rect.width()
                    try:
                        _smooth = Qt.TransformationMode.SmoothTransformation
                    except Exception:
                        _smooth = getattr(Qt, 'SmoothTransformation', 1)
                    scaled_header = header_pixmap.scaledToWidth(header_w, _smooth)
                    painter.drawPixmap((header_w - scaled_header.width()) // 2, 0, scaled_header)
                    y_offset = scaled_header.height() + 10
            elif include_header and header_pixmap and not header_pixmap.isNull():
                header_w = page_rect.width()
                try:
                    _smooth = Qt.TransformationMode.SmoothTransformation
                except Exception:
                    _smooth = getattr(Qt, 'SmoothTransformation', 1)
                scaled_header = header_pixmap.scaledToWidth(header_w, _smooth)
                painter.drawPixmap((header_w - scaled_header.width()) // 2, 0, scaled_header)
                y_offset = scaled_header.height() + 10
            avail_h = max(1, page_rect.height() - y_offset)
            scale = avail_h / rect.height()
            scaled_total_w = max(1.0, rect.width() * scale)
            cols = max(1, int(ceil(scaled_total_w / page_rect.width())))
            for col in range(cols):
                if col > 0:
                    printer.newPage()
                    # Repaint header on each page if enabled
                    if include_header:
                        if header_is_svg and header_svg_renderer:
                            target_w = page_rect.width()
                            header_h = _svg_header_h(header_svg_renderer, target_w)
                            if header_h:
                                target_rect = QRectF(0, 0, target_w, header_h)
                                header_svg_renderer.render(painter, target_rect)
                            elif header_pixmap and not header_pixmap.isNull():
                                header_w = page_rect.width()
                                try:
                                    _smooth = Qt.TransformationMode.SmoothTransformation
                                except Exception:
                                    _smooth = getattr(Qt, 'SmoothTransformation', 1)
                                scaled_header = header_pixmap.scaledToWidth(header_w, _smooth)
                                painter.drawPixmap((header_w - scaled_header.width()) // 2, 0, scaled_header)
                        elif header_pixmap and not header_pixmap.isNull():
                            header_w = page_rect.width()
                            try:
                                _smooth = Qt.TransformationMode.SmoothTransformation
                            except Exception:
                                _smooth = getattr(Qt, 'SmoothTransformation', 1)
                            scaled_header = header_pixmap.scaledToWidth(header_w, _smooth)
                            painter.drawPixmap((header_w - scaled_header.width()) // 2, 0, scaled_header)
                painter.save()
                painter.translate(0, y_offset)
                painter.scale(scale, scale)
                source_x = (col * page_rect.width()) / scale
                source = QRectF(source_x, 0, page_rect.width() / scale, rect.height())
                scene.render(painter, target=QRectF(0, 0, page_rect.width(), rect.height() * scale), source=source)
                painter.restore()
            painter.end()
            print(f"Exported PDF -> {path}")
            return
        # Raster export (PNG)
        # Create content pixmap with margins (convert mm to px using screen DPI)
        screen = QApplication.primaryScreen()
        dpi = screen.logicalDotsPerInch() if screen else 96.0
        def mm_to_px(mm):
            return int(round((mm / 25.4) * dpi))
        pad_l, pad_t, pad_r, pad_b = mm_to_px(ml), mm_to_px(mt), mm_to_px(mr), mm_to_px(mb)
        content_pixmap = QPixmap(rect.width() + pad_l + pad_r, rect.height() + pad_t + pad_b)
        content_pixmap.fill()
        painter = QPainter(content_pixmap)
        painter.translate(pad_l, pad_t)
        scene.render(painter)
        painter.end()
        # Compose header + content into a single pixmap (header centered)
        if include_header and header_is_svg and header_svg_renderer:
            target_w = content_pixmap.width()
            def _svg_header_h(renderer, target_w):
                try:
                    ds = renderer.defaultSize(); w, h = ds.width(), ds.height()
                    if w <= 0 or h <= 0:
                        vb = renderer.viewBoxF(); w, h = vb.width(), vb.height()
                    if w > 0 and h > 0 and target_w > 0:
                        return int(round(h * (target_w / w)))
                except Exception:
                    return None
                return None
            header_h = _svg_header_h(header_svg_renderer, target_w)
            if header_h:
                combined_width = max(target_w, content_pixmap.width())
                combined_height = header_h + content_pixmap.height()
                combined = QPixmap(combined_width, combined_height)
                combined.fill()
                painter = QPainter(combined)
                # Render SVG centered at top into target rect of width target_w
                from PyQt6.QtCore import QRectF
                target_rect = QRectF((combined_width - target_w) / 2, 0, target_w, header_h)
                header_svg_renderer.render(painter, target_rect)
                painter.drawPixmap(0, header_h, content_pixmap)
                painter.end()
                combined.save(path, 'PNG')
            elif header_pixmap and not header_pixmap.isNull():
                combined_width = max(header_pixmap.width(), content_pixmap.width())
                combined_height = header_pixmap.height() + content_pixmap.height()
                combined = QPixmap(combined_width, combined_height)
                combined.fill()
                painter = QPainter(combined)
                header_x = (combined_width - header_pixmap.width()) // 2
                painter.drawPixmap(header_x, 0, header_pixmap)
                painter.drawPixmap(0, header_pixmap.height(), content_pixmap)
                painter.end()
                combined.save(path, 'PNG')
            else:
                content_pixmap.save(path, 'PNG')
        elif include_header and header_pixmap and not header_pixmap.isNull():
            combined_width = max(header_pixmap.width(), content_pixmap.width())
            combined_height = header_pixmap.height() + content_pixmap.height()
            combined = QPixmap(combined_width, combined_height)
            combined.fill()
            painter = QPainter(combined)
            header_x = (combined_width - header_pixmap.width()) // 2
            painter.drawPixmap(header_x, 0, header_pixmap)
            painter.drawPixmap(0, header_pixmap.height(), content_pixmap)
            painter.end()
            combined.save(path, 'PNG')
        else:
            content_pixmap.save(path, 'PNG')
        print(f"Exported PNG -> {path}")
    def __init__(self):
        print("GanttChartView: __init__ called")
        super().__init__()
        layout = QVBoxLayout()
        toolbar = QHBoxLayout()
        title_lbl = QLabel("Gantt Chart")
        title_lbl.setStyleSheet("font-weight:600; padding:2px 4px;")
        toolbar.addWidget(title_lbl)
        self.view = ZoomableGraphicsView()
        self.scene = QGraphicsScene()
        self.view.setScene(self.scene)
        self.view.setSettingsKey("GanttZoom")
        layout.addWidget(self.view)

        self._init_filters()

        self.preview_label = QLabel()
        self.preview_label.setFixedHeight(140)
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.preview_label.setStyleSheet("border:1px solid #666; background:#222;")

        # Export / Settings / Refresh
        export_btn = QPushButton("Export")
        export_btn.setToolTip("Export Gantt Chart")
        export_btn.clicked.connect(self.export_gantt_chart)
        settings_btn = QPushButton("Settings…")
        def open_settings():
            try:
                dlg = ExportSettingsDialog(self)
                dlg.exec()
            except Exception as e:
                print(f"Open Export Settings failed: {e}")
        settings_btn.clicked.connect(open_settings)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setToolTip("Refresh Gantt Chart")
        refresh_btn.clicked.connect(self.refresh_gantt)
        toolbar.addWidget(export_btn)
        toolbar.addWidget(settings_btn)
        toolbar.addWidget(refresh_btn)

        from PyQt6.QtWidgets import QCheckBox, QComboBox
        self.hierarchy_checkbox = QCheckBox("Hierarchy")
        self.hierarchy_checkbox.setChecked(True)
        self.hierarchy_checkbox.stateChanged.connect(lambda _s: self.refresh_gantt())
        toolbar.addWidget(self.hierarchy_checkbox)

        self.include_legend = True

        self.critical_path_checkbox = QCheckBox("Critical Path")
        self.critical_path_checkbox.setChecked(False)
        self.critical_path_checkbox.stateChanged.connect(lambda _s: self.refresh_gantt())
        toolbar.addWidget(self.critical_path_checkbox)

        # Group predecessors toggle (persisted)
        self._group_predecessors = True
        try:
            from PyQt6.QtCore import QSettings as _QS_gp
            _ps_gp = _QS_gp('LSI','ProjectPlanner')
            gp_val = _ps_gp.value('Gantt/GroupPredecessors', 'true')
            def _b_gp(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return True
            self._group_predecessors = _b_gp(gp_val)
        except Exception:
            self._group_predecessors = True
        from PyQt6.QtWidgets import QCheckBox as _QCB_gp
        self.group_preds_checkbox = _QCB_gp("Group Preds")
        self.group_preds_checkbox.setToolTip("Cluster multiple predecessors just above their successor (within same parent)")
        self.group_preds_checkbox.setChecked(self._group_predecessors)
        def _on_gp_toggle():
            self._group_predecessors = self.group_preds_checkbox.isChecked()
            try:
                from PyQt6.QtCore import QSettings as _QS2_gp
                _QS2_gp('LSI','ProjectPlanner').setValue('Gantt/GroupPredecessors', self._group_predecessors)
            except Exception:
                pass
            self.refresh_gantt()
        self.group_preds_checkbox.stateChanged.connect(lambda _s: _on_gp_toggle())
        toolbar.addWidget(self.group_preds_checkbox)

        # Cross-parent grouping toggle (persisted, default off)
        self._group_cross_parent = False
        try:
            from PyQt6.QtCore import QSettings as _QS_gcp
            _ps_gcp = _QS_gcp('LSI','ProjectPlanner')
            gcp_val = _ps_gcp.value('Gantt/GroupCrossParent', 'false')
            def _b_gcp(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return False
            self._group_cross_parent = _b_gcp(gcp_val)
        except Exception:
            self._group_cross_parent = False
        from PyQt6.QtWidgets import QCheckBox as _QCB_gcp
        self.group_cross_parent_checkbox = _QCB_gcp("Cross-parent")
        self.group_cross_parent_checkbox.setToolTip("Allow clustering predecessors across different parents if they share the same top-level ancestor")
        self.group_cross_parent_checkbox.setChecked(self._group_cross_parent)
        def _on_gcp_toggle():
            self._group_cross_parent = self.group_cross_parent_checkbox.isChecked()
            try:
                from PyQt6.QtCore import QSettings as _QS2_gcp
                _QS2_gcp('LSI','ProjectPlanner').setValue('Gantt/GroupCrossParent', self._group_cross_parent)
            except Exception:
                pass
            self.refresh_gantt()
        self.group_cross_parent_checkbox.stateChanged.connect(lambda _s: _on_gcp_toggle())
        toolbar.addWidget(self.group_cross_parent_checkbox)

        # Grouping weight strategy dropdown
        from PyQt6.QtWidgets import QComboBox as _QCB_w, QPushButton as _QPB_w
        self.weight_combo = _QCB_w()
        self.weight_combo.addItems([
            "Weight: Successors",
            "Weight: Original",
            "Weight: Start Date",
            "Weight: Reverse Start Date",
            "Weight: Duration",
            "Weight: Criticality",
            "Weight: Criticality then Start Date"
        ]) 
        # Persisted selection
        try:
            from PyQt6.QtCore import QSettings as _QS_w
            sel = _QS_w('LSI','ProjectPlanner').value('Gantt/GroupWeightStrategy', 'Weight: Successors')
            if sel not in [
                "Weight: Successors",
                "Weight: Original",
                "Weight: Start Date",
                "Weight: Reverse Start Date",
                "Weight: Duration",
                "Weight: Criticality"
            ]:
                sel = "Weight: Successors"
            idx = self.weight_combo.findText(sel)
            if idx >= 0:
                self.weight_combo.setCurrentIndex(idx)
        except Exception:
            pass
        def _on_weight_change(_txt=None):
            try:
                from PyQt6.QtCore import QSettings as _QS2_w
                _QS2_w('LSI','ProjectPlanner').setValue('Gantt/GroupWeightStrategy', self.weight_combo.currentText())
            except Exception:
                pass
            self.refresh_gantt()
        self.weight_combo.currentTextChanged.connect(_on_weight_change)
        toolbar.addWidget(self.weight_combo)
        # Small help button explaining strategies
        help_btn = _QPB_w("?")
        help_btn.setFixedWidth(22)
        def _show_weight_help():
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Weighting Strategies",
                "Successors: Tasks depended on by more successors appear closer to the successor.\n"
                "Original: Preserve original order.\n"
                "Start Date: Earlier starts first.\n"
                "Reverse Start Date: Later starts first.\n"
                "Duration: Longer durations first.\n"
                "Criticality: Critical-path tasks first.\n"
                "Criticality then Start Date: Critical first; among equals, earlier starts first.")
        help_btn.clicked.connect(_show_weight_help)
        toolbar.addWidget(help_btn)

        # Links toggle (persisted, controls link indicators and click-to-open behavior)
        from PyQt6.QtWidgets import QCheckBox as _QCB_links
        try:
            from PyQt6.QtCore import QSettings as _QS_links
            _ps_links = _QS_links('LSI','ProjectPlanner')
            val_links = _ps_links.value('UI/ShowLinks', 'true')
            def _b_l(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return True
            self._show_links = _b_l(val_links)
        except Exception:
            self._show_links = True
        self.links_checkbox = _QCB_links('Links')
        self.links_checkbox.setToolTip('Show link indicators; click bars to open Pace Link')
        self.links_checkbox.setChecked(self._show_links)
        def _on_links_toggle():
            self._show_links = self.links_checkbox.isChecked()
            try:
                from PyQt6.QtCore import QSettings as _QS_links2
                _QS_links2('LSI','ProjectPlanner').setValue('UI/ShowLinks', self._show_links)
            except Exception:
                pass
            # Propagate to sibling views if present
            try:
                mw = self.window()
                if hasattr(mw, 'timeline_view') and mw.timeline_view:
                    mw.timeline_view._show_links = self._show_links
                    mw.timeline_view._sync_links_checkbox()
                    mw.timeline_view.render_timeline()
                if hasattr(mw, 'project_tree_view') and mw.project_tree_view:
                    mw.project_tree_view._show_links = self._show_links
                    mw.project_tree_view._sync_links_checkbox()
                    mw.project_tree_view.refresh()
            except Exception:
                pass
            self.refresh_gantt()
        self.links_checkbox.stateChanged.connect(lambda _s: _on_links_toggle())
        toolbar.addWidget(self.links_checkbox)

        # Unscheduled toggle
        self.unscheduled_checkbox = QCheckBox("Show Unscheduled")
        # Load persisted setting (default true)
        try:
            from PyQt6.QtCore import QSettings as _QS
            _ps = _QS('LSI','ProjectPlanner')
            val = _ps.value('Gantt/ShowUnscheduled', 'true')
            def _b(v):
                if isinstance(v, bool): return v
                if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                return True
            self._show_unscheduled = _b(val)
            self.unscheduled_checkbox.setChecked(self._show_unscheduled)
        except Exception:
            self._show_unscheduled = True
            self.unscheduled_checkbox.setChecked(True)
        def _toggle_unscheduled():
            self._show_unscheduled = self.unscheduled_checkbox.isChecked()
            try:
                from PyQt6.QtCore import QSettings as _QS2
                _QS2('LSI','ProjectPlanner').setValue('Gantt/ShowUnscheduled', self._show_unscheduled)
            except Exception:
                pass
            self.refresh_gantt()
        self.unscheduled_checkbox.stateChanged.connect(lambda _s: _toggle_unscheduled())
        toolbar.addWidget(self.unscheduled_checkbox)

        # Baseline controls
        baseline_row = QHBoxLayout()
        self.baseline_combo = QComboBox()
        self.baseline_combo.addItem("None")
        self.baseline_combo.setToolTip("Overlay a saved baseline snapshot")
        def _on_baseline_change(txt):
            self._selected_baseline_name = txt if txt and txt != "None" else None
            self.refresh_gantt()
        self.baseline_combo.currentTextChanged.connect(_on_baseline_change)
        def _save_baseline():
            from PyQt6.QtWidgets import QInputDialog
            name, ok = QInputDialog.getText(self, "Save Baseline", "Baseline name:")
            if ok and name:
                try:
                    if hasattr(self, 'model') and self.model:
                        self.model.save_baseline(name)
                        self._populate_baselines()
                        idx = self.baseline_combo.findText(name)
                        if idx >= 0:
                            self.baseline_combo.setCurrentIndex(idx)
                except Exception as e:
                    print(f"Save baseline failed: {e}")
        save_baseline_btn = QPushButton("Save Baseline…")
        save_baseline_btn.clicked.connect(_save_baseline)
        baseline_row.addWidget(QLabel("Baseline:"))
        baseline_row.addWidget(self.baseline_combo)
        baseline_row.addWidget(save_baseline_btn)
        toolbar.addLayout(baseline_row)
        layout.addLayout(toolbar)

        # Zoom controls
        zoom_layout = QHBoxLayout()
        zoom_in_btn = QPushButton("Zoom In")
        zoom_out_btn = QPushButton("Zoom Out")
        zoom_reset_btn = QPushButton("Reset Zoom")
        zoom_in_btn.clicked.connect(self.view.zoomIn)
        zoom_out_btn.clicked.connect(self.view.zoomOut)
        zoom_reset_btn.clicked.connect(self.reset_zoom)
        fit_all_btn = QPushButton("Fit All")
        fit_sel_btn = QPushButton("Fit Sel")
        def _fit_all():
            r = self.scene.itemsBoundingRect()
            if not r.isNull():
                self.view.fitInView(r, _keep_ar())
                try:
                    if hasattr(self.view, '_persist_zoom'):
                        self.view._persist_zoom()
                    sf = float(self.view.transform().m11())
                    from math import isfinite
                    if isfinite(sf):
                        try:
                            zoom_label.setText(f"{int(round(sf*100))}%")
                        except Exception:
                            pass
                except Exception:
                    pass
        fit_all_btn.clicked.connect(_fit_all)
        def _fit_sel():
            items = [it for it in self.scene.selectedItems()]
            from PyQt6.QtCore import QRectF
            if items:
                rect = QRectF()
                for it in items:
                    rect = rect.united(it.sceneBoundingRect())
                if not rect.isNull():
                    self.view.fitInView(rect, _keep_ar())
                    try:
                        if hasattr(self.view, '_persist_zoom'):
                            self.view._persist_zoom()
                        sf = float(self.view.transform().m11())
                        from math import isfinite
                        if isfinite(sf):
                            try:
                                zoom_label.setText(f"{int(round(sf*100))}%")
                            except Exception:
                                pass
                    except Exception:
                        pass
                    return
            name = getattr(self, '_locked_label', None)
            if name and name in getattr(self, '_name_to_rect', {}):
                rect = self._name_to_rect[name].sceneBoundingRect()
                self.view.fitInView(rect.adjusted(-40, -20, 40, 20), _keep_ar())
                try:
                    if hasattr(self.view, '_persist_zoom'):
                        self.view._persist_zoom()
                    sf = float(self.view.transform().m11())
                    from math import isfinite
                    if isfinite(sf):
                        try:
                            zoom_label.setText(f"{int(round(sf*100))}%")
                        except Exception:
                            pass
                except Exception:
                    pass
            else:
                _fit_all()
        fit_sel_btn.clicked.connect(_fit_sel)
        zoom_layout.addWidget(zoom_in_btn)
        zoom_layout.addWidget(zoom_out_btn)
        zoom_layout.addWidget(zoom_reset_btn)
        zoom_layout.addWidget(fit_all_btn)
        zoom_layout.addWidget(fit_sel_btn)
        # Live zoom percentage label
        try:
            zoom_label = QLabel("100%")
            zoom_label.setToolTip("Current zoom")
            def _set_zoom_label_g(sf):
                try:
                    from math import isfinite
                    val = float(sf)
                    if isfinite(val):
                        zoom_label.setText(f"{int(round(val*100))}%")
                except Exception:
                    pass
            # Connect signal and seed
            try:
                self.view.zoomChanged.connect(_set_zoom_label_g)
                _set_zoom_label_g(self.view.transform().m11())
            except Exception:
                pass
            zoom_layout.addWidget(zoom_label)
        except Exception:
            pass
        layout.addLayout(zoom_layout)
        layout.addWidget(self.preview_label)
        # Mini legend (hierarchy vs dependency vs highlight colors)
        self.legend_label = QLabel()
        self.legend_label.setFixedHeight(24)
        self.legend_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.legend_label.setStyleSheet("font-size:11px; padding:2px 6px; background:#1a1a1a; border:1px solid #444;")
        try:
            self.legend_label.setTextFormat(Qt.RichText)
        except Exception:
            pass
        layout.addWidget(self.legend_label)
        self.setLayout(layout)
        self._did_initial_fit = False

        # Keyboard shortcuts
        try:
            from PyQt6.QtWidgets import QShortcut
            from PyQt6.QtGui import QKeySequence
            QShortcut(QKeySequence.ZoomIn, self.view, activated=self.view.zoomIn)
            QShortcut(QKeySequence.ZoomOut, self.view, activated=self.view.zoomOut)
            QShortcut(QKeySequence("Ctrl+0"), self.view, activated=self.reset_zoom)
        except Exception:
            pass

    def _maybe_initial_fit(self):
        """Perform a one-time fit-to-view if no prior zoom preference was stored.
        We infer absence of stored zoom by checking if current scale ~= 1 and flag not yet set."""
        if self._did_initial_fit:
            return
        try:
            scale_x = self.view.transform().m11()
            if abs(scale_x - 1.0) < 0.001:
                from PyQt6.QtCore import QTimer
                def do_fit():
                    r = self.scene.itemsBoundingRect()
                    if not r.isNull():
                        self.view.fitInView(r, _keep_ar())
                QTimer.singleShot(0, do_fit)
            self._did_initial_fit = True
        except Exception:
            self._did_initial_fit = True

    def _populate_baselines(self):
        try:
            if not hasattr(self, 'model') or not self.model:
                return
            current = self.baseline_combo.currentText() if hasattr(self, 'baseline_combo') else "None"
            self.baseline_combo.blockSignals(True)
            self.baseline_combo.clear()
            self.baseline_combo.addItem("None")
            for b in self.model.list_baselines():
                self.baseline_combo.addItem(b)
            # restore selection
            idx = self.baseline_combo.findText(current)
            if idx >= 0:
                self.baseline_combo.setCurrentIndex(idx)
            self.baseline_combo.blockSignals(False)
        except Exception as e:
            print(f"Populate baselines failed: {e}")

    def reset_zoom(self):
        # Reset zoom via the view helper (also persists state)
        if hasattr(self, 'view') and self.view:
            try:
                self.view.resetZoom()
            except Exception:
                # Fallback for older code paths
                self.view.resetTransform()

    def refresh_gantt(self):
        if hasattr(self, 'model') and self.model:
            # keep baseline list up to date
            if hasattr(self, '_populate_baselines'):
                self._populate_baselines()
            self.render_gantt(self.model)

    # Unified connector + label highlighting
    # Restores original dynamic connector line highlighting AND integrates label font/background highlight.
    def _highlight_connectors(self, part_name, on):
        from PyQt6.QtGui import QPen, QColor, QFont, QBrush
        # 1. Connector lines
        if hasattr(self, '_connector_lines_map'):
            lines = self._connector_lines_map.get(part_name, [])
            if lines:
                for ln in lines:
                    try:
                        # Animated hierarchy connector support
                        if hasattr(ln, 'set_highlight') and callable(getattr(ln, 'set_highlight')):
                            ln.set_highlight(on)
                            continue
                        kind = ln.data(50)  # 'dep' or 'hier'
                        if on:
                            pen = QPen(QColor("#00BFFF"))
                            pen.setWidth(2)
                            ln.setPen(pen)
                        else:
                            if kind == 'dep':
                                pen = QPen(QColor("#FFAA28"))  # dependency base color
                                pen.setWidth(2)
                            else:  # hierarchy or default
                                pen = QPen(QColor(180,180,180))
                                pen.setWidth(1)
                            ln.setPen(pen)
                    except Exception:
                        pass
        # 2. Label font + background (preserve original orange bg when turning off)
        if hasattr(self, '_name_to_text_item'):
            ti = self._name_to_text_item.get(part_name)
            if ti:
                orig_font = ti.data(1)
                bg = ti.data(3)
                orig_brush = ti.data(4)  # stored original brush
                if on:
                    if isinstance(orig_font, QFont):
                        bold_font = QFont(orig_font)
                        bold_font.setBold(True)
                        ti.setFont(bold_font)
                    if bg:
                        base_col = QColor(255,130,0)
                        glow = QColor(base_col.red(), base_col.green(), base_col.blue(), 160)
                        bg.setBrush(QBrush(glow))
                else:
                    if isinstance(orig_font, QFont):
                        ti.setFont(orig_font)
                    if bg:
                        if orig_brush:
                            bg.setBrush(orig_brush)
                        else:
                            bg.setBrush(QBrush(QColor("#FF8200")))
    def render_gantt(self, model):
        self.model = model
        self.scene.clear()
        self.preview_label.clear()
        if not hasattr(model, 'rows'):
            return
        # Clear any stale unscheduled flags from prior renders
        try:
            for _r in getattr(model, 'rows', []) or []:
                if isinstance(_r, dict) and '_unscheduled' in _r:
                    _r.pop('_unscheduled', None)
        except Exception:
            pass
        # Performance guard rails and feature toggles
        # Allow external toggle of connector rendering via attribute set by MainWindow
        # Connector enable & mode (restored by MainWindow after construction)
        if not hasattr(self, '_enable_connectors'):
            self._enable_connectors = True  # default
        if not hasattr(self, '_connector_mode'):
            # 'all' (hierarchy+deps) or 'deps'
            self._connector_mode = 'all'
        # Dynamic thresholds: scale by approximate visible complexity
        # Base thresholds (for ~1200px width viewport & moderate rows)
        base_hard = 2500
        base_soft = 1200
        try:
            vw = self.view.viewport().width() if hasattr(self,'view') else 1200
            vh = self.view.viewport().height() if hasattr(self,'view') else 600
            area_factor = max(0.6, min(2.0, (vw*vh)/(1200*700)))  # clamp scaling
            row_factor = 1.0
            try:
                row_factor = max(0.5, min(2.0, (len(raw_rows)/200.0)))  # more rows increases chance of edges, tighten thresholds
            except Exception:
                row_factor = 1.0
            MAX_CONNECTOR_EDGES = int(base_hard * area_factor / row_factor)
            SOFT_WARN_EDGES = int(base_soft * area_factor / row_factor)
        except Exception:
            MAX_CONNECTOR_EDGES = 2500
            SOFT_WARN_EDGES = 1200
        raw_rows = model.rows
        # Precompute critical path set for filtering/highlighting
        self._current_critical_set = set()
        try:
            import datetime as _dt
            name_to_row_cp = {r.get("Project Part", ""): r for r in raw_rows}
            graph = {}
            for r in raw_rows:
                nm = r.get("Project Part", "")
                # Some deployments may only populate 'Dependencies' (downstream list) and leave legacy 'Predecessors' blank.
                # For visualization / critical path we treat Dependencies as predecessors (tasks that must finish before this one starts).
                preds_raw = (r.get("Predecessors") or "").strip()
                if not preds_raw:
                    deps_field = (r.get("Dependencies") or "").strip()
                    # Use dependencies as predecessors if predecessors absent
                    if deps_field:
                        preds_raw = deps_field
                preds = [p.strip() for p in preds_raw.split(',') if p.strip()] if preds_raw else []
                graph[nm] = preds
            visited = set(); order = []
            def dfs(n):
                if n in visited: return
                visited.add(n)
                for p in graph.get(n, []):
                    if p in name_to_row_cp:
                        dfs(p)
                order.append(n)
            for n in graph:
                dfs(n)
            es = {}; ef = {}
            for n in order:
                r = name_to_row_cp.get(n) or {}
                s_str = r.get("Start Date", "")
                try:
                    s_dt = _dt.datetime.strptime(s_str, "%m-%d-%Y") if s_str else None
                    dur = int(r.get("Duration (days)") or 0)
                except Exception:
                    s_dt = None; dur = 0
                pred_finishes = [ef[p] for p in graph.get(n, []) if p in ef]
                base = max(pred_finishes) if pred_finishes else s_dt
                if base is None:
                    base = _dt.datetime.today()
                es[n] = base
                ef[n] = base + _dt.timedelta(days=dur)
            if order:
                proj_finish = max(ef.values())
                ls = {}; lf = {}
                for n in reversed(order):
                    succs = [s for s, preds in graph.items() if n in preds]
                    if not succs:
                        lf[n] = proj_finish
                    else:
                        lf[n] = min(ls[s] for s in succs)
                    dur = (ef[n] - es[n]).days
                    ls[n] = lf[n] - _dt.timedelta(days=dur)
                self._current_critical_set = {n for n in order if abs((es[n]-ls[n]).days) <= 0}
        except Exception:
            self._current_critical_set = set()

        matched = []
        name_to_row = {}
        for r in raw_rows:
            name_to_row[r.get("Project Part", "")] = r
            if self._passes_filters(r):
                matched.append(r)
        if any([self._filter_statuses, self._filter_internal_external, self._filter_responsible_substr,
                self._filter_critical_only, self._filter_risk_only]):
            parent_names_needed = set()
            for r in matched:
                parent_name = r.get("Parent") or ""
                while parent_name:
                    if parent_name in parent_names_needed:
                        break
                    parent_names_needed.add(parent_name)
                    parent_row = next((x for x in raw_rows if x.get("Project Part") == parent_name), None)
                    if parent_row:
                        parent_name = parent_row.get("Parent") or ""
                    else:
                        break
            rows = [r for r in raw_rows if r in matched or r.get("Project Part") in parent_names_needed]
        else:
            rows = raw_rows

        # ---------- Helpers ----------
        def topo_sort(all_rows):
            name_to_row = {r.get("Project Part", ""): r for r in all_rows}
            visited = set()
            result = []
            def visit(r):
                name = r.get("Project Part", "")
                if name in visited:
                    return
                parent = r.get("Parent", "")
                if parent and parent in name_to_row:
                    visit(name_to_row[parent])
                visited.add(name)
                result.append(r)
            for r in all_rows:
                visit(r)
            return result

        def compute_parent_spans(all_rows):
            import datetime as _dt
            children = {}
            for r in all_rows:
                p = r.get("Parent", "")
                if p:
                    children.setdefault(p, []).append(r)
            def update_span(r, visited=None):
                if visited is None:
                    visited = set()
                name = r.get("Project Part", "")
                if name in visited:
                    return None, None
                visited.add(name)
                if name not in children:
                    try:
                        s = _dt.datetime.strptime(r.get("Start Date", ""), "%m-%d-%Y")
                        d = int(r.get("Duration (days)", 0))
                        e = s + _dt.timedelta(days=d)
                        return s, e
                    except Exception:
                        return None, None
                child_spans = [update_span(c, visited.copy()) for c in children[name]]
                starts = [s for s, e in child_spans if s]
                ends = [e for s, e in child_spans if e]
                if starts and ends:
                    r["_auto_start"] = min(starts)
                    r["_auto_end"] = max(ends)
                    return r["_auto_start"], r["_auto_end"]
                return None, None
            for r in all_rows:
                update_span(r)

        compute_parent_spans(rows)
        rows = topo_sort(rows)
        if not rows:
            return

        # ---------- Build bar data ----------
        import datetime
        bar_height = 24
        bar_gap = 10
        min_date = None
        max_date = None
        bars = []  # (name, start, duration, index, row_dict)
        # Flexible parsers to maximize inclusion of valid rows
        def _parse_date_flex(s: str):
            s = (s or "").strip()
            if not s:
                return None
            for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        def _parse_duration_flex(v):
            if v in (None, ""):
                return None
            try:
                return int(str(v).strip())
            except Exception:
                # Try float, or extract first numeric token (e.g., "10 d", "7.5")
                try:
                    import re
                    m = re.search(r"-?\d+(\.\d+)?", str(v))
                    if m:
                        return int(round(float(m.group(0))))
                except Exception:
                    pass
            return None
        for idx, r in enumerate(rows):
            if "_auto_start" in r and "_auto_end" in r:
                start = r["_auto_start"]
                end = r["_auto_end"]
                duration = (end - start).days
            else:
                # Be flexible: include parts that have Start+Duration OR End+Duration OR Start+End
                start = None; end = None; duration = None
                try:
                    start_str = (r.get("Start Date") or r.get("Actual Start Date") or r.get("Baseline Start Date") or "").strip()
                    end_str = (r.get("Calculated End Date") or r.get("Actual Finish Date") or r.get("Baseline End Date") or "").strip()
                    dur_val = r.get("Duration (days)")
                    # Flexible parsing for dates and duration
                    start = _parse_date_flex(start_str)
                    end = _parse_date_flex(end_str)
                    duration = _parse_duration_flex(dur_val)
                    # Derive one missing piece when possible
                    if start and duration is not None and duration >= 0:
                        end = start + datetime.timedelta(days=duration)
                    elif end and duration is not None and duration >= 0:
                        start = end - datetime.timedelta(days=duration)
                    elif start and end:
                        duration = max(0, (end - start).days)
                    elif start and not end:
                        # Default to 1-day bar when only a single boundary exists
                        duration = 1
                        end = start + datetime.timedelta(days=1)
                    elif end and not start:
                        duration = 1
                        start = end - datetime.timedelta(days=1)
                    else:
                        continue
                except Exception:
                    continue
            if not start:
                continue
            if min_date is None or start < min_date:
                min_date = start
            if max_date is None or end > max_date:
                max_date = end
            bars.append((r.get("Project Part", ""), start, duration, idx, r))

        # Debug: quick coverage summary for Gantt
        try:
            name_set = {n for (n, *_rest) in bars}
            total_rows = len(rows)
            missing = [r.get("Project Part", "") for r in rows if r.get("Project Part", "") not in name_set]
            sample_missing = ", ".join([m for m in missing if m][:8])
            if min_date and max_date:
                try:
                    print(f"[Gantt] built {len(bars)} bars from {total_rows} rows; date range {min_date.strftime('%m-%d-%Y')}..{max_date.strftime('%m-%d-%Y')}")
                except Exception:
                    print(f"[Gantt] built {len(bars)} bars from {total_rows} rows")
            else:
                print(f"[Gantt] built {len(bars)} bars from {total_rows} rows")
            if missing and sample_missing:
                print(f"[Gantt] missing (no sched) examples: {sample_missing}")
        except Exception:
            pass

        # Include unscheduled items as label-only markers at chart start for visibility (when enabled)
        try:
            if getattr(self, '_show_unscheduled', True):
                if min_date and rows:
                    name_set = {n for (n, *_rest) in bars}
                    for idx, r in enumerate(rows):
                        n = r.get("Project Part", "")
                        if not n or n in name_set:
                            continue
                        # Mark unscheduled; draw as tiny marker with label
                        r["_unscheduled"] = True
                        bars.append((n, min_date, 0, idx, r))
        except Exception:
            pass

        if not bars:
            return

        chart_min_date = min_date  # earliest start

        # --- Optional: Reorder rows to group predecessors, with cross-parent/weight options ---
        row_index_map = None
        try:
            if getattr(self, '_group_predecessors', False):
                # Build base order from current filtered/topo rows
                base_order = [r.get("Project Part", "") for r in rows]
                name_to_row_local = {r.get("Project Part", ""): r for r in rows}
                base_pos_map = {nm: idx for idx, nm in enumerate(base_order)}
                # Build deps map using Predecessors if present else Dependencies
                deps_map = {}
                for r in rows:
                    nm = r.get("Project Part", "")
                    preds_raw = (r.get("Predecessors") or "").strip()
                    if not preds_raw:
                        preds_raw = (r.get("Dependencies") or "").strip()
                    preds = [p.strip() for p in preds_raw.split(',') if p.strip()]
                    # keep only those in our current set
                    preds = [p for p in preds if p in name_to_row_local]
                    deps_map[nm] = preds
                # Build successors count for weighting: how many tasks depend on each node
                succ_count = {}
                for succ, preds in deps_map.items():
                    for p in preds:
                        succ_count[p] = succ_count.get(p, 0) + 1
                # Helper to find root ancestor (top-most parent)
                parent_of = {nm: (name_to_row_local.get(nm, {}).get("Parent") or "") for nm in name_to_row_local}
                def root_of(nm):
                    seen = set()
                    cur = nm
                    while True:
                        if not cur or cur in seen:
                            return cur
                        seen.add(cur)
                        par = parent_of.get(cur, "")
                        if not par:
                            return cur
                        cur = par
                # Build auxiliary maps for weighting: start date and duration per row
                import datetime as _dt_gp
                name_to_start = {}
                name_to_duration = {}
                for r in rows:
                    nm = r.get("Project Part", "")
                    try:
                        if "_auto_start" in r:
                            name_to_start[nm] = r["_auto_start"]
                            name_to_duration[nm] = max(0, (r.get("_auto_end") - r.get("_auto_start")).days)
                        else:
                            s_str = (r.get("Start Date") or r.get("Actual Start Date") or r.get("Baseline Start Date") or "").strip()
                            d_val = r.get("Duration (days)")
                            s_dt = None
                            if s_str:
                                for fmt in ("%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%Y/%m/%d"):
                                    try:
                                        s_dt = _dt_gp.datetime.strptime(s_str, fmt)
                                        break
                                    except Exception:
                                        pass
                            name_to_start[nm] = s_dt
                            try:
                                name_to_duration[nm] = int(str(d_val).strip()) if d_val not in (None,"") else 0
                            except Exception:
                                name_to_duration[nm] = 0
                    except Exception:
                        name_to_start[nm] = None
                        name_to_duration[nm] = 0
                # Weight strategy helper
                def _weight_key(p, strategy, base_order_map=base_pos_map):
                    if strategy == "Weight: Successors":
                        return (-succ_count.get(p, 0), base_order_map.get(p, 10**9))
                    if strategy == "Weight: Original":
                        return (0, base_order_map.get(p, 10**9))
                    if strategy == "Weight: Start Date":
                        sd = name_to_start.get(p)
                        # Earlier start first; None last
                        return (1, 10**9) if sd is None else (0, sd.toordinal() if hasattr(sd, 'toordinal') else base_order_map.get(p, 10**9))
                    if strategy == "Weight: Reverse Start Date":
                        sd = name_to_start.get(p)
                        # Later start first; None last
                        return (1, -10**9) if sd is None else (0, -sd.toordinal() if hasattr(sd, 'toordinal') else -base_order_map.get(p, 10**9))
                    if strategy == "Weight: Duration":
                        # Longer first
                        return (-name_to_duration.get(p, 0), base_order_map.get(p, 10**9))
                    if strategy == "Weight: Criticality":
                        # Critical tasks first; then successors, then base order
                        is_crit = 1 if p in getattr(self, '_current_critical_set', set()) else 0
                        # invert to put critical (1) first by sorting negative; use -is_crit
                        return (-is_crit, -succ_count.get(p, 0), base_order_map.get(p, 10**9))
                    if strategy == "Weight: Criticality then Start Date":
                        is_crit = 1 if p in getattr(self, '_current_critical_set', set()) else 0
                        sd = name_to_start.get(p)
                        # critical first, then earlier start first; None last
                        sd_key = (1, 10**9) if sd is None else (0, sd.toordinal() if hasattr(sd, 'toordinal') else base_order_map.get(p, 10**9))
                        return (-is_crit, sd_key, base_order_map.get(p, 10**9))
                    # Fallback
                    return (-succ_count.get(p, 0), base_order_map.get(p, 10**9))
                # Pick current strategy
                try:
                    strategy = self.weight_combo.currentText()
                except Exception:
                    strategy = "Weight: Successors"
                # Grouping pass: for each successor, cluster its multiple preds that share same parent or same root (option)
                seq = list(base_order)
                i = 0
                while i < len(seq):
                    s = seq[i]
                    if not s:
                        i += 1; continue
                    row_s = name_to_row_local.get(s) or {}
                    parent_s = (row_s.get("Parent") or "")
                    preds_all = deps_map.get(s, [])
                    if getattr(self, '_group_cross_parent', False):
                        root_s = root_of(s)
                        preds_eligible = [p for p in preds_all if p in seq and root_of(p) == root_s]
                    else:
                        preds_eligible = [p for p in preds_all if (name_to_row_local.get(p, {}).get("Parent") or "") == parent_s and p in seq]
                    # Only act when multiple preds exist
                    if len(preds_eligible) >= 2:
                        # Weighted, stable order by chosen strategy; tiebreaker is base order
                        preds_in_order = [x for x in seq if x in preds_eligible]
                        preds_sorted = sorted(preds_in_order, key=lambda p: _weight_key(p, strategy))
                        s_index = i
                        block_start = max(0, s_index - len(preds_sorted))
                        if seq[block_start:s_index] != preds_sorted:
                            # Remove these preds from wherever they are
                            seq = [x for x in seq if x not in preds_sorted]
                            # Recompute index of successor after removal
                            s_index = seq.index(s)
                            # Insert the predecessors as a contiguous block just before successor
                            seq[s_index:s_index] = preds_sorted
                            # Advance past successor to avoid reprocessing same window endlessly
                            i = seq.index(s) + 1
                            continue
                    i += 1
                # Build index map for bars that actually render
                bars_names = [n for (n, *_rest) in bars]
                ordered_names = [n for n in seq if n in bars_names]
                # Include any bars missing from seq (unlikely) at the end to ensure mapping exists
                for n in bars_names:
                    if n not in ordered_names:
                        ordered_names.append(n)
                row_index_map = {n: idx for idx, n in enumerate(ordered_names)}
        except Exception:
            row_index_map = None

    # ---------- Draw bars ----------
        from PyQt6.QtGui import QColor
        from PyQt6.QtWidgets import QGraphicsRectItem, QGraphicsItem
        gantt_color = QColor("#FF8200")

        # Optional critical path calculation
        critical_set = set()
        if hasattr(self, 'critical_path_checkbox') and self.critical_path_checkbox.isChecked():
            try:
                # Build dependency graph and durations
                name_to_row = {r.get("Project Part", ""): r for r in rows}
                graph = {}
                duration_map = {}
                import datetime as _dt_cp
                for r in rows:
                    name = r.get("Project Part", "")
                    deps = [d.strip() for d in (r.get("Dependencies", "") or "").split(',') if d.strip()]
                    graph[name] = deps
                    try:
                        if "_auto_start" in r and "_auto_end" in r:
                            duration_map[name] = (r["_auto_end"] - r["_auto_start"]).days
                        else:
                            duration_map[name] = int(r.get("Duration (days)", 0) or 0)
                    except Exception:
                        duration_map[name] = 0
                # Topological order (simple DFS; assumes no complex cycles)
                visited = set(); order = []
                def dfs(n):
                    if n in visited: return
                    for d in graph.get(n, []): dfs(d)
                    visited.add(n); order.append(n)
                for n in graph: dfs(n)
                # Forward pass: earliest finish
                earliest_finish = {}
                earliest_start = {}
                for n in order:
                    deps = graph.get(n, [])
                    if not deps:
                        # Use explicit start date if available for alignment; else zero
                        row = name_to_row.get(n, {})
                        try:
                            if "_auto_start" in row:
                                est = row["_auto_start"]
                            else:
                                est = datetime.datetime.strptime(row.get("Start Date", ""), "%m-%d-%Y")
                        except Exception:
                            est = min_date
                        earliest_start[n] = est
                    else:
                        # earliest start is max earliest finish of deps
                        ef_candidates = []
                        for d in deps:
                            if d in earliest_finish:
                                ef_candidates.append(earliest_finish[d])
                        base = max(ef_candidates) if ef_candidates else min_date
                        earliest_start[n] = base
                    earliest_finish[n] = earliest_start[n] + datetime.timedelta(days=duration_map.get(n,0))
                project_finish = max(earliest_finish.values()) if earliest_finish else None
                # Backward pass: latest start
                latest_start = {}; latest_finish = {}
                for n in reversed(order):
                    # Successors: tasks that depend on n
                    succs = [k for k,v in graph.items() if n in v]
                    if not succs:
                        latest_finish[n] = project_finish
                    else:
                        latest_finish[n] = min([latest_start[s] for s in succs]) if succs else project_finish
                    latest_start[n] = latest_finish[n] - datetime.timedelta(days=duration_map.get(n,0))
                # Critical tasks: zero total float (allow <= 0 days tolerance)
                for n in order:
                    if abs((earliest_start[n] - latest_start[n]).days) <= 0:
                        critical_set.add(n)
            except Exception as e:
                print(f"WARNING: Critical path calculation failed: {e}")

        class ClickableBar(QGraphicsRectItem):
            def __init__(self, x, y, w, h, row_dict, preview_label, gantt_view):
                super().__init__(x, y, w, h)
                self.row = row_dict
                self.preview_label = preview_label
                self.gantt_view = gantt_view
                self.setAcceptHoverEvents(True)
                # Link indicator + cursor if Pace Link present and links enabled
                try:
                    self._pace_link = (self.row.get("Pace Link") or "").strip()
                    has_link = self._pace_link.lower().startswith("http://") or self._pace_link.lower().startswith("https://")
                except Exception:
                    self._pace_link = ""; has_link = False
                if has_link and getattr(self.gantt_view, '_show_links', True):
                    try:
                        # Small chain icon at top-right of bar
                        from PyQt6.QtWidgets import QGraphicsSimpleTextItem
                        icon = QGraphicsSimpleTextItem("🔗", self)
                        icon.setBrush(QColor("white"))
                        # position relative to bar (padding 3px)
                        icon.setPos(max(2, w - 12), 2)
                        icon.setZValue(self.zValue() + 2)
                    except Exception:
                        pass
                    try:
                        # Tooltip with URL and pointing hand cursor
                        self.setToolTip(self._pace_link)
                        from PyQt6.QtGui import QCursor
                        self.setCursor(getattr(Qt,'CursorShape', Qt).PointingHandCursor)
                    except Exception:
                        pass
                # Set selectable flag (PyQt6 namespaced enums with fallback)
                try:
                    from PyQt6.QtWidgets import QGraphicsItem as _QGI
                    self.setFlag(_QGI.GraphicsItemFlag.ItemIsSelectable, True)
                except Exception:
                    try:
                        from PyQt6.QtWidgets import QGraphicsItem as _QGI2
                        if hasattr(_QGI2, 'ItemIsSelectable'):
                            self.setFlag(_QGI2.ItemIsSelectable, True)
                    except Exception:
                        pass

            # --- Attachment utilities ---
            def _attachments_list(self):
                import json
                raw = self.row.get("Attachments") or "[]"
                try:
                    lst = json.loads(raw)
                    if isinstance(lst, list):
                        return [p for p in lst if isinstance(p, str)]
                except Exception:
                    pass
                return []
            def _save_attachments_list(self, lst):
                import json
                self.row["Attachments"] = json.dumps(lst)
                # Persist via model if parent widget exposes save_model()
                pw = self.preview_label.parentWidget()
                if pw and hasattr(pw, 'model'):
                    try:
                        pw.model.save_to_db()
                    except Exception as e:
                        print(f"Attachment save failed: {e}")
            def contextMenuEvent(self, event):
                from PyQt6.QtWidgets import QMenu
                menu = QMenu()
                open_action = menu.addAction("Open Attachments…")
                add_action = menu.addAction("Add Attachment…")
                open_folder_action = menu.addAction("Open Attachments Folder")
                chosen = menu.exec(event.screenPos())
                if chosen == open_action:
                    self.show_attachments_dialog()
                elif chosen == add_action:
                    self.add_attachment_files()
                elif chosen == open_folder_action:
                    self.open_attachments_folder()
            def add_attachment_files(self):
                from PyQt6.QtWidgets import QFileDialog
                import os, shutil
                files, _ = QFileDialog.getOpenFileNames(None, "Select Attachment(s)")
                if not files:
                    return
                import sys
                base_dir = os.path.dirname(resolve_resource_path("."))
                attach_dir = os.path.join(base_dir, 'attachments')
                if not os.path.exists(attach_dir):
                    os.makedirs(attach_dir)
                current = self._attachments_list()
                for f in files:
                    name = os.path.basename(f)
                    dest = os.path.join(attach_dir, name)
                    root, ext = os.path.splitext(name)
                    counter = 1
                    while os.path.exists(dest):
                        dest = os.path.join(attach_dir, f"{root}_{counter}{ext}")
                        counter += 1
                    try:
                        shutil.copy2(f, dest)
                        rel = os.path.relpath(dest, base_dir)
                        current.append(rel)
                    except Exception as e:
                        print(f"Attachment copy failed: {e}")
                self._save_attachments_list(current)
            def open_attachments_folder(self):
                import os, sys, subprocess
                base_dir = os.path.dirname(resolve_resource_path("."))
                attach_dir = os.path.join(base_dir, 'attachments')
                if not os.path.exists(attach_dir):
                    os.makedirs(attach_dir)
                if sys.platform.startswith('win'):
                    os.startfile(attach_dir)
                elif sys.platform == 'darwin':
                    subprocess.Popen(['open', attach_dir])
                else:
                    subprocess.Popen(['xdg-open', attach_dir])
            def show_attachments_dialog(self):
                from PyQt6.QtWidgets import QDialog, QVBoxLayout, QListWidget, QPushButton, QHBoxLayout, QLabel
                import os, webbrowser
                dlg = QDialog()
                dlg.setWindowTitle(f"Attachments - {self.row.get('Project Part','')}")
                vbox = QVBoxLayout(dlg)
                info = QLabel("Double-click or Open to launch. Remove only deletes reference.")
                vbox.addWidget(info)
                lst = QListWidget(); vbox.addWidget(lst)
                thumb = QLabel(); thumb.setFixedHeight(110); vbox.addWidget(thumb)
                btn_row = QHBoxLayout()
                add_btn = QPushButton("Add…"); rem_btn = QPushButton("Remove"); open_btn = QPushButton("Open")
                btn_row.addWidget(add_btn); btn_row.addWidget(rem_btn); btn_row.addWidget(open_btn)
                vbox.addLayout(btn_row)
                for p in self._attachments_list():
                    lst.addItem(p)
                def refresh_thumb():
                    from PyQt6.QtGui import QPixmap
                    item = lst.currentItem()
                    if not item:
                        thumb.clear(); return
                    rel = item.text()
                    base_dir = os.path.dirname(resolve_resource_path("."))
                    full = os.path.join(base_dir, rel)
                    if os.path.exists(full) and os.path.splitext(full)[1].lower() in ('.png','.jpg','.jpeg','.bmp','.gif'):
                        pm = QPixmap(full)
                        if not pm.isNull():
                            try:
                                _smooth = Qt.TransformationMode.SmoothTransformation
                            except Exception:
                                _smooth = getattr(Qt, 'SmoothTransformation', 1)
                            thumb.setPixmap(pm.scaledToHeight(100, _smooth)); return
                    thumb.setText(os.path.basename(full))
                def do_add():
                    self.add_attachment_files(); lst.clear(); [lst.addItem(p) for p in self._attachments_list()]; refresh_thumb()
                def do_remove():
                    item = lst.currentItem();
                    if not item: return
                    rel = item.text()
                    remain = [p for p in self._attachments_list() if p != rel]
                    self._save_attachments_list(remain)
                    lst.takeItem(lst.currentRow()); refresh_thumb()
                def do_open():
                    item = lst.currentItem();
                    if not item: return
                    rel = item.text()
                    base_dir = os.path.dirname(resolve_resource_path("."))
                    full = os.path.join(base_dir, rel)
                    if os.path.exists(full):
                        webbrowser.open(full)
                lst.currentItemChanged.connect(lambda *_: refresh_thumb())
                lst.itemDoubleClicked.connect(lambda *_: do_open())
                add_btn.clicked.connect(do_add)
                rem_btn.clicked.connect(do_remove)
                open_btn.clicked.connect(do_open)
                refresh_thumb()
                dlg.exec()
            def _set_preview(self):
                img_path = self.row.get("Images", "")
                if img_path and str(img_path).strip():
                    from PyQt6.QtGui import QPixmap
                    img_path_full = resolve_resource_path(img_path)
                    pm = QPixmap(img_path_full)
                    if not pm.isNull():
                        try:
                            _smooth = Qt.TransformationMode.SmoothTransformation
                        except Exception:
                            _smooth = getattr(Qt, 'SmoothTransformation', 1)
                        self.preview_label.setPixmap(pm.scaledToHeight(90, _smooth))
                        self.preview_label.setText("")
                        return
                # Ensure QPixmap is imported when clearing
                from PyQt6.QtGui import QPixmap
                self.preview_label.setText("")
                self.preview_label.setPixmap(QPixmap())
            def mousePressEvent(self, event):
                try:
                    # If Pace Link present, open it on click
                    _left = getattr(Qt, 'MouseButton', Qt).LeftButton
                    if hasattr(event, 'button') and event.button() == _left and getattr(self.gantt_view, '_show_links', True):
                        url = (self.row.get('Pace Link') or '').strip()
                        if url and (url.lower().startswith('http://') or url.lower().startswith('https://')):
                            try:
                                from PyQt6.QtGui import QDesktopServices
                                from PyQt6.QtCore import QUrl
                                QDesktopServices.openUrl(QUrl(url))
                            except Exception:
                                import webbrowser; webbrowser.open(url)
                            return
                    # Fallback: original behavior (show edit dialog)
                    self._set_preview()
                    parent_widget = self.preview_label.parentWidget()
                    if parent_widget and hasattr(parent_widget, 'show_edit_dialog'):
                        parent_widget.show_edit_dialog(self.row)
                except Exception as e:
                    print(f"ERROR in ClickableBar.mousePressEvent: {e}")
            def hoverEnterEvent(self, event):
                self._set_preview()
                part = self.row.get("Project Part", "")
                if part:
                    self.gantt_view._highlight_connectors(part, True)
                parent = self.row.get("Parent", "")
                if parent:
                    self.gantt_view._highlight_connectors(parent, True)
                # Reverse lookup: also highlight dependency neighbors (incoming & outgoing)
                try:
                    if hasattr(self.gantt_view, '_connector_lines_map'):
                        # Incoming: tasks that depend on this part (edges src->part)
                        for name, lines in self.gantt_view._connector_lines_map.items():
                            if name == part:
                                continue
                            if any(getattr(ln, 'data', lambda *_: None)(50) == 'dep' and part == part for ln in lines):
                                # If shared line between name and part, highlight neighbor
                                self.gantt_view._highlight_connectors(name, True)
                except Exception:
                    pass
            def hoverLeaveEvent(self, event):
                self.preview_label.clear()
                part = self.row.get("Project Part", "")
                if part:
                    self.gantt_view._highlight_connectors(part, False)
                parent = self.row.get("Parent", "")
                if parent:
                    self.gantt_view._highlight_connectors(parent, False)
                # Remove dependency neighbor highlights
                try:
                    if hasattr(self.gantt_view, '_connector_lines_map'):
                        for name in list(self.gantt_view._connector_lines_map.keys()):
                            if name != part:
                                self.gantt_view._highlight_connectors(name, False)
                except Exception:
                    pass
                # Fallback to first image attachment preview if no explicit image assigned
                if not self.row.get("Images"):
                    atts = self._attachments_list()
                    if atts:
                        from PyQt6.QtGui import QPixmap
                        full = resolve_resource_path(atts[0])
                        if os.path.exists(full):
                            pm = QPixmap(full)
                            if not pm.isNull():
                                try:
                                    _smooth = Qt.TransformationMode.SmoothTransformation
                                except Exception:
                                    _smooth = getattr(Qt, 'SmoothTransformation', 1)
                                self.preview_label.setPixmap(pm.scaledToHeight(90, _smooth))
                                self.preview_label.setText("")

        name_to_bar = {}
        self._name_to_rect = {}
        bar_items = []
        # (Reverted) Previously labels were placed in a dedicated left column and bars were offset.
        # We now restore inline style with external labels to the right of bars.
        from PyQt6.QtGui import QFontMetrics, QFont
        font = self.font() if hasattr(self, 'font') else None
        fm = QFontMetrics(font) if font else None
        max_chars_fixed = 32  # keep truncation behavior
        left_margin = 60
        bar_offset_x = left_margin
        self._name_to_text_item = {}
        # Weekend/holiday background shading for full chart span
        try:
            from PyQt6.QtGui import QBrush
            from datetime import timedelta
            shade_wknd = QBrush(QColor(220,220,220,120))
            holidays = load_holiday_dates()
            shade_hol = QBrush(QColor(255,215,0,60))
            cur = chart_min_date
            while cur <= max_date:
                # weekend
                if cur.weekday() >= 5:
                    run_start = cur
                    while cur <= max_date and cur.weekday() >= 5:
                        cur += timedelta(days=1)
                    run_end = cur
                    x0 = (run_start - chart_min_date).days * 10 + bar_offset_x
                    x1 = (run_end - chart_min_date).days * 10 + bar_offset_x
                    self.scene.addRect(x0, 0, max(1, x1-x0), len(bars)*(bar_height+bar_gap)+80, pen=(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), brush=shade_wknd)
                else:
                    # holidays (single days)
                    if cur.date() in holidays:
                        x0 = (cur - chart_min_date).days * 10 + bar_offset_x
                        self.scene.addRect(x0, 0, 10, len(bars)*(bar_height+bar_gap)+80, pen=(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), brush=shade_hol)
                    cur += timedelta(days=1)
        except Exception:
            pass

        for name, start, duration, i, r in bars:
            x = (start - chart_min_date).days * 10 + bar_offset_x
            # Use grouped index if available
            j = row_index_map.get(name, i) if isinstance(row_index_map, dict) else i
            y = j * (bar_height + bar_gap) + 40
            # Unscheduled rows: draw a small tick and label only
            if r.get("_unscheduled"):
                width = 2
                from PyQt6.QtGui import QPen as _MarkerPen
                marker = self.scene.addLine(x, y, x, y + bar_height, _MarkerPen(QColor("#777777")))
                # Label to the right, italic to indicate unscheduled
                text_item = self.scene.addText(name)
                from PyQt6.QtGui import QFont as _QFontLab, QColor as _QColor, QBrush as _QBrush, QPen as _QPen
                f = text_item.font(); f.setItalic(True); text_item.setFont(f)
                text_item.setDefaultTextColor(_QColor("white"))
                ty = y + (bar_height - text_item.boundingRect().height())/2
                text_item.setPos(x + 6, ty)
                # Subtle pill background
                br = text_item.boundingRect().translated(text_item.pos())
                from PyQt6.QtGui import QPainterPath as _LblPath
                padded = br.adjusted(-3,-1,3,1)
                path = _LblPath(); path.addRoundedRect(padded, 6, 6)
                bg_brush = _QBrush(QColor(80,80,80,180))
                bg_rect = self.scene.addPath(path, _QPen(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), bg_brush)
                bg_rect.setZValue(text_item.zValue()-1)
                self._name_to_text_item[name] = text_item
                name_to_bar[name] = (x, y, width, bar_height)
                bar_items.append((marker, r))
                continue
            # Milestone diamond for zero-duration or explicit type
            is_milestone = False
            try:
                is_milestone = (duration == 0) or ((r.get('Type') or '').strip().lower() == 'milestone')
            except Exception:
                is_milestone = False
            width = max(duration * 10, 10)
            rect = None
            if is_milestone:
                # Draw a diamond centered at x with side equal to bar_height
                from PyQt6.QtGui import QPainterPath as _PathMS, QPen as _PenMS, QBrush as _BrushMS
                size = bar_height
                cx = x
                cy = y + bar_height/2
                path = _PathMS()
                path.moveTo(cx, cy - size/2)
                path.lineTo(cx + size/2, cy)
                path.lineTo(cx, cy + size/2)
                path.lineTo(cx - size/2, cy)
                path.closeSubpath()
                diamond = self.scene.addPath(path, _PenMS(QColor('#FF8200'), 2), _BrushMS(QColor('#333333')))
                # Make selectable and carry row
                try:
                    from PyQt6.QtWidgets import QGraphicsItem as _QGI_ms
                    diamond.setFlag(_QGI_ms.GraphicsItemFlag.ItemIsSelectable, True)
                except Exception:
                    pass
                diamond.row = r
                # Keep consistent API for selection handling
                rect = diamond
                width = size  # for label positioning below
            else:
                rect = ClickableBar(x, y, width, bar_height, r, self.preview_label, self)
                rect.setBrush(QColor("#333333"))
            from PyQt6.QtGui import QPen as _QPen4
            import datetime as _dt_ov
            overdue = False; at_risk = False
            try:
                if "_auto_end" in r:
                    scheduled_end = r["_auto_end"]
                else:
                    end_calc = r.get("Calculated End Date", "")
                    if end_calc:
                        scheduled_end = _dt_ov.datetime.strptime(end_calc, "%m-%d-%Y")
                    else:
                        scheduled_end = start + _dt_ov.timedelta(days=duration)
                today = _dt_ov.datetime.today()
                pc_val = int(r.get("% Complete") or 0)
                status_val = (r.get("Status") or "").strip()
                if pc_val < 100 and today.date() > scheduled_end.date():
                    overdue = True
                elif pc_val == 0 and status_val in ("Planned", "Blocked") and today.date() > start.date():
                    at_risk = True
            except Exception:
                pass
            outline_pen = _QPen4(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0))
            if overdue:
                outline_pen = _QPen4(QColor("red")); outline_pen.setWidth(2)
            elif at_risk:
                outline_pen = _QPen4(QColor("#FFA500")); outline_pen.setWidth(2)
            rect.setPen(outline_pen)
            self.scene.addItem(rect)
            self._name_to_rect[name] = rect
            try:
                pc = int(r.get("% Complete") or 0)
            except Exception:
                pc = 0
            if (not is_milestone) and pc > 0:
                prog_w = max(2, int(width * pc / 100))
                prog_color = QColor("#DAA520") if name in critical_set else gantt_color
                from PyQt6.QtGui import QPen as _QPen3
                prog_rect = self.scene.addRect(x, y, prog_w, bar_height, _QPen3(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), prog_color)
                # Disable mouse interaction: PyQt6 uses Qt.MouseButton.NoButton
                try:
                    prog_rect.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
                except Exception:
                    # Fallback for legacy style
                    nb = getattr(Qt, 'NoButton', 0)
                    prog_rect.setAcceptedMouseButtons(nb)
                prog_rect.setZValue(rect.zValue() + 1)
            full_name = name
            display_name = full_name
            # Paperclip if attachments present
            try:
                import json as _json_attlabel
                att_raw = r.get("Attachments") or "[]"
                att_list = _json_attlabel.loads(att_raw) if att_raw else []
                if isinstance(att_list, list) and len(att_list) > 0:
                    display_name = "\uD83D\uDCCE " + display_name  # paperclip emoji
            except Exception:
                pass
            if len(display_name) > max_chars_fixed:
                display_name = display_name[:max_chars_fixed-1] + "…"
            text_item = self.scene.addText(display_name)
            from PyQt6.QtGui import QColor as _QColor, QFont, QBrush, QPen
            text_item.setDefaultTextColor(_QColor("black"))
            orig_font = text_item.font()
            text_item.setData(1, orig_font)
            text_item.setData(2, full_name)  # store full for tooltip
            text_item.setToolTip(full_name)
            ty = y + (bar_height - text_item.boundingRect().height())/2
            # Place label just to the right of the bar with small gap
            gap = 6
            text_item.setPos(x + width + gap, ty)
            # Always-visible subtle contrasting background for readability
            br = text_item.boundingRect().translated(text_item.pos())
            from PyQt6.QtGui import QPen as _LblPen, QBrush as _LblBrush, QColor as _LblColor, QPainterPath as _LblPath
            bg_color = _LblColor("#FF8200")  # orange background
            padded = br.adjusted(-3,-1,3,1)
            path = _LblPath()
            radius = 6
            path.addRoundedRect(padded, radius, radius)
            bg_brush = _LblBrush(bg_color)
            bg_rect = self.scene.addPath(path, _LblPen(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), bg_brush)
            bg_rect.setZValue(text_item.zValue()-1)
            text_item.setData(3, bg_rect)  # store bg rect
            text_item.setData(4, bg_brush)  # store original brush
            self._name_to_text_item[name] = text_item
            name_to_bar[name] = (x, y, width, bar_height)
            bar_items.append((rect, r))

        # Baseline overlay (thin background lines)
        try:
            baseline_name = getattr(self, '_selected_baseline_name', None)
            if baseline_name:
                bmap = self.model.load_baseline_map(baseline_name)
                from PyQt6.QtGui import QPen
                pen = QPen(QColor(150,150,150))
                try:
                    pen.setStyle(Qt.PenStyle.DashLine)  # PyQt6 namespaced enum
                except Exception:
                    # Fallback for PyQt5
                    pen.setStyle(getattr(Qt, 'DashLine', 1))
                pen.setWidth(1)
                for name, pos in name_to_bar.items():
                    if name in bmap:
                        bs, be = bmap[name]
                        import datetime as _dtb
                        try:
                            if bs:
                                s = _dtb.datetime.strptime(bs, "%m-%d-%Y")
                            else:
                                continue
                            if be:
                                e = _dtb.datetime.strptime(be, "%m-%d-%Y")
                            else:
                                continue
                            x0 = (s - chart_min_date).days * 10 + bar_offset_x
                            x1 = (e - chart_min_date).days * 10 + bar_offset_x
                            y = pos[1] + pos[3]//2
                            self.scene.addLine(x0, y, x1, y, pen)
                        except Exception:
                            pass
        except Exception as e:
            print(f"Baseline overlay failed: {e}")

        # ---------- Selection handling ----------
        self._bar_rect_to_row = {}
        for rect, r in bar_items:
            try:
                from PyQt6.QtWidgets import QGraphicsItem as _QGI
                rect.setFlag(_QGI.GraphicsItemFlag.ItemIsSelectable, True)
            except Exception:
                try:
                    from PyQt6.QtWidgets import QGraphicsItem as _QGI2
                    if hasattr(_QGI2, 'ItemIsSelectable'):
                        rect.setFlag(_QGI2.ItemIsSelectable, True)
                except Exception:
                    pass
            self._bar_rect_to_row[rect] = r
        def on_selection_changed():
            selected = [it for it in self.scene.selectedItems() if it in self._bar_rect_to_row]
            if selected:
                bar = selected[0]
                try:
                    if bar.scene() is not None:
                        r = self._bar_rect_to_row[bar]
                        self.show_edit_dialog(r)
                        bar.setSelected(False)
                except RuntimeError:
                    pass
        try:
            self.scene.selectionChanged.disconnect()
        except TypeError:
            pass
        self.scene.selectionChanged.connect(on_selection_changed)

        # ---------- Axis ----------
        if min_date and max_date:
            axis_y = 30
            axis_x0 = bar_offset_x
            axis_x1 = (max_date - chart_min_date).days * 10 + bar_offset_x + 40
            self.scene.addLine(axis_x0, axis_y, axis_x1, axis_y)
            tick_interval = 7
            total_days = (max_date - chart_min_date).days
            import datetime as _dt2
            for d in range(0, total_days + 1, tick_interval):
                tick_x = axis_x0 + d * 10
                self.scene.addLine(tick_x, axis_y - 5, tick_x, axis_y + 5)
                tick_date = chart_min_date + _dt2.timedelta(days=d)
                tick_label = self.scene.addText(tick_date.strftime("%m-%d-%Y"))
                from PyQt6.QtGui import QColor as _QColor
                tick_label.setDefaultTextColor(_QColor("white"))
                tick_label.setPos(tick_x - 30, axis_y - 25)
            # Today line
            try:
                import datetime as _dt_today
                today = _dt_today.datetime.today().date()
                if chart_min_date.date() <= today <= max_date.date():
                    dx = (today - chart_min_date.date()).days
                    x_today = axis_x0 + dx * 10
                    from PyQt6.QtGui import QPen as _QPToday, QColor as _QCToday
                    pen_today = _QPToday(_QCToday(0, 200, 255))
                    pen_today.setWidth(2)
                    try:
                        pen_today.setStyle(Qt.PenStyle.DashLine)
                    except Exception:
                        pen_today.setStyle(getattr(Qt,'DashLine',1))
                    self.scene.addLine(x_today, 0, x_today, axis_y + len(bars)*(bar_height+bar_gap) + 80, pen_today)
                    lbl = self.scene.addText("Today")
                    lbl.setDefaultTextColor(_QCToday(0,200,255))
                    lbl.setPos(x_today + 4, 4)
            except Exception:
                pass

        # ---------- Dependency arrows (simple) ----------
        from PyQt6.QtGui import QPen, QColor as _QColor2
        import datetime as _dt3
        name_to_dates = {}
        for name, start, duration, i, r in bars:
            end = start + _dt3.timedelta(days=duration)
            name_to_dates[name] = (start, end)
        for name, start, duration, i, r in bars:
            deps = r.get("Dependencies", "")
            if not deps:
                continue
            dep_list = [d.strip() for d in deps.split(',') if d.strip()]
            for dep_name in dep_list:
                if dep_name not in name_to_bar:
                    continue
                dep_x, dep_y, dep_w, dep_h = name_to_bar[dep_name]
                this_x, this_y, this_w, this_h = name_to_bar.get(name, (None, None, None, None))
                if this_x is None:
                    continue
                dep_end = name_to_dates.get(dep_name, (None, None))[1]
                this_start = name_to_dates.get(name, (None, None))[0]
                conflict = dep_end and this_start and dep_end >= this_start
                if conflict:
                    pen = QPen(_QColor2("red"), 2)
                else:
                    # Critical path dependency (both tasks critical and not conflict) use gold
                    if dep_name in critical_set and name in critical_set:
                        pen = QPen(_QColor2("#DAA520"), 2)
                    else:
                        pen = QPen(_QColor2("#FF8200"), 2)
                start_x = dep_x + dep_w
                start_y = dep_y + dep_h/2
                end_x = this_x
                end_y = this_y + this_h/2
                # L-shaped routing (feature 6)
                self.scene.addLine(start_x, start_y, end_x, start_y, pen)
                self.scene.addLine(end_x, start_y, end_x, end_y, pen)
        # ---------- Parent-child connectors (hierarchical fan-out, animated) ----------
        from PyQt6.QtGui import QPen as _QPen, QColor as _QColor3
        from PyQt6.QtWidgets import QGraphicsLineItem
        from PyQt6.QtCore import QPropertyAnimation, pyqtProperty
        draw_hierarchy = True
        if hasattr(self, 'hierarchy_checkbox'):
            try:
                draw_hierarchy = self.hierarchy_checkbox.isChecked()
            except Exception:
                draw_hierarchy = True
        class _AnimatedConnector(QGraphicsLineItem):
            def __init__(self, x1, y1, x2, y2, base_pen, highlight_pen, style):
                super().__init__(x1, y1, x2, y2)
                self._base_pen = _QPen(base_pen)
                self._highlight_pen = _QPen(highlight_pen)
                if style == 'trunk':
                    try:
                        self._base_pen.setStyle(Qt.PenStyle.DashLine)
                    except Exception:
                        self._base_pen.setStyle(getattr(Qt, 'DashLine', 1))
                    self._base_pen.setWidth(2)
                    self._highlight_pen.setWidth(3)
                else:
                    self._base_pen.setWidth(1)
                    self._highlight_pen.setWidth(2)
                self.setPen(self._base_pen)
                self._opacity = 0.55
                self._anim = None
                self._apply_opacity()
            def _apply_opacity(self):
                p = self.pen()
                c = p.color()
                c.setAlphaF(self._opacity)
                p.setColor(c)
                self.setPen(p)
            def getOpacity(self):
                return self._opacity
            def setOpacity(self, val):
                self._opacity = val
                self._apply_opacity()
            opacity = pyqtProperty(float, fget=getOpacity, fset=setOpacity)
            def fade(self, target, duration):
                if self._anim:
                    self._anim.stop()
                self._anim = QPropertyAnimation(self, b"opacity")
                self._anim.setDuration(duration)
                self._anim.setStartValue(self._opacity)
                self._anim.setEndValue(target)
                self._anim.start()
            def set_highlight(self, on):
                if on:
                    self.setPen(self._highlight_pen)
                    self.fade(1.0, 180)
                else:
                    self.setPen(self._base_pen)
                    self.fade(0.55, 260)
        self._connector_lines_map = {}
        # Estimate potential edge count early (parent-child + dependency edges) to decide skip
        edge_estimate = 0
        try:
            # parent-child count ~ number of rows with a parent
            edge_estimate += sum(1 for r in raw_rows if r.get('Parent'))
            # dependencies count
            for r in raw_rows:
                deps_txt = r.get('Dependencies') or ''
                if deps_txt.strip():
                    edge_estimate += len([d for d in deps_txt.split(',') if d.strip()])
        except Exception:
            pass
        skip_connectors = False
        soft_warn = False
        if edge_estimate > MAX_CONNECTOR_EDGES:
            skip_connectors = True
        elif edge_estimate > SOFT_WARN_EDGES:
            soft_warn = True
        if soft_warn and hasattr(self, 'preview_label'):
            try:
                self.preview_label.setText(f"Connector density high ({edge_estimate} edges) – drawn with reduced styling")
            except Exception:
                pass
        if skip_connectors and hasattr(self, 'preview_label'):
            try:
                self.preview_label.setText(f"Connectors skipped (edge est {edge_estimate} > {MAX_CONNECTOR_EDGES})")
            except Exception:
                pass
        if self._enable_connectors and not skip_connectors:
            base_color = _QColor3(180,180,180)         # hierarchy child lines
            trunk_color = _QColor3(150,150,150)        # hierarchy trunk
            dep_color = _QColor3(255,170,40)           # dependency arrows (amber/orange)
            highlight_color = _QColor3('#00BFFF')
            parent_children = {}
            for name, start, duration, i, r in bars:
                parent_name = r.get("Parent", "") or ""
                if parent_name and parent_name in name_to_bar and name in name_to_bar:
                    parent_children.setdefault(parent_name, []).append(name)
            def _register(part, item):
                self._connector_lines_map.setdefault(part, []).append(item)
            # Draw hierarchy connectors only if mode is 'all' and hierarchy toggle present
            if self._connector_mode == 'all' and draw_hierarchy:
                for parent, children in parent_children.items():
                    if not children:
                        continue
                    px, py, pw, ph = name_to_bar[parent]
                    parent_mid_x = px + pw/2
                    parent_bottom_y = py + ph
                    child_positions = []
                    for child in children:
                        cx, cy, cw, ch = name_to_bar[child]
                        child_positions.append((child, cx + cw/2, cy))
                    child_positions.sort(key=lambda t: t[2])
                    trunk_top = parent_bottom_y
                    trunk_bottom = child_positions[-1][2]
                    trunk = _AnimatedConnector(parent_mid_x, trunk_top, parent_mid_x, trunk_bottom,
                                                base_pen=_QPen(trunk_color), highlight_pen=_QPen(highlight_color), style='trunk')
                    trunk.setZValue(-1)
                    self.scene.addItem(trunk)
                    _register(parent, trunk)
                    for child, cmx, cty in child_positions:
                        h_line = _AnimatedConnector(min(parent_mid_x, cmx), cty, max(parent_mid_x, cmx), cty,
                                                    base_pen=_QPen(base_color), highlight_pen=_QPen(highlight_color), style='child')
                        h_line.setZValue(-1)
                        self.scene.addItem(h_line)
                        v_line = _AnimatedConnector(cmx, cty, cmx, cty,
                                                    base_pen=_QPen(base_color), highlight_pen=_QPen(highlight_color), style='child')
                        v_line.setZValue(-1)
                        self.scene.addItem(v_line)
                        _register(parent, h_line); _register(child, h_line)
                        _register(child, v_line); _register(parent, v_line)
            # Draw dependency arrows (Dependencies field) when mode == 'deps' OR mode == 'all'
            if self._connector_mode in ('deps','all'):
                from math import atan2, cos, sin, pi
                try:
                    # Build quick lookup: name -> (x,y,w,h)
                    name_rect = {n: name_to_bar.get(n) for n in name_to_bar}
                    # Collect dependency edges (A depends on B => arrow B -> A)
                    dep_edges = []  # list of (src,dst)
                    for r in raw_rows:
                        a = r.get('Project Part','')
                        if not a:
                            continue
                        deps_txt = r.get('Dependencies') or ''
                        for d in [x.strip() for x in deps_txt.split(',') if x.strip()]:
                            if d and d in name_rect and a in name_rect and d != a:
                                dep_edges.append((d, a))
                    if len(dep_edges) > MAX_CONNECTOR_EDGES:
                        if hasattr(self, 'preview_label'):
                            try:
                                prev = self.preview_label.text() or ''
                                self.preview_label.setText(prev + f"  (dep edges {len(dep_edges)} skipped)")
                            except Exception:
                                pass
                    else:
                        # Collision-aware lane assignment: spread mid_x offsets per destination
                        from collections import defaultdict
                        edges_by_dst = defaultdict(list)
                        for s, d in dep_edges:
                            edges_by_dst[d].append(s)
                        dep_pen_template = _QPen(dep_color); dep_pen_template.setWidth(2)
                        lane_spacing = 26  # px separation between vertical lanes
                        for dst, srcs in edges_by_dst.items():
                            srcs_sorted = sorted(srcs, key=lambda n: name_rect[n][1])  # order by y to keep stable
                            count = len(srcs_sorted)
                            for idx_src, src in enumerate(srcs_sorted):
                                sx, sy, sw, sh = name_rect[src]
                                dx, dy, dw, dh = name_rect[dst]
                                start_x = sx + sw
                                start_y = sy + sh / 2
                                end_x = dx
                                end_y = dy + dh / 2
                                base_mid = (start_x + end_x) / 2
                                offset = (idx_src - (count - 1) / 2.0) * lane_spacing
                                mid_x = base_mid + offset
                                # Build segmented path (could be curved later)
                                from PyQt6.QtGui import QPainterPath, QPen as _QPenDep
                                path = QPainterPath()
                                path.moveTo(start_x, start_y)
                                path.lineTo(mid_x, start_y)
                                path.lineTo(mid_x, end_y)
                                path.lineTo(end_x, end_y)
                                dep_pen = _QPenDep(dep_color); dep_pen.setWidth(2)
                                path_item = self.scene.addPath(path, dep_pen)
                                # Mark as dependency for legend/highlight restore
                                try:
                                    path_item.setData(50, 'dep')
                                    path_item.setToolTip(f"{src} → {dst}")
                                except Exception:
                                    pass
                                _register(src, path_item); _register(dst, path_item)
                                # Arrow head (centered on final segment)
                                try:
                                    angle = atan2(0, end_x - mid_x)  # horizontal final segment
                                    ah = 8.0
                                    a1 = angle + pi - 0.45
                                    a2 = angle + pi + 0.45
                                    p1x = end_x; p1y = end_y
                                    p2x = end_x + ah * cos(a1); p2y = end_y + ah * sin(a1)
                                    p3x = end_x + ah * cos(a2); p3y = end_y + ah * sin(a2)
                                    from PyQt6.QtGui import QPolygonF
                                    from PyQt6.QtCore import QPointF
                                    poly = QPolygonF([QPointF(p1x, p1y), QPointF(p2x, p2y), QPointF(p3x, p3y)])
                                    arrow_item = self.scene.addPolygon(poly, dep_pen, dep_color)
                                    try:
                                        arrow_item.setData(50, 'dep')
                                        arrow_item.setToolTip(f"{src} → {dst}")
                                    except Exception:
                                        pass
                                    _register(src, arrow_item); _register(dst, arrow_item)
                                except Exception:
                                    pass
                        # Legend / mode message
                        if hasattr(self, 'preview_label'):
                            try:
                                prev = self.preview_label.text() or ''
                                mode_note = f"Mode: {self._connector_mode} (hierarchy {'on' if self._connector_mode=='all' else 'off'}, deps {len(dep_edges)})"
                                if prev:
                                    self.preview_label.setText(prev + '  ' + mode_note)
                                else:
                                    self.preview_label.setText(mode_note)
                            except Exception:
                                pass
                except Exception:
                    pass
            # (Future) Could draw dependency arrows separately when _connector_mode == 'deps'
            # Currently dependency visualization is label highlighting; expand later for explicit edges.

        # Update legend label
        try:
            if hasattr(self, 'legend_label'):
                if self._enable_connectors and not skip_connectors:
                    self.legend_label.setVisible(True)
                    legend_html = (
                        "Legend: "
                        "<span style='display:inline-block; padding:2px 6px; background:#b4b4b4; color:#000; border-radius:4px;'>Hierarchy</span> "
                        "<span style='display:inline-block; padding:2px 6px; background:#FFAA28; color:#000; border-radius:4px;'>Dependency</span> "
                        "<span style='display:inline-block; padding:2px 6px; background:#00BFFF; color:#000; border-radius:4px;'>Highlight</span>"
                    )
                    # Show Unscheduled chip when toggle enabled
                    try:
                        if getattr(self, '_show_unscheduled', True):
                            legend_html += " " + "<span style='display:inline-block; padding:2px 6px; background:#888; color:#fff; border-radius:4px;'>Unscheduled</span>"
                    except Exception:
                        pass
                    self.legend_label.setText(legend_html)
                else:
                    self.legend_label.setVisible(False)
        except Exception:
            pass

        # ---------- Scene rect ----------
        self.view.setSceneRect(0, 0, 800, max(300, len(bars)*(bar_height+bar_gap)+60))
        # Adjust scene width to fit axis_x1 if larger
        current_rect = self.view.sceneRect()
        if current_rect.width() < axis_x1 + 100:
            self.view.setSceneRect(0, 0, axis_x1 + 100, current_rect.height())
        # One-time initial auto-fit (Option B) if user has not previously zoomed
        try:
            self._maybe_initial_fit()
        except Exception:
            pass

    # Click-to-lock highlight support
    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        # Determine if a bar was clicked -> lock its label highlight until another click
        pos = event.pos()
        if hasattr(self, 'view'):
            scene_pos = self.view.mapToScene(pos)
            items = self.scene.items(scene_pos)
            target_name = None
            for it in items:
                if hasattr(it, 'row') and isinstance(it.row, dict):
                    target_name = it.row.get("Project Part", "")
                    break
            if target_name:
                # Clear previous lock
                if hasattr(self, '_locked_label') and self._locked_label and self._locked_label != target_name:
                    self._highlight_connectors(self._locked_label, False)
                self._locked_label = target_name
                self._highlight_connectors(target_name, True)
        event.accept()

class CalendarView(QWidget):
    def __init__(self, model=None):
        super().__init__()
        self.model = model
        from PyQt6.QtWidgets import QCalendarWidget, QListWidget, QMessageBox, QPushButton, QHBoxLayout
        from PyQt6.QtGui import QTextCharFormat, QBrush, QColor
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Calendar (Click a date to see tasks)"))
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        layout.addWidget(self.calendar)
        self.task_list = QListWidget()
        layout.addWidget(self.task_list)
        # Add 'Today' and 'Export Calendar' buttons
        btn_layout = QHBoxLayout()
        today_btn = QPushButton("Today")
        today_btn.clicked.connect(self.go_to_today)
        btn_layout.addWidget(today_btn)
        export_btn = QPushButton("Export Calendar (.ics)")
        export_btn.clicked.connect(self.export_calendar_ics)
        btn_layout.addWidget(export_btn)
        layout.addLayout(btn_layout)
        self.setLayout(layout)
        self.calendar.selectionChanged.connect(self.update_task_list)
        self.calendar.clicked.connect(self.show_task_details)
        self.highlight_task_dates()
        self.update_task_list()

    def export_calendar_ics(self):
        """Export all tasks as iCalendar (.ics) file, including Pace Link, Responsible, and Type."""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import datetime
        if not self.model or not hasattr(self.model, 'rows'):
            QMessageBox.warning(self, "Export Failed", "No data to export.")
            return
        # Ask for file path
        path, _ = QFileDialog.getSaveFileName(self, "Export Calendar", "project_calendar.ics", "iCalendar Files (*.ics)")
        if not path:
            return
        # Build .ics content
        def escape_ics(text):
            return str(text).replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")
        ics_lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//Aja au Grimace//Project Calendar//EN"
        ]
        import re
        email_regex = r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+"
        for row in self.model.rows:
            part = row.get("Project Part", "(Unnamed)")
            start_str = row.get("Start Date", "")
            duration = row.get("Duration (days)", "")
            notes = row.get("Notes", "")
            pace_link = row.get("Pace Link", "")
            responsible = row.get("Responsible", "")
            type_ = row.get("Type", "")
            if not start_str or not duration:
                continue
            try:
                start_dt = datetime.datetime.strptime(start_str, "%m-%d-%Y")
                days = int(duration)
                # End date is exclusive in iCalendar, so add 1 day
                end_dt = start_dt + datetime.timedelta(days=days)
                dtstart = start_dt.strftime("%Y%m%d")
                dtend = end_dt.strftime("%Y%m%d")
            except Exception:
                continue
            # Compose description with all requested fields
            desc_lines = []
            if notes:
                desc_lines.append(notes)
            if pace_link:
                desc_lines.append(f"Pace Link: {pace_link}")
            if responsible:
                desc_lines.append(f"Responsible: {responsible}")
            if type_:
                desc_lines.append(f"Type: {type_}")
            description = "\n".join(desc_lines)
            # Find all email addresses in Responsible field
            attendee_lines = []
            if responsible:
                emails = re.findall(email_regex, responsible)
                for email in emails:
                    attendee_lines.append(f"ATTENDEE;CN={email}:mailto:{email}")
            ics_lines.extend([
                "BEGIN:VEVENT",
                f"SUMMARY:{escape_ics(part)}",
                f"DTSTART;VALUE=DATE:{dtstart}",
                f"DTEND;VALUE=DATE:{dtend}",
                f"DESCRIPTION:{escape_ics(description)}",
            ] + attendee_lines + [
                "END:VEVENT"
            ])
        ics_lines.append("END:VCALENDAR")
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\r\n".join(ics_lines))
            QMessageBox.information(self, "Export Complete", f"Calendar exported to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write file: {e}")

    def highlight_task_dates(self):
        if not self.model:
            return
        from PyQt6.QtGui import QTextCharFormat, QBrush, QColor
        fmt = QTextCharFormat()
        fmt.setBackground(QBrush(QColor("#ffe082")))  # Light yellow
        # Clear previous highlights
        self.calendar.setDateTextFormat(self.calendar.minimumDate(), QTextCharFormat())
        self.calendar.setDateTextFormat(self.calendar.maximumDate(), QTextCharFormat())
        # Highlight all dates with tasks
        dates_with_tasks = set()
        for row in getattr(self.model, 'rows', []):
            date_str = row.get("Start Date", "")
            if date_str:
                from PyQt6.QtCore import QDate
                date = QDate.fromString(date_str, "MM-dd-yyyy")
                if date.isValid():
                    dates_with_tasks.add(date)
        for date in dates_with_tasks:
            self.calendar.setDateTextFormat(date, fmt)

    def update_task_list(self):
        self.task_list.clear()
        if not self.model:
            return
        selected_date = self.calendar.selectedDate()
        date_str = selected_date.toString("MM-dd-yyyy")
        for row in getattr(self.model, 'rows', []):
            if row.get("Start Date", "") == date_str:
                part = row.get("Project Part", "(Unnamed)")
                self.task_list.addItem(part)

    def show_task_details(self, qdate):
        if not self.model:
            return
        date_str = qdate.toString("MM-dd-yyyy")
        tasks = [row for row in getattr(self.model, 'rows', []) if row.get("Start Date", "") == date_str]
        if not tasks:
            return
        msg = "Tasks for {}:\n".format(date_str)
        for row in tasks:
            msg += f"- {row.get('Project Part', '(Unnamed)')}\n"
        QMessageBox.information(self, "Tasks on {}".format(date_str), msg)

    def go_to_today(self):
        from PyQt6.QtCore import QDate
        self.calendar.setSelectedDate(QDate.currentDate())
        self.update_task_list()
        self.calendar.showSelectedDate()

class TimelineView(QWidget):
    def __init__(self, model=None):
        super().__init__()
        self.model = model
        layout = QVBoxLayout()
        layout.addWidget(QLabel("Project Timeline (Read-Only)"))
        from PyQt6.QtWidgets import QGraphicsScene
        self.scene = QGraphicsScene()
        self.view = ZoomableGraphicsView()
        self.view.setScene(self.scene)
        self.view.setSettingsKey("TimelineZoom")
        layout.addWidget(self.view)
        # Export buttons
        from PyQt6.QtWidgets import QHBoxLayout, QPushButton
        export_row = QHBoxLayout()
        export_png_btn = QPushButton("Export Timeline (PNG/PDF)")
        def _do_export():
            # Reuse GanttChartView helper through lightweight wrapper
            try:
                # Local import to avoid circular issues
                helper = getattr(self, '_export_helper', None)
                if helper is None:
                    helper = GanttChartView()  # temporary helper just for exporter
                    self._export_helper = helper
                helper._export_scene_with_header(self.scene, title="Timeline")
            except Exception as e:
                print(f"Timeline export failed: {e}")
        export_png_btn.clicked.connect(_do_export)
        export_row.addWidget(export_png_btn)
        # Zoom / Fit controls
        zoom_in_btn = QPushButton("Zoom In")
        zoom_out_btn = QPushButton("Zoom Out")
        zoom_reset_btn = QPushButton("Reset Zoom")
        fit_all_btn = QPushButton("Fit to View")
        zoom_in_btn.clicked.connect(self.view.zoomIn)
        zoom_out_btn.clicked.connect(self.view.zoomOut)
        zoom_reset_btn.clicked.connect(self.view.resetZoom)
        def _fit_all_tl():
            r = self.scene.itemsBoundingRect()
            if not r.isNull():
                self.view.fitInView(r, _keep_ar())
                try:
                    if hasattr(self.view, '_persist_zoom'):
                        self.view._persist_zoom()
                    # Update zoom label if available
                    try:
                        sf = float(self.view.transform().m11())
                        from math import isfinite
                        if isfinite(sf):
                            tl_zoom_label.setText(f"{int(round(sf*100))}%")
                    except Exception:
                        pass
                except Exception:
                    pass
        fit_all_btn.clicked.connect(_fit_all_tl)
        export_row.addWidget(zoom_in_btn)
        export_row.addWidget(zoom_out_btn)
        export_row.addWidget(zoom_reset_btn)
        export_row.addWidget(fit_all_btn)
        # Live zoom percentage label
        try:
            tl_zoom_label = QLabel("100%")
            tl_zoom_label.setToolTip("Current zoom")
            def _set_zoom_label_tl(sf):
                try:
                    from math import isfinite
                    val = float(sf)
                    if isfinite(val):
                        tl_zoom_label.setText(f"{int(round(val*100))}%")
                except Exception:
                    pass
            try:
                self.view.zoomChanged.connect(_set_zoom_label_tl)
                _set_zoom_label_tl(self.view.transform().m11())
            except Exception:
                pass
            export_row.addWidget(tl_zoom_label)
        except Exception:
            pass
        layout.addLayout(export_row)
        self.preview_label = QLabel()
        self.preview_label.setFixedHeight(200)
        self.preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.preview_label)
        self.setLayout(layout)
        # Keyboard shortcuts for zoom
        try:
            from PyQt6.QtWidgets import QShortcut
            from PyQt6.QtGui import QKeySequence
            QShortcut(QKeySequence.ZoomIn, self.view, activated=self.view.zoomIn)
            QShortcut(QKeySequence.ZoomOut, self.view, activated=self.view.zoomOut)
            QShortcut(QKeySequence("Ctrl+0"), self.view, activated=self.view.resetZoom)
        except Exception:
            pass
        self.render_timeline()

    def render_timeline(self):
        # --- Critical Path Calculation ---
        def find_critical_path(rows):
            import datetime
            # Build graph: node = part name, edges = dependencies
            name_to_row = {row.get("Project Part", ""): row for row in rows}
            graph = {}
            for row in rows:
                name = row.get("Project Part", "")
                deps = row.get("Dependencies", "")
                dep_list = [d.strip() for d in deps.split(",") if d.strip()]
                graph[name] = dep_list
            # Topo sort
            visited = set()
            order = []
            def visit(n):
                if n in visited:
                    return
                for dep in graph.get(n, []):
                    visit(dep)
                visited.add(n)
                order.append(n)
            for n in graph:
                visit(n)
            # Calculate earliest start/finish
            est = {}
            eft = {}
            for n in order:
                row = name_to_row.get(n, {})
                duration = int(row.get("Duration (days)", 0) or 0)
                if not duration:
                    continue
                deps = graph.get(n, [])
                if not deps:
                    start = row.get("Start Date", "")
                    if start:
                        est[n] = datetime.datetime.strptime(start, "%m-%d-%Y")
                    else:
                        est[n] = datetime.datetime.min
                else:
                    est[n] = max([eft.get(dep, datetime.datetime.min) for dep in deps])
                eft[n] = est[n] + datetime.timedelta(days=duration)
            # Calculate latest finish/start
            lft = {n: max(eft.values()) for n in order}
            lst = {}
            for n in reversed(order):
                row = name_to_row.get(n, {})
                duration = int(row.get("Duration (days)", 0) or 0)
                deps = graph.get(n, [])
                if not deps:
                    lft[n] = lft[n]
                else:
                    lft[n] = min([lst.get(dep, lft[n]) for dep in deps])
                lst[n] = lft[n] - datetime.timedelta(days=duration)
            # Critical path: nodes where est==lst
            critical = set(n for n in order if est.get(n) == lst.get(n))
            return critical
        import datetime
        from PyQt6.QtGui import QBrush, QColor
        from PyQt6.QtCore import QDate
        self.scene.clear()
        if not self.model or not hasattr(self.model, 'rows'):
            return
        rows = [row for row in self.model.rows if row.get("Start Date") and row.get("Duration (days)")]
        def topo_sort(rows):
            name_to_row = {row.get("Project Part", ""): row for row in rows}
            visited = set()
            result = []
            def visit(row):
                name = row.get("Project Part", "")
                if name in visited:
                    return
                parent = row.get("Parent", "")
                if parent and parent in name_to_row:
                    visit(name_to_row[parent])
                visited.add(name)
                result.append(row)
            for row in rows:
                visit(row)
            return result
        def compute_parent_spans(rows):
            import datetime
            name_to_row = {row.get("Project Part", ""): row for row in rows}
            children = {}
            for row in rows:
                parent = row.get("Parent", "")
                if parent:
                    children.setdefault(parent, []).append(row)
            def update_span(row, visited=None):
                if visited is None:
                    visited = set()
                name = row.get("Project Part", "")
                if name in visited:
                    # Cycle detected, break recursion
                    return None, None
                visited.add(name)
                if name not in children:
                    try:
                        start = datetime.datetime.strptime(row.get("Start Date", ""), "%m-%d-%Y")
                        duration = int(row.get("Duration (days)", 0))
                        end = start + datetime.timedelta(days=duration)
                        return start, end
                    except Exception:
                        return None, None
                else:
                    child_spans = [update_span(child, visited.copy()) for child in children[row.get("Project Part")]]
                    child_starts = [s for s, e in child_spans if s]
                    child_ends = [e for s, e in child_spans if e]
                    if child_starts and child_ends:
                        min_start = min(child_starts)
                        max_end = max(child_ends)
                        row["_auto_start"] = min_start
                        row["_auto_end"] = max_end
                        return min_start, max_end
                    return None, None
            for row in rows:
                update_span(row)
        compute_parent_spans(rows)
        rows = topo_sort(rows)
        if not rows:
            return
        # Parse dates and durations
        bars = []
        name_to_idx = {}
        for idx, row in enumerate(rows):
            if "_auto_start" in row and "_auto_end" in row:
                start = row["_auto_start"]
                end = row["_auto_end"]
                duration = (end - start).days
            else:
                start_str = row.get("Start Date", "")
                duration = row.get("Duration (days)", 0)
                try:
                    start = datetime.datetime.strptime(start_str, "%m-%d-%Y")
                    duration = int(duration)
                    end = start + datetime.timedelta(days=duration)
                except Exception:
                    continue
            bars.append((row.get("Project Part", "(Unnamed)"), start, duration, row, idx))
            name_to_idx[row.get("Project Part", "(Unnamed)")] = idx
        if not bars:
            return
        # Find min and max dates
        min_date = min([b[1] for b in bars])
        max_date = max([b[1] + datetime.timedelta(days=b[2]) for b in bars])
        total_days = (max_date - min_date).days
        # Compute critical path set for highlighting
        try:
            critical_path = find_critical_path(rows)
        except Exception:
            critical_path = set()
        # Draw bars and record their positions for connectors
        bar_height = 24
        bar_gap = 12
        # (Reverted) Remove left-column label layout; use a fixed bar offset and put labels on bars.
        from PyQt6.QtGui import QFontMetrics
        font = self.font() if hasattr(self, 'font') else None
        fm = QFontMetrics(font) if font else None
        max_chars_fixed_tl = 32
        left_margin = 60
        bar_offset_x = left_margin
        y = 40
        bar_positions = {}  # idx -> (x, y, width)
        self._timeline_name_to_text = {}
        for name, start, duration, row, idx in bars:
            x = bar_offset_x + (start - min_date).days * 8
            width = max(8, duration * 8)
            # Highlight critical path bars in red
            color = QColor("red") if name in critical_path else QColor("#FF8200")
            # Add hoverable rect for image preview
            from PyQt6.QtWidgets import QGraphicsRectItem
            class HoverableTimelineBar(QGraphicsRectItem):
                def __init__(self, x, y, width, height, row, timeline_view):
                    super().__init__(x, y, width, height)
                    self.row = row
                    self.timeline_view = timeline_view
                    self.setAcceptHoverEvents(True)
                    # Link indicator for Pace Link (honors links toggle)
                    try:
                        self._pace_link = (self.row.get('Pace Link') or '').strip()
                        has_link = self._pace_link.lower().startswith('http://') or self._pace_link.lower().startswith('https://')
                    except Exception:
                        self._pace_link = ''; has_link = False
                    if has_link and getattr(self.timeline_view, '_show_links', True):
                        try:
                            from PyQt6.QtWidgets import QGraphicsSimpleTextItem
                            icon = QGraphicsSimpleTextItem('🔗', self)
                            icon.setBrush(QColor('white'))
                            icon.setPos(max(2, width - 12), 2)
                            icon.setZValue(self.zValue() + 2)
                        except Exception:
                            pass
                        try:
                            self.setToolTip(self._pace_link)
                            self.setCursor(getattr(Qt,'CursorShape', Qt).PointingHandCursor)
                        except Exception:
                            pass
                def get_preview_label(self):
                    # Try to get preview_label from parent widget
                    parent = self.timeline_view.parent()
                    if hasattr(parent, 'preview_label'):
                        return parent.preview_label
                    # Fallback: try timeline_view itself
                    if hasattr(self.timeline_view, 'preview_label'):
                        return self.timeline_view.preview_label
                    return None
                def mousePressEvent(self, event):
                    try:
                        _left = getattr(Qt, 'MouseButton', Qt).LeftButton
                        if hasattr(event, 'button') and event.button() == _left and getattr(self.timeline_view, '_show_links', True):
                            url = (self.row.get('Pace Link') or '').strip()
                            if url and (url.lower().startswith('http://') or url.lower().startswith('https://')):
                                try:
                                    from PyQt6.QtGui import QDesktopServices
                                    from PyQt6.QtCore import QUrl
                                    QDesktopServices.openUrl(QUrl(url))
                                except Exception:
                                    import webbrowser; webbrowser.open(url)
                                return
                    except Exception:
                        pass
                    super().mousePressEvent(event)
                def hoverEnterEvent(self, event):
                    preview_label = self.get_preview_label()
                    if preview_label is None:
                        super().hoverEnterEvent(event)
                        return
                    img_path = self.row.get("Images", "")
                    if img_path and str(img_path).strip():
                        from PyQt6.QtGui import QPixmap
                        img_path_full = resolve_resource_path(img_path)
                        pixmap = QPixmap(img_path_full)
                        if not pixmap.isNull():
                            try:
                                _smooth = Qt.TransformationMode.SmoothTransformation
                            except Exception:
                                _smooth = getattr(Qt, 'SmoothTransformation', 1)
                            preview_label.setPixmap(pixmap.scaledToHeight(180, _smooth))
                            preview_label.setText("")
                        else:
                            preview_label.setText("[Image not found]")
                            preview_label.setPixmap(QPixmap())
                    else:
                        preview_label.clear()
                    super().hoverEnterEvent(event)
                def hoverLeaveEvent(self, event):
                    preview_label = self.get_preview_label()
                    if preview_label is not None:
                        preview_label.clear()
                    super().hoverLeaveEvent(event)
            # Milestone diamond for zero-duration or explicit type
            is_milestone = False
            try:
                is_milestone = (duration == 0) or ((row.get('Type') or '').strip().lower() == 'milestone')
            except Exception:
                is_milestone = False
            if is_milestone:
                from PyQt6.QtGui import QPainterPath as _PathTL, QPen as _PenTL, QBrush as _BrushTL
                size = bar_height
                cx = x
                cy = y + bar_height/2
                path = _PathTL()
                path.moveTo(cx, cy - size/2)
                path.lineTo(cx + size/2, cy)
                path.lineTo(cx, cy + size/2)
                path.lineTo(cx - size/2, cy)
                path.closeSubpath()
                bar_item = self.scene.addPath(path, _PenTL(color, 2), _BrushTL(QColor('#333333')))
                # Attach row for hover handlers wiring below
                bar_item.row = row
            else:
                bar_item = HoverableTimelineBar(x, y, width, bar_height, row, self)
                bar_item.setBrush(QBrush(color))
            self.scene.addItem(bar_item)
            full_name = name
            display_name = full_name
            if len(display_name) > max_chars_fixed_tl:
                display_name = display_name[:max_chars_fixed_tl-1] + "…"
            text_item = self.scene.addText(display_name)
            from PyQt6.QtGui import QFont, QPen, QBrush
            text_item.setDefaultTextColor(QColor("white"))
            orig_font = text_item.font()
            text_item.setData(1, orig_font)
            text_item.setData(2, full_name)
            text_item.setToolTip(full_name)
            # Center vertically and place label to the right of the bar
            ty = y + (bar_height - text_item.boundingRect().height())/2
            gap = 6
            text_item.setPos(x + width + gap, ty)
            # Always-visible subtle contrasting background
            from PyQt6.QtGui import QPen as _LblPen2, QBrush as _LblBrush2, QColor as _LblColor2
            br = text_item.boundingRect().translated(text_item.pos())
            bg_color = _LblColor2(0,0,0,160)
            padded = br.adjusted(-3,-1,3,1)
            from PyQt6.QtGui import QPainterPath as _LblPath2
            path = _LblPath2()
            radius = 6
            path.addRoundedRect(padded, radius, radius)
            bg_rect = self.scene.addPath(path, _LblPen2(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), _LblBrush2(bg_color))
            bg_rect.setZValue(text_item.zValue()-1)
            text_item.setData(3, bg_rect)
            self._timeline_name_to_text[name] = text_item
            bar_positions[idx] = (x, y, width)
            y += bar_height + bar_gap
        # Draw connector lines for parent-child relationships
        for name, start, duration, row, idx in bars:
            parent_name = row.get("Parent", "")
            if parent_name and parent_name in name_to_idx:
                parent_idx = name_to_idx[parent_name]
                if parent_idx in bar_positions and idx in bar_positions:
                    px, py, pwidth = bar_positions[parent_idx]
                    cx, cy, cwidth = bar_positions[idx]
                    parent_mid_x = px + pwidth // 2
                    child_mid_x = cx + cwidth // 2
                    parent_bottom = py + bar_height
                    child_top = cy
                    # Highlight critical path connectors in red
                    from PyQt6.QtGui import QPen
                    pen = QPen(QColor("red"), 2) if name in critical_path and parent_name in critical_path else QPen(QColor("#FF8200"), 2)
                    # Vertical line from parent to horizontal level
                    self.scene.addLine(parent_mid_x, parent_bottom, parent_mid_x, (parent_bottom + child_top) // 2, pen)
                    # Horizontal line to child
                    self.scene.addLine(parent_mid_x, (parent_bottom + child_top) // 2, child_mid_x, (parent_bottom + child_top) // 2, pen)
                    # Vertical line down to child
                    self.scene.addLine(child_mid_x, (parent_bottom + child_top) // 2, child_mid_x, child_top, pen)
        # Weekend shading (Sat/Sun) for readability
        try:
            from PyQt6.QtGui import QBrush, QColor
            from datetime import timedelta
            shade = QBrush(QColor(220, 220, 220, 120))
            cur = min_date
            while cur <= max_date:
                if cur.weekday() >= 5:
                    run_start = cur
                    while cur <= max_date and cur.weekday() >= 5:
                        cur += timedelta(days=1)
                    run_end = cur
                    x0 = bar_offset_x + (run_start - min_date).days * 8
                    x1 = bar_offset_x + (run_end - min_date).days * 8
                    self.scene.addRect(x0, 0, max(1, x1 - x0), y + 30, pen=(Qt.PenStyle.NoPen if hasattr(Qt,'PenStyle') else getattr(Qt,'NoPen',0)), brush=shade)
                else:
                    cur += timedelta(days=1)
        except Exception:
            pass
        # Draw x-axis with date marks every 7 days
        axis_y = 20
        axis_x0 = bar_offset_x
        axis_x1 = bar_offset_x + total_days * 8 + 40
        self.scene.addLine(axis_x0, axis_y, axis_x1, axis_y)
        for d in range(0, total_days + 1, 7):
            tick_x = axis_x0 + d * 8
            self.scene.addLine(tick_x, axis_y - 5, tick_x, axis_y + 5)
            tick_date = min_date + datetime.timedelta(days=d)
            tick_label = self.scene.addText(tick_date.strftime("%m-%d-%Y"))
            tick_label.setDefaultTextColor(QColor("white"))
            tick_label.setPos(tick_x - 30, axis_y - 25)
        # Today line
        try:
            today = datetime.datetime.today().date()
            if min_date.date() <= today <= max_date.date():
                dx = (today - min_date.date()).days
                x_today = axis_x0 + dx * 8
                from PyQt6.QtGui import QPen as _QPToday2, QColor as _QCToday2
                pen_today = _QPToday2(_QCToday2(0, 200, 255)); pen_today.setWidth(2)
                try:
                    pen_today.setStyle(Qt.PenStyle.DashLine)
                except Exception:
                    pen_today.setStyle(getattr(Qt,'DashLine',1))
                self.scene.addLine(x_today, 0, x_today, y + 40, pen_today)
                lbl = self.scene.addText("Today"); lbl.setDefaultTextColor(_QCToday2(0,200,255))
                lbl.setPos(x_today + 4, 2)
        except Exception:
            pass
        self.view.setSceneRect(0, 0, axis_x1 + 40, max(300, y + 40))
        # Extend hover bars to highlight labels (monkey patch hover events)
        for item in self.scene.items():
            if hasattr(item, 'row') and hasattr(item, 'hoverEnterEvent'):
                original_enter = item.hoverEnterEvent
                original_leave = item.hoverLeaveEvent
                def make_enter(orig, bar_item=item):
                    def _enter(ev):
                        try:
                            name = bar_item.row.get("Project Part", "")
                            ti = self._timeline_name_to_text.get(name)
                            if ti:
                                f = QFont(ti.font())
                                f.setBold(True)
                                ti.setFont(f)
                                bg = ti.data(3)
                                if bg:
                                    from PyQt6.QtGui import QColor as _QColorTL, QBrush as _QBrushTL
                                    bg.setBrush(_QBrushTL(_QColorTL(255,255,255,50)))
                        except Exception:
                            pass
                        return orig(ev)
                    return _enter
                def make_leave(orig_leave, bar_item=item):
                    def _leave(ev):
                        try:
                            name = bar_item.row.get("Project Part", "")
                            ti = self._timeline_name_to_text.get(name)
                            if ti:
                                base_font = ti.data(1)
                                if isinstance(base_font, QFont):
                                    ti.setFont(base_font)
                                bg = ti.data(3)
                                if bg:
                                    from PyQt6.QtGui import QBrush as _QBrushTL2, QColor as _QColorTL2
                                    # PyQt6-safe transparent brush (no Qt.transparent constant)
                                    bg.setBrush(_QBrushTL2(_QColorTL2(0, 0, 0, 0)))
                        except Exception:
                            pass
                        return orig_leave(ev)
                    return _leave
                item.hoverEnterEvent = make_enter(original_enter)
                item.hoverLeaveEvent = make_leave(original_leave)
        # Click-to-lock for timeline (reusing view mouse events)
        def lock_click_event(event):
            try:
                _left_btn2 = Qt.MouseButton.LeftButton
            except Exception:
                _left_btn2 = getattr(Qt, 'LeftButton', 1)
            if event.button() == _left_btn2:
                scene_pos = self.view.mapToScene(event.pos())
                for it in self.scene.items(scene_pos):
                    if hasattr(it, 'row'):
                        name = it.row.get("Project Part", "")
                        # Clear existing lock
                        if hasattr(self, '_timeline_locked') and self._timeline_locked and self._timeline_locked != name:
                            # Reset previous locked label
                            prev_ti = self._timeline_name_to_text.get(self._timeline_locked)
                            if prev_ti:
                                orig = prev_ti.data(1)
                                if isinstance(orig, QFont):
                                    prev_ti.setFont(orig)
                                bg = prev_ti.data(3)
                                if bg:
                                    from PyQt6.QtGui import QBrush, QColor as _QColorTL3
                                    # PyQt6-safe transparent brush (no Qt.transparent constant)
                                    bg.setBrush(QBrush(_QColorTL3(0, 0, 0, 0)))
                        self._timeline_locked = name
                        ti = self._timeline_name_to_text.get(name)
                        if ti:
                            f = QFont(ti.font()); f.setBold(True); ti.setFont(f)
                            bg = ti.data(3)
                            if bg:
                                from PyQt6.QtGui import QColor, QBrush
                                bg.setBrush(QBrush(QColor(255,255,255,70)))
                        break
            return original_mouse_press(event)
        if not hasattr(self.view, '_lock_click_installed'):
            original_mouse_press = self.view.mousePressEvent
            self.view.mousePressEvent = lock_click_event
            self.view._lock_click_installed = True


# New DatabaseView class
from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem

from PyQt6.QtWidgets import QDateEdit
from PyQt6.QtCore import QDate

class DatabaseView(QWidget):
    DATE_FIELDS = {"Start Date", "Calculated End Date"}
    DROPDOWN_FIELDS = {
        "Internal/External": ["Internal", "External"],
    "Type": ["Milestone", "Phase", "Feature", "Item"],
    # Progress status field handled similarly
    "Status": ["Planned", "In Progress", "Blocked", "Done", "Deferred"]
    }
    PROGRESS_STATUSES = ["Planned", "In Progress", "Blocked", "Done", "Deferred"]

    def __init__(self, model, on_data_changed=None):
        super().__init__()
        self.model = model
        self.on_data_changed = on_data_changed
        # Honor app-level read-only flag if present
        self._read_only = bool(getattr(self.model, 'read_only', False))
        layout = QVBoxLayout()
        # Subtle read-only note at top
        top_row = QHBoxLayout()
        title = QLabel("Database View")
        self.ro_banner = QLabel("read-only")
        self.ro_banner.setStyleSheet("color:#777; font-style:italic; padding:2px 4px;")
        self.ro_banner.setVisible(self._read_only)
        top_row.addWidget(title)
        top_row.addStretch(1)
        top_row.addWidget(self.ro_banner)
        layout.addLayout(top_row)
        self.table = QTableWidget()
        self.table.setColumnCount(len(ProjectDataModel.COLUMNS))
        self.table.setHorizontalHeaderLabels(ProjectDataModel.COLUMNS)
        # Set tooltip for Duration column
        duration_col = ProjectDataModel.COLUMNS.index("Duration (days)")
        self.table.horizontalHeaderItem(duration_col).setToolTip("Duration is in days")
        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        self.add_btn = QPushButton("Add Row")
        self.add_btn.clicked.connect(self.add_row)
        btn_layout.addWidget(self.add_btn)
        self.del_btn = QPushButton("Delete Row")
        self.del_btn.clicked.connect(self.delete_row)
        btn_layout.addWidget(self.del_btn)
        self.export_btn = QPushButton("Export Database")
        self.export_btn.clicked.connect(self.export_database)
        btn_layout.addWidget(self.export_btn)
        self.import_btn = QPushButton("Import Data")
        self.import_btn.clicked.connect(self.import_data)
        btn_layout.addWidget(self.import_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)
        self.refresh_table()
        self.table.cellChanged.connect(self.cell_edited)

    def set_read_only(self, read_only: bool):
        """Enable/disable editing and mutating operations in the Database view."""
        self._read_only = bool(read_only)
        try:
            if hasattr(self, 'ro_banner') and self.ro_banner is not None:
                self.ro_banner.setVisible(self._read_only)
        except Exception:
            pass
        try:
            from PyQt6.QtWidgets import QAbstractItemView
            if self._read_only:
                self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            else:
                self.table.setEditTriggers(QAbstractItemView.AllEditTriggers)
        except Exception:
            pass
        # Buttons: disable mutating actions in read-only; export stays enabled
        try:
            self.add_btn.setEnabled(not self._read_only)
            self.del_btn.setEnabled(not self._read_only)
            self.import_btn.setEnabled(not self._read_only)
        except Exception:
            pass
        # Rebuild widgets to reflect enabled/disabled state
        try:
            self.refresh_table()
        except Exception:
            pass

    def import_data(self):
        if getattr(self, '_read_only', False):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Read-Only", "Import is disabled in read-only mode.")
            return
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import csv
        path, _ = QFileDialog.getOpenFileName(self, "Import Data", "", "CSV Files (*.csv)")
        if not path:
            return
        try:
            with open(path, "r", encoding='utf-8') as f:
                reader = csv.DictReader(f)
                imported_rows = []
                for row in reader:
                    imported_row = {col: row.get(col, "") for col in ProjectDataModel.COLUMNS}
                    imported_rows.append(imported_row)
            # Replace current data with imported data
            self.model.rows = imported_rows
            self.model.save_to_db()
            self.refresh_table()
            QMessageBox.information(self, "Import Successful", f"Imported {len(imported_rows)} rows from {path}")
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", f"Error importing data: {e}")
    def export_database(self):
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import csv
        path, _ = QFileDialog.getSaveFileName(self, "Export Database", "database_export.csv", "CSV Files (*.csv)")
        if not path:
            return
        try:
            with open(path, "w", newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(ProjectDataModel.COLUMNS)
                for row in self.model.rows:
                    writer.writerow([row.get(col, "") for col in ProjectDataModel.COLUMNS])
            QMessageBox.information(self, "Export Successful", f"Database exported to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Error exporting database: {e}")

    def refresh_table(self):
        self.table.blockSignals(True)
        self.table.setRowCount(len(self.model.rows))
        for row, rowdata in enumerate(self.model.rows):
            # Determine if this row is a parent (has at least one child referencing its Project Part)
            part_name = rowdata.get("Project Part", "")
            has_children = any(r.get("Parent", "") == part_name for r in self.model.rows if r is not rowdata)
            for col, colname in enumerate(ProjectDataModel.COLUMNS):
                # Only use QDateEdit for editable date fields, not Calculated End Date
                if colname in self.DATE_FIELDS and colname != "Calculated End Date":
                    date_val = rowdata.get(colname, "")
                    date_edit = QDateEdit()
                    date_edit.setCalendarPopup(True)
                    min_blank = QDate(1753, 1, 1)
                    date_edit.setMinimumDate(min_blank)
                    date_edit.setSpecialValueText("")
                    # Prevent wheel events unless focused (clicked)
                    def block_wheel(event):
                        if not date_edit.hasFocus():
                            event.ignore()
                        else:
                            QDateEdit.wheelEvent(date_edit, event)
                    date_edit.wheelEvent = block_wheel
                    if date_val:
                        try:
                            date = QDate.fromString(date_val, "MM-dd-yyyy")
                            if date.isValid() and date != QDate(1752, 9, 14) and date != min_blank:
                                date_edit.setDate(date)
                            else:
                                if colname == "Start Date":
                                    date_edit.setDate(QDate.currentDate())
                                else:
                                    date_edit.clear()
                        except Exception:
                            if colname == "Start Date":
                                date_edit.setDate(QDate.currentDate())
                            else:
                                date_edit.clear()
                    else:
                        if colname == "Start Date":
                            date_edit.setDate(QDate.currentDate())
                        else:
                            date_edit.clear()
                    date_edit.dateChanged.connect(lambda d, r=row, c=col: self.date_changed(r, c, d))
                    if getattr(self, '_read_only', False):
                        date_edit.setEnabled(False)
                    self.table.setCellWidget(row, col, date_edit)
                    # Show blank in the table cell if value is empty or minimum blank
                    if not date_val or date_val == min_blank.toString("MM-dd-yyyy"):
                        self.table.setItem(row, col, QTableWidgetItem(""))
                    else:
                        self.table.setItem(row, col, QTableWidgetItem(date_val))
                elif colname == "Calculated End Date":
                    # Show as read-only text
                    val = rowdata.get(colname, "")
                    self.table.setItem(row, col, QTableWidgetItem(val))
                elif colname == "% Complete":
                    from PyQt6.QtWidgets import QSpinBox
                    spin = QSpinBox()
                    spin.setRange(0, 100)
                    try:
                        spin.setValue(int(rowdata.get(colname) or 0))
                    except Exception:
                        spin.setValue(0)
                    # Prevent wheel without focus
                    def block_wheel_spin(event, sb=spin):
                        if not sb.hasFocus():
                            event.ignore()
                        else:
                            QSpinBox.wheelEvent(sb, event)
                    spin.wheelEvent = block_wheel_spin
                    if has_children:
                        spin.setEnabled(False)
                        spin.setToolTip("Parent progress is rolled up automatically from children.")
                    else:
                        spin.valueChanged.connect(lambda val, r=row, c=col: self.percent_changed(r, c, val))
                    if getattr(self, '_read_only', False):
                        spin.setEnabled(False)
                    self.table.setCellWidget(row, col, spin)
                    self.table.setItem(row, col, QTableWidgetItem(str(spin.value())))
                elif colname in self.DROPDOWN_FIELDS or colname == "Parent":
                    from PyQt6.QtWidgets import QComboBox
                    combo = QComboBox()
                    # Prevent wheel events unless focused (clicked)
                    def block_wheel_combo(event):
                        if not combo.hasFocus():
                            event.ignore()
                        else:
                            QComboBox.wheelEvent(combo, event)
                    combo.wheelEvent = block_wheel_combo
                    if colname == "Parent":
                        # List all other project part names except this row
                        part_names = [self.model.rows[i]["Project Part"] for i in range(len(self.model.rows)) if i != row]
                        combo.addItem("")  # Allow no parent
                        combo.addItems(part_names)
                        current_val = rowdata.get("Parent", "")
                        if current_val in part_names:
                            combo.setCurrentText(current_val)
                        combo.currentTextChanged.connect(lambda val, r=row, c=col: self.dropdown_changed(r, c, val))
                        if getattr(self, '_read_only', False):
                            combo.setEnabled(False)
                        self.table.setCellWidget(row, col, combo)
                        self.table.setItem(row, col, QTableWidgetItem(combo.currentText()))
                    else:
                        combo.addItems(self.DROPDOWN_FIELDS[colname])
                        current_val = rowdata.get(colname, "")
                        if current_val in self.DROPDOWN_FIELDS[colname]:
                            combo.setCurrentText(current_val)
                        if colname == "Status":
                            if has_children:
                                combo.setEnabled(False)
                                combo.setToolTip("Parent status is derived from child statuses.")
                            else:
                                combo.currentTextChanged.connect(lambda val, r=row, c=col: self.status_changed(r, c, val))
                        else:
                            combo.currentTextChanged.connect(lambda val, r=row, c=col: self.dropdown_changed(r, c, val))
                            if getattr(self, '_read_only', False):
                                combo.setEnabled(False)
                        self.table.setCellWidget(row, col, combo)
                        self.table.setItem(row, col, QTableWidgetItem(combo.currentText()))
                elif colname == "Images":
                    img_widget = ImageCellWidget(self, row, col, self.model, self.on_data_changed)
                    if getattr(self, '_read_only', False) and hasattr(img_widget, 'btn'):
                        try:
                            img_widget.btn.setEnabled(False)
                        except Exception:
                            pass
                    self.table.setCellWidget(row, col, img_widget)
                    img_widget.refresh()  # Ensure preview is updated after loading
                    img_val = rowdata.get(colname, "")
                    if img_val:
                        self.table.setItem(row, col, QTableWidgetItem(img_val.split("/")[-1] or img_val.split("\\")[-1]))
                    else:
                        self.table.setItem(row, col, QTableWidgetItem(""))
                elif colname == "Children":
                    # Read-only: list all project parts whose parent is this part
                    this_part = rowdata.get("Project Part", "")
                    children = [r["Project Part"] for r in self.model.rows if r.get("Parent", "") == this_part]
                    self.table.setItem(row, col, QTableWidgetItem(", ".join(children)))
                elif colname == "Pace Link":
                    link = rowdata.get(colname, "")
                    if link and (link.startswith("http://") or link.startswith("https://")):
                        label = QLabel(f'<a href="{link}">{link}</a>')
                        label.setOpenExternalLinks(True)
                        self.table.setCellWidget(row, col, label)
                        self.table.setItem(row, col, QTableWidgetItem(link))
                    else:
                        line_edit = QLineEdit(link)
                        def on_edit_finished(row=row, col=col, edit=line_edit):
                            val = edit.text()
                            self.model.rows[row][colname] = val
                            self.model.save_to_db()
                            self.refresh_table()
                            if self.on_data_changed:
                                self.on_data_changed()
                        line_edit.editingFinished.connect(on_edit_finished)
                        if getattr(self, '_read_only', False):
                            line_edit.setEnabled(False)
                        self.table.setCellWidget(row, col, line_edit)
                        self.table.setItem(row, col, QTableWidgetItem(link))
                else:
                    self.table.setItem(row, col, QTableWidgetItem(rowdata.get(colname, "")))
        self.table.blockSignals(False)
        # Ensure all image widgets are refreshed after table is populated
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                colname = ProjectDataModel.COLUMNS[col]
                if colname == "Images":
                    widget = self.table.cellWidget(row, col)
                    if widget and hasattr(widget, "refresh"):
                        widget.refresh()

        # Automatically resize Project Part column to fit contents
        part_col = ProjectDataModel.COLUMNS.index("Project Part")
        self.table.resizeColumnToContents(part_col)

    def add_row(self):
        if getattr(self, '_read_only', False):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Read-Only", "Add Row is disabled in read-only mode.")
            return
        data = []
        for col in ProjectDataModel.COLUMNS:
            if col == "Duration (days)":
                data.append("1")
            elif col == "Internal/External":
                data.append("Internal")
            elif col == "% Complete":
                data.append("0")
            elif col == "Status":
                data.append("Planned")
            # Removed Deadline field
            else:
                data.append("")
        idx = self.model.add_row(data)
        self.model.save_to_db()
        self.refresh_table()
        if self.on_data_changed:
            self.on_data_changed()

    def delete_row(self):
        if getattr(self, '_read_only', False):
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Read-Only", "Delete Row is disabled in read-only mode.")
            return
        row = self.table.currentRow()
        if row >= 0:
            self.model.delete_row(row)
            self.model.save_to_db()
            self.refresh_table()
            if self.on_data_changed:
                self.on_data_changed()


    def cell_edited(self, row, col):
        colname = ProjectDataModel.COLUMNS[col]
        if colname in self.DATE_FIELDS:
            widget = self.table.cellWidget(row, col)
            if widget:
                date_val = widget.date().toString("MM-dd-yyyy")
                self.model.rows[row][colname] = date_val
        elif colname in self.DROPDOWN_FIELDS or colname == "Parent":
            widget = self.table.cellWidget(row, col)
            if widget:
                self.model.rows[row][colname] = widget.currentText()
        elif colname == "Pace Link":
            widget = self.table.cellWidget(row, col)
            if isinstance(widget, QLineEdit):
                val = widget.text()
                self.model.rows[row][colname] = val

        else:
            val = self.table.item(row, col).text()
            self.model.rows[row][colname] = val
            self.model.update_calculated_end_dates()
            self.model.save_to_db()
            self.refresh_table()
        if self.on_data_changed:
            self.on_data_changed()

    def dropdown_changed(self, row, col, value):
        colname = ProjectDataModel.COLUMNS[col]
        try:
            part_name = self.model.rows[row].get('Project Part','')
            # Prepare optimistic update
            expected_version = self.model.rows[row].get('row_version', 0)
            new_values = {colname: value}
            ok, info = self.model.update_part_values(part_name, new_values, expected_version)
            if not ok and info == 'Conflict':
                try:
                    log_event('conflict','detected', part=part_name, field=colname)
                except Exception:
                    pass
                remote = self.model.get_row_snapshot(part_name) or {}
                original = dict(self.model.rows[row])
                dlg = ConflictResolutionDialog(part_name, original=original, pending=new_values, remote=remote, parent=self)
                if dlg.exec():
                    if dlg.choice == 'keep':
                        try: log_event('conflict','keep_remote', part=part_name)
                        except Exception: pass
                        # Reload remote into memory row
                        if remote:
                            self.model.rows[row].update(remote)
                    elif dlg.choice == 'overwrite':
                        # Overwrite: use latest remote version as expected
                        latest_ver = remote.get('row_version', expected_version)
                        force_values = dict(new_values)
                        # Attempt forced update with new version
                        ok2, info2 = self.model.update_part_values(part_name, force_values, latest_ver)
                        try: log_event('conflict','overwrite_attempt', part=part_name, success=ok2)
                        except Exception: pass
                        if not ok2:
                            # Could still race; just refresh row
                            fresh = self.model.get_row_snapshot(part_name)
                            if fresh:
                                self.model.rows[row].update(fresh)
                    elif dlg.choice == 'merge':
                        merged = dlg.merged
                        latest_ver = remote.get('row_version', expected_version)
                        # Remove concurrency keys
                        merged_clean = {k:v for k,v in merged.items() if k in self.model.COLUMNS and k != 'Project Part'}
                        ok2, info2 = self.model.update_part_values(part_name, merged_clean, latest_ver)
                        try: log_event('conflict','merge_attempt', part=part_name, success=ok2, fields=list(merged_clean.keys()))
                        except Exception: pass
                        if not ok2:
                            fresh = self.model.get_row_snapshot(part_name)
                            if fresh:
                                self.model.rows[row].update(fresh)
                else:
                    # Cancel: discard local change and refresh
                    fresh = self.model.get_row_snapshot(part_name)
                    if fresh:
                        self.model.rows[row].update(fresh)
            elif not ok:
                try: log_event('conflict','other_update_failure', part=part_name, reason=info)
                except Exception: pass
            else:
                # Successful optimistic update; update in-memory version already done by method
                pass
            self.refresh_table()
            if self.on_data_changed:
                self.on_data_changed()
        except Exception as e:
            print(f"ERROR in dropdown_changed: {e}")
    def date_changed(self, row, col, qdate):
        colname = ProjectDataModel.COLUMNS[col]
        min_blank = QDate(1753, 1, 1)
        try:
            part_name = self.model.rows[row].get('Project Part','')
            date_val = "" if qdate == min_blank else qdate.toString("MM-dd-yyyy")
            expected_version = self.model.rows[row].get('row_version', 0)
            ok, info = self.model.update_part_values(part_name, {colname: date_val}, expected_version)
            if not ok and info == 'Conflict':
                try: log_event('conflict','detected', part=part_name, field=colname)
                except Exception: pass
                remote = self.model.get_row_snapshot(part_name) or {}
                original = dict(self.model.rows[row])
                dlg = ConflictResolutionDialog(part_name, original=original, pending={colname: date_val}, remote=remote, parent=self)
                if dlg.exec():
                    if dlg.choice == 'keep':
                        if remote:
                            self.model.rows[row].update(remote)
                            try: log_event('conflict','keep_remote', part=part_name)
                            except Exception: pass
                    elif dlg.choice == 'overwrite':
                        latest_ver = remote.get('row_version', expected_version)
                        ok2, info2 = self.model.update_part_values(part_name, {colname: date_val}, latest_ver)
                        try: log_event('conflict','overwrite_attempt', part=part_name, success=ok2)
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                    elif dlg.choice == 'merge':
                        # For date single-field merge same as overwrite local selection outcome
                        latest_ver = remote.get('row_version', expected_version)
                        ok2, info2 = self.model.update_part_values(part_name, {colname: date_val}, latest_ver)
                        try: log_event('conflict','merge_attempt', part=part_name, success=ok2, fields=[colname])
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                else:
                    # Cancel -> leave remote
                    if remote:
                        self.model.rows[row].update(remote)
            elif not ok:
                try: log_event('conflict','other_update_failure', part=part_name, reason=info)
                except Exception: pass
            self.refresh_table()
            if self.on_data_changed:
                self.on_data_changed()
        except Exception as e:
            print(f"ERROR in date_changed: {e}")

    # --- Progress field handlers ---
    def percent_changed(self, row, col, value):
        try:
            part_name = self.model.rows[row].get('Project Part','')
            updates = {"% Complete": int(value)}
            if int(value) >= 100 and self.model.rows[row].get("Status") != "Done":
                updates["Status"] = "Done"
                import datetime
                if not self.model.rows[row].get("Actual Finish Date"):
                    updates["Actual Finish Date"] = datetime.datetime.today().strftime("%m-%d-%Y")
                if not self.model.rows[row].get("Actual Start Date"):
                    updates["Actual Start Date"] = datetime.datetime.today().strftime("%m-%d-%Y")
            expected_version = self.model.rows[row].get('row_version', 0)
            ok, info = self.model.update_part_values(part_name, updates, expected_version)
            if not ok and info == 'Conflict':
                try: log_event('conflict','detected', part=part_name, field='% Complete')
                except Exception: pass
                remote = self.model.get_row_snapshot(part_name) or {}
                original = dict(self.model.rows[row])
                dlg = ConflictResolutionDialog(part_name, original=original, pending=updates, remote=remote, parent=self)
                if dlg.exec():
                    choice_fields = updates.keys()
                    if dlg.choice == 'keep':
                        if remote:
                            self.model.rows[row].update(remote)
                            try: log_event('conflict','keep_remote', part=part_name)
                            except Exception: pass
                    elif dlg.choice == 'overwrite':
                        latest_ver = remote.get('row_version', expected_version)
                        ok2, info2 = self.model.update_part_values(part_name, updates, latest_ver)
                        try: log_event('conflict','overwrite_attempt', part=part_name, success=ok2)
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                    elif dlg.choice == 'merge':
                        merged = dlg.merged
                        latest_ver = remote.get('row_version', expected_version)
                        merged_clean = {k:v for k,v in merged.items() if k in self.model.COLUMNS and k != 'Project Part'}
                        ok2, info2 = self.model.update_part_values(part_name, merged_clean, latest_ver)
                        try: log_event('conflict','merge_attempt', part=part_name, success=ok2, fields=list(merged_clean.keys()))
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                else:
                    if remote:
                        self.model.rows[row].update(remote)
            elif not ok:
                try: log_event('conflict','other_update_failure', part=part_name, reason=info)
                except Exception: pass
            self.refresh_table()
            if self.on_data_changed:
                self.on_data_changed()
        except Exception as e:
            print(f"ERROR in percent_changed: {e}")

    def status_changed(self, row, col, value):
        try:
            part_name = self.model.rows[row].get('Project Part','')
            import datetime
            today_str = datetime.datetime.today().strftime("%m-%d-%Y")
            updates = {"Status": value}
            if value == "In Progress":
                if not self.model.rows[row].get("Actual Start Date"):
                    updates["Actual Start Date"] = today_str
            elif value == "Done":
                updates["% Complete"] = 100
                if not self.model.rows[row].get("Actual Start Date"):
                    updates["Actual Start Date"] = today_str
                if not self.model.rows[row].get("Actual Finish Date"):
                    updates["Actual Finish Date"] = today_str
            expected_version = self.model.rows[row].get('row_version', 0)
            ok, info = self.model.update_part_values(part_name, updates, expected_version)
            if not ok and info == 'Conflict':
                try: log_event('conflict','detected', part=part_name, field='Status')
                except Exception: pass
                remote = self.model.get_row_snapshot(part_name) or {}
                original = dict(self.model.rows[row])
                dlg = ConflictResolutionDialog(part_name, original=original, pending=updates, remote=remote, parent=self)
                if dlg.exec():
                    if dlg.choice == 'keep':
                        if remote:
                            self.model.rows[row].update(remote)
                            try: log_event('conflict','keep_remote', part=part_name)
                            except Exception: pass
                    elif dlg.choice == 'overwrite':
                        latest_ver = remote.get('row_version', expected_version)
                        ok2, info2 = self.model.update_part_values(part_name, updates, latest_ver)
                        try: log_event('conflict','overwrite_attempt', part=part_name, success=ok2)
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                    elif dlg.choice == 'merge':
                        merged = dlg.merged
                        latest_ver = remote.get('row_version', expected_version)
                        merged_clean = {k:v for k,v in merged.items() if k in self.model.COLUMNS and k != 'Project Part'}
                        ok2, info2 = self.model.update_part_values(part_name, merged_clean, latest_ver)
                        try: log_event('conflict','merge_attempt', part=part_name, success=ok2, fields=list(merged_clean.keys()))
                        except Exception: pass
                        if not ok2 and remote:
                            self.model.rows[row].update(remote)
                else:
                    if remote:
                        self.model.rows[row].update(remote)
            elif not ok:
                try: log_event('conflict','other_update_failure', part=part_name, reason=info)
                except Exception: pass
            self.refresh_table()
            if self.on_data_changed:
                self.on_data_changed()
        except Exception as e:
            print(f"ERROR in status_changed: {e}")


class MainWindow(QMainWindow):
    # --- Desktop Shortcut Helpers (OneDrive deployment convenience) ---
    def _ensure_app_icon(self):
        """Render header.svg into a multi-size .ico (16,32,48,64,128,256) if available.
        Stores header.ico adjacent to executable/script. Returns path or None."""
        try:
            import os, time
            svg_path = resolve_resource_path('header.svg')
            png_fallback = resolve_resource_path('header.png')
            use_svg = svg_path and os.path.exists(svg_path)
            if (not use_svg) and (not (png_fallback and os.path.exists(png_fallback))):
                return None
            # Target icon path beside script/exe
            base_dir = os.path.dirname(resolve_resource_path('.'))
            ico_path = os.path.join(base_dir, 'header.ico')
            # Rebuild if missing or svg newer
            rebuild = True
            try:
                if os.path.exists(ico_path):
                    src_mtime = os.path.getmtime(svg_path if use_svg else png_fallback)
                    rebuild = src_mtime > os.path.getmtime(ico_path)
            except Exception:
                rebuild = True
            if rebuild:
                try:
                    from PyQt6.QtGui import QImage, QPainter, QPixmap
                    if use_svg:
                        from PyQt6.QtSvg import QSvgRenderer  # type: ignore
                        renderer = QSvgRenderer(svg_path)
                        if not renderer.isValid():
                            use_svg = False
                    sizes = [16,32,48,64,128,256]
                    images = []
                    for sz in sizes:
                        img = QImage(sz, sz, QImage.Format_ARGB32)
                        img.fill(0)
                        p = QPainter(img)
                        if use_svg:
                            renderer.render(p)
                        else:
                            # raster fallback scale
                            pm = QPixmap(png_fallback)
                            if not pm.isNull():
                                try:
                                    _smooth = Qt.TransformationMode.SmoothTransformation
                                except Exception:
                                    _smooth = getattr(Qt, 'SmoothTransformation', 1)
                                pm_scaled = pm.scaled(sz, sz, _keep_ar(), _smooth)
                                p.drawPixmap(0,0, pm_scaled)
                        p.end()
                        images.append(img)
                    # Save first size then append others (PyQt lacks direct multi-icon save; fallback to largest only)
                    # Attempt to use PIL if available for true multi-size ICO
                    saved = False
                    try:
                        from PIL import Image  # type: ignore
                        pil_images = []
                        for img in images:
                            ptr = img.bits(); ptr.setsize(img.width()*img.height()*4)
                            data = bytes(ptr)
                            pil = Image.frombuffer('RGBA', (img.width(), img.height()), data, 'raw', 'BGRA', 0, 1)
                            pil_images.append(pil)
                        pil_images[0].save(ico_path, sizes=[(im.width, im.height) for im in pil_images])
                        saved = True
                    except Exception:
                        # Fallback: save 256 png then convert single-size ICO via qt (largest only)
                        try:
                            images[-1].save(ico_path)
                            saved = True
                        except Exception:
                            pass
                    if not saved:
                        return None
                except Exception:
                    return None
            return ico_path if os.path.exists(ico_path) else None
        except Exception:
            return None
    def _desktop_path(self):
        import os
        # Prefer USERPROFILE/Desktop, fallback to expanduser
        try:
            home = os.path.expanduser('~')
            candidates = []
            if 'USERPROFILE' in os.environ:
                candidates.append(os.path.join(os.environ['USERPROFILE'], 'Desktop'))
            candidates.append(os.path.join(home, 'Desktop'))
            for c in candidates:
                if c and os.path.isdir(c):
                    return c
            return home
        except Exception:
            return os.path.expanduser('~')
    def _shortcut_exists(self, base_name: str) -> bool:
        import os
        desk = self._desktop_path()
        stem = base_name
        return (os.path.exists(os.path.join(desk, stem + '.lnk')) or
                os.path.exists(os.path.join(desk, stem + '.url')))
    def _create_desktop_shortcut(self, base_name: str = 'Vols Signage'):
        """Create a desktop shortcut to the current executable or launcher script.
        Attempts .lnk via COM (win32com) first; falls back to .url if COM not available.
        Safe no-op on non-Windows.
        Returns (True, path) or (False, reason)."""
        import sys, os, platform, traceback
        desk = self._desktop_path()
        target = None
        # Determine launch target: packaged exe OR python main script
        try:
            if getattr(sys, 'frozen', False):
                target = sys.executable
            else:
                target = os.path.abspath(sys.argv[0])
        except Exception:
            target = sys.executable
        # Prefer generated header.ico if present
        icon_path = None
        try:
            icon_path = self._ensure_app_icon() or None
        except Exception:
            icon_path = None
        if not icon_path:
            icon_path = target  # fallback to exe/script
        shortcut_base = os.path.join(desk, base_name)
        if platform.system().lower().startswith('win'):
            # Try .lnk first
            try:
                import win32com.client  # type: ignore
                shell = win32com.client.Dispatch('WScript.Shell')
                lnk_path = shortcut_base + '.lnk'
                sc = shell.CreateShortcut(lnk_path)
                sc.TargetPath = target
                sc.WorkingDirectory = os.path.dirname(target)
                sc.IconLocation = icon_path
                sc.Description = 'Launch Vols Signage'
                sc.save()
                return True, lnk_path
            except Exception:
                # Fall back to .url (works for scripts/executables)
                try:
                    url_path = shortcut_base + '.url'
                    with open(url_path, 'w', encoding='utf-8') as f:
                        norm_target = target.replace('\\', '/')
                        f.write('[InternetShortcut]\n')
                        f.write(f'URL=file:///{norm_target}\n')
                        f.write('IconIndex=0\n')
                        if icon_path and os.path.exists(icon_path):
                            f.write(f'IconFile={icon_path}\n')
                    return True, url_path
                except Exception as e2:
                    return False, f'Fallback .url failed: {e2}'
        else:
            # Non-Windows: create a .desktop file if on Linux (optional)
            try:
                if platform.system().lower() == 'linux':
                    desktop_file = shortcut_base + '.desktop'
                    with open(desktop_file, 'w', encoding='utf-8') as f:
                        f.write('[Desktop Entry]\n')
                        f.write('Type=Application\n')
                        f.write(f'Name={base_name}\n')
                        f.write(f'Exec="{target}"\n')
                        f.write('Terminal=false\n')
                    os.chmod(desktop_file, 0o755)
                    return True, desktop_file
            except Exception as e:
                return False, f'Desktop file failed: {e}'
        return False, 'Unsupported platform or creation failed'
    def _maybe_offer_shortcut(self):
        """Prompt user once (per machine) to add a desktop shortcut if absent.
        Stored flag: QSettings('LSI','ProjectApp')['Shortcut/prompted'] = True."""
        try:
            from PyQt6.QtCore import QSettings, QTimer
            from PyQt6.QtWidgets import QMessageBox
            s = QSettings('LSI','ProjectApp')
            already_prompted = s.value('Shortcut/prompted', False)
            if isinstance(already_prompted, str):
                already_prompted = already_prompted.lower() in ('1','true','yes','on')
            # Only prompt if not prompted before AND no shortcut exists
            base_name = 'Vols Signage'
            if already_prompted or self._shortcut_exists(base_name):
                return
            def _ask():
                try:
                    ret = QMessageBox.question(self, 'Add Desktop Shortcut?',
                        'Would you like to add a Vols Signage shortcut to your Desktop?',
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                    if ret == QMessageBox.Yes:
                        ok, info = self._create_desktop_shortcut(base_name)
                        if self.statusBar():
                            if ok:
                                self.statusBar().showMessage(f'Shortcut created: {info}', 4000)
                            else:
                                self.statusBar().showMessage(f'Shortcut failed: {info}', 6000)
                    s.setValue('Shortcut/prompted', True)
                except Exception:
                    pass
            # Defer until UI settles
            QTimer.singleShot(1500, _ask)
        except Exception:
            pass
    def _open_holidays_manager(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QListWidget, QLineEdit, QMessageBox
        dlg = QDialog(self)
        dlg.setWindowTitle("Manage Holidays")
        v = QVBoxLayout(dlg)
        lst = QListWidget()
        # Load existing holidays
        dates = []
        try:
            for d in sorted(load_holiday_dates()):
                dates.append(d.strftime("%m-%d-%Y"))
        except Exception:
            pass
        for ds in dates:
            lst.addItem(ds)
        # Input row
        row = QHBoxLayout()
        inp = QLineEdit(); inp.setPlaceholderText("MM-dd-YYYY")
        add_btn = QPushButton("Add")
        rem_btn = QPushButton("Remove Selected")
        def add_date():
            val = inp.text().strip()
            if not val:
                return
            import datetime as _dt
            try:
                _dt.datetime.strptime(val, "%m-%d-%Y")
            except Exception:
                QMessageBox.warning(dlg, "Invalid date", "Use format MM-dd-YYYY")
                return
            # Avoid duplicates
            for i in range(lst.count()):
                if lst.item(i).text() == val:
                    return
            lst.addItem(val)
            inp.clear()
        def remove_sel():
            for it in lst.selectedItems():
                row = lst.row(it)
                lst.takeItem(row)
        add_btn.clicked.connect(add_date)
        rem_btn.clicked.connect(remove_sel)
        row.addWidget(inp); row.addWidget(add_btn); row.addWidget(rem_btn)
        v.addWidget(lst)
        v.addLayout(row)
        # Save/Close
        btns = QHBoxLayout(); ok = QPushButton("Save"); cancel = QPushButton("Close")
        def do_save():
            vals = [lst.item(i).text() for i in range(lst.count())]
            save_holiday_dates(vals)
            # Refresh views to apply shading
            try:
                if hasattr(self, 'gantt_chart_view'):
                    self.gantt_chart_view.render_gantt(self.model)
                if hasattr(self, 'timeline_view'):
                    self.timeline_view.render_timeline()
                if self.statusBar():
                    self.statusBar().showMessage("Holidays saved", 2500)
            except Exception:
                pass
        ok.clicked.connect(do_save)
        cancel.clicked.connect(dlg.close)
        btns.addWidget(ok); btns.addWidget(cancel)
        v.addLayout(btns)
        dlg.setLayout(v)
        dlg.exec()
    def on_data_changed(self):
        # Refresh all views when data changes
        if hasattr(self, 'project_tree_view'):
            self.project_tree_view.refresh()
        if hasattr(self, 'gantt_chart_view'):
            self.gantt_chart_view.render_gantt(self.model)
        if hasattr(self, 'timeline_view'):
            self.timeline_view.render_timeline()
        if hasattr(self, 'database_view'):
            self.database_view.refresh_table()
        if hasattr(self, 'progress_dashboard'):
            # Refresh metrics summary
            self.progress_dashboard.refresh()
        # Update DB status banner
        try:
            self._update_db_status()
        except Exception:
            pass
    def display_view(self, index):
        self.views.setCurrentIndex(index)
        if index == 0:
            self.project_tree_view.refresh()
        elif index == 1:
            self.gantt_chart_view.render_gantt(self.model)
        elif index == 4:
            self.database_view.refresh_table()
        elif index == 5 and hasattr(self, 'progress_dashboard'):
            self.progress_dashboard.refresh()
        elif index == 6 and hasattr(self, 'cost_estimates_view'):
            self.cost_estimates_view.refresh()
    def _on_jump_to_gantt_from_tree(self, part_name):
        try:
            # Switch to Gantt tab
            try:
                if self.sidebar.currentRow() != 1:
                    self.sidebar.setCurrentRow(1)
            except Exception:
                pass
            # Render & highlight
            if hasattr(self, 'gantt_chart_view'):
                self.gantt_chart_view.render_gantt(self.model)
                if hasattr(self.gantt_chart_view, 'highlight_bar'):
                    self.gantt_chart_view.highlight_bar(part_name)
            if self.statusBar():
                self.statusBar().showMessage(f"Jumped to '{part_name}' in Gantt", 2500)
        except Exception as e:
            print(f"Jump to gantt failed: {e}")
    def __init__(self, model):
        try:
            super().__init__()
            # Title with optional VERSION suffix
            version_suffix = ""
            try:
                ver_path = resolve_resource_path("VERSION")
                if ver_path and os.path.exists(ver_path):
                    with open(ver_path, 'r', encoding='utf-8') as vf:
                        _ver = vf.read().strip()
                        if _ver:
                            version_suffix = f" v{_ver}"
            except Exception:
                pass
            self.setWindowTitle(f"Project Management App{version_suffix}")
            self.resize(1200, 700)
            # Set window icon from generated header.ico if possible
            try:
                from PyQt6.QtGui import QIcon
                ico_path = self._ensure_app_icon()
                if ico_path:
                    self.setWindowIcon(QIcon(ico_path))
            except Exception:
                pass
            # Set global stylesheet for background and foreground colors
            self.setStyleSheet("""
                QWidget {
                    background-color: #4B4B4B;
                    color: #FF8200;
                }
                QLineEdit, QTableWidget, QTreeWidget, QComboBox, QDateEdit, QHeaderView::section {
                    background-color: #333333;
                    color: #FF8200;
                    border: 1px solid #FF8200;
                }
                QPushButton {
                    background-color: #333333;
                    color: #FF8200;
                    border: 1px solid #FF8200;
                }
                QTableWidget QTableCornerButton::section {
                    background-color: #333333;
                }
            """)

            self.model = model
            # File-based edit lock ownership flag
            self._own_lock = False

            # Header with centered header.svg; PNG fallback if SVG missing/invalid
            header_layout = QHBoxLayout()
            # Eliminate extra margins/spacing around the header row
            header_layout.setContentsMargins(0, 0, 0, 0)
            header_layout.setSpacing(0)
            header_layout.addStretch(1)
            header_widget = None
            # Keep references for dynamic resize
            self._header_widget = None
            self._header_is_svg = False
            self._header_svg_renderer = None
            self._header_png_pixmap = None  # if set, use PNG-based header in UI
            self._header_aspect = None  # width / height (unused after trim logic)
            self._header_label = None  # QLabel used to display trimmed pixmap
            try:
                svg_path = resolve_resource_path("header.svg")
                used_svg = False
                if svg_path and os.path.exists(svg_path):
                    try:
                        from PyQt6.QtSvg import QSvgRenderer  # type: ignore
                        renderer = QSvgRenderer(svg_path)
                        if renderer.isValid():
                            try:
                                print(f"[UI] Using header.svg -> {svg_path}")
                            except Exception:
                                pass
                            # Use a QLabel + pixmap rendered from SVG and trimmed of transparent margins
                            lbl = QLabel()
                            try:
                                lbl.setStyleSheet("background: transparent; margin:0; padding:0; border:0;")
                                lbl.setAttribute(Qt.WA_TranslucentBackground, True)
                            except Exception:
                                pass
                            header_widget = lbl
                            used_svg = True
                            # Save for dynamic resizing
                            self._header_widget = lbl
                            self._header_label = lbl
                            self._header_is_svg = True
                            self._header_svg_renderer = renderer
                    except Exception:
                        try:
                            print("[UI] QSvgRenderer load failed; SVG header disabled")
                        except Exception:
                            pass
                        used_svg = False
                if not used_svg:
                    # Try PNG fallback before giving up
                    try:
                        png_path = resolve_resource_path("header.png")
                        from PyQt6.QtGui import QPixmap as _QPM
                        if png_path and os.path.exists(png_path):
                            pm = _QPM(png_path)
                            if not pm.isNull():
                                try:
                                    print(f"[UI] Using header.png fallback -> {png_path}")
                                except Exception:
                                    pass
                                lbl = QLabel()
                                try:
                                    lbl.setStyleSheet("background: transparent; margin:0; padding:0; border:0;")
                                    lbl.setAttribute(Qt.WA_TranslucentBackground, True)
                                except Exception:
                                    pass
                                header_widget = lbl
                                # Save for dynamic resizing
                                self._header_widget = lbl
                                self._header_label = lbl
                                self._header_is_svg = False
                                self._header_png_pixmap = pm
                    except Exception:
                        pass
                    # If still no header, show a small placeholder text
                    if header_widget is None:
                        try:
                            _msg = f"[UI] header.svg missing or invalid at path: {svg_path if svg_path else '(none)'}"
                            print(_msg)
                        except Exception:
                            pass
                        lbl = QLabel("[header.svg not found]")
                        header_widget = lbl
                        # Save for dynamic sizing
                        self._header_widget = lbl
                        self._header_label = lbl
                        self._header_is_svg = False
                        self._header_aspect = 4.0
            except Exception:
                # Final fallback
                lbl = QLabel("[header load error]")
                header_widget = lbl
                self._header_widget = lbl
                self._header_is_svg = False
                self._header_aspect = 4.0
            # Ensure the header widget itself has no padding/margins/border
            try:
                header_widget.setStyleSheet("margin:0px; padding:0px; border:0px;")
            except Exception:
                pass
            # Center column: logo only (remove inline menu bar under the logo)
            try:
                center_col = QVBoxLayout()
                center_col.setContentsMargins(0, 0, 0, 0)
                center_col.setSpacing(0)
                center_col.addWidget(header_widget, alignment=Qt.AlignmentFlag.AlignCenter)
                header_layout.addLayout(center_col)
            except Exception:
                # Fallback to just adding the header widget centered
                header_layout.addWidget(header_widget, alignment=Qt.AlignmentFlag.AlignCenter)
            header_layout.addStretch(1)

            # Controls row (separate from header row so the logo stays perfectly centered)
            controls_layout = QHBoxLayout()
            controls_layout.setContentsMargins(0, 0, 0, 0)
            controls_layout.setSpacing(6)
            controls_layout.addStretch(1)

            # Search field (affects Gantt view)
            from PyQt6.QtWidgets import QLineEdit, QPushButton
            self.search_input = QLineEdit()
            self.search_input.setPlaceholderText("Jump to part (substring)...")
            self.search_input.setFixedWidth(260)
            def do_jump():
                text = self.search_input.text().strip()
                if not text:
                    return
                # Find first matching part name (case-insensitive)
                lower = text.lower()
                match_name = None
                for r in self.model.rows:
                    name = r.get("Project Part", "")
                    if lower in name.lower():
                        match_name = name
                        break
                if match_name and hasattr(self.gantt_chart_view, 'highlight_bar'):
                    # Ensure Gantt view visible
                    if self.sidebar.currentRow() != 1:
                        self.sidebar.setCurrentRow(1)
                    self.gantt_chart_view.highlight_bar(match_name)
            self.search_input.returnPressed.connect(do_jump)
            controls_layout.addWidget(self.search_input)
            
            # Define reload action logic to use in Tools menu
            def do_reload():
                try:
                    # Re-load from disk to pick up synced changes
                    self.model.load_from_db()
                    self.on_data_changed()
                    if self.statusBar():
                        self.statusBar().showMessage("Reloaded from disk", 3000)
                except Exception as e:
                    print(f"Reload failed: {e}")
                finally:
                    try:
                        self._update_db_status()
                    except Exception:
                        pass
            # Keep a reference for other components (e.g., file change watcher)
            self._do_reload = do_reload

            # Tools menu button (moved here from under the header)
            try:
                from PyQt6.QtWidgets import QToolButton, QMenu, QAction, QInputDialog
                self.tools_btn = QToolButton()
                self.tools_btn.setText("Tools")
                self.tools_btn.setToolTip("App tools and utilities")
                self.tools_btn.setPopupMode(QToolButton.InstantPopup)
                tmenu = QMenu(self)
                # Actions
                act_jump = tmenu.addAction("Jump to Part")
                act_jump.triggered.connect(do_jump)
                tmenu.addSeparator()
                act_reload = tmenu.addAction("Reload Data")
                act_reload.triggered.connect(do_reload)
                tmenu.addSeparator()
                # Onboarding settings toggle
                from PyQt6.QtCore import QSettings
                s_on = QSettings('LSI','ProjectApp')
                hide_flag = s_on.value('Onboarding/hide_empty_dialog', False)
                if isinstance(hide_flag, str):
                    hide_flag = hide_flag.lower() in ('1','true','yes','on')
                act_onboarding = tmenu.addAction("Show First-Run Dialog on Empty DB")
                act_onboarding.setCheckable(True)
                act_onboarding.setChecked(not bool(hide_flag))
                def toggle_onboarding():
                    try:
                        new_show = act_onboarding.isChecked()
                        # store inverse (hide flag)
                        s_on.setValue('Onboarding/hide_empty_dialog', not new_show)
                        if self.statusBar():
                            self.statusBar().showMessage("First-run dialog {}".format("enabled" if new_show else "disabled"), 2500)
                    except Exception:
                        pass
                act_onboarding.toggled.connect(toggle_onboarding)
                
                act_open_folder = tmenu.addAction("Open Data Folder")
                act_open_folder.triggered.connect(self.open_data_folder)
                act_manage_holidays = tmenu.addAction("Manage Holidays…")
                act_manage_holidays.triggered.connect(self._open_holidays_manager)
                tmenu.addSeparator()
                act_switch_db = tmenu.addAction("Switch Data File…")
                def do_switch_db():
                    try:
                        path, _ = QFileDialog.getOpenFileName(self, "Choose Database File", os.path.dirname(os.path.abspath(self.model.DB_FILE)), "SQLite DB (*.db);;All Files (*.*)")
                        if not path:
                            return
                        # Persist to db_path.txt and QSettings
                        try:
                            with open(os.path.join(os.getcwd(), "db_path.txt"), "w", encoding="utf-8") as f:
                                f.write(path)
                        except Exception:
                            pass
                        try:
                            from PyQt6.QtCore import QSettings
                            QSettings('LSI','ProjectApp').setValue('DB/path', path)
                        except Exception:
                            pass
                        # Point model to new DB and reload
                        self.model.DB_FILE = path
                        self.model.ensure_schema()
                        self.model.load_from_db()
                        self.on_data_changed()
                        if self.statusBar():
                            self.statusBar().showMessage("Switched DB and reloaded", 3000)
                    except Exception as e:
                        print(f"Switch DB failed: {e}")
                act_switch_db.triggered.connect(do_switch_db)
                act_backup_db = tmenu.addAction("Backup Database…")
                def do_backup_db():
                    import datetime
                    try:
                        base = os.path.abspath(self.model.DB_FILE)
                        folder = os.path.dirname(base)
                        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                        stem = os.path.splitext(os.path.basename(base))[0]
                        dest = os.path.join(folder, f"{stem}_{ts}.db")
                        import shutil
                        shutil.copy2(base, dest)
                        # Copy WAL/SHM if present
                        for ext in ("-wal","-shm"):
                            src = base + ext
                            if os.path.exists(src):
                                shutil.copy2(src, dest + ext)
                        try:
                            # Record last backup time in QSettings and log
                            from PyQt6.QtCore import QSettings
                            QSettings('LSI','ProjectApp').setValue('Backup/last_backup_utc', datetime.datetime.utcnow().isoformat(timespec='seconds')+'Z')
                            log_event('backup','manual_backup', dest=dest)
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage(f"Backup created: {dest}", 3000)
                    except Exception as e:
                        try: log_event('backup','backup_failed', error=str(e))
                        except Exception: pass
                        print(f"Backup failed: {e}")
                act_backup_db.triggered.connect(do_backup_db)
                # --- Reports submenu ---
                reports_menu = QMenu("Reports", self)
                # Health Snapshot (PDF)
                act_health_pdf = reports_menu.addAction("Project Health Snapshot (PDF)")
                def _do_health_pdf():
                    try:
                        self._report_health_snapshot()
                    except Exception as e:
                        print(f"Health snapshot failed: {e}")
                act_health_pdf.triggered.connect(_do_health_pdf)
                # Baseline Variance (CSV)
                act_var_csv = reports_menu.addAction("Baseline Variance (CSV)")
                def _do_var_csv():
                    try:
                        self._report_baseline_variance_csv()
                    except Exception as e:
                        print(f"Baseline variance CSV failed: {e}")
                act_var_csv.triggered.connect(_do_var_csv)
                # Baseline Variance (PDF)
                act_var_pdf = reports_menu.addAction("Baseline Variance (PDF)")
                def _do_var_pdf():
                    try:
                        self._report_baseline_variance_pdf()
                    except Exception as e:
                        print(f"Baseline variance PDF failed: {e}")
                act_var_pdf.triggered.connect(_do_var_pdf)
                # Milestone Digest (PDF)
                act_milestone_pdf = reports_menu.addAction("Milestone Digest (PDF)")
                def _do_milestone_pdf():
                    try:
                        self._report_milestone_digest_pdf()
                    except Exception as e:
                        print(f"Milestone digest failed: {e}")
                act_milestone_pdf.triggered.connect(_do_milestone_pdf)
                tmenu.addMenu(reports_menu)
                # Export Settings
                act_export_settings = tmenu.addAction("Export Settings…")
                def _open_export_settings_inline():
                    try:
                        dlg = ExportSettingsDialog(self)
                        dlg.exec()
                    except Exception as e:
                        print(f"Open Export Settings failed: {e}")
                act_export_settings.triggered.connect(_open_export_settings_inline)
                # Global toggles
                tmenu.addSeparator()
                from PyQt6.QtCore import QSettings as _QS_links_tm
                _ps_links_tm = _QS_links_tm('LSI','ProjectPlanner')
                v_links = _ps_links_tm.value('UI/ShowLinks', 'true')
                def _b_l_tm(v):
                    if isinstance(v, bool): return v
                    if isinstance(v, str): return v.lower() in ('1','true','yes','on')
                    return True
                current_links = _b_l_tm(v_links)
                act_links = tmenu.addAction("Show Link Indicators")
                act_links.setCheckable(True)
                act_links.setChecked(current_links)
                def _toggle_links_menu(checked):
                    try:
                        _ps_links_tm.setValue('UI/ShowLinks', bool(checked))
                    except Exception:
                        pass
                    # Propagate to views
                    try:
                        if hasattr(self, 'gantt_chart_view') and self.gantt_chart_view:
                            self.gantt_chart_view._show_links = bool(checked)
                            if hasattr(self.gantt_chart_view, 'links_checkbox'):
                                self.gantt_chart_view.links_checkbox.setChecked(bool(checked))
                            self.gantt_chart_view.refresh_gantt()
                        if hasattr(self, 'timeline_view') and self.timeline_view:
                            self.timeline_view._show_links = bool(checked)
                            if hasattr(self.timeline_view, '_sync_links_checkbox'):
                                self.timeline_view._sync_links_checkbox()
                            self.timeline_view.render_timeline()
                        if hasattr(self, 'project_tree_view') and self.project_tree_view:
                            self.project_tree_view._show_links = bool(checked)
                            if hasattr(self.project_tree_view, '_sync_links_checkbox'):
                                self.project_tree_view._sync_links_checkbox()
                            self.project_tree_view.refresh()
                    except Exception:
                        pass
                act_links.toggled.connect(_toggle_links_menu)
                # --- Migration / Backfill Utilities ---
                tmenu.addSeparator()
                act_backfill_preds = tmenu.addAction("Backfill Predecessors from Dependencies")
                def do_backfill_preds():
                    changed = 0
                    try:
                        for r in self.model.rows:
                            if (not r.get('Predecessors')) and r.get('Dependencies'):
                                r['Predecessors'] = r.get('Dependencies')
                                changed += 1
                        if changed:
                            try:
                                self.model.save_to_db()
                            except Exception:
                                pass
                            self.on_data_changed()
                        if self.statusBar():
                            self.statusBar().showMessage(f"Backfill complete – updated {changed} rows", 3500)
                    except Exception as e:
                        print(f"Backfill failed: {e}")
                act_backfill_preds.triggered.connect(do_backfill_preds)
                # Toggle dependency/parent connectors
                # Connector visibility toggle + mode submenu
                from PyQt6.QtCore import QSettings as _QS
                qs = _QS('LSI','ProjectPlanner')
                stored_enabled = qs.value('Gantt/connectors_enabled', True)
                if isinstance(stored_enabled, str):
                    stored_enabled = stored_enabled.lower() in ('1','true','yes','on')
                stored_mode = qs.value('Gantt/connectors_mode', 'all')
                if stored_mode not in ('all','deps'):
                    stored_mode = 'all'
                act_toggle_connectors = tmenu.addAction("Show Connectors")
                act_toggle_connectors.setCheckable(True)
                act_toggle_connectors.setChecked(bool(stored_enabled))
                # Mode submenu
                mode_menu = QMenu("Connector Mode", tmenu)
                act_mode_all = mode_menu.addAction("Hierarchy + Dependencies")
                act_mode_deps = mode_menu.addAction("Dependencies Only")
                act_mode_all.setCheckable(True); act_mode_deps.setCheckable(True)
                if stored_mode == 'all':
                    act_mode_all.setChecked(True)
                else:
                    act_mode_deps.setChecked(True)
                tmenu.addMenu(mode_menu)
                # Apply stored to existing view
                try:
                    if hasattr(self, 'gantt_chart_view'):
                        self.gantt_chart_view._enable_connectors = bool(stored_enabled)
                        self.gantt_chart_view._connector_mode = stored_mode
                except Exception:
                    pass
                def do_toggle_connectors(checked):
                    try:
                        qs.setValue('Gantt/connectors_enabled', bool(checked))
                        if hasattr(self, 'gantt_chart_view'):
                            self.gantt_chart_view._enable_connectors = bool(checked)
                            self.gantt_chart_view.render_gantt(self.model)
                        if self.statusBar():
                            self.statusBar().showMessage("Connectors " + ("enabled" if checked else "disabled"), 2000)
                    except Exception:
                        pass
                act_toggle_connectors.toggled.connect(do_toggle_connectors)
                def set_mode(mode):
                    try:
                        qs.setValue('Gantt/connectors_mode', mode)
                        if hasattr(self, 'gantt_chart_view'):
                            self.gantt_chart_view._connector_mode = mode
                            self.gantt_chart_view.render_gantt(self.model)
                        if self.statusBar():
                            self.statusBar().showMessage(f"Connector mode: {mode}", 2000)
                    except Exception:
                        pass
                def pick_all():
                    act_mode_all.setChecked(True); act_mode_deps.setChecked(False); set_mode('all')
                def pick_deps():
                    act_mode_all.setChecked(False); act_mode_deps.setChecked(True); set_mode('deps')
                act_mode_all.triggered.connect(pick_all)
                act_mode_deps.triggered.connect(pick_deps)
                # Allow clicking the title of submenu parent to cycle modes (optional future enhancement)
                # Future: could add a settings dialog for thresholds
                # Create Shared Folder (OneDrive template)
                def do_create_shared_folder():
                    try:
                        from PyQt6.QtWidgets import QFileDialog, QInputDialog, QMessageBox
                        import os, shutil
                        base_dir = QFileDialog.getExistingDirectory(self, "Choose OneDrive location for shared folder", os.path.expanduser("~"))
                        if not base_dir:
                            return
                        name, ok = QInputDialog.getText(self, "Folder Name", "Shared folder name:", text="ProjectPlanner-Shared")
                        if not ok or not name.strip():
                            return
                        target = os.path.join(base_dir, name.strip())
                        if not os.path.exists(target):
                            os.makedirs(target)
                        # Create subfolders
                        for sub in ("images", "attachments", "backups"):
                            try:
                                os.makedirs(os.path.join(target, sub), exist_ok=True)
                            except Exception:
                                pass
                        # Copy holidays.json if present
                        try:
                            hp = _holidays_path()
                            if hp and os.path.exists(hp):
                                shutil.copy2(hp, os.path.join(target, "holidays.json"))
                        except Exception:
                            pass
                        # Offer to copy current DB
                        dest_db = os.path.join(target, "project_data.db")
                        do_copy = False
                        try:
                            if os.path.exists(self.model.DB_FILE):
                                ret = QMessageBox.question(self, "Copy Database?", f"Copy current database to shared folder as\n{dest_db}?", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                                do_copy = (ret == QMessageBox.Yes)
                        except Exception:
                            pass
                        if do_copy:
                            try:
                                shutil.copy2(self.model.DB_FILE, dest_db)
                                for ext in ("-wal","-shm"):
                                    side = self.model.DB_FILE + ext
                                    if os.path.exists(side):
                                        shutil.copy2(side, dest_db + ext)
                            except Exception:
                                pass
                        # Write helper docs
                        try:
                            readme = os.path.join(target, "README_SHARED.md")
                            with open(readme, "w", encoding="utf-8") as f:
                                f.write("""# Vols Signage Shared Folder (OneDrive)\n\nThis folder is designed to be placed in OneDrive so a small team can view/edit the same project data.\n\nRecommended contents:\n- `project_data.db` – SQLite database used by the desktop app\n- `holidays.json` – Shared holidays used for weekend/holiday shading\n- `images/` – Any task-linked images\n- `attachments/` – Optional linked files\n- `backups/` – Destination for timestamped database backups (optional)\n\nThe app may also create/manage:\n- `project_data.db.lock.json` – Lightweight edit lock file\n- `project_data.db-wal` / `project_data.db-shm` – SQLite WAL sidecar files\n\nHow to wire it up:\n1. Each teammate runs the desktop app locally (not from this folder).\n2. In the app: Tools → Switch Data File… → select `project_data.db` here.\n3. Viewers use Read-Only Mode; editors toggle it off to acquire the edit lock.\n""")
                        except Exception:
                            pass
                        try:
                            g = os.path.join(target, "GETTING_STARTED_SHARED.md")
                            with open(g, "w", encoding="utf-8") as f:
                                f.write("""# Getting Started (Shared Data)\n\nPlace this folder in OneDrive and share it with your team. Then either:\n\nOption A – Point your local app at this DB (recommended)\n1) Launch the desktop app locally\n2) Tools → Switch Data File… → select this folder’s `project_data.db`\n3) Toggle Tools → Read-Only Mode ON if viewing; OFF to edit (takes lock)\n\nOption B – Environment variable\n- Set `PROJECT_DB_PATH` to the full path of `project_data.db` in this folder before launching the app\n\nNotes\n- The app uses an edit-lock file to coordinate a single active editor. If a lock looks stale, you may be prompted to take over (configurable).\n- Backups: use Tools → Backup Database…; you can keep them in `backups/`.\n- Don’t put your `.venv` here; keep virtualenvs local.\n""")
                        except Exception:
                            pass
                        # Write a db_path.txt in the shared folder to document the canonical path
                        try:
                            with open(os.path.join(target, "db_path.txt"), "w", encoding="utf-8") as f:
                                f.write(dest_db)
                        except Exception:
                            pass
                        # Ask to switch app to new DB
                        try:
                            ret = QMessageBox.question(self, "Switch to Shared DB?", "Switch the app to use the new shared database now?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                            if ret == QMessageBox.Yes and os.path.exists(dest_db):
                                # Persist to db_path.txt in current working dir and QSettings
                                try:
                                    with open(os.path.join(os.getcwd(), "db_path.txt"), "w", encoding="utf-8") as wf:
                                        wf.write(dest_db)
                                except Exception:
                                    pass
                                try:
                                    from PyQt6.QtCore import QSettings
                                    QSettings('LSI','ProjectApp').setValue('DB/path', dest_db)
                                except Exception:
                                    pass
                                self.model.DB_FILE = dest_db
                                self.model.ensure_schema()
                                self.model.load_from_db()
                                self.on_data_changed()
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage(f"Shared folder created at {target}", 4000)
                    except Exception as e:
                        print(f"Create Shared Folder failed: {e}")
                act_create_shared = tmenu.addAction("Create Shared Folder…")
                act_create_shared.triggered.connect(do_create_shared_folder)
                tmenu.addSeparator()
                act_toggle_ro = tmenu.addAction("Read-Only Mode")
                act_toggle_ro.setCheckable(True)
                # Initialize from model
                try:
                    act_toggle_ro.setChecked(bool(getattr(self.model, 'read_only', False)))
                except Exception:
                    pass
                def do_toggle_ro(checked):
                    try:
                        # When switching to editing (checked=False), try to acquire lock
                        if not checked:
                            if not self._acquire_edit_lock():
                                # Revert to read-only and notify
                                try:
                                    act_toggle_ro.blockSignals(True)
                                    act_toggle_ro.setChecked(True)
                                    act_toggle_ro.blockSignals(False)
                                except Exception:
                                    pass
                                self.model.read_only = True
                                try:
                                    from PyQt6.QtWidgets import QMessageBox
                                    info = self._read_edit_lock() or {}
                                    holder = info.get('owner', 'someone else')
                                    when = info.get('when', '')
                                    QMessageBox.information(self, "Edit Lock Held",
                                        f"Another user holds the edit lock (owner: {holder}{' @ ' + when if when else ''}).\n"
                                        "Continue in Read-Only mode or try again later.")
                                except Exception:
                                    pass
                            else:
                                self.model.read_only = False
                        else:
                            # Enabling read-only -> release lock if owned
                            try:
                                self._release_edit_lock()
                            except Exception:
                                pass
                            self.model.read_only = True
                        from PyQt6.QtCore import QSettings
                        QSettings('LSI','ProjectApp').setValue('DB/read_only', bool(self.model.read_only))
                        # Update DatabaseView editability if available
                        if hasattr(self, 'database_view') and hasattr(self.database_view, 'set_read_only'):
                            self.database_view.set_read_only(bool(self.model.read_only))
                        # Update subtle read-only badge visibility
                        try:
                            self._update_read_only_indicator()
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage("Read-Only {}".format("On" if self.model.read_only else "Off"), 2500)
                    except Exception as e:
                        print(f"Toggle Read-Only failed: {e}")
                act_toggle_ro.toggled.connect(do_toggle_ro)
                self._act_toggle_ro = act_toggle_ro

                # --- Filters submenu (consolidated from right-side dock) ---
                self._init_filter_state()
                filters_menu = QMenu("Filters", self)
                # Status submenu
                status_menu = QMenu("Status", self)
                self._filter_actions = {"status": {}, "ie": {}, "flags": {}, "summary": None}
                for st in ["Planned", "In Progress", "Blocked", "Done"]:
                    a = QAction(st, self)
                    a.setCheckable(True)
                    a.toggled.connect(lambda checked, s=st: self._on_filter_status_toggled(s, checked))
                    status_menu.addAction(a)
                    self._filter_actions["status"][st] = a
                filters_menu.addMenu(status_menu)
                # Internal/External submenu
                ie_menu = QMenu("Internal / External", self)
                a_int = QAction("Internal", self); a_int.setCheckable(True)
                a_int.toggled.connect(lambda checked: self._on_filter_ie_toggled("Internal", checked))
                ie_menu.addAction(a_int); self._filter_actions["ie"]["Internal"] = a_int
                a_ext = QAction("External", self); a_ext.setCheckable(True)
                a_ext.toggled.connect(lambda checked: self._on_filter_ie_toggled("External", checked))
                ie_menu.addAction(a_ext); self._filter_actions["ie"]["External"] = a_ext
                filters_menu.addMenu(ie_menu)
                # Responsible contains (opens prompt)
                def set_responsible_substr():
                    cur = self._filter_state.get("responsible_substr") or ""
                    text, ok = QInputDialog.getText(self, "Responsible Contains", "Substring:", text=cur)
                    if ok:
                        self._filter_state["responsible_substr"] = text.strip() or None
                        self._update_filter_summary()
                        try:
                            self._apply_filters()
                        except Exception:
                            pass
                filters_menu.addAction("Responsible Contains…", set_responsible_substr)
                # Flags
                a_crit = QAction("Critical Path Only", self); a_crit.setCheckable(True)
                a_crit.toggled.connect(lambda checked: self._on_filter_flag_toggled("critical_only", checked))
                filters_menu.addAction(a_crit); self._filter_actions["flags"]["critical_only"] = a_crit
                a_risk = QAction("Risk (Overdue / At-Risk) Only", self); a_risk.setCheckable(True)
                a_risk.toggled.connect(lambda checked: self._on_filter_flag_toggled("risk_only", checked))
                filters_menu.addAction(a_risk); self._filter_actions["flags"]["risk_only"] = a_risk
                filters_menu.addSeparator()
                # Summary (disabled)
                sum_act = QAction("No filters active", self)
                sum_act.setEnabled(False)
                filters_menu.addAction(sum_act)
                self._filter_actions["summary"] = sum_act
                filters_menu.addSeparator()
                # Apply / Reset
                filters_menu.addAction("Apply Filters", lambda: self._apply_filters())
                filters_menu.addAction("Reset Filters", lambda: self._reset_filters())

                # Attach Filters submenu and put menu on the button
                tmenu.addMenu(filters_menu)
                # --- Sync submenu for OneDrive file watching ---
                sync_menu = QMenu("Sync", self)
                # Load persisted settings
                try:
                    from PyQt6.QtCore import QSettings
                    s = QSettings('LSI','ProjectApp')
                    self._sync_auto_reload_readonly = bool(s.value('Sync/auto_reload_readonly', True, type=bool))
                    self._sync_prompt_reload_editing = bool(s.value('Sync/prompt_reload_editing', True, type=bool))
                    self._sync_watch_ms = int(s.value('Sync/watch_interval_ms', 2000))
                except Exception:
                    self._sync_auto_reload_readonly = True
                    self._sync_prompt_reload_editing = True
                    self._sync_watch_ms = 2000
                # Actions
                a_auto = QAction("Auto-Reload on Sync (Read-Only)", self)
                a_auto.setCheckable(True)
                a_auto.setChecked(bool(self._sync_auto_reload_readonly))
                def on_auto(t):
                    self._sync_auto_reload_readonly = bool(t)
                    try:
                        from PyQt6.QtCore import QSettings
                        QSettings('LSI','ProjectApp').setValue('Sync/auto_reload_readonly', bool(t))
                    except Exception:
                        pass
                a_auto.toggled.connect(on_auto)
                sync_menu.addAction(a_auto)
                a_prompt = QAction("Prompt to Reload on Sync (Editing)", self)
                a_prompt.setCheckable(True)
                a_prompt.setChecked(bool(self._sync_prompt_reload_editing))
                def on_prompt(t):
                    self._sync_prompt_reload_editing = bool(t)
                    try:
                        from PyQt6.QtCore import QSettings
                        QSettings('LSI','ProjectApp').setValue('Sync/prompt_reload_editing', bool(t))
                    except Exception:
                        pass
                a_prompt.toggled.connect(on_prompt)
                sync_menu.addAction(a_prompt)
                sync_menu.addSeparator()
                def change_interval():
                    from PyQt6.QtWidgets import QInputDialog
                    cur_sec = max(1, int(round(self._sync_watch_ms/1000)))
                    sec, ok = QInputDialog.getInt(self, "Change Watch Interval", "Seconds:", cur_sec, 1, 60, 1)
                    if ok:
                        self._sync_watch_ms = int(sec * 1000)
                        try:
                            if hasattr(self, '_db_watch_timer') and self._db_watch_timer is not None:
                                self._db_watch_timer.setInterval(self._sync_watch_ms)
                        except Exception:
                            pass
                        try:
                            from PyQt6.QtCore import QSettings
                            QSettings('LSI','ProjectApp').setValue('Sync/watch_interval_ms', int(self._sync_watch_ms))
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage(f"Watch interval set to {sec}s", 2500)
                sync_menu.addAction("Change Watch Interval…", change_interval)
                tmenu.addMenu(sync_menu)
                # --- Edit Lock submenu ---
                lock_menu = QMenu("Edit Lock", self)
                def on_acquire_lock():
                    if self._acquire_edit_lock():
                        try:
                            self._act_toggle_ro.blockSignals(True)
                            self._act_toggle_ro.setChecked(False)
                            self._act_toggle_ro.blockSignals(False)
                        except Exception:
                            pass
                        self.model.read_only = False
                        if hasattr(self, 'database_view') and hasattr(self.database_view, 'set_read_only'):
                            self.database_view.set_read_only(False)
                        try:
                            self._update_read_only_indicator()
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage("Edit lock acquired", 2500)
                    else:
                        try:
                            from PyQt6.QtWidgets import QMessageBox
                            info = self._read_edit_lock() or {}
                            holder = info.get('owner', 'someone else')
                            when = info.get('when', '')
                            QMessageBox.information(self, "Edit Lock Held",
                                f"Another user holds the edit lock (owner: {holder}{' @ ' + when if when else ''}).")
                        except Exception:
                            pass
                def on_release_lock():
                    self._release_edit_lock()
                    try:
                        self._act_toggle_ro.blockSignals(True)
                        self._act_toggle_ro.setChecked(True)
                        self._act_toggle_ro.blockSignals(False)
                    except Exception:
                        pass
                    self.model.read_only = True
                    if hasattr(self, 'database_view') and hasattr(self.database_view, 'set_read_only'):
                        self.database_view.set_read_only(True)
                    try:
                        self._update_read_only_indicator()
                    except Exception:
                        pass
                    if self.statusBar():
                        self.statusBar().showMessage("Edit lock released", 2500)
                lock_menu.addAction("Acquire Edit Lock", on_acquire_lock)
                lock_menu.addAction("Release Edit Lock", on_release_lock)
                # Settings inside Edit Lock submenu
                try:
                    from PyQt6.QtCore import QSettings
                    s = QSettings('LSI','ProjectApp')
                    pt_val = s.value('Lock/prompt_takeover', True)
                    if isinstance(pt_val, str):
                        pt_val = pt_val.lower() in ('1','true','yes','on')
                    act_prompt = lock_menu.addAction("Prompt to Take Over Stale Lock")
                    act_prompt.setCheckable(True)
                    act_prompt.setChecked(bool(pt_val))
                    def toggle_prompt():
                        try:
                            s.setValue('Lock/prompt_takeover', bool(act_prompt.isChecked()))
                        except Exception:
                            pass
                    act_prompt.toggled.connect(toggle_prompt)
                    def change_timeout():
                        try:
                            from PyQt6.QtWidgets import QInputDialog
                            cur = int(s.value('Lock/stale_minutes', 30))
                            minutes, ok = QInputDialog.getInt(self, 'Stale Lock Timeout', 'Consider lock stale after (minutes):', value=max(1,cur), min=1, max=1440, step=1)
                            if ok:
                                s.setValue('Lock/stale_minutes', int(minutes))
                                self._update_lock_status()
                        except Exception:
                            pass
                    lock_menu.addAction("Change Stale Timeout…", change_timeout)
                except Exception:
                    pass
                tmenu.addMenu(lock_menu)
                self.tools_btn.setMenu(tmenu)
                controls_layout.addWidget(self.tools_btn)
                # Quick Filters dialog button
                try:
                    from PyQt6.QtWidgets import QToolButton
                    btn_filters = QToolButton()
                    btn_filters.setText("Filters…")
                    def _open_filters_dialog():
                        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QCheckBox, QDialogButtonBox, QGroupBox
                        dlg = QDialog(self)
                        dlg.setWindowTitle("Quick Filters")
                        v = QVBoxLayout(dlg)
                        # Status
                        grp_status = QGroupBox("Status")
                        gsv = QVBoxLayout(grp_status)
                        cb_pl = QCheckBox("Planned"); cb_ip = QCheckBox("In Progress"); cb_bl = QCheckBox("Blocked"); cb_dn = QCheckBox("Done")
                        gsv.addWidget(cb_pl); gsv.addWidget(cb_ip); gsv.addWidget(cb_bl); gsv.addWidget(cb_dn)
                        v.addWidget(grp_status)
                        # Internal/External
                        grp_ie = QGroupBox("Internal / External")
                        giev = QVBoxLayout(grp_ie)
                        cb_in = QCheckBox("Internal"); cb_ex = QCheckBox("External")
                        giev.addWidget(cb_in); giev.addWidget(cb_ex)
                        v.addWidget(grp_ie)
                        # Responsible contains
                        row_resp = QHBoxLayout(); row_resp.addWidget(QLabel("Responsible contains:"))
                        le_resp = QLineEdit(); row_resp.addWidget(le_resp, 1)
                        v.addLayout(row_resp)
                        # Flags
                        grp_flags = QGroupBox("Flags")
                        gfv = QVBoxLayout(grp_flags)
                        cb_crit = QCheckBox("Critical Path only")
                        cb_risk = QCheckBox("Risk (Overdue / At-Risk) only")
                        gfv.addWidget(cb_crit); gfv.addWidget(cb_risk)
                        v.addWidget(grp_flags)
                        # Seed current state
                        try:
                            st = self._filter_state.get("statuses", set())
                            cb_pl.setChecked("Planned" in st)
                            cb_ip.setChecked("In Progress" in st)
                            cb_bl.setChecked("Blocked" in st)
                            cb_dn.setChecked("Done" in st)
                            ie = self._filter_state.get("ie", set())
                            cb_in.setChecked("Internal" in ie)
                            cb_ex.setChecked("External" in ie)
                            le_resp.setText(self._filter_state.get("responsible_substr") or "")
                            cb_crit.setChecked(bool(self._filter_state.get("critical_only")))
                            cb_risk.setChecked(bool(self._filter_state.get("risk_only")))
                        except Exception:
                            pass
                        # Buttons
                        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
                        def _apply_and_close():
                            try:
                                st_new = set()
                                if cb_pl.isChecked(): st_new.add("Planned")
                                if cb_ip.isChecked(): st_new.add("In Progress")
                                if cb_bl.isChecked(): st_new.add("Blocked")
                                if cb_dn.isChecked(): st_new.add("Done")
                                self._filter_state["statuses"] = st_new
                                ie_new = set()
                                if cb_in.isChecked(): ie_new.add("Internal")
                                if cb_ex.isChecked(): ie_new.add("External")
                                self._filter_state["ie"] = ie_new
                                self._filter_state["responsible_substr"] = (le_resp.text() or "").strip() or None
                                self._filter_state["critical_only"] = bool(cb_crit.isChecked())
                                self._filter_state["risk_only"] = bool(cb_risk.isChecked())
                                # Sync menu checks and apply
                                try:
                                    self._sync_tools_filter_checks_from_state()
                                except Exception:
                                    pass
                                try:
                                    self._apply_filters()
                                except Exception:
                                    pass
                            except Exception:
                                pass
                            dlg.accept()
                        btns.accepted.connect(_apply_and_close)
                        btns.rejected.connect(dlg.reject)
                        v.addWidget(btns)
                        dlg.exec()
                    btn_filters.clicked.connect(_open_filters_dialog)
                    controls_layout.addWidget(btn_filters)
                except Exception:
                    pass
                # Manual shortcut creation action
                try:
                    act_shortcut = tmenu.addAction('Create Desktop Shortcut Now')
                    def _do_shortcut():
                        ok, info = self._create_desktop_shortcut('Vols Signage')
                        if self.statusBar():
                            self.statusBar().showMessage(('Shortcut created: ' if ok else 'Shortcut failed: ') + info, 5000)
                    act_shortcut.triggered.connect(_do_shortcut)
                except Exception:
                    pass

                # Sync menu with current (loaded) settings later in init
            except Exception:
                pass

            # Initialize filter storage in gantt view if available
            # (Filters UI is now in Tools → Filters; no right-side dock)

            # Sidebar for view selection (create and add to layout first)
            self.sidebar = QListWidget()
            self.sidebar.addItems([
                "Project Tree",
                "Gantt Chart",
                "Calendar",
                "Project Timeline",
                "Database",
                "Progress Dashboard",
                "Cost Estimates"
            ])

            # Stacked widget for views
            self.project_tree_view = ProjectTreeView(
                self.model,
                on_part_selected=self.on_tree_part_selected,
                on_jump_to_gantt=self._on_jump_to_gantt_from_tree
            )
            self.gantt_chart_view = GanttChartView()
            self.calendar_view = CalendarView(self.model)
            self.timeline_view = TimelineView(self.model)
            self.database_view = DatabaseView(self.model, on_data_changed=self.on_data_changed)
            # Enforce exclusive editing based on existing lock at startup
            try:
                info = self._read_edit_lock() or {}
                owner = info.get('owner')
                is_stale = self._is_lock_stale(info) if info else False
                me = self._whoami()
                if owner and not is_stale:
                    if owner == me:
                        # We appear to own a current lock from a previous session – honor it
                        self._own_lock = True
                        self.model.read_only = False
                        try:
                            if hasattr(self, '_act_toggle_ro'):
                                self._act_toggle_ro.blockSignals(True)
                                self._act_toggle_ro.setChecked(False)
                                self._act_toggle_ro.blockSignals(False)
                        except Exception:
                            pass
                    else:
                        # Someone else holds the lock – force read-only
                        self.model.read_only = True
                        try:
                            if hasattr(self, '_act_toggle_ro'):
                                self._act_toggle_ro.blockSignals(True)
                                self._act_toggle_ro.setChecked(True)
                                self._act_toggle_ro.blockSignals(False)
                        except Exception:
                            pass
                        try:
                            from PyQt6.QtCore import QSettings
                            QSettings('LSI','ProjectApp').setValue('DB/read_only', True)
                        except Exception:
                            pass
            except Exception:
                pass
            try:
                # Initialize database view read-only state from (possibly adjusted) model setting
                if hasattr(self.database_view, 'set_read_only'):
                    self.database_view.set_read_only(bool(getattr(self.model, 'read_only', False)))
            except Exception:
                pass
            self.progress_dashboard = ProgressDashboard(self.model)
            self.cost_estimates_view = CostEstimatesView(self.model)

            self.views = QStackedWidget()
            self.views.addWidget(self.project_tree_view)
            self.views.addWidget(self.gantt_chart_view)
            self.views.addWidget(self.calendar_view)
            self.views.addWidget(self.timeline_view)
            self.views.addWidget(self.database_view)
            self.views.addWidget(self.progress_dashboard)
            self.views.addWidget(self.cost_estimates_view)

            # --- Global Preview Panel setting (applies to all views with preview labels) ---
            def _read_preview_setting_default_true():
                try:
                    from PyQt6.QtCore import QSettings
                    s = QSettings('LSI','ProjectApp')
                    v = s.value('UI/ShowPreviewPanel', None)
                    if v is None:
                        # Back-compat with older per-tree key
                        v = s.value('TreeShowPreviewPanel', 'true')
                    if isinstance(v, str):
                        v = v.lower() in ('1','true','yes','on')
                    return True if v is None else bool(v)
                except Exception:
                    return True
            def _persist_preview_setting(val: bool):
                try:
                    from PyQt6.QtCore import QSettings
                    s = QSettings('LSI','ProjectApp')
                    s.setValue('UI/ShowPreviewPanel', bool(val))
                    # Keep tree key in sync for back-compat with internal tree logic
                    s.setValue('TreeShowPreviewPanel', bool(val))
                except Exception:
                    pass
            def _apply_preview_to_widget(w, visible: bool):
                try:
                    if hasattr(w, 'preview_label') and getattr(w, 'preview_label') is not None:
                        lab = w.preview_label
                        lab.setVisible(bool(visible))
                        # Collapse when hidden to reclaim space
                        try:
                            if bool(visible):
                                lab.setMaximumHeight(16777215)
                            else:
                                lab.setMaximumHeight(0)
                        except Exception:
                            pass
                except Exception:
                    pass
            self._show_preview_panels = _read_preview_setting_default_true()
            # Apply initial state to all views
            for vw in (self.project_tree_view, self.gantt_chart_view, self.calendar_view, self.timeline_view):
                _apply_preview_to_widget(vw, bool(self._show_preview_panels))
            # Sync Project Tree checkbox with global and propagate on change
            try:
                if hasattr(self.project_tree_view, 'preview_panel_cb'):
                    cb = self.project_tree_view.preview_panel_cb
                    # Initialize checkbox to global value without firing loops
                    try:
                        cb.blockSignals(True)
                        cb.setChecked(bool(self._show_preview_panels))
                        cb.blockSignals(False)
                    except Exception:
                        pass
                    def _on_tree_preview_toggled(_state):
                        val = bool(cb.isChecked())
                        self._show_preview_panels = val
                        _persist_preview_setting(val)
                        for vw in (self.project_tree_view, self.gantt_chart_view, self.calendar_view, self.timeline_view):
                            _apply_preview_to_widget(vw, val)
                    cb.stateChanged.connect(_on_tree_preview_toggled)
            except Exception:
                pass

            # Layout
            main_layout = QVBoxLayout()
            # Eliminate extra margins/spacing in the main layout as well
            main_layout.setContentsMargins(0, 0, 0, 0)
            main_layout.setSpacing(0)
            main_layout.addLayout(header_layout)
            # Add controls row under the centered logo/menu
            main_layout.addLayout(controls_layout)
            content_layout = QHBoxLayout()
            content_layout.setContentsMargins(0, 0, 0, 0)
            content_layout.setSpacing(0)
            content_layout.addWidget(self.sidebar)
            content_layout.addWidget(self.views, 1)
            main_layout.addLayout(content_layout)

            # Footer
            footer_label = QLabel("Copyright 2025 © LSI Graphics, LLC. All Rights Reserved.")
            footer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            footer_label.setStyleSheet("color: #888; font-size: 11px; margin-top: 8px;")
            main_layout.addWidget(footer_label)

            container = QWidget()
            container.setLayout(main_layout)
            self.setCentralWidget(container)

            # Initial header sizing based on current window size
            try:
                self._resize_header()
            except Exception:
                pass
            # Offer desktop shortcut once
            try:
                self._maybe_offer_shortcut()
            except Exception:
                pass

            # Status bar with DB info + quick action
            from PyQt6.QtWidgets import QStatusBar, QPushButton as _QBtn
            sb = QStatusBar()
            self.setStatusBar(sb)
            self.db_status_label = QLabel()
            self.db_status_label.setStyleSheet("color:#ccc; font-size:11px")
            # Last sync/update detected label
            self.db_sync_label = QLabel("Last: —")
            self.db_sync_label.setStyleSheet("color:#aaa; font-size:11px")
            # Last local code update (proxy for last sync of shared source)
            self.code_sync_label = QLabel("Code: —")
            self.code_sync_label.setStyleSheet("color:#aaa; font-size:11px")
            # Edit lock indicator label
            self.lock_label = QLabel("Lock: —")
            self.lock_label.setStyleSheet("color:#aaa; font-size:11px")
            self.db_warning_label = QLabel()
            self.db_warning_label.setStyleSheet("color:#FFD166; font-size:11px")
            # Read-only indicator label
            self.db_ro_label = QLabel("READ-ONLY")
            # Subtle badge styling: muted colors, thin border
            self.db_ro_label.setStyleSheet(
                "color:#555; background-color:#efefef; border:1px solid #cfcfcf;"
                "font-size:10px; padding:1px 6px; border-radius:3px;"
            )
            self.open_folder_btn = _QBtn("Open Data Folder")
            self.open_folder_btn.setStyleSheet("font-size:11px")
            self.open_folder_btn.clicked.connect(self.open_data_folder)
            sb.addPermanentWidget(self.db_status_label, 1)
            sb.addPermanentWidget(self.db_sync_label, 0)
            sb.addPermanentWidget(self.code_sync_label, 0)
            sb.addPermanentWidget(self.lock_label, 0)
            sb.addPermanentWidget(self.db_warning_label, 0)
            sb.addPermanentWidget(self.db_ro_label, 0)
            sb.addPermanentWidget(self.open_folder_btn, 0)
            # Active view/zoom label
            self.view_status_label = QLabel("Zoom: —")
            self.view_status_label.setStyleSheet("color:#ccc; font-size:11px")
            sb.addPermanentWidget(self.view_status_label, 0)
            self._update_db_status()
            # Initialize code sync status once at startup
            try:
                self._update_code_status()
            except Exception:
                pass
            # Initialize lock status once at startup
            try:
                self._update_lock_status()
            except Exception:
                pass
            # Initialize read-only indicator visibility
            try:
                self._update_read_only_indicator()
            except Exception:
                pass
            # Initialize DB change watcher (detect OneDrive sync updates)
            try:
                from PyQt6.QtCore import QTimer
                self._db_last_mtime = self._get_db_mtime()
                self._db_change_prompt_at = 0.0
                self._db_watch_timer = QTimer(self)
                # Respect persisted interval
                try:
                    from PyQt6.QtCore import QSettings
                    self._sync_watch_ms = int(QSettings('LSI','ProjectApp').value('Sync/watch_interval_ms', 2000))
                except Exception:
                    self._sync_watch_ms = 2000
                self._db_watch_timer.setInterval(int(self._sync_watch_ms))
                self._db_watch_timer.timeout.connect(self._check_db_changed)
                self._db_watch_timer.start()
                # Initialize code sync tracking for periodic refresh
                try:
                    self._last_code_mtime = self._get_code_mtime()
                except Exception:
                    self._last_code_mtime = 0.0
                self._code_check_accum = 0
                self._lock_tick = 0
            except Exception:
                pass

            # Basic Tools menu (keep hidden to avoid extra top padding – use inline menu below the logo)
            mb = self.menuBar(); tools_menu = mb.addMenu("Tools")
            act = tools_menu.addAction("Open Data Folder")
            act.triggered.connect(self.open_data_folder)
            act2 = tools_menu.addAction("Manage Holidays…")
            act2.triggered.connect(self._open_holidays_manager)
            act_export_settings2 = tools_menu.addAction("Export Settings…")
            def _open_export_settings_menu():
                try:
                    dlg = ExportSettingsDialog(self)
                    dlg.exec()
                except Exception as e:
                    print(f"Export Settings dialog failed: {e}")
            act_export_settings2.triggered.connect(_open_export_settings_menu)
            act_pricing = tools_menu.addAction("Pricing Settings…")
            def _open_pricing_settings():
                try:
                    dlg = PricingSettingsDialog(self)
                    dlg.exec()
                except Exception as e:
                    print(f"Pricing settings dialog failed: {e}")
            act_pricing.triggered.connect(_open_pricing_settings)
            # Sample Data action (available anytime)
            act_sample = tools_menu.addAction("Create Sample Data…")
            def do_sample():
                try:
                    from PyQt6.QtWidgets import QMessageBox
                    if self.model.rows:
                        resp = QMessageBox.question(self, "Append Sample Data?", "Existing tasks detected. Append sample tasks anyway?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                        if resp != QMessageBox.Yes:
                            return
                    self.model.load_sample_data()
                    self.on_data_changed()
                    if self.statusBar():
                        self.statusBar().showMessage("Sample data created", 3000)
                except Exception as e:
                    print(f"Sample data creation failed: {e}")
            act_sample.triggered.connect(do_sample)
            try:
                mb.setVisible(False)
            except Exception:
                pass

            # --- Integrity check and backup reminder timers ---
            try:
                from PyQt6.QtCore import QTimer, QSettings
                # Perform a quick integrity check shortly after startup (asynchronous so UI shows quickly)
                def run_integrity_check():
                    try:
                        import sqlite3
                        with self.model._connect() as _c:
                            cur = _c.cursor()
                            cur.execute('PRAGMA quick_check')
                            rows = cur.fetchall()
                            ok = all(r[0] == 'ok' for r in rows)
                            log_event('integrity','quick_check_result', ok=ok, details=[r[0] for r in rows])
                            if not ok and self.statusBar():
                                self.statusBar().showMessage('DB integrity issues detected – consider restoring a backup', 8000)
                    except Exception as e:
                        try: log_event('integrity','quick_check_error', error=str(e))
                        except Exception: pass
                QTimer.singleShot(3000, run_integrity_check)
                # Daily reminder (every 6 hours tick) if last backup older than 7 days
                def backup_reminder_tick():
                    try:
                        s = QSettings('LSI','ProjectApp')
                        last = s.value('Backup/last_backup_utc', '')
                        import datetime
                        stale = True
                        if last:
                            try:
                                # Accept both with/without trailing Z
                                if last.endswith('Z'): last = last[:-1]
                                dt = datetime.datetime.fromisoformat(last)
                                age_days = (datetime.datetime.utcnow() - dt).days
                                stale = age_days >= 7
                            except Exception:
                                stale = True
                        if stale and self.statusBar():
                            self.statusBar().showMessage('No recent backup (>=7 days) – use Tools → Backup Database…', 10000)
                            log_event('backup','reminder_shown')
                    except Exception:
                        pass
                self._backup_reminder_timer = QTimer(self)
                self._backup_reminder_timer.setInterval(6 * 3600 * 1000)  # 6 hours
                self._backup_reminder_timer.timeout.connect(backup_reminder_tick)
                self._backup_reminder_timer.start()
                # Run first reminder check after 10s
                QTimer.singleShot(10000, backup_reminder_tick)
            except Exception:
                pass

            # Finalize Filters: initialize gantt filter storage, load persisted settings, sync menu checks, and apply
            try:
                if hasattr(self, 'gantt_chart_view') and hasattr(self.gantt_chart_view, '_init_filters'):
                    self.gantt_chart_view._init_filters()
                # Load settings into state and sync menu checks
                self.load_filter_settings()
                self._sync_tools_filter_checks_from_state()
                # Apply once after construction
                self._apply_filters()
            except Exception:
                pass

            # Now that all views are constructed, connect sidebar signals and set current row
            self.sidebar.currentRowChanged.connect(self.display_view)
            self.sidebar.setCurrentRow(4)  # Start on Database view for editing (Dashboard is index 5)
            # If Gantt tab is selected at startup, render it
            if self.sidebar.currentRow() == 1:
                if hasattr(self.gantt_chart_view, 'scene') and self.gantt_chart_view.scene is not None:
                    self.gantt_chart_view.render_gantt(self.model)
            # Wire zoom updates to status bar
            try:
                def _set_zoom_label_from_view(view, name):
                    try:
                        sf = float(view.transform().m11())
                        from math import isfinite
                        pct = f"{int(round(sf*100))}%" if isfinite(sf) else "—"
                    except Exception:
                        pct = "—"
                    sel = 0
                    try:
                        scn = view.scene()
                        if scn is not None:
                            sel = len([it for it in scn.selectedItems()])
                    except Exception:
                        pass
                    txt = f"{name}: {pct}"
                    if sel:
                        txt += f"  (selected: {sel})"
                    self.view_status_label.setText(txt)
                # Connect signals
                if hasattr(self, 'project_tree_view') and hasattr(self.project_tree_view, 'view'):
                    self.project_tree_view.view.zoomChanged.connect(lambda sf: _set_zoom_label_from_view(self.project_tree_view.view, 'Tree'))
                if hasattr(self, 'gantt_chart_view') and hasattr(self.gantt_chart_view, 'view'):
                    self.gantt_chart_view.view.zoomChanged.connect(lambda sf: _set_zoom_label_from_view(self.gantt_chart_view.view, 'Gantt'))
                if hasattr(self, 'timeline_view') and hasattr(self.timeline_view, 'view'):
                    self.timeline_view.view.zoomChanged.connect(lambda sf: _set_zoom_label_from_view(self.timeline_view.view, 'Timeline'))
                # Update on tab change
                def _on_tab_change(idx):
                    name = {0:'Tree',1:'Gantt',2:'Timeline',3:'',4:'DB',5:'Dashboard',6:'Pricing'}.get(idx, '')
                    v = None
                    if idx == 0 and hasattr(self.project_tree_view,'view'): v = self.project_tree_view.view
                    elif idx == 1 and hasattr(self.gantt_chart_view,'view'): v = self.gantt_chart_view.view
                    elif idx == 2 and hasattr(self.timeline_view,'view'): v = self.timeline_view.view
                    if v is not None:
                        _set_zoom_label_from_view(v, name)
                    else:
                        self.view_status_label.setText(f"{name or 'View'}")
                self.sidebar.currentRowChanged.connect(_on_tab_change)
                # Seed with current
                _on_tab_change(self.sidebar.currentRow())
            except Exception:
                pass

            # --- First-run onboarding dialog (empty DB) ---
            try:
                from PyQt6.QtCore import QSettings
                s = QSettings('LSI','ProjectApp')
                hide = s.value('Onboarding/hide_empty_dialog', False)
                if isinstance(hide, str):
                    hide = hide.lower() in ('1','true','yes','on')
                if not hide and not self.model.rows:
                    dlg = FirstRunDialog(self)
                    dlg.exec()
                    if dlg.hide_future():
                        s.setValue('Onboarding/hide_empty_dialog', True)
                    if dlg.selected_action == 'sample':
                        try:
                            self.model.load_sample_data()
                            self.on_data_changed()
                        except Exception as e:
                            print(f"Onboarding sample data failed: {e}")
                    elif dlg.selected_action == 'switch':
                        # Reuse existing switch DB logic (Tools button action)
                        try:
                            for act in self.tools_btn.menu().actions():
                                if act.text().startswith('Switch Data File'):
                                    act.trigger()
                                    break
                        except Exception:
                            pass
                    elif dlg.selected_action == 'open':
                        try:
                            self.open_data_folder()
                        except Exception:
                            pass
            except Exception as e:
                print(f"Onboarding dialog error: {e}")
        except Exception as e:
            import traceback
            print("EXCEPTION in MainWindow.__init__:", e)
            traceback.print_exc()
    def _update_read_only_indicator(self):
        try:
            ro = bool(getattr(self.model, 'read_only', False))
        except Exception:
            ro = False
        try:
            if hasattr(self, 'db_ro_label') and self.db_ro_label is not None:
                self.db_ro_label.setVisible(ro)
        except Exception:
            pass
        # Keep Tools button text unchanged and avoid modifying the window title for a subtler signal
    def open_data_folder(self):
        import os, sys, subprocess
        base_dir = os.path.dirname(os.path.abspath(self.model.DB_FILE))
        try:
            if sys.platform.startswith('win'):
                os.startfile(base_dir)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', base_dir])
            else:
                subprocess.Popen(['xdg-open', base_dir])
        except Exception as e:
            print(f"Failed to open data folder: {e}")
    def _update_db_status(self):
        import os, time
        try:
            db_path = os.path.abspath(self.model.DB_FILE)
        except Exception:
            db_path = self.model.DB_FILE
        exists = os.path.exists(db_path)
        if exists:
            ts = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(db_path)))
            text = f"DB: {db_path}  |  Last Modified: {ts}"
            # Staleness warning if older than 24h and basic conflict hint
            try:
                age_sec = time.time() - os.path.getmtime(db_path)
                warn = []
                if age_sec > 24*3600:
                    warn.append("DB older than 24h (check sync)")
                base_dir = os.path.dirname(db_path)
                conf = [n for n in os.listdir(base_dir) if n.lower().endswith('.db') and n.startswith('project_data') and n != 'project_data.db']
                if conf:
                    warn.append(f"conflict copies: {len(conf)}")
                self.db_warning_label.setText(" | ".join(warn))
            except Exception:
                pass
        else:
            text = f"DB: {db_path} (missing)"
            try:
                self.db_warning_label.setText("DB missing")
            except Exception:
                pass
        if hasattr(self, 'db_status_label') and self.db_status_label is not None:
            self.db_status_label.setText(text)
        # Update cached mtime baseline for watcher if needed
        try:
            self._db_last_mtime = self._get_db_mtime()
        except Exception:
            pass
    def _get_db_mtime(self):
        import os
        try:
            base = os.path.abspath(self.model.DB_FILE)
        except Exception:
            base = self.model.DB_FILE
        mtimes = []
        try:
            if os.path.exists(base):
                mtimes.append(os.path.getmtime(base))
            for ext in ('-wal','-shm'):
                side = base + ext
                if os.path.exists(side):
                    mtimes.append(os.path.getmtime(side))
        except Exception:
            pass
        return max(mtimes) if mtimes else 0.0
    def _get_code_mtime(self):
        # Proxy for last sync time: most recent mtime among source files in the app folder
        import os, time
        root = os.path.dirname(os.path.abspath(__file__))
        latest = 0.0
        skip_dirs = {'.git','build','dist','__pycache__','images','web/static/vendor'}
        allow_ext = {'.py','.spec','.md','.json','.ini','.yaml','.yml','.qss'}
        for base, dirs, files in os.walk(root):
            # prune skip dirs
            pruned = []
            for d in list(dirs):
                rel = os.path.relpath(os.path.join(base, d), root).replace('\\','/')
                if d in skip_dirs or rel in skip_dirs:
                    pruned.append(d)
            for d in pruned:
                dirs.remove(d)
            for f in files:
                _, ext = os.path.splitext(f)
                if ext.lower() in allow_ext:
                    p = os.path.join(base, f)
                    try:
                        m = os.path.getmtime(p)
                        if m > latest:
                            latest = m
                    except Exception:
                        pass
        return latest
    def _update_code_status(self):
        import time
        try:
            m = self._get_code_mtime()
            if m:
                ts = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(m))
                text = f"Code: {ts}"
            else:
                text = "Code: —"
            if hasattr(self, 'code_sync_label') and self.code_sync_label is not None:
                self.code_sync_label.setText(text)
        except Exception:
            pass
    # -------- Edit Lock helpers --------
    def _lock_path(self):
        import os
        try:
            base = os.path.abspath(self.model.DB_FILE)
        except Exception:
            base = self.model.DB_FILE
        return base + ".lock.json"
    def _whoami(self):
        import socket, getpass
        try:
            user = getpass.getuser()
        except Exception:
            user = "unknown"
        try:
            host = socket.gethostname()
        except Exception:
            host = "host"
        return f"{user}@{host}"
    def _read_edit_lock(self):
        import os, json
        p = self._lock_path()
        try:
            if os.path.exists(p):
                with open(p, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return None
    def _write_edit_lock(self):
        import json, time, os
        # When called for acquisition, we want atomic create; when called as heartbeat, we only update if already owner
        def _write(force=False, exclusive=False):
            info = {"owner": self._whoami(), "when": time.strftime('%Y-%m-%d %H:%M:%S'), "pid": os.getpid()}
            p = self._lock_path()
            try:
                if exclusive:
                    # Atomic create: fail if exists
                    flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY
                    with os.fdopen(os.open(p, flags), 'w', encoding='utf-8') as f:
                        json.dump(info, f)
                    return True
                else:
                    # Update only if we are current owner unless force=True (used for stale takeover)
                    try:
                        cur = None
                        if os.path.exists(p):
                            cur = json.load(open(p, 'r', encoding='utf-8'))
                    except Exception:
                        cur = None
                    if force or (isinstance(cur, dict) and cur.get('owner') == info['owner']):
                        with open(p, 'w', encoding='utf-8') as f:
                            json.dump(info, f)
                        return True
                    return False
            except Exception:
                return False
        # Default behavior: heartbeat update (non-exclusive)
        return _write(force=False, exclusive=False)
    def _get_lock_settings(self):
        # Returns (stale_minutes:int, prompt_takeover:bool)
        try:
            from PyQt6.QtCore import QSettings
            s = QSettings('LSI','ProjectApp')
            stale_minutes = int(s.value('Lock/stale_minutes', 30))
            pt = s.value('Lock/prompt_takeover', True)
            if isinstance(pt, str):
                pt = pt.lower() in ('1','true','yes','on')
            return max(1, stale_minutes), bool(pt)
        except Exception:
            return 30, True
    def _is_lock_stale(self, info):
        try:
            if not info:
                return False
            when = info.get('when')
            if not when:
                return False
            import time
            try:
                ts = time.mktime(time.strptime(when, '%Y-%m-%d %H:%M:%S'))
            except Exception:
                return False
            stale_minutes, _ = self._get_lock_settings()
            return (time.time() - ts) > (stale_minutes * 60)
        except Exception:
            return False
    def _acquire_edit_lock(self):
        existing = self._read_edit_lock()
        if existing and existing.get('owner') != self._whoami():
            # If stale, optionally prompt for takeover
            if self._is_lock_stale(existing):
                _, prompt = self._get_lock_settings()
                if prompt:
                    try:
                        from PyQt6.QtWidgets import QMessageBox
                        owner = existing.get('owner') or 'unknown'
                        when = existing.get('when') or ''
                        m = QMessageBox(self)
                        m.setIcon(QMessageBox.Warning)
                        m.setWindowTitle('Stale Edit Lock')
                        m.setText('An existing edit lock appears stale.')
                        m.setInformativeText(f"Owner: {owner}{' @ ' + when if when else ''}\nTake over the lock?")
                        m.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                        m.setDefaultButton(QMessageBox.No)
                        choice = m.exec()
                        if choice == QMessageBox.Yes:
                            # Forcefully take over (overwrite existing)
                            try:
                                import json, time, os
                                info = {"owner": self._whoami(), "when": time.strftime('%Y-%m-%d %H:%M:%S'), "pid": os.getpid()}
                                with open(self._lock_path(), 'w', encoding='utf-8') as f:
                                    json.dump(info, f)
                                ok = True
                            except Exception:
                                ok = False
                            if ok:
                                self._own_lock = True
                                self._update_lock_status()
                            return ok
                    except Exception:
                        pass
            self._own_lock = False
            self._update_lock_status(existing)
            return False
        # Try atomic create to avoid races between two editors
        try:
            ok = False
            # Use the internal writer with exclusive create path
            import types
            # Recreate the inner function call path
            import json, time, os
            info = {"owner": self._whoami()}
            # attempt exclusive create by emulating the helper
            p = self._lock_path()
            try:
                flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY
                with os.fdopen(os.open(p, flags), 'w', encoding='utf-8') as f:
                    info.update({"when": time.strftime('%Y-%m-%d %H:%M:%S'), "pid": os.getpid()})
                    json.dump(info, f)
                ok = True
            except Exception:
                ok = False
        except Exception:
            ok = False
        if ok:
            self._own_lock = True
            self._update_lock_status()
        return ok
    def _release_edit_lock(self):
        import os
        p = self._lock_path()
        try:
            info = self._read_edit_lock()
            if info and info.get('owner') != self._whoami():
                self._own_lock = False
                self._update_lock_status(info)
                return False
            if os.path.exists(p):
                os.remove(p)
            self._own_lock = False
            self._update_lock_status()
            return True
        except Exception:
            return False
    def _update_lock_status(self, info=None):
        try:
            info = info if info is not None else (self._read_edit_lock() or {})
            if info:
                owner = info.get('owner') or 'unknown'
                when = info.get('when') or ''
                stale_mark = ' (stale)' if self._is_lock_stale(info) else ''
                txt = f"Lock: {owner}{' @ ' + when if when else ''}{stale_mark}"
            else:
                txt = "Lock: —"
            if hasattr(self, 'lock_label') and self.lock_label is not None:
                self.lock_label.setText(txt)
        except Exception:
            pass
    def _check_db_changed(self):
        import time
        try:
            # Skip heavy file/lock polling during automated tests to avoid race/access issues
            import os as _os_guard
            if _os_guard.environ.get('PYTEST_CURRENT_TEST'):
                return
            # Refresh lock status every tick; heartbeat our lock every ~10 ticks
            try:
                self._update_lock_status()
                self._lock_tick = getattr(self, '_lock_tick', 0) + 1
                if self._lock_tick >= 10 and getattr(self, '_own_lock', False):
                    self._lock_tick = 0
                    self._write_edit_lock()
            except Exception:
                pass
            # If another user holds a fresh lock and we're in editing mode, switch to read-only
            try:
                info = self._read_edit_lock() or {}
                owner = info.get('owner')
                if owner and owner != self._whoami() and not self._is_lock_stale(info):
                    if not bool(getattr(self.model, 'read_only', False)):
                        # Flip to read-only and reflect in UI
                        self.model.read_only = True
                        try:
                            if hasattr(self, '_act_toggle_ro'):
                                self._act_toggle_ro.blockSignals(True)
                                self._act_toggle_ro.setChecked(True)
                                self._act_toggle_ro.blockSignals(False)
                        except Exception:
                            pass
                        try:
                            if hasattr(self, 'database_view') and hasattr(self.database_view, 'set_read_only'):
                                self.database_view.set_read_only(True)
                        except Exception:
                            pass
                        try:
                            self._update_read_only_indicator()
                        except Exception:
                            pass
                        if self.statusBar():
                            self.statusBar().showMessage("Another user acquired the edit lock — switching to Read-Only", 3500)
            except Exception:
                pass
            # Periodically refresh code sync label (every ~5 ticks)
            try:
                self._code_check_accum = getattr(self, '_code_check_accum', 0) + 1
                if self._code_check_accum >= 5:
                    self._code_check_accum = 0
                    m = self._get_code_mtime()
                    if m and m != getattr(self, '_last_code_mtime', 0.0):
                        self._last_code_mtime = m
                        self._update_code_status()
            except Exception:
                pass

            cur = self._get_db_mtime()
            last = getattr(self, '_db_last_mtime', 0.0)
            if cur and (cur > last + 0.5):
                # File changed on disk – update baseline and act
                self._db_last_mtime = cur
                # Compute timestamp once for status label
                tstr = None
                try:
                    tstr = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                except Exception:
                    tstr = None
                # Auto-reload if read-only (safe), otherwise prompt (cooldown 10s)
                if bool(getattr(self.model, 'read_only', False)):
                    # Auto-reload only if enabled in settings
                    if bool(getattr(self, '_sync_auto_reload_readonly', True)):
                        if hasattr(self, '_do_reload') and callable(self._do_reload):
                            self._do_reload()
                        # Update label with outcome
                        try:
                            if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                base = f"Last: {tstr}" if tstr else "Last"
                                self.db_sync_label.setText(base + " — auto")
                        except Exception:
                            pass
                    else:
                        if self.statusBar():
                            self.statusBar().showMessage("Update detected (read-only) – auto-reload disabled", 4000)
                        # Update label with outcome
                        try:
                            if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                base = f"Last: {tstr}" if tstr else "Last"
                                self.db_sync_label.setText(base + " — off")
                        except Exception:
                            pass
                else:
                    if bool(getattr(self, '_sync_prompt_reload_editing', True)):
                        now = time.time()
                        if now - getattr(self, '_db_change_prompt_at', 0.0) >= 10.0:
                            self._db_change_prompt_at = now
                            try:
                                from PyQt6.QtWidgets import QMessageBox
                                resp = QMessageBox.question(self, "Database Updated",
                                    "The database changed on disk (e.g., via OneDrive sync). Reload now?",
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                                if resp == QMessageBox.Yes:
                                    if hasattr(self, '_do_reload') and callable(self._do_reload):
                                        self._do_reload()
                                    # Update label with outcome
                                    try:
                                        if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                            base = f"Last: {tstr}" if tstr else "Last"
                                            self.db_sync_label.setText(base + " — reload")
                                    except Exception:
                                        pass
                                else:
                                    if self.statusBar():
                                        self.statusBar().showMessage("Update detected – use Tools → Reload Data to refresh", 5000)
                                    # Update label with outcome
                                    try:
                                        if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                            base = f"Last: {tstr}" if tstr else "Last"
                                            self.db_sync_label.setText(base + " — later")
                                    except Exception:
                                        pass
                            except Exception:
                                # Fallback: just show a status message
                                if self.statusBar():
                                    self.statusBar().showMessage("Database updated on disk – Reload available", 4000)
                                try:
                                    if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                        base = f"Last: {tstr}" if tstr else "Last"
                                        self.db_sync_label.setText(base + " — seen")
                                except Exception:
                                    pass
                    else:
                        if self.statusBar():
                            self.statusBar().showMessage("Update detected (editing) – prompts disabled", 4000)
                        # Update label with outcome
                        try:
                            if hasattr(self, 'db_sync_label') and self.db_sync_label is not None:
                                base = f"Last: {tstr}" if tstr else "Last"
                                self.db_sync_label.setText(base + " — off")
                        except Exception:
                            pass
        except Exception:
            pass
    def closeEvent(self, event):
        try:
            if getattr(self, '_own_lock', False):
                self._release_edit_lock()
        except Exception:
            pass
        try:
            super().closeEvent(event)
        except Exception:
            pass
    def on_tree_part_selected(self, part_name):
        # No automatic view switching. Optionally, highlight in Gantt if already there.
        if self.sidebar.currentRow() == 1 and hasattr(self.gantt_chart_view, 'highlight_bar'):
            self.gantt_chart_view.highlight_bar(part_name)
    # ---------------- Filters (menu-based) ----------------
    def _init_filter_state(self):
        # Initialize minimal state store for filters used by the menu
        self._filter_state = {
            "statuses": set(),               # e.g., {"Planned", "Done"}
            "ie": set(),                     # {"Internal","External"}
            "responsible_substr": None,      # lowercased substring or None
            "critical_only": False,
            "risk_only": False,
        }
    def _on_filter_status_toggled(self, status, checked):
        if checked:
            self._filter_state["statuses"].add(status)
        else:
            self._filter_state["statuses"].discard(status)
        self._update_filter_summary()
    def _on_filter_ie_toggled(self, which, checked):
        if checked:
            self._filter_state["ie"].add(which)
        else:
            self._filter_state["ie"].discard(which)
        self._update_filter_summary()
    def _on_filter_flag_toggled(self, flag, checked):
        self._filter_state[flag] = bool(checked)
        self._update_filter_summary()
    def _update_filter_summary(self):
        try:
            parts = []
            st = self._filter_state.get("statuses") or set()
            ie = self._filter_state.get("ie") or set()
            resp = self._filter_state.get("responsible_substr") or ""
            if st: parts.append(f"Status={len(st)}")
            if ie: parts.append("IE=" + ",".join(sorted(ie)))
            if resp: parts.append(f"Resp~{resp}")
            if self._filter_state.get("critical_only"): parts.append("Critical")
            if self._filter_state.get("risk_only"): parts.append("Risk")
            text = " | ".join(parts) if parts else "No filters active"
            act = getattr(self, "_filter_actions", {}).get("summary")
            if act:
                act.setText(text)
        except Exception:
            pass
    def _sync_tools_filter_checks_from_state(self):
        try:
            acts = getattr(self, "_filter_actions", {})
            for st, a in acts.get("status", {}).items():
                a.blockSignals(True); a.setChecked(st in self._filter_state["statuses"]); a.blockSignals(False)
            for ie, a in acts.get("ie", {}).items():
                a.blockSignals(True); a.setChecked(ie in self._filter_state["ie"]); a.blockSignals(False)
            for flag, a in acts.get("flags", {}).items():
                a.blockSignals(True); a.setChecked(bool(self._filter_state.get(flag))); a.blockSignals(False)
            self._update_filter_summary()
        except Exception:
            pass
    def _apply_filters(self):
        try:
            statuses = sorted(self._filter_state["statuses"]) or None
            ie = sorted(self._filter_state["ie"]) or None
            resp = self._filter_state.get("responsible_substr") or None
            crit = bool(self._filter_state.get("critical_only"))
            risk = bool(self._filter_state.get("risk_only"))
            if hasattr(self, 'gantt_chart_view'):
                self.gantt_chart_view.set_filters(
                    statuses=statuses,
                    internal_external=ie,
                    responsible_substr=resp,
                    critical_only=crit,
                    risk_only=risk
                )
            # Persist after applying
            self.save_filter_settings()
            self._update_filter_summary()
        except Exception:
            pass
    def _reset_filters(self):
        try:
            self._filter_state["statuses"].clear()
            self._filter_state["ie"].clear()
            self._filter_state["responsible_substr"] = None
            self._filter_state["critical_only"] = False
            self._filter_state["risk_only"] = False
            self._sync_tools_filter_checks_from_state()
            self._apply_filters()
        except Exception:
            pass
    # ---------------- Filter Settings Persistence ----------------
    def load_filter_settings(self):
        from PyQt6.QtCore import QSettings, QTimer
        s = QSettings("LSI", "ProjectPlanner")
        # Statuses
        st_sel = set()
        for st in ["Planned", "In Progress", "Blocked", "Done"]:
            if s.value(f"filters/status/{st}", False, type=bool):
                st_sel.add(st)
        self._filter_state["statuses"] = st_sel
        # IE
        ie_sel = set()
        if s.value("filters/internal", False, type=bool):
            ie_sel.add("Internal")
        if s.value("filters/external", False, type=bool):
            ie_sel.add("External")
        self._filter_state["ie"] = ie_sel
        # Others
        self._filter_state["responsible_substr"] = (s.value("filters/responsible_substr", "", type=str) or "").strip() or None
        self._filter_state["critical_only"] = s.value("filters/critical_only", False, type=bool)
        self._filter_state["risk_only"] = s.value("filters/risk_only", False, type=bool)
        # Apply after UI settles
        QTimer.singleShot(50, lambda: self._apply_filters())
    def save_filter_settings(self):
        from PyQt6.QtCore import QSettings
        s = QSettings("LSI", "ProjectPlanner")
        for st in ["Planned", "In Progress", "Blocked", "Done"]:
            s.setValue(f"filters/status/{st}", st in self._filter_state["statuses"])
        s.setValue("filters/internal", "Internal" in self._filter_state["ie"])
        s.setValue("filters/external", "External" in self._filter_state["ie"])
        s.setValue("filters/responsible_substr", self._filter_state.get("responsible_substr") or "")
        s.setValue("filters/critical_only", bool(self._filter_state.get("critical_only")))
        s.setValue("filters/risk_only", bool(self._filter_state.get("risk_only")))
    def closeEvent(self, event):
        try:
            self.save_filter_settings()
        except Exception:
            pass
        super().closeEvent(event)

    # ---------------- Reports ----------------
    def _draw_report_header(self, painter, printer, title_text: str) -> int:
        """Draw the standard report header (logo) and title, return y-coordinate to continue content.
        Respects Export/include_header (QSettings LSI/ProjectPlanner)."""
        from PyQt6.QtGui import QFont
        from PyQt6.QtCore import QRectF, QSettings
        import os
        y = 0
        # Check whether to include branded header graphic
        include_header = True
        try:
            s = QSettings('LSI', 'ProjectPlanner')
            v = s.value('Export/include_header', True)
            if isinstance(v, str):
                include_header = v.lower() in ('1','true','yes','on')
            else:
                include_header = bool(v)
        except Exception:
            include_header = True
        try:
            if include_header:
                svg_path = resolve_resource_path("header.svg")
                header_svg_renderer = None; header_is_svg = False
                if os.path.exists(svg_path):
                    from PyQt6.QtSvg import QSvgRenderer
                    r = QSvgRenderer(svg_path)
                    if r.isValid():
                        header_is_svg = True; header_svg_renderer = r
                page_rect = printer.pageRect()
                if header_is_svg and header_svg_renderer:
                    ds = header_svg_renderer.defaultSize(); w, h = ds.width(), ds.height()
                    if w <= 0 or h <= 0:
                        vb = header_svg_renderer.viewBoxF(); w, h = vb.width(), vb.height()
                    if w > 0 and h > 0:
                        target_w = page_rect.width()
                        target_h = int(round(h * (target_w / w)))
                        target_rect = QRectF(0, 0, target_w, target_h)
                        header_svg_renderer.render(painter, target_rect)
                        y = target_h + 12
                else:
                    # PNG fallback
                    try:
                        from PyQt6.QtGui import QPixmap
                        p = resolve_resource_path('header.png')
                        if os.path.exists(p):
                            pm = QPixmap(p)
                            if not pm.isNull():
                                scaled = pm.scaledToWidth(printer.pageRect().width())
                                painter.drawPixmap((printer.pageRect().width() - scaled.width())//2, 0, scaled)
                                y = scaled.height() + 12
                    except Exception:
                        pass
        except Exception:
            y = 0
        title_font = QFont(); title_font.setPointSize(14); title_font.setBold(True)
        painter.setFont(title_font)
        painter.drawText(40, y + 10 + 18, title_text)
        return y + 10 + 34

    def _elided_text(self, painter, text: str, max_width: int) -> str:
        from PyQt6.QtCore import Qt
        fm = painter.fontMetrics()
        if max_width <= 0:
            return text
        return fm.elidedText(text or "", Qt.TextElideMode.ElideRight, max_width)

    class MilestoneDigestOptionsDialog(QDialog):
        def __init__(self, parent=None):
            super().__init__(parent)
            from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QSpinBox, QDialogButtonBox
            from PyQt6.QtCore import QSettings
            self.setWindowTitle("Milestone Digest Options")
            s = QSettings("LSI", "ProjectPlanner")
            default_upcoming = int(s.value("reports/milestone_digest/upcoming_days", 60, type=int))
            default_recent = int(s.value("reports/milestone_digest/recent_days", 30, type=int))
            layout = QVBoxLayout(self)
            row1 = QHBoxLayout(); row2 = QHBoxLayout()
            row1.addWidget(QLabel("Upcoming horizon (days):"))
            self.spin_upcoming = QSpinBox(); self.spin_upcoming.setRange(1, 365); self.spin_upcoming.setValue(default_upcoming)
            row1.addWidget(self.spin_upcoming)
            row2.addWidget(QLabel("Recent completed window (days):"))
            self.spin_recent = QSpinBox(); self.spin_recent.setRange(1, 365); self.spin_recent.setValue(default_recent)
            row2.addWidget(self.spin_recent)
            layout.addLayout(row1); layout.addLayout(row2)
            btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            btns.accepted.connect(self.accept)
            btns.rejected.connect(self.reject)
            layout.addWidget(btns)
        def values(self):
            return self.spin_upcoming.value(), self.spin_recent.value()

    def _report_health_snapshot(self):
        """Generate a one-page Project Health Snapshot (PDF) with header and KPIs."""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import datetime
        path, _ = QFileDialog.getSaveFileName(self, "Export Health Snapshot", "health_snapshot.pdf", "PDF Files (*.pdf)")
        if not path:
            return
        # Compute KPIs
        rows = getattr(self.model, 'rows', []) or []
        total = len(rows)
        by_status = {}
        overdue = 0
        at_risk = 0
        done = 0
        avg_pct = 0.0
        today = datetime.datetime.today()
        def parse_date(s):
            s = (s or '').strip()
            for fmt in ("%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        for r in rows:
            st = (r.get('Status') or '').strip() or 'Planned'
            by_status[st] = by_status.get(st, 0) + 1
            try:
                pc = int(r.get('% Complete') or 0)
            except Exception:
                pc = 0
            avg_pct += pc
            if st == 'Done':
                done += 1
            # overdue / at-risk logic similar to Gantt
            try:
                start = parse_date(r.get('Start Date')) or parse_date(r.get('Actual Start Date'))
                end_calc = parse_date(r.get('Calculated End Date')) or parse_date(r.get('Actual Finish Date'))
                dur = None
                try:
                    dur = int(r.get('Duration (days)') or 0)
                except Exception:
                    dur = None
                if start and dur is not None and end_calc is None:
                    end_calc = start + datetime.timedelta(days=dur)
                if pc < 100 and end_calc and today.date() > end_calc.date():
                    overdue += 1
                elif pc == 0 and (st in ('Planned','Blocked')) and start and today.date() > start.date():
                    at_risk += 1
            except Exception:
                pass
        avg_pct = (avg_pct / total) if total else 0.0
        # Build PDF page using QPrinter + QPainter; reuse header render helper from Gantt exporter
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QPainter, QFont
            from PyQt6.QtCore import QSettings, QMarginsF
            # Apply export settings
            s = QSettings('LSI','ProjectPlanner')
            page_size = s.value('Export/page_size','Letter')
            orientation = s.value('Export/orientation','Portrait')
            ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            size_map={'A4':QPrinter.PaperSize.A4,'Letter':QPrinter.PaperSize.Letter,'Legal':QPrinter.PaperSize.Legal,'Tabloid':QPrinter.PaperSize.Tabloid}
            printer.setPaperSize(size_map.get(page_size, QPrinter.PaperSize.Letter))
            printer.setOrientation(QPrinter.Orientation.Portrait if orientation=='Portrait' else QPrinter.Orientation.Landscape)
            try:
                printer.setPageMargins(QMarginsF(ml,mt,mr,mb))
            except Exception:
                pass
            painter = QPainter(printer)
            y = self._draw_report_header(painter, printer, "Project Health Snapshot")
            sub_font = QFont(); sub_font.setPointSize(9)
            painter.setFont(sub_font)
            painter.drawText(40, y, f"As of {today.strftime('%Y-%m-%d')}  |  Items: {total}  |  Avg %: {avg_pct:.1f}%")
            # Status breakdown
            painter.drawText(40, y + 18, "Status distribution:")
            row_y = y + 34
            x0 = 40
            for st in ["Planned","In Progress","Blocked","Done","Deferred"]:
                val = by_status.get(st, 0)
                painter.drawText(x0, row_y, f"{st}: {val}")
                row_y += 16
            # Risk
            row_y += 8
            painter.drawText(40, row_y, f"Overdue: {overdue}    At-Risk: {at_risk}")
            # Completion
            row_y += 18
            painter.drawText(40, row_y, f"Completed: {done}    Remaining: {max(0, total - done)}")
            painter.end()
            if self.statusBar():
                self.statusBar().showMessage(f"Exported: {path}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write PDF: {e}")

    def _report_baseline_variance_csv(self):
        """Export baseline vs current schedule variance as CSV."""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import csv, datetime
        path, _ = QFileDialog.getSaveFileName(self, "Export Baseline Variance (CSV)", "baseline_variance.csv", "CSV Files (*.csv)")
        if not path:
            return
        # Build baseline map: choose selected baseline from Gantt view if any
        baseline_name = None
        try:
            if hasattr(self, 'gantt_chart_view'):
                baseline_name = getattr(self.gantt_chart_view, '_selected_baseline_name', None)
        except Exception:
            baseline_name = None
        bmap = {}
        try:
            if baseline_name:
                bmap = self.model.load_baseline_map(baseline_name)
        except Exception:
            bmap = {}
        rows = getattr(self.model, 'rows', []) or []
        def to_dt(s):
            s = (s or '').strip()
            for fmt in ("%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(["Project Part","Baseline Start","Baseline End","Current Start","Current End","Start Slip (d)","Finish Slip (d)"])
                for r in rows:
                    name = r.get('Project Part','')
                    b_start = b_end = None
                    if name in bmap:
                        bs, be = bmap[name]
                        b_start = to_dt(bs) if bs else None
                        b_end = to_dt(be) if be else None
                    c_start = to_dt(r.get('Start Date'))
                    c_end = to_dt(r.get('Calculated End Date'))
                    if not c_end and c_start:
                        try:
                            d = int(r.get('Duration (days)') or 0)
                        except Exception:
                            d = 0
                        c_end = c_start + datetime.timedelta(days=d)
                    slip_s = (c_start - b_start).days if (b_start and c_start) else ''
                    slip_e = (c_end - b_end).days if (b_end and c_end) else ''
                    def fmt(d):
                        return d.strftime('%m-%d-%Y') if d else ''
                    w.writerow([name, fmt(b_start), fmt(b_end), fmt(c_start), fmt(c_end), slip_s, slip_e])
            if self.statusBar():
                self.statusBar().showMessage(f"Exported: {path}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write CSV: {e}")

    def _report_baseline_variance_pdf(self):
        """Export baseline vs current schedule variance as a simple one-page PDF table (top N)."""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        import datetime
        path, _ = QFileDialog.getSaveFileName(self, "Export Baseline Variance (PDF)", "baseline_variance.pdf", "PDF Files (*.pdf)")
        if not path:
            return
        # Load baseline as above
        baseline_name = None
        try:
            if hasattr(self, 'gantt_chart_view'):
                baseline_name = getattr(self.gantt_chart_view, '_selected_baseline_name', None)
        except Exception:
            baseline_name = None
        bmap = {}
        try:
            if baseline_name:
                bmap = self.model.load_baseline_map(baseline_name)
        except Exception:
            bmap = {}
        rows = getattr(self.model, 'rows', []) or []
        def to_dt(s):
            s = (s or '').strip()
            for fmt in ("%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        # Prepare data with slips; take top 30 by absolute finish slip
        data = []
        for r in rows:
            name = r.get('Project Part','')
            bs = be = None
            if name in bmap:
                bs_str, be_str = bmap[name]
                bs = to_dt(bs_str) if bs_str else None
                be = to_dt(be_str) if be_str else None
            cs = to_dt(r.get('Start Date'))
            ce = to_dt(r.get('Calculated End Date'))
            if not ce and cs:
                try:
                    d = int(r.get('Duration (days)') or 0)
                except Exception:
                    d = 0
                ce = cs + datetime.timedelta(days=d)
            slip_s = (cs - bs).days if (bs and cs) else None
            slip_e = (ce - be).days if (be and ce) else None
            data.append((name, bs, be, cs, ce, slip_s, slip_e))
        data.sort(key=lambda t: abs(t[6]) if isinstance(t[6], int) else 0, reverse=True)
        data = data[:30]
        # Render to PDF
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QPainter, QFont
            from PyQt6.QtCore import QSettings, QMarginsF
            s = QSettings('LSI','ProjectPlanner')
            page_size = s.value('Export/page_size','Letter')
            orientation = s.value('Export/orientation','Portrait')
            ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            size_map={'A4':QPrinter.PaperSize.A4,'Letter':QPrinter.PaperSize.Letter,'Legal':QPrinter.PaperSize.Legal,'Tabloid':QPrinter.PaperSize.Tabloid}
            printer.setPaperSize(size_map.get(page_size, QPrinter.PaperSize.Letter))
            printer.setOrientation(QPrinter.Orientation.Portrait if orientation=='Portrait' else QPrinter.Orientation.Landscape)
            try:
                printer.setPageMargins(QMarginsF(ml,mt,mr,mb))
            except Exception:
                pass
            painter = QPainter(printer)
            y = self._draw_report_header(painter, printer, "Baseline Variance")
            body_font = QFont(); body_font.setPointSize(9)
            painter.setFont(body_font)
            y0 = y
            # Table headers
            headers = ["Project Part","Base Start","Base End","Cur Start","Cur End","Slip S (d)","Slip E (d)"]
            cols_x = [40, 220, 310, 400, 490, 580, 650]
            painter.drawText(cols_x[0], y0, headers[0])
            for i in range(1, len(headers)):
                painter.drawText(cols_x[i], y0, headers[i])
            y_cur = y0 + 16
            def fmt(d):
                return d.strftime('%m-%d-%Y') if d else ''
            for name, bs, be, cs, ce, ss, se in data:
                # elide long text to fit column widths
                width0 = cols_x[1] - cols_x[0] - 6
                painter.drawText(cols_x[0], y_cur, self._elided_text(painter, name, width0))
                painter.drawText(cols_x[1], y_cur, fmt(bs))
                painter.drawText(cols_x[2], y_cur, fmt(be))
                painter.drawText(cols_x[3], y_cur, fmt(cs))
                painter.drawText(cols_x[4], y_cur, fmt(ce))
                painter.drawText(cols_x[5], y_cur, '' if ss is None else str(ss))
                painter.drawText(cols_x[6], y_cur, '' if se is None else str(se))
                y_cur += 14
            painter.end()
            if self.statusBar():
                self.statusBar().showMessage(f"Exported: {path}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write PDF: {e}")

    def _report_milestone_digest_pdf(self):
        """Generate a Milestone Digest (PDF) with sections for Upcoming and Recently Completed milestones."""
        from PyQt6.QtWidgets import QFileDialog, QMessageBox
        from PyQt6.QtGui import QFont, QColor
        import datetime
        path, _ = QFileDialog.getSaveFileName(self, "Export Milestone Digest (PDF)", "milestone_digest.pdf", "PDF Files (*.pdf)")
        if not path:
            return
        # Options dialog
        dlg = self.MilestoneDigestOptionsDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        horizon_upcoming, horizon_completed = dlg.values()
        # persist
        try:
            from PyQt6.QtCore import QSettings
            s = QSettings("LSI", "ProjectPlanner")
            s.setValue("reports/milestone_digest/upcoming_days", horizon_upcoming)
            s.setValue("reports/milestone_digest/recent_days", horizon_completed)
        except Exception:
            pass
        rows = getattr(self.model, 'rows', []) or []
        today = datetime.datetime.today()
        def to_dt(s):
            s = (s or '').strip()
            for fmt in ("%m-%d-%Y","%m/%d/%Y","%Y-%m-%d","%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        upcoming = []  # (date, name, resp, status, notes)
        recent_done = []  # (date_done, name, resp, status, notes)
        for r in rows:
            name = (r.get('Project Part') or '').strip()
            status = (r.get('Status') or '').strip() or 'Planned'
            notes = (r.get('Notes') or '').strip()
            resp = (r.get('Responsible') or '').strip()
            # Determine if milestone
            is_ms = False
            try:
                typ = (r.get('Type') or '').strip().lower()
                if typ == 'milestone':
                    is_ms = True
            except Exception:
                pass
            if not is_ms:
                try:
                    d = r.get('Duration (days)')
                    d = int(d) if d not in (None, '') else None
                    if d == 0:
                        is_ms = True
                except Exception:
                    pass
            if not is_ms:
                continue
            # Dates
            start = to_dt(r.get('Start Date'))
            end_calc = to_dt(r.get('Calculated End Date'))
            act_finish = to_dt(r.get('Actual Finish Date'))
            use_date = end_calc or start or act_finish
            try:
                pc = int(r.get('% Complete') or 0)
            except Exception:
                pc = 0
            # Upcoming: within next horizon_upcoming days, not done
            if use_date and pc < 100 and status != 'Done':
                if today.date() <= use_date.date() <= (today + datetime.timedelta(days=horizon_upcoming)).date():
                    upcoming.append((use_date, name, resp, status, notes))
            # Recently completed: finished within last horizon_completed days
            done_date = act_finish or end_calc or start
            if (pc >= 100 or status == 'Done') and done_date:
                if (today - datetime.timedelta(days=horizon_completed)).date() <= done_date.date() <= today.date():
                    recent_done.append((done_date, name, resp, status, notes))
        # Sort
        upcoming.sort(key=lambda t: t[0])
        recent_done.sort(key=lambda t: t[0], reverse=True)
        # Render
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QPainter
            from PyQt6.QtCore import QSettings, QMarginsF
            s = QSettings('LSI','ProjectPlanner')
            page_size = s.value('Export/page_size','Letter')
            orientation = s.value('Export/orientation','Portrait')
            ml = float(s.value('Export/margin_left_mm',8.0)); mt = float(s.value('Export/margin_top_mm',8.0)); mr = float(s.value('Export/margin_right_mm',8.0)); mb = float(s.value('Export/margin_bottom_mm',8.0))
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFileName(path)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            size_map={'A4':QPrinter.PaperSize.A4,'Letter':QPrinter.PaperSize.Letter,'Legal':QPrinter.PaperSize.Legal,'Tabloid':QPrinter.PaperSize.Tabloid}
            printer.setPaperSize(size_map.get(page_size, QPrinter.PaperSize.Letter))
            printer.setOrientation(QPrinter.Orientation.Portrait if orientation=='Portrait' else QPrinter.Orientation.Landscape)
            try:
                printer.setPageMargins(QMarginsF(ml,mt,mr,mb))
            except Exception:
                pass
            painter = QPainter(printer)
            page_rect = printer.pageRect()
            y = self._draw_report_header(painter, printer, "Milestone Digest")
            body_font = QFont(); body_font.setPointSize(9)
            painter.setFont(body_font)

            def new_page_with_section(section_title: str) -> int:
                printer.newPage()
                y0 = self._draw_report_header(painter, printer, "Milestone Digest")
                sub_font = QFont(); sub_font.setPointSize(11); sub_font.setBold(True)
                painter.setFont(sub_font)
                painter.drawText(40, y0, section_title)
                painter.setFont(body_font)
                return y0 + 18

            # Upcoming section
            sub_font = QFont(); sub_font.setPointSize(11); sub_font.setBold(True)
            painter.setFont(sub_font)
            painter.drawText(40, y, f"Upcoming Milestones (next {horizon_upcoming} days)")
            painter.setFont(body_font)
            y += 18
            cols_x = [40, 135, 360, 520, 650]  # Date, Milestone, Responsible, Status, Notes
            headers = ["Date", "Milestone", "Responsible", "Status", "Notes"]
            for i, htxt in enumerate(headers):
                painter.drawText(cols_x[i], y, htxt)
            y += 14
            def fmt(d):
                return d.strftime('%m-%d-%Y') if d else ''
            for dt, name, resp, status, notes in upcoming:
                if y > page_rect.height() - 60:
                    y = new_page_with_section("Upcoming Milestones (cont.)")
                    for i, htxt in enumerate(headers):
                        painter.drawText(cols_x[i], y, htxt)
                    y += 14
                # overdue highlighting
                is_overdue = dt and (dt.date() < today.date())
                if is_overdue:
                    painter.setPen(QColor(200, 0, 0))
                painter.drawText(cols_x[0], y, fmt(dt) + ("  (overdue)" if is_overdue else ""))
                if is_overdue:
                    painter.setPen(QColor(0, 0, 0))
                # elide columns to fit widths
                w1 = cols_x[2] - cols_x[1] - 6
                w2 = cols_x[3] - cols_x[2] - 6
                w3 = cols_x[4] - cols_x[3] - 6
                w4 = page_rect.width() - cols_x[4] - 20
                painter.drawText(cols_x[1], y, self._elided_text(painter, name, w1))
                painter.drawText(cols_x[2], y, self._elided_text(painter, resp, w2))
                painter.drawText(cols_x[3], y, self._elided_text(painter, status, w3))
                note_txt = (notes or '').replace('\n', ' ').strip()
                painter.drawText(cols_x[4], y, self._elided_text(painter, note_txt, w4))
                y += 14

            # Recently Completed section
            if y > page_rect.height() - 120:
                y = new_page_with_section("Recently Completed Milestones (last 30 days)")
            else:
                sub_font = QFont(); sub_font.setPointSize(11); sub_font.setBold(True)
                painter.setFont(sub_font)
                painter.drawText(40, y + 12, f"Recently Completed Milestones (last {horizon_completed} days)")
                painter.setFont(body_font)
                y += 12 + 18
            # headers reuse
            for i, htxt in enumerate(headers):
                painter.drawText(cols_x[i], y, htxt)
            y += 14
            for dt, name, resp, status, notes in recent_done:
                if y > page_rect.height() - 60:
                    y = new_page_with_section("Recently Completed (cont.)")
                    for i, htxt in enumerate(headers):
                        painter.drawText(cols_x[i], y, htxt)
                    y += 14
                painter.drawText(cols_x[0], y, fmt(dt))
                w1 = cols_x[2] - cols_x[1] - 6
                w2 = cols_x[3] - cols_x[2] - 6
                w3 = cols_x[4] - cols_x[3] - 6
                w4 = page_rect.width() - cols_x[4] - 20
                painter.drawText(cols_x[1], y, self._elided_text(painter, name, w1))
                painter.drawText(cols_x[2], y, self._elided_text(painter, resp, w2))
                painter.drawText(cols_x[3], y, self._elided_text(painter, status, w3))
                note_txt = (notes or '').replace('\n', ' ').strip()
                painter.drawText(cols_x[4], y, self._elided_text(painter, note_txt, w4))
                y += 14

            painter.end()
            if self.statusBar():
                self.statusBar().showMessage(f"Exported: {path}", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write PDF: {e}")

    # --- Dynamic header resize to fit window and eliminate cushion ---
    def _resize_header(self):
        try:
            if not getattr(self, '_header_widget', None):
                return
            # Allow the logo to use nearly full window width
            max_w = max(1, int(self.width() * 0.98))
            # Make the logo taller without distorting: use a larger fraction of window height (cap kept reasonable)
            target_h = max(128, min(int(self.height() * 0.22), 360))
            try:
                print(f"[UI] _resize_header: win=({self.width()}x{self.height()}) max_w={max_w} target_h={target_h} svg={self._header_is_svg}")
            except Exception:
                pass
            # Helper: trim transparent borders
            def trim_transparent(pixmap: QPixmap) -> QPixmap:
                try:
                    if pixmap.isNull():
                        return pixmap
                    from PyQt6.QtGui import QImage
                    img = pixmap.toImage().convertToFormat(QImage.Format_ARGB32_Premultiplied)
                    w, h = img.width(), img.height()
                    left, right, top, bottom = 0, w - 1, 0, h - 1
                    # scan top
                    found = False
                    for y in range(h):
                        for x in range(w):
                            if img.pixelColor(x, y).alpha() > 0:
                                top = y; found = True; break
                        if found: break
                    # scan bottom
                    found = False
                    for y in range(h - 1, -1, -1):
                        for x in range(w):
                            if img.pixelColor(x, y).alpha() > 0:
                                bottom = y; found = True; break
                        if found: break
                    # scan left
                    found = False
                    for x in range(w):
                        for y in range(h):
                            if img.pixelColor(x, y).alpha() > 0:
                                left = x; found = True; break
                        if found: break
                    # scan right
                    found = False
                    for x in range(w - 1, -1, -1):
                        for y in range(h):
                            if img.pixelColor(x, y).alpha() > 0:
                                right = x; found = True; break
                        if found: break
                    if right >= left and bottom >= top:
                        from PyQt6.QtCore import QRect
                        cropped = img.copy(QRect(left, top, right - left + 1, bottom - top + 1))
                        pm2 = QPixmap.fromImage(cropped)
                        return pm2
                except Exception:
                    pass
                return pixmap
            # Helper: trim uniform color margins (e.g., solid white padding) with tolerance, without touching content
            def trim_uniform_color(pixmap: QPixmap, tol: int = 10) -> QPixmap:
                try:
                    if pixmap.isNull():
                        return pixmap
                    from PyQt6.QtGui import QImage
                    img = pixmap.toImage().convertToFormat(QImage.Format_ARGB32_Premultiplied)
                    w, h = img.width(), img.height()
                    if w <= 2 or h <= 2:
                        return pixmap
                    # Determine background color from corners (require near-equality across corners)
                    c00 = img.pixelColor(0, 0)
                    c10 = img.pixelColor(w - 1, 0)
                    c01 = img.pixelColor(0, h - 1)
                    c11 = img.pixelColor(w - 1, h - 1)
                    cols = [c00, c10, c01, c11]
                    def close(a, b):
                        return abs(a.red() - b.red()) <= tol and abs(a.green() - b.green()) <= tol and abs(a.blue() - b.blue()) <= tol and abs(a.alpha() - b.alpha()) <= tol
                    if not (close(c00, c10) and close(c00, c01) and close(c00, c11)):
                        return pixmap  # not a uniform border color
                    bg = c00
                    def is_bg(col):
                        return abs(col.red() - bg.red()) <= tol and abs(col.green() - bg.green()) <= tol and abs(col.blue() - bg.blue()) <= tol and abs(col.alpha() - bg.alpha()) <= tol
                    # Scan top
                    top = 0
                    for y in range(h):
                        if any(not is_bg(img.pixelColor(x, y)) for x in range(w)):
                            top = y
                            break
                    # Scan bottom
                    bottom = h - 1
                    for y in range(h - 1, -1, -1):
                        if any(not is_bg(img.pixelColor(x, y)) for x in range(w)):
                            bottom = y
                            break
                    # Scan left
                    left = 0
                    for x in range(w):
                        if any(not is_bg(img.pixelColor(x, y)) for y in range(h)):
                            left = x
                            break
                    # Scan right
                    right = w - 1
                    for x in range(w - 1, -1, -1):
                        if any(not is_bg(img.pixelColor(x, y)) for y in range(h)):
                            right = x
                            break
                    # Keep a tiny guard margin of 1px to avoid clipping anti-aliased edges
                    left = max(0, left - 1)
                    top = max(0, top - 1)
                    right = min(w - 1, right + 1)
                    bottom = min(h - 1, bottom + 1)
                    if right > left and bottom > top:
                        from PyQt6.QtCore import QRect
                        cropped = img.copy(QRect(left, top, right - left + 1, bottom - top + 1))
                        return QPixmap.fromImage(cropped)
                except Exception:
                    return pixmap
                return pixmap
            if self._header_is_svg:
                # Render SVG to a pixmap at target height, trim transparent, then optionally crop percent
                try:
                    from PyQt6.QtGui import QPainter
                    r = self._header_svg_renderer
                    ds = r.defaultSize(); w, h = ds.width(), ds.height()
                    if w <= 0 or h <= 0:
                        vb = r.viewBoxF(); w, h = vb.width(), vb.height()
                    if w <= 0 or h <= 0:
                        w, h = 800, 200
                    scale = target_h / float(h)
                    render_w = max(1, int(round(w * scale)))
                    render_h = max(1, int(round(h * scale)))
                    pm = QPixmap(render_w, render_h)
                    # PyQt6: use QColor(0,0,0,0) for transparent fill
                    from PyQt6.QtGui import QColor as _QCol
                    pm.fill(_QCol(0, 0, 0, 0))
                    p = QPainter(pm)
                    try:
                        from PyQt6.QtCore import QRectF
                    except Exception:
                        QRectF = None  # type: ignore
                    if QRectF is not None:
                        r.render(p, QRectF(0, 0, float(render_w), float(render_h)))
                    else:
                        # Fallback if QRectF import fails
                        r.render(p)
                    p.end()
                    # Trim transparent borders only; avoid uniform-color trim to prevent over-cropping
                    pm_trim = trim_transparent(pm)
                    try:
                        print(f"[UI] header render: base=({w}x{h}) scaled=({render_w}x{render_h}) after-trim=({pm_trim.width()}x{pm_trim.height()})")
                    except Exception:
                        pass
                    # Fallback to PNG if SVG rasterization produced an unexpectedly tiny/empty image
                    if pm_trim.isNull() or pm_trim.width() < 8 or pm_trim.height() < 8:
                        try:
                            png_path = resolve_resource_path("header.png")
                            from PyQt6.QtGui import QPixmap as _QPM
                            if png_path and os.path.exists(png_path):
                                alt = _QPM(png_path)
                                if not alt.isNull():
                                    try:
                                        _smooth = Qt.TransformationMode.SmoothTransformation
                                    except Exception:
                                        _smooth = getattr(Qt, 'SmoothTransformation', 1)
                                    pm_trim = alt.scaledToHeight(target_h, _smooth)
                                    print(f"[UI] header fallback to PNG -> {png_path}")
                        except Exception as _e_fallback:
                            try:
                                print(f"[UI] header PNG fallback error: {_e_fallback}")
                            except Exception:
                                pass
                    # Enforce width cap
                    if pm_trim.width() > max_w:
                        try:
                            _smooth = Qt.TransformationMode.SmoothTransformation
                        except Exception:
                            _smooth = getattr(Qt, 'SmoothTransformation', 1)
                        pm_trim = pm_trim.scaled(max_w, target_h, _keep_ar(), _smooth)
                    if getattr(self, '_header_label', None):
                        try:
                            from PyQt6.QtWidgets import QSizePolicy
                            self._header_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                            self._header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        except Exception:
                            pass
                        self._header_label.setPixmap(pm_trim)
                        self._header_label.setFixedHeight(target_h)
                        try:
                            print(f"[UI] label set: pix=({pm_trim.width()}x{pm_trim.height()}) fixed_h={target_h}")
                        except Exception:
                            pass
                except Exception as e:
                    try:
                        print(f"[UI] header render error: {e}")
                    except Exception:
                        pass
                    # Fallback: just set height of widget
                    self._header_widget.setFixedHeight(target_h)
            else:
                # PNG fallback or placeholder
                try:
                    if getattr(self, '_header_png_pixmap', None) is not None and getattr(self, '_header_label', None) is not None:
                        # Scale PNG to target height while keeping aspect ratio, cap width
                        try:
                            _smooth = Qt.TransformationMode.SmoothTransformation
                        except Exception:
                            _smooth = getattr(Qt, 'SmoothTransformation', 1)
                        pm = self._header_png_pixmap
                        scaled = pm.scaledToHeight(target_h, _smooth)
                        if scaled.width() > max_w:
                            scaled = scaled.scaled(max_w, target_h, _keep_ar(), _smooth)
                        try:
                            from PyQt6.QtWidgets import QSizePolicy
                            self._header_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                            self._header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        except Exception:
                            pass
                        self._header_label.setPixmap(scaled)
                        self._header_label.setFixedHeight(target_h)
                        try:
                            print(f"[UI] header PNG set: pix=({scaled.width()}x{scaled.height()}) fixed_h={target_h}")
                        except Exception:
                            pass
                    else:
                        # Placeholder label: just update height for consistent spacing
                        self._header_widget.setFixedHeight(target_h)
                except Exception:
                    pass
        except Exception:
            pass

    def resizeEvent(self, event):
        try:
            self._resize_header()
        except Exception:
            pass
        return super().resizeEvent(event)

# ------------------------------------------------------------
# Application Entry Point (was missing; caused immediate exit)
# ------------------------------------------------------------
if __name__ == "__main__":
    try:
        import sys
        from PyQt6.QtWidgets import QApplication
        app = QApplication(sys.argv)
        model = ProjectDataModel()
        window = MainWindow(model)
        window.show()
        exit_code = app.exec()
        sys.exit(exit_code)
    except Exception as e:
        import traceback, sys
        print("FATAL: Unhandled exception during startup:", e)
        traceback.print_exc()
        sys.exit(1)